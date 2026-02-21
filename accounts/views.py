import traceback
from urllib import request
from django.contrib import messages
from django.shortcuts import get_object_or_404, render, redirect
from copy import deepcopy
from shutil import copy2
from pathlib import Path
import shutil
from calendar import monthrange
from django.urls import reverse
from django.utils.dateformat import DateFormat
from django.utils.formats import get_format
from django.db import connections
from django.db.models import Count, Sum, Case, When, IntegerField, Q
from django.db.models import Max
from .models import AccountBank, Client, DataEntry, PdfConvertFailure, UserData, TaxSuspension, IncomeTaxReturn ,MailLog
from django.template.loader import render_to_string

 # ✅ Import your custom model      
import os
from django.http import FileResponse
from urllib.parse import quote
from .utills import get_fiscal_year_from_date


from django.views.decorators.http import require_GET

from .forms import ClientForm,AccountForm, GroupForm,DataEntryForm,UserForm,TaxSuspensionForm ,IncomeTaxReturnForm, fy_to_start_date
from collections import defaultdict
ALLOWED_ACCOUNT_GROUPS = ["Bank Accounts", "Credit Card","Bank O/D Account"]

from accounts import models

# Create your views here.
def home(request):
    show_welcome = request.session.pop('show_welcome', False)
    return render(request, 'home.html', {'show_welcome': show_welcome})

# def group_form(request):
#     popup = request.GET.get("popup")

#     # Determine if we're in modify mode
#     is_modify_mode = 'mode=modify' in request.get_full_path() or '/group/modify/' in request.path

#     # Get selected group ID from GET or POST
#     selected_group_id = request.GET.get('group_id') or request.POST.get('group_id')

#     # Initialize instance to None
#     instance = None
#     if is_modify_mode and selected_group_id:
#         instance = get_object_or_404(Group, id=selected_group_id)

#     if request.method == 'POST':
#         try:
#             with transaction.atomic():
#                 form = GroupForm(request.POST, instance=instance)
#                 if form.is_valid():
#                     form.save()
#                     messages.success(
#                         request,
#                         f"Group {'updated' if is_modify_mode else 'added'} successfully"
#                     )
#                     # return redirect('group_form')
#                     if request.headers.get("x-requested-with") == "XMLHttpRequest":
#                         return JsonResponse({"success": True})
#                     return redirect("group_form")

#                 else:
#                     messages.error(request, "Please correct the errors below")
#         except Exception as e:
#             messages.error(request, f"An error occurred: {str(e)}")
#     else:
#         form = GroupForm(instance=instance)

#     groups = Group.objects.all() if is_modify_mode else None
#     return render(request, 'group_form.html', {
#         'form': form,
#         'is_modify_mode': is_modify_mode,
#         'groups': groups,
#         'title': 'Modify Group' if is_modify_mode else 'Add Group',
#         'selected_group_id': int(selected_group_id) if selected_group_id else None,
#     }) 

# def group_form(request):
#     popup = request.GET.get("popup")

#     is_modify_mode = 'mode=modify' in request.get_full_path() or '/group/modify/' in request.path
#     selected_group_id = request.GET.get('group_id') or request.POST.get('group_id')

#     instance = None
#     if is_modify_mode and selected_group_id:
#         instance = get_object_or_404(Group, id=selected_group_id)

#     if request.method == 'POST':
#         print("GROUP POST:", request.POST)  # DEBUG

#         form = GroupForm(request.POST, instance=instance)

#         if form.is_valid():
#             group = form.save()

#             # ✅ JSON if modal
#             if popup or request.headers.get("x-requested-with") == "XMLHttpRequest":
#                 return JsonResponse({
#                     "success": True,
#                     "id": group.id,
#                     "name": group.group_name
#                 })

#             messages.success(request, "Group saved")
#             return redirect("group_form")

#         # ❌ invalid
#         if popup or request.headers.get("x-requested-with") == "XMLHttpRequest":
#             html = render_to_string("group_form.html", {
#                 "form": form,
#                 "is_modify_mode": is_modify_mode
#             }, request=request)

#             return JsonResponse({"success": False, "html": html})

#     else:
#         form = GroupForm(instance=instance)

#     return render(request, "group_form.html", {
#         "form": form,
#         "is_modify_mode": is_modify_mode
#     })


# def group_form(request):
#     popup = request.GET.get("popup")

#     is_modify_mode = 'mode=modify' in request.get_full_path() or '/group/modify/' in request.path
#     selected_group_id = request.GET.get('group_id') or request.POST.get('group_id')

#     instance = None
#     if is_modify_mode and selected_group_id:
#         instance = get_object_or_404(Group, id=selected_group_id)

#     if request.method == 'POST':
#         print("GROUP POST:", request.POST)

#         form = GroupForm(request.POST, instance=instance)

#         if form.is_valid():
#             group = form.save()

#             if popup or request.headers.get("x-requested-with") == "XMLHttpRequest":
#                 return JsonResponse({
#                     "success": True,
#                     "id": group.id,
#                     "name": group.group_name
#                 })

#             messages.success(request, "Group saved")
#             # return redirect("group_form")
#             if is_modify_mode:
#                 return redirect(f"{reverse('group_form')}?mode=modify&group_id={group.id}")
#             else:
#                 return redirect("group_form")


#     else:
#         form = GroupForm(instance=instance)

#     # ✅ LOAD ALL GROUPS FOR DROPDOWN
#     groups = Group.objects.all().order_by("group_name")

#     return render(request, "group_form.html", {
#         "form": form,
#         "is_modify_mode": is_modify_mode,
#         "groups": groups,                      # ✅ REQUIRED
#         "selected_group_id": int(selected_group_id) if selected_group_id else None,
#         "cancel_url": "/",  
#         "popup": popup,                   # ✅ for cancel button safety
#     })

def group_form(request):
    popup = request.GET.get("popup")

    # ✅ keep modify mode independent
    is_modify_mode = request.GET.get("mode") == "modify" or request.POST.get("mode") == "modify"

    # ✅ detect selected group
    if request.method == "POST":
        gid = request.POST.get("group_id")
    else:
        gid = request.GET.get("group_id")

    instance = None
    if gid:
        instance = Group.objects.filter(id=gid).first()

    selected_group_id = gid

    if request.method == "POST":
        print("GROUP POST:", request.POST)
        print("USING INSTANCE:", instance)

        form = GroupForm(request.POST, instance=instance)

        if form.is_valid():
            group = form.save()

            if popup or request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({
                    "success": True,
                    "id": group.id,
                    "name": group.group_name
                })

            messages.success(request, "Group updated" if instance else "Group created")

            return redirect(
                f"{reverse('group_form')}?mode=modify&group_id={group.id}"
                if is_modify_mode else "group_form"
            )

    else:
        form = GroupForm(instance=instance)

    groups = Group.objects.all().order_by("group_name")

    return render(request, "group_form.html", {
        "form": form,
        "is_modify_mode": is_modify_mode,
        "groups": groups,
        "selected_group_id": selected_group_id,
        "popup": popup,
        "cancel_url": "/",
    })




#############
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Client,Group
from .forms import ClientForm
import json
from django.db import transaction

def client_form(request):

    # ✅ HARD BLOCK if no groups exist
    if not Group.objects.exists():
        messages.error(request, "No Group exists. Create Group first before adding client.")
        return redirect('group_form')   # change if URL name is different
    
    # Determine if we're in modify mode and get client ID
    is_modify_mode = 'mode=modify' in request.get_full_path() or '/client/modify/' in request.path
    selected_client_id = request.GET.get('client_id') or request.POST.get('client_id')

    if request.method == 'POST':
        try:
            with transaction.atomic():
                instance = None
                # Get existing client in modify mode
                if is_modify_mode and selected_client_id:
                    try:
                        instance = Client.objects.get(client_id=selected_client_id)
                    except Client.DoesNotExist:
                        messages.error(request, "Selected client not found")
                        return redirect(request.path)

                # Process the form
                form = ClientForm(request.POST, instance=instance)
                if form.is_valid():
                    client = form.save(commit=False)
                    
                    # Save all GST data regardless of checkbox state
                    client.trade_name = form.cleaned_data.get('trade_name')
                    client.gst_scheme = form.cleaned_data.get('gst_scheme')
                    client.period = form.cleaned_data.get('period')
                    client.reg_date = form.cleaned_data.get('reg_date')
                    client.cancel_date = form.cleaned_data.get('cancel_date')
                    client.status = form.cleaned_data.get('status')
                    client.user_id = form.cleaned_data.get('user_id')
                    client.password = form.cleaned_data.get('password')
                    client.gst_data = form.cleaned_data.get('gst_data')
                    client.sale_define = form.cleaned_data.get('sale_define')
                    client.bank = form.cleaned_data.get('bank')
                    
                    # Save the Yes/No values for return types
                    client.it_return = form.cleaned_data.get('it_return', 'No')
                    client.gst_return = form.cleaned_data.get('gst_return', 'No')
                    client.tds_return = form.cleaned_data.get('tds_return', 'No')
                    client.tcs_return = form.cleaned_data.get('tcs_return', 'No')
                    client.it_alloted_to = form.cleaned_data.get('it_alloted_to')
                    client.audit_status = form.cleaned_data.get('audit_status')
                    client.it_start_date = form.cleaned_data.get('it_start_date')

                    
                    client.save()
                    form.save_m2m()
                    
                    messages.success(request, f"Client {'updated' if is_modify_mode else 'added'} successfully")
                    return redirect('client_form')
                else:
                    error_messages = []
                    for field, errors in form.errors.items():
                        for error in errors:
                            # Human-readable field label
                            field_label = form.fields[field].label if field in form.fields else field
                            error_messages.append(f"{field_label}: {error}")

                    # ✅ Join all errors into one message
                    messages.error(request, "❌ Form has errors:\n" + "\n".join(error_messages))
                  
        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")

    else:  # GET request
        instance = None
        if is_modify_mode and selected_client_id:
            try:
                instance = Client.objects.get(client_id=selected_client_id)
                # Initialize form with instance data
                form = ClientForm(instance=instance)
            except Client.DoesNotExist:
                messages.error(request, "Selected client not found")
                return redirect(request.path)
        else:
            form = ClientForm()
    
    # Prepare client list for dropdown (modify mode only)
    clients = Client.objects.all() if is_modify_mode else None
    # ✅ NEW: where to go when clicking Cancel
    cancel_url = _safe_back_url(request, default_name="home")
    group_count = Group.objects.count()

    return render(request, 'client_form.html', {
        'form': form,
        'is_modify_mode': is_modify_mode,
        'clients': clients,
        'title': 'Modify Client' if is_modify_mode else 'Add Client',
        'selected_client_id': selected_client_id,
        'cancel_url': cancel_url, 
        'group_count': group_count 
    })

def client_list(request):
    clients = Client.objects.all().order_by('client_name')
    return render(request, 'client_list.html', {'clients': clients})

def delete_client(request, client_id):
    if request.method == 'GET':
        client = get_object_or_404(Client.objects.using('default'), id=client_id)

        fiscal_dbs = [db for db in settings.DATABASES if db != 'default']

        # Step 1: Delete DataEntry & IncomeTaxReturn from fiscal DBs
        for db in fiscal_dbs:
            try:
                DataEntry.objects.using(db).filter(client_id=client.id).delete()
            except Exception as e:
                print(f"[{db}] Skipping DataEntry delete: {e}")
            try:
                IncomeTaxReturn.objects.using(db).filter(client_id=client.id).delete()
            except Exception as e:
                print(f"[{db}] Skipping IncomeTaxReturn delete: {e}")

        # Step 2: Delete related rows in default DB
        try:
            with connections['default'].cursor() as cursor:
                # AccountBank uses client.client_id (string), so match that!
                cursor.execute("DELETE FROM accounts_accountbank WHERE client_id = %s", [client.client_id])

                # TaxSuspension uses FK to Client.id
                cursor.execute("DELETE FROM accounts_taxsuspension WHERE client_id = %s", [client.id])

                # Finally delete the client itself
                cursor.execute("DELETE FROM accounts_client WHERE id = %s", [client.id])
        except Exception as e:
            print(f"❌ Failed to delete client {client_id}: {e}")

        return redirect('client_list')

# bulk client  Delete
from django.views.decorators.csrf import csrf_exempt
@csrf_exempt
def bulk_delete_clients(request):
    if request.method == "POST":
        data = json.loads(request.body)
        ids = data.get("ids", [])

        Client.objects.filter(id__in=ids).delete()

        return JsonResponse({"success": True})

    return JsonResponse({"error": "Invalid request"}, status=400)

    
from .models import AccountBank, DataEntry
from django.db import connections\

def delete_account(request, account_id):
    if request.method == 'GET':
        # Step 1: Get the AccountBank record (from default DB)
        account = get_object_or_404(AccountBank.objects.using('default'), id=account_id)
        account_id_str = account.account_id  # This is the string identifier used in DataEntry

        # Step 2: Loop through all fiscal DBs and delete related DataEntry entries + attached files
        fiscal_dbs = [db for db in settings.DATABASES if db != 'default']

        for db in fiscal_dbs:
            try:
                entries = DataEntry.objects.using(db).filter(account_id=account_id_str)
                for entry in entries:
                    # Delete attached file
                    if entry.attach_file and entry.attach_file.name:
                        file_path = os.path.join(settings.MEDIA_ROOT, entry.attach_file.name)
                        if os.path.exists(file_path):
                            os.remove(file_path)
                    entry.delete(using=db)
            except Exception as e:
                print(f"[{db}] Skipping DataEntry delete for account {account_id_str}: {e}")

        # Step 3: Delete related TaxSuspension and AccountBank from default DB via raw SQL
        try:
            with connections['default'].cursor() as cursor:
                # TaxSuspension linked via FK to AccountBank.id
                cursor.execute("DELETE FROM accounts_taxsuspension WHERE account_id = %s", [account_id])

                # Delete the AccountBank itself
                cursor.execute("DELETE FROM accounts_accountbank WHERE id = %s", [account_id])
        except Exception as e:
            print(f"❌ Failed to delete AccountBank from default: {e}")

        return redirect('account_list')

def delete_group(request, group_id):
    if request.method == 'GET':
        # Step 1: Get group object from default DB
        group = get_object_or_404(Group.objects.using('default'), id=group_id)

        # Step 2: Get all clients in this group (based on group.id)
        from .models import Client  # import here to avoid circular issues
        clients = Client.objects.using('default').filter(group_id=group.id)

        fiscal_dbs = [db for db in settings.DATABASES if db != 'default']

        for client in clients:
            # Delete DataEntry and IncomeTaxReturn across all fiscal DBs
            for db in fiscal_dbs:
                try:
                    DataEntry.objects.using(db).filter(client_id=client.id).delete()
                except Exception as e:
                    print(f"[{db}] Skipping DataEntry delete for client {client.client_id}: {e}")
                try:
                    from .models import IncomeTaxReturn
                    IncomeTaxReturn.objects.using(db).filter(client_id=client.id).delete()
                except Exception as e:
                    print(f"[{db}] Skipping IncomeTaxReturn delete for client {client.client_id}: {e}")

            # Step 3: Delete related items from default DB via raw SQL
            try:
                with connections['default'].cursor() as cursor:
                    # Delete AccountBank linked via client_id (string)
                    cursor.execute("DELETE FROM accounts_accountbank WHERE client_id = %s", [client.client_id])

                    # Delete TaxSuspension linked via client.id
                    cursor.execute("DELETE FROM accounts_taxsuspension WHERE client_id = %s", [client.id])

                    # Finally delete client
                    cursor.execute("DELETE FROM accounts_client WHERE id = %s", [client.id])
            except Exception as e:
                print(f"❌ Failed to delete client {client.id}: {e}")

        # Step 4: Delete group itself
        try:
            with connections['default'].cursor() as cursor:
                cursor.execute("DELETE FROM accounts_group WHERE id = %s", [group_id])
        except Exception as e:
            print(f"❌ Failed to delete group {group_id}: {e}")

        return redirect('group_list')
    
from django.http import JsonResponse
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import AccountBank, Client
from .forms import AccountForm

def account_form(request):
    print("🔍 REQUEST METHOD:", request.method)
    print("🔍 RAW POST:", request.POST.dict())
    print("🔍 mode from POST:", request.POST.get("mode"))
    print("🔍 account_id from POST:", request.POST.get("account_id"))


    account_id = request.POST.get('account_id') or request.GET.get('account_id')
    # account_pk = request.POST.get('account_pk') or request.GET.get('account_pk')
    is_modify_mode = 'modify' in [request.POST.get('mode'), request.GET.get('mode')]


    # Debug prints
    print(f"Account ID received: {account_id}")
    # print(f"Account ID received: {account_pk}")
    print(f"Modify mode: {is_modify_mode}")

    instance = None
    if is_modify_mode and account_id:
        print("📌 Available AccountBank IDs:", list(AccountBank.objects.values_list('id', flat=True)))

        try:
       
            # instance = AccountBank.objects.get(
            #     Q(id=account_id) | Q(account_id=account_id)
            # )
        

            instance = AccountBank.objects.get(account_id=account_id)
            # instance = AccountBank.objects.get(id=account_pk)

            print("✅ Loaded account:", instance)
            print(f"Found account: {instance.account} (ID: {instance.id})")
        except AccountBank.DoesNotExist:
            messages.error(request, "Account not found")
        except AccountBank.MultipleObjectsReturned:
            messages.error(request, "Multiple accounts found")
    if request.method == 'POST':
        print("🟢 Before saving: instance is", "NOT None" if instance else "None")
        form = AccountForm(request.POST, instance=instance)
        if form.is_valid():
            obj = form.save()
           

            print("🟢 Saved object ID:", obj.id)
            if instance:
                messages.success(request, " Account updated successfully")
            else:
                messages.success(request, "✅ Account created successfully")
            return redirect('account_form')  # Or redirect back with modify params
        else:
            errors = "; ".join([f"{f}: {', '.join(e)}" for f, e in form.errors.items()])
            messages.error(request, f"Form errors → {errors}")
    else:
        form = AccountForm(instance=instance)
        #  Pre-fill client if provided in URL
       
        # ✅ AUTO-SELECT CLIENT FROM URL (if exists)
        client_id = request.GET.get('client_id')

        # if client_id and not instance:
        #     try:
        #         client_obj = Client.objects.get(client_id=client_id)

        #         # ✅ IMPORTANT: set BOTH initial AND field value
        #         form.initial['client'] = client_obj.client_id

        #     except Client.DoesNotExist:
        #         pass


    # ✅ Generate next Account ID (ONLY for new form)
    if not is_modify_mode:
        last_id = AccountBank.objects.aggregate(mx=Max("account_id"))["mx"]

        if last_id is None:
            next_account_id = 1
        else:
            next_account_id = int(last_id) + 1   # ✅ force convert to number

        # last_id = AccountBank.objects.aggregate(mx=Max("account_id"))["mx"]
        # next_account_id = (last_id + 1) if last_id else 1
    else:
        next_account_id = None


    # For dropdowns
    all_clients = Client.objects.all()
    client_id_for_accounts = instance.client.client_id if instance and instance.client else request.GET.get('client_id')
    # client_id_for_accounts = (
    #     instance.client.client_id
    #     if instance and instance.client
    #     else request.GET.get('client_id')
    # )

    accounts_queryset = (
        AccountBank.objects.filter(client__client_id=client_id_for_accounts)
        if client_id_for_accounts
        else AccountBank.objects.none()
    )

    accounts_queryset = AccountBank.objects.filter(client__client_id=client_id_for_accounts) if client_id_for_accounts else []
    if client_id_for_accounts:
        accounts_queryset = AccountBank.objects.filter(client__client_id=client_id_for_accounts)
    else:
        accounts_queryset = AccountBank.objects.all()  # fallback to all
     # ✅ NEW: where to go when clicking Cancel
    cancel_url = _safe_back_url(request, default_name="home")

    return render(request, 'account_form.html', {
        'form': form,
        'clients': all_clients,
        'accounts': accounts_queryset,
        'is_modify_mode': is_modify_mode,
        # 'selected_account_id': instance,
        'selected_account_id': instance.id if instance else None,
        # 'selected_account_id': instance.account_id if instance else None,
        # 'selected_account_pk': instance.id if instance else None,
        'next_account_id': next_account_id,
        'title': 'Modify Account' if is_modify_mode else 'Add Account',
        'cancel_url': cancel_url,  
    })

# # account_form ###############
# def get_client_accounts(request):
#     client_id = request.GET.get('client_id')
#     print(">>> Client ID received in view:", client_id)

#     if not client_id:
#         return JsonResponse({'error': 'No client_id provided'}, status=400)

#     try:
#         client = Client.objects.get(client_id=client_id)
#         accounts = client.accounts.all().values('id', 'account', 'account_id')
#         return JsonResponse({'accounts': list(accounts)})
#     except Client.DoesNotExist:
#         return JsonResponse({'error': 'Client not found'}, status=404)
#     except Exception as e:
#         return JsonResponse({'error': str(e)}, status=500)

# account_form ###############
from django.http import JsonResponse
from django.db.models import Q

def get_client_accounts(request):
    client_id = (request.GET.get('client_id') or "").strip()
    if not client_id:
        return JsonResponse({'error': 'No client_id provided'}, status=400)

    allowed_groups = ["Bank Accounts", "Credit Card","Bank O/D Account"]

    try:
        client = Client.objects.get(client_id=client_id)

        qs = (client.accounts
                    .filter(
                        Q(account_group__in=allowed_groups) &
                        ~Q(account_group__isnull=True) &
                        ~Q(account_group__exact="")
                    )
                    .order_by("account"))

        accounts = qs.values('id', 'account', 'account_id', 'account_group')

        return JsonResponse({'accounts': list(accounts)})
    except Client.DoesNotExist:
        return JsonResponse({'error': 'Client not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)





from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def get_account_data(request):
    client_id = request.GET.get('client_id')
    account_id = request.GET.get('account_id')
    
    try:
        if account_id:
            account = AccountBank.objects.get(id=account_id)
            # account = AccountBank.objects.get(account_id=account_id)

        elif client_id:
            account = AccountBank.objects.filter(client__client_id=client_id).latest('id')
        else:
            return JsonResponse({'error': 'No client or account specified'}, status=400)
            
        data = {
            'account_id': account.account_id,
            'data_entry': account.data_entry,
            'account': account.account,
            'account_group': account.account_group,
            'acc_mail_id': account.acc_mail_id,
            'bank_name': account.bank_name,
            'branch': account.branch,
            'account_no': account.account_no,
            'busyacccode':account.busyacccode,
            'ifsc_code': account.ifsc_code,
            'acc_type': account.acc_type,
            'full_match': account.full_match,
            'e_statement': account.e_statement,
            'two_match': account.two_match,
            'contact_no': account.contact_no,
            'pw': account.pw,
            'cif_number': account.cif_number or '',
            'closing_date': account.closing_date.strftime('%Y-%m-%d') if account.closing_date else '',
            'bank_mail_id': account.bank_mail_id or '',
            'branch_mail_id': account.branch_mail_id or '',
            'branch_mobile_no': account.branch_mobile_no or '',
            'statement_frequency': account.statement_frequency or '',
            'reminder_date': account.reminder_date.strftime('%Y-%m-%d') if account.reminder_date else '',
        }
        return JsonResponse(data)
    except AccountBank.DoesNotExist:
        return JsonResponse({'error': 'No account found'}, status=404)
    
    
    

from django.http import JsonResponse



def account_list(request):
    """View to list all accounts"""
    accounts = AccountBank.objects.all().select_related('client')
    return render(request, 'account_list.html', {'accounts': accounts})

def group_list(request):
    group = Group.objects.all().order_by('group_name')
    return render(request, 'group_list.html', {'groups': group})



import json
from django.db import transaction
from django.contrib import messages
from django.shortcuts import render, redirect
from .models import DataEntry, Client, AccountBank
from .forms import DataEntryForm

from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from .models import DataEntry, Client, AccountBank
from .forms import DataEntryForm
from django.contrib.auth.models import User

def get_recent_fiscal_dbs(years_back=1, years_forward=1):
    today = date.today()
    current_fy = today.year if today.month >= 4 else today.year - 1
    fiscal_years = []
    for offset in range(-years_back, years_forward + 1):
        fy_start = current_fy + offset
        fy_label = f"{fy_start}_{str(fy_start + 1)[-2:]}"
        fiscal_years.append(f"fy_{fy_label}")
    return fiscal_years

# perfect view #######
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import DataEntry, Client, AccountBank, UserData
from .forms import DataEntryForm
from .utills import get_fiscal_year_from_date
from datetime import datetime,timedelta, time 
from pathlib import Path
import os
import shutil

# move back to last access page
from django.urls import reverse
from django.utils.http import url_has_allowed_host_and_scheme

def _safe_back_url(request, default_name="home"):
    """Prefer ?next=..., then HTTP_REFERER, else fallback to a safe default."""
    host = request.get_host()
    next_url = request.GET.get("next") or request.POST.get("next")
    if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={host}):
        return next_url

    referer = request.META.get("HTTP_REFERER")
    if referer and url_has_allowed_host_and_scheme(referer, allowed_hosts={host}):
        # Avoid loop if referer is this page
        if request.build_absolute_uri(request.get_full_path()) != referer:
            return referer

    return reverse(default_name)

def data_entry(request):
    is_modify_mode = request.GET.get('mode') == 'modify' or request.POST.get('mode') == 'modify'
    entry_id = request.GET.get('entry_id') or request.POST.get('original_entry_id')
    fy = request.GET.get("fy") or request.POST.get("fy")

    instance = None
    db_found = None

    # === LOAD INSTANCE ===
    if is_modify_mode and entry_id:
        requested_fy = request.GET.get("fy") or request.POST.get("fy")
        if requested_fy:
            normalized_fy = requested_fy.replace("-", "_")
            db_name = f"fy_{normalized_fy}"
            if db_name in settings.DATABASES:
                try:
                    instance = DataEntry.objects.using(db_name).get(id=entry_id)
                    db_found = db_name
                    print(f"✅ Loaded entry {entry_id} from {db_name}")

                    # 🔹 Auto-clean if file missing on disk
                    if instance.attach_file:
                        try:
                            if not instance.attach_file.storage.exists(instance.attach_file.name):
                                instance.attach_file = None
                                instance.save(using=db_name, update_fields=["attach_file"])
                        except Exception as e:
                            print("⚠️ Could not verify existing attachment:", e)

                except DataEntry.DoesNotExist:
                    messages.error(request, f"Entry {entry_id} not found in {db_name}")
                    return redirect('data_entry')
            else:
                messages.error(request, f"Database {db_name} not configured.")
                return redirect('data_entry')

    # === POST ===
    if request.method == 'POST':
        form = DataEntryForm(request.POST, request.FILES, instance=instance)
        form.fields['alloted_to'].required = False
        form.fields['status'].required = False

        if form.is_valid():
            entry = form.save(commit=False)
            account_id = request.POST.get('account')
             # 🔹 If a file is attached, force format = "Soft Copy"
            uploaded_file = request.FILES.get('attach_file')
            if uploaded_file:
                entry.format = "Soft Copy"   # <-- use the exact value from your choices
            remove_file = bool(request.POST.get('remove_file'))

            if account_id in ['1', '2']:
                entry.account = None
                entry.virtual_account_type = account_id
            else:
                try:
                    entry.account_id = int(account_id)
                except:
                    form.add_error('account', 'Invalid account')
                    return render(request, 'data_entry.html', {'form': form})

            fiscal_start = entry.from_date
            fiscal_end = entry.last_date
            if not fiscal_start or not fiscal_end:
                form.add_error(None, 'Both From Date and Last Date are required.')
                return render(request, 'data_entry.html', {'form': form})

            primary_fy = get_fiscal_year_from_date(fiscal_start)
            primary_db = db_found if is_modify_mode and db_found else f"fy_{primary_fy}"
            dbs_used = {primary_db}

            if instance:
                entry.id = instance.id

            # uploaded_file = request.FILES.get('attach_file')
            if uploaded_file:
                # ✅ If a new file is uploaded, optionally remove old physical file
                if instance and instance.attach_file:
                    try:
                        old_path = instance.attach_file.path
                        if os.path.exists(old_path):
                            os.remove(old_path)
                    except Exception as e:
                        print("⚠️ Could not delete old attachment:", e)

                entry.attach_file = None
                entry.save(using=primary_db)
                ext = Path(uploaded_file.name).suffix
                client_part = entry.client.client_id.strip()
                account_part = entry.account.account_id.strip() if entry.account else "unknown-account"
                new_filename = f"{entry.id}{ext}"
                new_rel_path = os.path.join(primary_db, client_part, account_part, new_filename)
                new_full_path = os.path.join(settings.MEDIA_ROOT, new_rel_path)
                os.makedirs(os.path.dirname(new_full_path), exist_ok=True)
                with open(new_full_path, 'wb+') as destination:
                    for chunk in uploaded_file.chunks():
                        destination.write(chunk)
                entry.attach_file.name = new_rel_path
                entry.save(using=primary_db, update_fields=["attach_file"])
            elif remove_file and instance and instance.attach_file:
                # ✅ User wants to remove the existing file, and no new file was uploaded
                try:
                    old_path = instance.attach_file.path
                    if os.path.exists(old_path):
                        os.remove(old_path)
                except Exception as e:
                    print("⚠️ Could not delete attachment on remove:", e)

                entry.attach_file = None
                # NOTE: here we don't touch entry.format in backend, because
                # the user / JS already set the desired value in the form.
                entry.save(using=primary_db, update_fields=["attach_file", "format"])

            else:
                entry.save(using=primary_db)
        
            #  Optional replication to future years
            if not is_modify_mode:
                current = datetime(fiscal_start.year + 1, 4, 1).date()
                while current <= fiscal_end:
                    fy = get_fiscal_year_from_date(current)
                    db_name = f"fy_{fy}"
                    if db_name not in dbs_used and db_name in settings.DATABASES:
                        entry_copy = deepcopy(entry)
                        entry_copy.pk = None
                        entry_copy.attach_file = None
                        entry_copy.save(using=db_name)
                        if entry.attach_file:
                            ext = Path(entry.attach_file.name).suffix
                            source_path = entry.attach_file.path
                            client_part = entry.client.client_id.strip()
                            account_part = entry.account.account_id.strip() if entry.account else "unknown-account"
                            new_filename = f"{entry_copy.id}{ext}"
                            new_rel_path = os.path.join(db_name, client_part, account_part, new_filename)
                            new_full_path = os.path.join(settings.MEDIA_ROOT, new_rel_path)
                            os.makedirs(os.path.dirname(new_full_path), exist_ok=True)
                            shutil.copy2(source_path, new_full_path)
                            entry_copy.attach_file.name = new_rel_path
                            entry_copy.save(using=db_name, update_fields=["attach_file"])
                        dbs_used.add(db_name)
                    current = datetime(current.year + 1, 4, 1).date()
                #  ADD MESSAGE HERE after replication is done
            if is_modify_mode:
                messages.success(
                    request,
                    f"✅ Entry {entry.id} updated successfully in {primary_db}"
                )
            else:
                # Sort DBs for cleaner display
                all_dbs_text = ", ".join(sorted(dbs_used))
                if len(dbs_used) > 1:
                    messages.success(
                        request,
                        f"✅ Entry {entry.id} created successfully in DBs: {all_dbs_text}"
                    )
                else:
                    messages.success(
                        request,
                        f"✅ Entry {entry.id} created successfully in {primary_db}"
                    )
            # ✅ Final redirect with preserved filters
            redirect_filters = {}
            for key in [
                "group_id", "client_id", "account_id", "from_month", "to_month",
                "status", "alloted", "entry_type", "report_type",
                "busy_tran", "acstkameet", "e_stmt", "account_groups"
            ]:
                val = request.POST.get(key)
                if val:
                    redirect_filters[key] = val

            redirect_filters["highlight_id"] = entry.id
            redirect_filters["fy"] = primary_fy

            query_string = "&".join([f"{k}={v}" for k, v in redirect_filters.items()])
            return redirect(f"/reports/?{query_string}")
            # if request.headers.get("x-requested-with") == "XMLHttpRequest":
            #     print(f"✅ Returning entry_id {entry.id} in JSON response")
            #     return JsonResponse({
            #         "success": True,
            #         "entry_id": entry.id,
            #         "redirect_url": "/reports/",
            #         "fiscal_year": primary_fy,  # ✅ add this key

            #     })
            
            # else:
            #     return redirect(f"/data_entry/?{query_string}")

        else:
            messages.error(request, "Please correct the errors below.")
            print(request.POST.dict())

    else:
        form = DataEntryForm(instance=instance)
        form.fields['alloted_to'].required = False
        if request.GET.get('client_id') and not is_modify_mode:
            try:
                client_obj = Client.objects.get(client_id=request.GET.get('client_id'))
                form.fields['client'].initial = client_obj
            except Client.DoesNotExist:
                pass

    all_clients = Client.objects.all()
    all_users = UserData.objects.all()
    default_accounts = [
        {'id': 1, 'account': 'Sales', 'account_id': '1', 'client': None},
        {'id': 2, 'account': 'Purchase', 'account_id': '2', 'client': None}
    ]
    client_accounts = AccountBank.objects.all().select_related('client').values(
        'id', 'account', 'account_id', 'client_id'
    )
    cancel_url = _safe_back_url(request, default_name="home")
    # --- Prefill when opened from "missing" row ---
    # --- Prefill when opened from "missing" row ---
    prefill = {}
    if request.GET.get("prefill") == "1":
        fy_in = request.GET.get("fy") or ""
        try:
            db_alias, _, _ = _normalize_fy(fy_in)
        except Exception:
            db_alias = "default"

        c_pk  = request.GET.get("client_pk")
        c_id  = request.GET.get("client_id")          # business ID
        a_pk  = request.GET.get("account_pk")
        a_id  = request.GET.get("account_id")         # account code
        d_from = request.GET.get("from")
        d_to   = request.GET.get("to")

        # Try FY DB first, then default
        def get_client():
            for alias in (db_alias, "default"):
                if c_pk:
                    obj = Client.objects.using(alias).filter(pk=c_pk).first()
                    if obj: return obj
                if c_id:
                    obj = Client.objects.using(alias).filter(client_id=c_id).first()
                    if obj: return obj
            return None

        def get_account(client_obj):
            for alias in (db_alias, "default"):
                if a_pk:
                    obj = AccountBank.objects.using(alias).filter(pk=a_pk).first()
                    if obj: return obj
                if a_id and client_obj:
                    obj = AccountBank.objects.using(alias).filter(client=client_obj, account_id=a_id).first()
                    if obj: return obj
            return None

        client  = get_client()
        account = get_account(client)

        prefill = {
            "client_id":     (client.client_id if client else (c_id or "")),
            "client_name":   (client.client_name if client else ""),  # may be blank
            "client_group":  (getattr(getattr(client, "group", None), "group_name", "") if client else ""),
            "account_pk":    str(account.pk) if account else (a_pk or ""),
            "account_code":  (account.account_id if account else (a_id or "")),
            "account_name":  (account.account if account else ""),
            "from":          d_from or "",
            "to":            d_to or "",
        }

    context = {
        'form': form,
        'clients': all_clients,
        'users': all_users,
        'entries': None,
        'is_modify_mode': is_modify_mode,
        'selected_entry_id': entry_id,
        'title': 'Modify Data Entry' if is_modify_mode else 'Add Data Entry',
        'default_accounts': default_accounts,
        'all_client_accounts': list(client_accounts),
        "fy": fy,
        "cancel_url": cancel_url,
        "prefill": prefill, 
    }
    return render(request, 'data_entry.html', context)

# AJAX view to get client data
def get_client_data(request):
    client_id = request.GET.get('client_id')
    if not client_id:
        return JsonResponse({'error': 'Client ID required'}, status=400)
    
    try:
        client = Client.objects.get(client_id=client_id)
        data = {
            'client_id': client.client_id,
            'client_name': client.client_name,
            # Add other client fields you need
        }
        return JsonResponse(data)
    except Client.DoesNotExist:
        return JsonResponse({'error': 'Client not found'}, status=404)


# AJAX view to get account data
def get_account_data_for_entry(request):
    account_id = request.GET.get('account_id')
    if not account_id:
        return JsonResponse({'error': 'Account ID required'}, status=400)
    
    try:
        account = AccountBank.objects.get(id=account_id)
        data = {
            'account_id': account.id,
            'account_name': account.account,
            'bank_name': account.bank_name,
            # Add other account fields you need
        }
        return JsonResponse(data)
    except AccountBank.DoesNotExist:
        return JsonResponse({'error': 'Account not found'}, status=404)
    

from collections import defaultdict
from django.http import JsonResponse
from .models import AccountBank
from collections import defaultdict, OrderedDict
ACCOUNT_GROUP_LABELS = {
    'bankaccount': 'Bank Account',
    'creditcard': 'Credit Card',
    'investment': 'Investment',
    'securedloan': 'Secured Loan',
    'Bank O/D Account':'Bank O/D Account'
}

from collections import defaultdict, OrderedDict

ACCOUNT_GROUP_DISPLAY_ORDER = [
    'Default Account',
    'Bank Accounts',
    'Credit Card',
    'Investments',
    'Secured Loans',
    'Securities & Deposits (Asset)',
    'Do Not Display'
]

def get_client_accounts_grouped_add(request):
    client_id = request.GET.get('client_id')
    if not client_id:
        return JsonResponse({'error': 'Client ID required'}, status=400)

    print(">>> Grouped Account View - client_id (PAN):", client_id)

    try:
        accounts = AccountBank.objects.filter(client__client_id=client_id).values(
            'id', 'account', 'account_id', 'account_group'
        )

        temp_grouped = defaultdict(list)

        # Add virtual accounts to Default Account
        temp_grouped['Default Account'].extend([
            {'id': '1', 'account': 'Sales', 'account_id': '1'},
            {'id': '2', 'account': 'Purchase', 'account_id': '2'}
        ])

        # Group real accounts
        for acc in accounts:
            raw_group = (acc['account_group'] or '').strip().lower()
            group_label = ACCOUNT_GROUP_LABELS.get(raw_group, raw_group.title() or "Others")

            temp_grouped[group_label].append({
                'id': acc['id'],
                'account': acc['account'],
                'account_id': acc['account_id']
            })

        # Order result
        grouped_accounts = OrderedDict()

        # Add specified priority groups first
        for label in ACCOUNT_GROUP_DISPLAY_ORDER:
            if label in temp_grouped:
                grouped_accounts[label] = temp_grouped.pop(label)

        # Add the remaining groups sorted alphabetically
        for label in sorted(temp_grouped):
            grouped_accounts[label] = temp_grouped[label]

        print(">>> Returning grouped accounts:", dict(grouped_accounts))
        return JsonResponse({'grouped_accounts': grouped_accounts})
    except Exception as e:
        print(">>> ERROR in get_client_accounts_grouped:", e)
        return JsonResponse({'error': str(e)}, status=500)

from collections import defaultdict, OrderedDict

ACCOUNT_GROUP_DISPLAY_ORDER = [
    'Default Account',
    'Bank Accounts',
    'Credit Card',
    'Investments',
    'Secured Loans',
    'Securities & Deposits (Asset)',
    'Do Not Display'

]

def get_client_accounts_grouped_edit(request):
    client_id = request.GET.get('client_id')
    if not client_id:
        return JsonResponse({'error': 'Client ID required'}, status=400)

    print(">>> Grouped Account View - client_id (numeric):", client_id)

    try:
        accounts = AccountBank.objects.filter(client__id=client_id).values(
            'id', 'account', 'account_id', 'account_group'
        )

        temp_grouped = defaultdict(list)

        # No virtual accounts added in edit mode (optional: add if needed)

        for acc in accounts:
            raw_group = (acc['account_group'] or '').strip().lower()
            group_label = ACCOUNT_GROUP_LABELS.get(raw_group, raw_group.title() or "Others")

            temp_grouped[group_label].append({
                'id': acc['id'],
                'account': acc['account'],
                'account_id': acc['account_id']
            })

        grouped_accounts = OrderedDict()

        # First add accounts in the specified display order
        for label in ACCOUNT_GROUP_DISPLAY_ORDER:
            if label in temp_grouped:
                grouped_accounts[label] = temp_grouped.pop(label)

        # Add remaining groups in alphabetical order
        for label in sorted(temp_grouped):
            grouped_accounts[label] = temp_grouped[label]

        print(">>> Returning grouped accounts:", dict(grouped_accounts))
        return JsonResponse({'grouped_accounts': grouped_accounts})
    except Exception as e:
        print(">>> ERROR in get_client_accounts_grouped_edit:", e)
        return JsonResponse({'error': str(e)}, status=500)

def user_form(request):
    # Determine if we're in modify mode
    is_modify_mode = 'mode=modify' in request.get_full_path() or '/user/modify/' in request.path

    # Get selected user ID from GET or POST
    selected_user_id = request.GET.get('user_id') or request.POST.get('user_id')

    # Initialize instance to None
    instance = None
    if is_modify_mode and selected_user_id:
        instance = get_object_or_404(UserData, id=selected_user_id)

    if request.method == 'POST':
        try:
            with transaction.atomic():
                form = UserForm(request.POST, instance=instance)
                if form.is_valid():
                    user = form.save(commit=False)
                    # If this is a new user or password was changed, save the plain text password
                    if not instance or 'password' in form.changed_data:
                        user.password = form.cleaned_data['password']
                    user.save()
                    
                    messages.success(
                        request,
                        f"User {'updated' if is_modify_mode else 'added'} successfully"
                    )
                    return redirect('user_form')
                else:
                    messages.error(request, "Please correct the errors below")
        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")
    else:
        form = UserForm(instance=instance)

    users = UserData.objects.all() if is_modify_mode else None
    return render(request, 'user_form.html', {
        'form': form,
        'is_modify_mode': is_modify_mode,
        'users': users,
        'title': 'Modify User' if is_modify_mode else 'Add User',
        'selected_user_id': int(selected_user_id) if selected_user_id else None,
    })


def user_list(request):
    users = UserData.objects.all().order_by('username')
    print(f"Users count: {users.count()}")
    return render(request,'user_list.html',{'users': users})

def delete_user(request, user_id):
    user = get_object_or_404(UserData, id=user_id)
    user.delete()
    messages.success(request, "User deleted successfully.")
    return redirect('user_list')  # Adjust if your URL name is different

from django.shortcuts import render
from django.contrib import messages
from .models import Client, AccountBank, UserData, DataEntry
from django.shortcuts import render
from django.contrib import messages
from .models import Client, AccountBank, UserData, DataEntry
import os
from django.template import loader
from accounts.utills import get_db_for_fy,fiscal_year_range
# from accounts.yearwise_router import fiscal_year_range  # ✅ correct

def data_entry_list(request):
    fiscal_years = ['2023_24', '2024_25', '2025_26']
    fy_param = request.GET.get('fy')
    data_entries = []

    for entry in data_entries:
        print(f"ID: {entry.id}, client_id: {entry.client_id}, account_id: {entry.account_id}")

    # Auto-select the latest FY with data if fy not provided
    if not fy_param:
        for year in reversed(fiscal_years):
            db_name = f"fy_{year}"
            try:
                count = DataEntry.objects.using(db_name).count()
                if count > 0:
                    fy_param = year
                    break
            except Exception as e:
                print(f"Skipping {db_name} due to error: {e}")

    # Load entries from selected fiscal year
    if fy_param:
        db_name = f"fy_{fy_param}"
        try:
            data_entries = list(
                DataEntry.objects.using(db_name).all()
                # .select_related('client', 'account', 'alloted_to')
                
                # .all()
            )
            print("\n=== DATA VERIFICATION ===")
            print(f"Using database: {db_name}")
            print(f"Total entries: {len(data_entries)}")
        except Exception as e:
            print(f"\n=== ERROR: {str(e)} ===")
            messages.error(request, f"Failed to load data from DB '{db_name}': {e}")

    print("TEMPLATE PATH:")
    print(os.path.abspath(loader.get_template('data_entry_list.html').origin.name))

    context = {
        'clients': Client.objects.all(),
        'accounts': AccountBank.objects.all(),
        'users': UserData.objects.all(),
        'fiscal_years': fiscal_years,
        'fy_selected': fy_param,
        'data_entries': data_entries,
    }
    return render(request, 'data_entry_list.html', context)

def delete_data_entry(request, entry_id):
    fy = request.GET.get("fy")  # Expected: "2025_26"
    if not fy:
        messages.error(request, "❌ Fiscal year (fy) not specified in URL.")
        return redirect('data_entry_list')

    db = f'fy_{fy}'  # ✅ Matches keys in settings.DATABASES
    print(f"\n=== DELETE DEBUG ===\nTarget DB: {db}\nEntry ID: {entry_id}")

    if db not in settings.DATABASES:
        print(f"❌ ERROR: Database '{db}' not found in settings.DATABASES.")
        messages.error(request, f"Database '{db}' not configured.")
        return redirect(f'/data_entry_list/?fy={fy}')

    try:
        entry = get_object_or_404(DataEntry.objects.using(db), id=entry_id)
        print(f"✅ Found entry: {entry}")

        # Delete file if exists
        if entry.attach_file and entry.attach_file.name:
            file_path = os.path.join(settings.MEDIA_ROOT, entry.attach_file.name)
            if os.path.exists(file_path):
                os.remove(file_path)
                print(f"🗑️ Deleted file: {file_path}")
            else:
                print(f"⚠️ File not found: {file_path}")

        # Delete entry
        entry.delete(using=db)
        print(f"✅ Entry ID {entry_id} deleted from {db}.")
        messages.success(request, "✅ Data entry deleted successfully.")

    except Exception as e:
        print(f"❌ Exception during deletion: {e}")
        messages.error(request, f"Failed to delete entry: {e}")

    return redirect(f'/data_entry_list/?fy={fy}')


from django.http import JsonResponse
from .models import DataEntry
from .utills import get_db_for_fy
import traceback
# FIXED: remove 'entry_id' argument from here
from django.http import JsonResponse
import traceback

def delete_report_data_entry(request):
    if request.method == "POST":
        try:
            entry_id = request.POST.get("entry_id")
            fy = request.POST.get("fy")

            print("🧪 DELETE View received → entry_id:", entry_id, "fy:", fy)

            if not entry_id or not fy:
                return JsonResponse({"success": False, "error": "Missing ID or FY."})

            normalized_fy = fy.replace("-", "_")
            db = f"fy_{normalized_fy}"
            print("✅ Resolved DB:", db)

            if db not in settings.DATABASES:
                return JsonResponse({"success": False, "error": f"Database '{db}' is not configured."})

            # ✅ Use raw SQL to delete to avoid FK DB mismatch
            from django.db import connections
            with connections[db].cursor() as cursor:
                cursor.execute("DELETE FROM accounts_dataentry WHERE id = %s", [entry_id])
                print("✅ Raw SQL delete successful from", db)

            return JsonResponse({"success": True})
        
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "Invalid request."})


from django.views.decorators.http import require_http_methods
@csrf_exempt
def get_entry(request, entry_id):
    print("=== get_entry CALLED ===", entry_id)

    fy = request.GET.get("fy")
    if not fy:
        return JsonResponse({"error": "Missing fiscal year"}, status=400)

    db = get_db_for_fy(fy)
    print("Resolved DB:", db)

    try:
        entry = DataEntry.objects.using(db).get(id=entry_id)
        client_accounts = AccountBank.objects.filter(client=entry.client) if entry.client else AccountBank.objects.none()

        virtual_accounts = {
            '1': 'Sales (Default)',
            '2': 'Purchase (Default)',
        }

        if entry.virtual_account_type in virtual_accounts:
            account_data = {
                'id': entry.virtual_account_type,
                'name': virtual_accounts[entry.virtual_account_type]
            }
        elif entry.account:
            account_data = {
                'id': str(entry.account.id),
                'name': entry.account.account
            }
        else:
            account_data = {
                'id': '',
                'name': ''
            }

        data = {
            'client': {
                'id': entry.client.id if entry.client else '',
                'name': entry.client.client_name if entry.client else ''
            },
            'account': account_data,
            'virtual_account_type': entry.virtual_account_type,
            'status': entry.status,
            'format': entry.format,
            'received_by': entry.received_by,
            'rec_date': entry.rec_date.isoformat() if entry.rec_date else '',
            'from_date': entry.from_date.isoformat() if entry.from_date else '',
            'last_date': entry.last_date.isoformat() if entry.last_date else '',
            'alloted_date': entry.alloted_date.isoformat() if entry.alloted_date else '',
            'done_date': entry.done_date.isoformat() if entry.done_date else '',
            'alloted_to': entry.alloted_to.id if entry.alloted_to else '',
            'query': entry.query,
            'remark': entry.remark,
            'attach_file': entry.attach_file.url if entry.attach_file else '',
            'is_nil': entry.is_nil,
            'accounts': [
                {'id': acc.id, 'account': acc.account}
                for acc in client_accounts
            ]
        }

        print("Returning data:", data)
        return JsonResponse(data)

    except DataEntry.DoesNotExist:
        print("❌ Entry not found:", entry_id)
        return JsonResponse({'error': 'Entry not found'}, status=404)

    except Exception as e:
        print("❌ Exception:", str(e))
        return JsonResponse({'error': str(e)}, status=500)

from datetime import date, datetime

#modal - dataentry
@csrf_exempt
def update_data_entry(request, entry_id):
    try:
        def parse_date(date_str):
            try:
                return datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else None
            except:
                return None

        data = request.POST
        files = request.FILES

        # Determine fiscal year DB
        from_date = parse_date(data.get("from_date"))
        last_date = parse_date(data.get("last_date"))
        base_date = from_date or last_date
        if not base_date:
            return JsonResponse({"success": False, "error": "Either from_date or last_date is required."})

        fiscal_year = get_fiscal_year_from_date(base_date)
        db = f"fy_{fiscal_year}"

        # Load entry from that DB
        try:
            entry = DataEntry.objects.using(db).get(id=entry_id)
        except DataEntry.DoesNotExist:
            return JsonResponse({"success": False, "error": "Entry not found in database."})
        

        # ForeignKey fields
        client_id = data.get("client")
        if not client_id:
            return JsonResponse({"success": False, "error": "Client is required."})
        try:
            entry.client = Client.objects.get(pk=client_id)
        except Client.DoesNotExist:
            return JsonResponse({"success": False, "error": "Invalid client ID."})

        account_id = data.get("account")
        if account_id and account_id not in ['1', '2']:
            try:
                entry.account = AccountBank.objects.get(pk=account_id)
                entry.virtual_account_type = None
            except AccountBank.DoesNotExist:
                entry.account = None
        else:
            entry.account = None
            entry.virtual_account_type = account_id

        alloted_to_id = data.get("alloted_to")
        if alloted_to_id:
            try:
                entry.alloted_to = UserData.objects.get(pk=alloted_to_id)
            except UserData.DoesNotExist:
                entry.alloted_to = None
        else:
            entry.alloted_to = None

        # Normal fields
        entry.status = data.get("status", "")
        entry.format = data.get("format", "")
        entry.received_by = data.get("received_by", "")
        entry.query = data.get("query", "")
        entry.rec_date = parse_date(data.get("rec_date"))
        entry.from_date = from_date
        entry.last_date = last_date
        entry.alloted_date = parse_date(data.get("alloted_date"))
        entry.done_date = parse_date(data.get("done_date"))
        entry.remark = data.get("remark", "")
        entry.is_nil = data.get("is_nil") == "true"

       # === File Handling (Save as entry_id.pdf) ===
        if "attach_file" in files:
            uploaded = files["attach_file"]

            client_part = entry.client.client_id.strip()
            if entry.account:
                account_part = entry.account.account_id.strip()
            elif entry.virtual_account_type == "1":
                account_part = "sale_1"
            elif entry.virtual_account_type == "2":
                account_part = "purchase_2"
            else:
                account_part = "unknown_account"

            filename = f"{entry_id}.pdf"  # ✅ Force filename to entry_id.pdf
            subdir = os.path.join(f"fy_{fiscal_year}", client_part, account_part)
            full_dir = os.path.join(settings.MEDIA_ROOT, subdir)
            os.makedirs(full_dir, exist_ok=True)
            full_path = os.path.join(full_dir, filename)

            # Save file manually
            with open(full_path, "wb+") as dest:
                for chunk in uploaded.chunks():
                    dest.write(chunk)

            # Save relative path to model field
            entry.attach_file.name = os.path.join(subdir, filename)


        entry.save(using=db)
        return JsonResponse({"success": True})

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


from django.shortcuts import render, redirect
from django.contrib import messages
from .models import UserData
from .forms import LoginForm

def login_view(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username'].strip()
            password = form.cleaned_data['password'].strip()

            try:
                user = UserData.objects.get(username=username, password=password)
                request.session['user_id'] = user.id
                request.session['username'] = user.username
                # messages.success(request, f"Welcome {user.username}")
                messages.success(request, "")
                
                # Set session flag to show welcome popup
                request.session['show_welcome'] = True
                return redirect('home')
            except UserData.DoesNotExist:
                # messages.error(request, "Invalid username or password")
                messages.error(request, "")
    else:
        form = LoginForm()
    
    return render(request, 'login.html', {'form': form})

def logout_view(request):
    request.session.flush()
    # messages.success(request, "You have been logged out.")
    messages.success(request, "")
    return redirect('login')



# Only bank account and creidt card displayed in report page(account dropdown)
# reports page accounts endpoint (name may differ)
from django.http import JsonResponse
from django.db.models import Q
ALLOWED_ACCOUNT_GROUPS = ["Bank Accounts", "Credit Card","Bank O/D Account"]

def get_client_accounts_for_reports(request):
    client_id = (request.GET.get("client_id") or "").strip()
    if not client_id:
        return JsonResponse({"ok": False, "error": "client_id required"}, status=400)

    try:
        client = Client.objects.get(client_id=client_id)
    except Client.DoesNotExist:
        return JsonResponse({"ok": False, "error": "Client not found"}, status=404)

    qs = (
        client.accounts
        .filter(
            Q(account_group__in=ALLOWED_ACCOUNT_GROUPS)
            & ~Q(account_group__isnull=True)
            & ~Q(account_group__exact="")
        )
        .order_by("account")
    )

    data = [{
        "id": a.id,
        "account": a.account,
        "account_id": a.account_id,
        "account_group": a.account_group,
    } for a in qs]

    return JsonResponse({"ok": True, "accounts": data})

def reports(request):
    # fiscal_years = ['2024-25', '2023-24', '2022-23']
    current_year = datetime.now().year
    start_fy = 2024
    fiscal_years = [f"{y}_{str(y+1)[-2:]}" for y in range(start_fy, current_year + 2)]
    account_groups = ['Bank Accounts', 'Credit Card', 'Secured Loans', 'Investments']
    allowed_groups = ["Bank Accounts", "Credit Card"]

    months = [ "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec","Jan", "Feb", "Mar",]
    short_months = [m[:3] for m in months]
    entry_fields = [
    ("trans_id",   "Trans ID"),
    ("rec_date",   "Rec Date"),
    ("format",     "Format"),
    ("rec_by",     "Rec By"),
    # ("balance",    "Balance"),
    ("alloted_to", "Alloted To"),
    ("alloted_dt", "Alloted Dt"),
    ("done_dt",    "Done Dt"),
    ("query",      "Query"),
]

    context = {
        "accounts": [],
        "fiscal_years": fiscal_years,
        "account_groups":['Bank Accounts', 'Credit Card',"Bank O/D Account"],
        "months": short_months,
        # "entry_fields": ["Trans ID", "Rec Date", "Format", "Rec By", "Balance",
        #                  "Alloted To", "Alloted Dt", "Done Dt", "Query"],
        "entry_fields": entry_fields,     #  ← now a list of tuples
        "employees": UserData.objects.order_by("username"),
        "groups": Group.objects.order_by("group_name"),        # all groups for first render
        "clients": Client.objects.order_by("client_name"),        # all clients  – will be filtered later
        "accounts": AccountBank.objects.order_by("account"), # all accounts – will be filtered later
        #   ⬇️ Only Bank Accounts + Credit Card
        "accounts": AccountBank.objects.filter(
            account_group__in=ALLOWED_ACCOUNT_GROUPS
        ).exclude(
            account_group__isnull=True
        ).exclude(
            account_group__exact=""
        ).order_by("account"),

    }
    return render(request, "reports.html", context)


# ---------- AJAX helpers ---------- #
def ajax_clients(request):
    """
    Return clients for a given group_id
    ?group_id=123   → only that group’s clients
    (no param)      → all clients
    """
    gid = request.GET.get("group_id")
    # qs = Client.objects.filter(group_id=gid) if gid else Client.objects.all()
    qs = Client.objects.filter(group_id=gid).order_by("client_name") if gid else Client.objects.all().order_by("client_name")

    data = list(qs.values("id", "client_id", "client_name"))
    return JsonResponse(data, safe=False)

def ajax_accounts(request):
    print("HIT ajax_accounts", request.GET.dict(), ALLOWED_ACCOUNT_GROUPS)
    """
    ?client_id=ADR... → that client’s accounts
    (no param)        → all accounts (but still only allowed groups)
    """
    cid = (request.GET.get("client_id") or "").strip()
    gid = (request.GET.get("group_id") or "").strip()


    # base = AccountBank.objects.all()
     # base queryset: only allowed account groups
    base = AccountBank.objects.filter(
        account_group__in=ALLOWED_ACCOUNT_GROUPS
    ).exclude(
        account_group__isnull=True
    ).exclude(
        account_group__exact=""
    )
    
    # priority: client -> group -> all

    if cid:
        base = base.filter(client__client_id=cid)
    elif gid:
        base = base.filter(client__group_id=gid)

    # qs = (base
    #       .filter(account_group__in=ALLOWED_ACCOUNT_GROUPS)
    #       .exclude(account_group__isnull=True)
    #       .exclude(account_group__exact="")
    #       .order_by("account"))
    qs = base.order_by("account")


     # DEBUG: what groups survived the filter?
    print("groups in qs:", list(qs.values_list("account_group", flat=True).distinct()))
    print("count in qs:", qs.count())
    
    print("ajax_accounts sample:", list(qs.values("id", "account_id", "account")[:5]))
    data = list(qs.values("id", "account_id", "account", "account_group"))
    return JsonResponse(data, safe=False)

from django.http import JsonResponse
from .models import DataEntry
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from .models import DataEntry
from accounts.utills import get_fiscal_year_from_date, get_db_for_fy

def entry_detail(request):
    pk = request.GET.get("id")
    fy = request.GET.get("fy", "").replace("-", "_")

    if not pk or not pk.isdigit():
        return JsonResponse({"error": "Invalid ID"}, status=400)

    if not fy or f"fy_{fy}" not in settings.DATABASES:
        return JsonResponse({"error": "Invalid or missing fiscal year"}, status=400)

    db = f"fy_{fy}"

    from accounts.models import DataEntry

    try:
        entry = DataEntry.objects.using(db).get(pk=pk)
    except DataEntry.DoesNotExist:
        return JsonResponse({"error": "Entry not found"}, status=404)

    data = {
        "trans_id": entry.id,
        "rec_date": entry.rec_date.strftime("%d-%m-%Y") if entry.rec_date else "",
        "format": entry.format,
        "rec_by": entry.received_by,
        "balance": entry.account.full_match if entry.account else "",
        "alloted_to": entry.alloted_to.username if entry.alloted_to else "",
        "alloted_dt": entry.alloted_date.strftime("%d-%m-%Y") if entry.alloted_date else "",
        "done_dt": entry.done_date.strftime("%d-%m-%Y") if entry.done_date else "",
        "query": entry.query or "",
    }

    return JsonResponse(data)

# views.py
from datetime import datetime
from django.db.models import Q
from django.db.models.lookups import Lookup
from django.http import JsonResponse
from .models import DataEntry
from datetime import date
from django.http import JsonResponse
from .models import DataEntry
from datetime import datetime, timedelta

def fiscal_year_range(fy):
    try:
        start_year = int(fy.split("-")[0])
        return datetime(start_year, 4, 1).date(), datetime(start_year + 1, 3, 31).date()
    except:
        return None, None

def get_db_for_fy(fy):
    normalized_fy = fy.replace("-", "_")  # 2025-26 → 2025_26
    return f"fy_{normalized_fy}"
   
from calendar import monthrange
from datetime import timedelta
from dateutil.relativedelta import relativedelta

def get_fy_month_ranges(start, end):
    month_ranges = []
    current = start
    while current <= end:
        last_day = monthrange(current.year, current.month)[1]
        month_start = current.replace(day=1)
        month_end = current.replace(day=last_day)

        month_ranges.append((month_start, month_end))
        current += relativedelta(months=1)
    return month_ranges


def subtract_covered_ranges(month_start, month_end, covered):
    """
    Given a full month range and list of covered ranges,
    return a list of missing gaps.
    """
    gaps = []
    current = month_start
    for a, b in sorted(covered):
        if b < month_start or a > month_end:
            continue  # skip ranges outside this month
        if a > current:
            gaps.append((current, min(b, month_end)))
        current = max(current, b + timedelta(days=1))
    if current <= month_end:
        gaps.append((current, month_end))
    return gaps


from collections import defaultdict

def get_month_colors(entries):
    month_status = defaultdict(set)
    for e in entries:
        from_date = e.get("from_date")
        status = e.get("status", "").strip().lower()
        if from_date:
            try:
                dt = datetime.strptime(from_date, "%d-%m-%Y")
                month = dt.strftime("%b")
                if status == "done":
                    month_status[month].add("green")
                else:
                    month_status[month].add("red")
            except:
                continue

    final_colors = {}
    for month, statuses in month_status.items():
        if "green" in statuses:
            final_colors[month] = "green"
        elif "red" in statuses:
            final_colors[month] = "red"
    return final_colors
# 🛑 Filter out suspended entries
from django.db.models import Q
from datetime import date
from collections import defaultdict


def is_suspended(entry, group_susp, client_susp, account_susp):
    acc_id = entry.account.id if entry.account else None
    client_id = entry.client.id if entry.client else None
    group_id = entry.client.group.id if entry.client and entry.client.group else None

    from_date = entry.from_date
    last_date = entry.last_date

    def overlaps(susp_list):
        for start, end in susp_list:
            # Overlap condition: any intersection between entry date range and suspension date range
            if start <= last_date and (end is None or from_date <= end):
                return True
        return False

    # ✅ Priority 1: Account-level
    if acc_id and overlaps(account_susp.get(acc_id, [])):
        print(f"🚫 Suspended by account: {acc_id}")
        return True

    # ✅ Priority 2: Client-level
    if client_id and overlaps(client_susp.get(client_id, [])):
        print(f"🚫 Suspended by client: {client_id}")
        return True

    # ✅ Priority 3: Group-level
    if group_id and overlaps(group_susp.get(group_id, [])):
        print(f"🚫 Suspended by group: {group_id}")
        return True

    return False


# missing client list
def normalize_ranges(ranges):
    """
    Merge overlapping or touching ranges [(a,b), ...] into minimal list.
    """
    if not ranges:
        return []
    ranges = sorted(ranges, key=lambda x: x[0])
    merged = [ranges[0]]
    for s,e in ranges[1:]:
        last_s, last_e = merged[-1]
        # Touching or overlapping
        if s <= (last_e + timedelta(days=1)):
            if e > last_e:
                merged[-1] = (last_s, e)
        else:
            merged.append((s,e))
    return merged



# summary report
from datetime import date

def merge_ranges(ranges):
    """Merge overlapping or contiguous [a,b] date ranges."""
    if not ranges:
        return []
    ranges = sorted(ranges, key=lambda ab: ab[0])
    merged = [list(ranges[0])]
    for a, b in ranges[1:]:
        la, lb = merged[-1]
        if a <= (lb + timedelta(days=1)):
            merged[-1][1] = max(lb, b)
        else:
            merged.append([a, b])
    return [(a, b) for a, b in merged]

def invert_ranges(window_start, window_end, covered):
    """Return the gaps within [window_start, window_end] given covered ranges."""
    covered = [(max(window_start, a), min(window_end, b))
               for a, b in covered if b >= window_start and a <= window_end]
    covered = merge_ranges(covered)
    gaps = []
    cur = window_start
    for a, b in covered:
        if cur < a:
            gaps.append((cur, a - timedelta(days=1)))
        cur = max(cur, b + timedelta(days=1))
    if cur <= window_end:
        gaps.append((cur, window_end))
    return gaps

def rng_list_to_str(ranges):
    """Format as 'dd-mm-YYYY to dd-mm-YYYY , ...'"""
    return " , ".join(f"{a.strftime('%d-%m-%Y')} to {b.strftime('%d-%m-%Y')}" for a, b in ranges)


def resolve_account_display(e, acc):
    """
    Returns (account_name, account_id)
    """
    if acc:
        return acc.account, acc.account_id

    # 🔁 Virtual fallback
    if e.virtual_account_type == "1":
        return "Sales", "VIRTUAL_SALES"
    elif e.virtual_account_type == "2":
        return "Purchase", "VIRTUAL_PURCHASE"

    return "", ""

# def pick_best_attachment(attachments):
#     if not attachments:
#         return None

#     PRIORITY = [
#         "application/pdf",
#         "application/vnd.ms-excel",
#         "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#         "application/zip",
#         "text/html",
#         "image/png",
#         "image/jpeg",
#     ]

#     for mime in PRIORITY:
#         for att in attachments:
#             ct = att.get("content_type") or att.get("mimeType") or ""
#             if ct.lower() == mime:
#                 return att

#     return attachments[0]



def ajax_report_rows(request):
    from calendar import monthrange
    from datetime import timedelta
    # Add this at the start of your view
    print("🔍 Request parameters:")
    for key, value in request.GET.items():
        print(f"  {key}: {value}")
    # today = date.today()
    print("🚨 ajax_report_rows called")
    print("📥 Request GET:", request.GET)
    g  = request.GET.get
    gl = request.GET.getlist
    fy = g("fy", "").replace("_", "-")  # Convert 2024_25 to 2024-25

    # ✅ Normalize FY once; get a proper DB alias with underscores
    fy_input = g("fy") or ""
    db, fy_display, fy_us = _normalize_fy(fy_input)   # e.g. db='fy_2024_25', fy_display='2024-25'
    if not db:
        return JsonResponse([], safe=False)

    from django.conf import settings
    if db not in settings.DATABASES:
        print(f"🚨 Error using DB alias: {db} (from '{fy_input}') → not configured")
        return JsonResponse([], safe=False)

    # ✅ Define months BEFORE any try/except uses them
    MONTH_MAP = {
        "Apr": 4, "May": 5, "Jun": 6, "Jul": 7,
        "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11,
        "Dec": 12, "Jan": 1, "Feb": 2, "Mar": 3,
    }
    from_month = (g("from_month") or "Apr").title()
    to_month   = (g("to_month") or "Mar").title()

    # ✅ One consistent date window for Apr→Mar
    from datetime import datetime
    from calendar import monthrange

    try:
        # If you already have fiscal_year_range(fy) that returns (start,end),
        # keep using it; pass the human-friendly '2024-25'
        fiscal_start, fiscal_end = fiscal_year_range(fy_display)
        y1, y2 = fiscal_start.year, fiscal_end.year

        if from_month and to_month:
            fm = MONTH_MAP[from_month]
            tm = MONTH_MAP[to_month]
            if fm <= tm:
                start_date = datetime(y1 if fm >= 4 else y2, fm, 1).date()
                end_date   = datetime(y1 if tm >= 4 else y2, tm, 1).date()
            else:
                start_date = datetime(y1, fm, 1).date()
                end_date   = datetime(y2, tm, 1).date()
            end_day  = monthrange(end_date.year, end_date.month)[1]
            end_date = end_date.replace(day=end_day)
        else:
            start_date, end_date = fiscal_start, fiscal_end
    except Exception as e:
        print("❌ Date window build error:", e)
        return JsonResponse({"rows": [], "error": f"Date filtering error: {e}"}, status=400)

    month_round_enabled = g("month_round") == "1"
    print("🟡 Month Round checkbox enabled:", month_round_enabled)

    try:
        group_id = g("group_id")
        client_id_str = g("client_id")
        account_id_str = g("account_id")
        account_groups = gl("account_groups")
        emp_id = g("emp")  # Get employee ID from request

        full_qs = DataEntry.objects.using(db).filter(
            from_date__isnull=False,
            last_date__isnull=False
        )

        # 📥 Apply same filters to full_qs as applied to qs (IMPORTANT for Only Missing logic)
        if group_id:
            client_ids = list(Client.objects.filter(group_id=group_id).values_list("id", flat=True))
            full_qs = full_qs.filter(client_id__in=client_ids)

        if client_id_str:
            client_obj = Client.objects.filter(client_id=client_id_str).first()
            if client_obj:
                full_qs = full_qs.filter(client_id=client_obj.id)

        if account_id_str:
            account_obj = AccountBank.objects.filter(account_id=account_id_str).first()
            if account_obj:
                full_qs = full_qs.filter(account_id=account_obj.id)

        if gl("account_groups"):
            valid_acc_ids = list(AccountBank.objects.filter(account_group__in=gl("account_groups")).values_list("id", flat=True))
            full_qs = full_qs.filter(account_id__in=valid_acc_ids)

         # Add employee filter if emp_id is provided
        if emp_id:
            # full_qs = full_qs.filter(alloted_to_id=emp_id)
            full_qs = full_qs.filter(alloted_to__id=emp_id)  # if alloted_to is ForeignKey


        # 📅 Apply date filtering to full_qs same as qs
        try:
            if from_month and to_month:
                fm = MONTH_MAP[from_month]
                tm = MONTH_MAP[to_month]
                fiscal_start, fiscal_end = fiscal_year_range(fy)
                y1, y2 = fiscal_start.year, fiscal_end.year

                if fm <= tm:
                    start_date = datetime(y1 if fm >= 4 else y2, fm, 1).date()
                    end_date = datetime(y1 if tm >= 4 else y2, tm, 1).date()
                else:
                    start_date = datetime(y1, fm, 1).date()
                    end_date = datetime(y2, tm, 1).date()

                end_day = monthrange(end_date.year, end_date.month)[1]
                end_date = end_date.replace(day=end_day)
            else:
                start_date, end_date = fiscal_year_range(fy)

            full_qs = full_qs.filter(from_date__gte=start_date, last_date__lte=end_date)
        except Exception as e:
            print("❌ Date filtering error for full_qs:", e)

        print("🔍 Total filtered full_qs count (used for missing rows):", full_qs.count())

        # Use filtered full_qs for both entries_by_pair and backend month coloring
        qs = full_qs
        print("✅ Using DB:", db)
        print("🔍 Total DataEntry count:", full_qs.count())
        qs = full_qs
    except Exception as e:
        print("🚨 Error using DB:", db, e)
        return JsonResponse([], safe=False)
    entry_id = g("entry_id")
    highlight_id = g("highlight_id")

    if entry_id and not highlight_id:
        qs = qs.filter(id=entry_id)
        print("🔎 entry_id provided: skipping other filters")

    else:
        group_id = g("group_id")
        client_id_str = g("client_id")
        account_id_str = g("account_id")

        print("📥 Received filters → Group:", group_id, "| Client:", client_id_str, "| Account:", account_id_str)

        # Filter by Group
        if group_id:
            client_ids = list(Client.objects.filter(group_id=group_id).values_list("id", flat=True))
            qs = qs.filter(client_id__in=client_ids)
            print("🔎 After group filter:", qs.count())
        if client_id_str:
            client_obj = Client.objects.filter(client_id=client_id_str).first()

            if client_obj:
                qs = qs.filter(client_id=client_obj.id)
                print("🔎 After client filter:", qs.count())
            else:
                print("❌ Client not found")
        
        # Filter by Account
        if account_id_str:
            account_obj = AccountBank.objects.filter(account_id=account_id_str).first()
            if account_obj:
                qs = qs.filter(account_id=account_obj.id)
                print("🔎 After account filter:", qs.count())
            else:
                print("❌ Account not found")

        # DATE RANGE FILTER
        MONTH_MAP = {
            "Apr": 4, "May": 5, "Jun": 6, "Jul": 7,
            "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11,
            "Dec": 12, "Jan": 1, "Feb": 2, "Mar": 3,
        }

        from_month = g("from_month")
        to_month = g("to_month")

        try:
            if from_month and to_month:
                fm = MONTH_MAP[from_month]
                tm = MONTH_MAP[to_month]
                fiscal_start, fiscal_end = fiscal_year_range(fy)
                y1, y2 = fiscal_start.year, fiscal_end.year

                if fm <= tm:
                    start_date = datetime(y1 if fm >= 4 else y2, fm, 1).date()
                    end_date = datetime(y1 if tm >= 4 else y2, tm, 1).date()
                else:
                    start_date = datetime(y1, fm, 1).date()
                    end_date = datetime(y2, tm, 1).date()

                end_day = monthrange(end_date.year, end_date.month)[1]
                end_date = end_date.replace(day=end_day)
            else:
                start_date, end_date = fiscal_year_range(fy)

            print(f"📅 Date range: {start_date} to {end_date}")
            qs = qs.filter(from_date__isnull=False, last_date__isnull=False)
            qs = qs.filter(from_date__gte=start_date, last_date__lte=end_date)

        except Exception as e:
            print("❌ Date filtering error:", e)

        # EXTRA FILTERS
        if g("alloted") == "Alloted":
            qs = qs.exclude(alloted_to__isnull=True)
        elif g("alloted") == "Not Alloted":
            qs = qs.filter(alloted_to__isnull=True)

        # if g("status") == "Done":
        #     qs = qs.exclude(done_date__isnull=True)
        # elif g("status") == "Pending":
        #     qs = qs.filter(done_date__isnull=True)

        if g("status") == "Done":
            qs = qs.filter(done_date__isnull=False)

        elif g("status") == "Pending":
            qs = qs.filter(done_date__isnull=True).exclude(status="Hold")


    
        

    
        # # ⭐ SPECIAL CASE: No group AND no client selected → show ALL pending entries
        # if not group_id and not client_id_str:
        #     print("🟡 No group or client selected → Show ALL pending entries")
        #     qs = full_qs  # reset to all filtered entries
        #     if g("status") == "Pending":
        #         qs = qs.filter(done_date__isnull=True).exclude(status="Hold")
        #     elif g("status") == "Done":
        #         qs = qs.exclude(done_date__isnull=True)
        #     elif g("status") == "Hold":
        #         qs = qs.filter(status="Hold")
            # If status = All → do nothing

        # status_filter = g("status")  # "All" / "Done" / "Pending" / "Hold"

        # if status_filter == "Done":
        #     # Done = has done_date
        #     # qs = qs.exclude(done_date__isnull=True)
        #     qs = qs.filter(status="Done")

        # elif status_filter == "Pending":
        #     # Pending = not done yet AND not on Hold
        #     # qs = qs.filter(done_date__isnull=True).exclude(status="Hold")
        #     qs = qs.filter(status="Pending")

        # elif status_filter == "Hold":
        #     # Only entries explicitly marked as Hold
        #     qs = qs.filter(status="Hold")
        # else: "All" → no extra filter

        if g("entry_type") == "Import":
            qs = qs.filter(imported=True)
        elif g("entry_type") == "Manual":
            qs = qs.filter(imported=False)

        if g("busy_tran") == "true":
            qs = qs.exclude(received_by="")

        if g("acstkameet") == "true":
            qs = qs.filter(received_by="Acstkameet")

        if g("e_stmt") == "true":
            qs = qs.filter(account__e_statement=True)

        # report_type = g("report_type")
        # include_real_rows = True


        # if report_type == "Only Received":
        #     qs = qs.filter(rec_date__isnull=False)
        #     include_virtual_rows = False  # Block red virtual rows

        # elif report_type == "Only Missing":
        #     # ✅ clear already collected green rows
        #     data = []
        #     include_real_rows = False
        #     include_virtual_rows = True
        # else:
        #     include_virtual_rows = True

        report_type = g("report_type")

        # ✅ Default behavior
        include_real_rows = True
        include_virtual_rows = True
        include_full_red = True

        # ✅ ONLY RECEIVED
        if report_type == "Only Received":
            qs = qs.filter(rec_date__isnull=False)
            include_virtual_rows = False
            include_full_red = False

        # ✅ ONLY MISSING
        elif report_type == "Only Missing":
            include_real_rows = False   # Hide real
            include_virtual_rows = True
            include_full_red = True

        # ✅ ALL
        else:
            include_real_rows = True
            include_virtual_rows = True
            include_full_red = True


        account_groups = gl("account_groups")
        print("➡️ Received account_groups[]:", account_groups)

        if account_groups:
            valid_acc_ids = list(AccountBank.objects.filter(account_group__in=account_groups).values_list("id", flat=True))
            qs = qs.filter(account_id__in=valid_acc_ids)
            print("🔎 After account group filter:", qs.count())


        # NEW: pending detail filters from modal
        # pending_mode   = request.GET.get("pending_mode")
        # client_ids_str = request.GET.get("client_ids", "").strip()
        # account_ids_str = request.GET.get("account_ids", "").strip()
        # # column_select  = request.GET.get("column_select", "all")
        # column_select = request.GET.get("column_select")

        # if pending_mode == "1":
        #     if client_ids_str:
        #         client_ids = [c.strip() for c in client_ids_str.split(",") if c.strip()]
        #         qs = qs.filter(account__client__client_id__in=client_ids)

        #     if account_ids_str:
        #         account_ids = [int(a) for a in account_ids_str.split(",") if a.strip()]
        #         qs = qs.filter(account_id__in=account_ids)

        #     # column_select -> decide which rows to show
        #     if column_select == "received":
        #         qs = qs.filter(rec_date__isnull=False)

        #         # qs = qs.filter(rec_date__isnull=False).exclude(rec_date="")
        #         # qs = qs.filter(received__isnull=False).exclude(received="")
        #     elif column_select == "missing":
        #         qs = qs.filter(rec_date__isnull=True)

                # qs = qs.filter(missings__isnull=False).exclude(missings="")



    # ✅ At this point qs has all filters applied → NOW apply suspend_date
    print(f"🔍 Before suspend filter: {qs.count()} rows")

    group_susp = defaultdict(list)
    client_susp = defaultdict(list)
    account_susp = defaultdict(list)

    for s in TaxSuspension.objects.filter(is_active=True):
        start = s.from_date
        end = s.to_date

        if s.account_id:
            account_susp[s.account_id].append((start, end))
        elif s.client_id:
            client_susp[s.client_id].append((start, end))
        elif s.group_id:
            group_susp[s.group_id].append((start, end))

    # ✅ Apply the suspension filter
    print(f"🚫 After suspension filter: {len(qs)} rows remain")
    
    print("✅ Final DB:", db)
   
    # qs = [
    #     e for e in qs
    #     if not is_suspended(e, group_susp, client_susp, account_susp)
    # ]

    if g("gap_report") != "1":
        qs = [
            e for e in qs
            if not is_suspended(e, group_susp, client_susp, account_susp)
        ]

    # BUILD RESPONSE
    data = []
    existing_months = set()

    # entries_by_pair = {}
    # account_month_status = {}
    entries_by_pair = defaultdict(list)
    account_month_status = defaultdict(dict)

    
    # for e in full_qs:
        
    #     if not e.from_date or not e.last_date:
    #         continue

    #     acc = e.account
    #     acc_id = acc.account_id if acc else None
    #     if not acc_id:
    #         continue

    for e in full_qs:

        if not e.from_date or not e.last_date:
            continue

        # 🔴 SAFE FK ACCESS
        try:
            acc = e.account
        except AccountBank.DoesNotExist:
            # continue
              acc = None        
        # if not acc:
        #     continue

        # acc_id = acc.account_id
        account_name, acc_id = resolve_account_display(e, acc)
        if not acc_id:
                continue
        # acc_id = acc.account_id


        # loop through all months covered
        from_date = e.from_date
        last_date = e.last_date
        loop_date = from_date.replace(day=1)
        end_month = last_date.replace(day=1)

        while loop_date <= end_month:
            month_name = loop_date.strftime("%b")
            full_month_days = monthrange(loop_date.year, loop_date.month)[1]
            this_month_end = loop_date.replace(day=full_month_days)

            # full coverage?
            covers_full_month = (
                loop_date >= from_date.replace(day=1) and
                this_month_end <= last_date
            )

            if covers_full_month and (e.rec_date or e.status == "Done"):
                account_month_status[acc_id][month_name] = "green"
            elif covers_full_month:
                if account_month_status[acc_id].get(month_name) != "green":
                    account_month_status[acc_id][month_name] = "red"
            else:
                if account_month_status[acc_id].get(month_name) not in ("green", "red"):
                    account_month_status[acc_id][month_name] = "yellow"

            # move to next month
            if loop_date.month == 12:
                loop_date = loop_date.replace(year=loop_date.year + 1, month=1)
            else:
                loop_date = loop_date.replace(month=loop_date.month + 1)

        
            # ✅ Store real covered date ranges for later missing-row calculation
        # if e.client and e.account:
        #     entries_by_pair[(e.client.id, e.account.id)].append((from_date, last_date))
        try:
            client = e.client
            account = e.account
        except AccountBank.DoesNotExist:
            continue

        if client and account:
            entries_by_pair[(client.id, account.id)].append((from_date, last_date))


    # if report_type != "Only Received":
    if include_full_red:

        # 1️⃣ Get filtered clients based on filters (group / client)
        filtered_clients = Client.objects.all()

        group_id = g("group_id")
        client_id_str = g("client_id")

        if group_id:
            filtered_clients = filtered_clients.filter(group_id=group_id)

        if client_id_str:
            filtered_clients = filtered_clients.filter(client_id=client_id_str)

        # 2️⃣ Clients that HAVE entries
        clients_with_entries = {cid for (cid, _) in entries_by_pair.keys()}

        # 3️⃣ Clients WITHOUT entries
        clients_without_entries = filtered_clients.exclude(id__in=clients_with_entries)

        print("🟥 Clients WITHOUT entries this FY:",
            list(clients_without_entries.values_list("client_name", flat=True)))

        # 4️⃣ Generate FULL RED rows
        for client in clients_without_entries:

            accounts = AccountBank.objects.filter(client=client)

            for acc in accounts:

                # ✅ Mark all months RED
                for m in ["Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec","Jan","Feb","Mar"]:
                    account_month_status[acc.account_id][m] = "red"

                # ✅ Add a single full-year missing row
                data.append({
                    "id": f"missing_full_{client.id}_{acc.id}",
                    "client_pk": client.id,
                    "account_pk": acc.id,
                    "client": client.client_name,
                    "account": acc.account,
                    "account_id": acc.account_id,
                    "account_group": acc.account_group,
                    "group": client.group.group_name if client.group else "",
                    "received_date": "",
                    "format": "",
                    "rece_by": "",
                    "from_date": start_date.strftime("%d-%m-%Y"),
                    "last_date": end_date.strftime("%d-%m-%Y"),
                    "status": "",
                    "e_stmt": "Yes" if acc.e_statement else "No",
                    "alloted_to": "",
                    "file_path": "",
                    "file_name": "",
                    "folder_path": "",
                    "password": acc.stms_pws or "",
                    "file_url": "",
                    "fy": fy,
                    "alloted_date": "",
                    "done_date": "",
                    "query": ""
                })

                print(f"🟥 FULL RED → {client.client_name} / {acc.account}")


    
    # GAP REPORT (≥ min_gap_days, default 30)
    # -----------------------------
    if g("gap_report") == "1":
        try:
            min_gap_days = int(g("min_gap_days") or "30")
        except ValueError:
            min_gap_days = 30

        # Use the same date window you already computed
        window_start = start_date
        # Don't look beyond "today"
        today = date.today()
        window_end = min(end_date, today)

        gaps_payload = []

        # (Optional) speed-up maps to avoid N+1
        acc_ids = {acc_id for (_, acc_id) in entries_by_pair.keys()}
        cli_ids = {cli_id for (cli_id, _) in entries_by_pair.keys()}
        client_map = {c.id: c for c in Client.objects.filter(id__in=cli_ids)}
        account_map = {a.id: a for a in AccountBank.objects.filter(id__in=acc_ids)}

                # -------------------------------------------------
        # 🟥 FULL YEAR GAP FOR CLIENT / ACCOUNT WITH NO ENTRY
        # -------------------------------------------------

        # 1️⃣ Get all relevant clients (respect filters)
        all_clients = Client.objects.all()

        if g("group_id"):
            all_clients = all_clients.filter(group_id=g("group_id"))

        if g("client_id"):
            all_clients = all_clients.filter(client_id=g("client_id"))

        # 2️⃣ Get all relevant accounts
        all_accounts = AccountBank.objects.filter(client__in=all_clients)

        if gl("account_groups"):
            all_accounts = all_accounts.filter(account_group__in=gl("account_groups"))

        # 3️⃣ Existing (client, account) pairs that HAVE entries
        existing_pairs = set(entries_by_pair.keys())

        # 4️⃣ Find pairs with NO entries at all
        for acc in all_accounts:
            client = acc.client
            if not client:
                continue

            pair_key = (client.id, acc.id)

            if pair_key in existing_pairs:
                continue  # already handled above

            # ✅ FULL YEAR GAP
            gaps_payload.append({
                "client": client.client_name,
                "group": client.group.group_name if client.group else "",
                "account": acc.account,
                "account_id": acc.account_id,
                "from_date": window_start.strftime("%d-%m-%Y"),
                "to_date": window_end.strftime("%d-%m-%Y"),
                "from_month": window_start.strftime("%b"),
                "to_month": window_end.strftime("%b"),
                "days_missing": (window_end - window_start).days + 1,
                "fy": fy_display,
                "_from_sort": window_start.isoformat(),
            })


        for (client_pk, account_pk), covered_ranges in entries_by_pair.items():
            # Normalize coverage into window
            covered = normalize_ranges([
                (max(window_start, a), min(window_end, b))
                for a, b in covered_ranges
                if b >= window_start and a <= window_end
            ])

            # Compute gaps in [window_start, window_end]
            if not covered:
                raw_gaps = [(window_start, window_end)]
            else:
                raw_gaps = []
                cursor = window_start
                for a, b in covered:
                    if cursor < a:
                        raw_gaps.append((cursor, a - timedelta(days=1)))
                    cursor = max(cursor, b + timedelta(days=1))
                if cursor <= window_end:
                    raw_gaps.append((cursor, window_end))

            # Keep only long-enough gaps
            long_gaps = [(s, e) for (s, e) in raw_gaps if (e - s).days + 1 >= min_gap_days]
            if not long_gaps:
                continue

            client = client_map.get(client_pk)
            account = account_map.get(account_pk)
            # 🔴 VERY IMPORTANT FIX
            if not client or not account:
                continue
            client_name  = client.client_name if client else ""
            account_name = account.account if account else ""
            account_id   = account.account_id if account else ""
            group_name   = client.group.group_name if (client and client.group) else ""

            for s, e in long_gaps:
                gaps_payload.append({
                    "client": client_name,
                    "group": group_name,
                    "account": account_name,
                    "account_id": account_id,
                    "from_date": s.strftime("%d-%m-%Y"),
                    "to_date": e.strftime("%d-%m-%Y"),
                    "from_month": s.strftime("%b"),
                    "to_month": e.strftime("%b"),
                    "days_missing": (e - s).days + 1,
                    "fy": fy_display,       # or fy, but be consistent
                    "_from_sort": s.isoformat(),  # for correct sorting
                })

        gaps_payload.sort(key=lambda d: (d["group"], d["client"], d["account"], d["_from_sort"]))
        for row in gaps_payload:
            row.pop("_from_sort", None)

        return JsonResponse({"gaps": gaps_payload})
    
    # -----------------------------
    # Table Rows → Only filtered qs
    # -----------------------------
    # data = []
    for e in qs:

        print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
        print("ENTRY ID:", e.id)
        print("FORMAT:", e.format)
        print("RAW MSG_ID:", e.msg_id)

        print("CLEAN:", (e.msg_id or "").strip().lower())

        print("DB USED:", db)

        mtest = MailLog.objects.using(db).filter(msg_id__icontains=(e.msg_id or "")).first()
        print("DIRECT TEST MAIL:", mtest)

        # print("FINAL STATEMENT LINK:", statement)

          # ✅ Skip real rows if Only Missing
        # if report_type == "Only Missing":
        #     continue
        if not include_real_rows:
            continue

        # ✅ Double protection: Only Received must NOT include black rows
        if report_type == "Only Received" and not e.rec_date:
            continue
        # acc = e.account
        try:
            acc = e.account
        except AccountBank.DoesNotExist:
            acc = None

        account_name, acc_id = resolve_account_display(e, acc)
        if not acc_id:
            continue
            # continue
        raw = (e.msg_id or "")
      
        import re

        # statement = ""

        # raw = (e.msg_id or "").strip()
        # clean = raw.strip("<>").lower()

        # print("🔍 RAW MSG_ID:", raw)
        # print("✅ CLEAN MSG_ID:", clean)

        # # exact match first
        # m = MailLog.objects.using(db).filter(
        #     msg_id__iexact=clean
        # ).order_by("-fetched_at").first()

        # # fallback: stripped match
        # if not m:
        #     m = MailLog.objects.using(db).filter(
        #         msg_id__icontains=clean
        #     ).order_by("-fetched_at").first()

        # print("📄 MAIL FOUND:", m)

        # if m:
        #     statement = m.statement_link or ""
        statement = ""

        raw = (e.msg_id or "").strip()
        clean = raw.strip("<>").lower()

        print("🔎 RAW MSG_ID:", raw)
        print("✅ CLEAN MSG_ID:", clean)

        m = MailLog.objects.using(db).filter(
            msg_id__icontains=clean
        ).exclude(statement_link__isnull=True).exclude(statement_link="") \
        .order_by("-fetched_at") \
        .first()

        print("📄 FINAL MAIL:", m)

        if m:
            statement = m.statement_link.strip()
            print("✅ FINAL LINK:", statement)
        else:
            print("❌ No MailLog row WITH statement_link found")

        account_name, acc_id = resolve_account_display(e, acc)

        data.append({
            "id": e.id,
            "client_pk": e.client.id if e.client else None,
            "account_pk": e.account.id if e.account else None,
            "client": e.client.client_name if e.client else "",
            "account": account_name,            # ✅ Sales / Purchase
            "account_id": acc_id,                # ✅ VIRTUAL_SALES / VIRTUAL_PURCHASE
            "virtual_account_type": e.virtual_account_type or "",

            "account_group": acc.account_group if acc else "",
            "group": e.client.group.group_name if e.client and e.client.group else "",
            "received_date": e.rec_date.strftime("%d-%m-%Y") if e.rec_date else "",
            "format": e.format,
            "rece_by": e.received_by or "",
            "from_date": e.from_date.strftime("%d-%m-%Y") if e.from_date else "",
            "last_date": e.last_date.strftime("%d-%m-%Y") if e.last_date else "",
            "status": e.status or "",
            "e_stmt": "Yes" if acc and acc.e_statement else "No",
            "alloted_to": e.alloted_to.username if e.alloted_to else "",
            "file_path": str(e.attach_file) if e.attach_file else "",
            "file_name": os.path.basename(str(e.attach_file)) if e.attach_file else "",
            "folder_path": os.path.dirname(str(e.attach_file)) if e.attach_file else "",
            "file_url": os.path.basename(str(e.attach_file)) if e.attach_file else "",
            "password": acc.stms_pws if acc else "",
            "fy": fy,
            "alloted_date": e.alloted_date.strftime("%d-%m-%Y") if e.alloted_date else "",
            "done_date": e.done_date.strftime("%d-%m-%Y") if e.done_date else "",
            "query": e.query or "",
            "msg_id": e.msg_id or "",
            "busy_code": e.client.busy_code if e.client else "",
            "busyacccode": acc.busyacccode if acc else "",
            "statement_link": statement,
        })          
    # ✅ Clean virtual block generator
    # if include_virtual_rows:
    if include_virtual_rows and report_type != "Only Received":

        full_months = get_fy_month_ranges(start_date, end_date)

        for (client_id, account_id), actual_ranges in entries_by_pair.items():
            covered = actual_ranges
            client = Client.objects.filter(id=client_id).first()
            account = AccountBank.objects.filter(id=account_id).first()
            if not month_round_enabled:
            # ✅ Add missing partial block per month
                for from_date, to_date in covered:
                    try:
                        month_start = from_date.replace(day=1)
                        month_end = from_date.replace(day=monthrange(from_date.year, from_date.month)[1])
                        gaps = subtract_covered_ranges(month_start, month_end, covered)

                        for gap_start, gap_end in gaps:
                            already_exists = any(
                                d["from_date"] == gap_start.strftime("%d-%m-%Y") and
                                d["last_date"] == gap_end.strftime("%d-%m-%Y") and
                                d["client"] == (client.client_name if client else "") and
                                d["account_id"] == (account.account_id if account else "")
                                for d in data
                            )
                            if not already_exists:
                                data.append({
                                    "id": f"missing_partial_{client_id}_{account_id}_{gap_start.month}_{gap_start.day}",
                                    "client_pk": client.id if client else None,
                                    "account_pk": account.id if account else None,
                                    "client": client.client_name if client else "",
                                    "account": account.account if account else "",
                                    "account_id": account.account_id if account else "",
                                    "account_group": account.account_group if account else "",
                                    "group": client.group.group_name if client and client.group else "",
                                    "received_date": "",
                                    "format": "",
                                    "rece_by": "",
                                    "from_date": gap_start.strftime("%d-%m-%Y"),
                                    "last_date": gap_end.strftime("%d-%m-%Y"),
                                    "status": "",
                                    "e_stmt": "Yes" if account and account.e_statement else "No",
                                    "alloted_to": "",
                                    "file_path": "",
                                    "file_name": "",
                                    "folder_path": "",
                                    "password": account.stms_pws if account else "",
                                    "file_url": "",
                                    "fy": fy,
                                    "alloted_date": "",
                                    "done_date": "",
                                    "query": ""
                                })
                                    # ✅ Also mark this gap’s month as red
                                account_month_status[account.account_id][gap_start.strftime("%b")] = "red"


                    except Exception as e:
                        print("❌ Partial month calc error:", e)

            # ✅ Add full-month missing blocks
            missing_blocks = []
            temp_start = None
            previous_mend = None

            for mstart, mend in full_months:
                # ✅ If nothing is covered, entire month is missing
                if not covered:
                    month_gaps = [(mstart, mend)]
                else:
                    month_gaps = subtract_covered_ranges(mstart, mend, covered)
                # month_gaps = subtract_covered_ranges(mstart, mend, covered)
                is_covered = len(month_gaps) == 0
                if not is_covered:
                    # Skip full month red block if any partial red rows already added
                    partial_already_exists = any(
                        d["from_date"] == gap_start.strftime("%d-%m-%Y") and
                        d["last_date"] == gap_end.strftime("%d-%m-%Y") and
                        d["account_id"] == (account.account_id if account else "")
                        for gap_start, gap_end in month_gaps
                        for d in data
                    )
                    if partial_already_exists and not month_round_enabled:
                        print(f"⚠️ Skipping full red row for {mstart}–{mend} (partials already present and month_round OFF)")
                        continue

                    if temp_start is None:
                        temp_start = mstart
                    previous_mend = mend

                else:
                    if temp_start:
                        # 🔁 Before adding, double check entire range isn't covered
                        is_block_covered = any(
                            a <= temp_start and b >= previous_mend
                            for a, b in covered
                        )
                        print(f"🔎 Final merge block check {temp_start} to {previous_mend} → {'COVERED' if is_block_covered else 'ADD MISSING'}")
                        if not is_block_covered:
                            missing_blocks.append((temp_start, previous_mend))
                        temp_start = None
                        previous_mend = None

            # After loop: check if last collected block should be added
            if temp_start:
                is_block_covered = any(
                    a <= temp_start and b >= previous_mend
                    for a, b in covered
                )
                print(f"🔎 Final trailing block check {temp_start} to {previous_mend} → {'COVERED' if is_block_covered else 'ADD MISSING'}")
                if not is_block_covered:
                    missing_blocks.append((temp_start, previous_mend))

            # ✅ Generate red data rows for missing blocks
            for mstart, mend in missing_blocks:
                already_exists = any(
                    d["from_date"] == mstart.strftime("%d-%m-%Y") and
                    d["last_date"] == mend.strftime("%d-%m-%Y") and
                    d["client"] == (client.client_name if client else "") and
                    d["account_id"] == (account.account_id if account else "")
                    for d in data
                )
                if already_exists:
                    print(f"⚠️ Skipping duplicate virtual row for {mstart} to {mend}")
                    continue

                print(f"🟥 Adding virtual red row: {mstart} → {mend}")

                data.append({
                    "id": f"missing_{client_id}_{account_id}_{mstart.month}",
                    "client_pk": client.id if client else None,
                    "account_pk": account.id if account else None,
                    "client": client.client_name if client else "",
                    "account": account.account if account else "",
                    "account_id": account.account_id if account else "",
                    "account_group": account.account_group if account else "",
                    "group": client.group.group_name if client and client.group else "",
                    "received_date": "",
                    "format": "",
                    "rece_by": "",
                    "from_date": mstart.strftime("%d-%m-%Y"),
                    "last_date": mend.strftime("%d-%m-%Y"),
                    "status": "",
                    "e_stmt": "Yes" if account and account.e_statement else "No",
                    "alloted_to": "",
                    "file_path": "",
                    "file_name": "",
                    "folder_path": "",
                    "password": account.stms_pws if account else "",
                    "file_url": "",
                    "fy": fy,
                    "alloted_date": "",
                    "done_date": "",
                    "query": ""
                })
                account_month_status[account.account_id][mstart.strftime("%b")] = "red"

       # ✅ After processing qs and virtual rows ...

    # ✅ Ensure all 12 months exist for each account
    all_fy_months = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]
    for acc_id, months in account_month_status.items():
        for m in all_fy_months:
            if m not in months:
                months[m] = "red"

            # ✅ DEBUG: See final backend colors before sending
    import json
    # Suppose you have a dict mapping account_id → account name and client name
    account_names = {acc.id: acc.account for acc in AccountBank.objects.all()}
    client_names = {acc.id: acc.client.client_name for acc in AccountBank.objects.select_related('client')}
    print("📤 Final month_status being sent to frontend:")
    readable_status = {}
    for acc_id, months in account_month_status.items():
        readable_status[f"{client_names.get(acc_id, 'Unknown Client')} → {account_names.get(acc_id, 'Unknown Account')}"] = months
    print(json.dumps(readable_status, indent=2))
    
    # summary report
    # -----------------------------
    # SUMMARY VIEW
    # -----------------------------
    view_kind = g("view", "detail")
    if view_kind == "summary":
        # Window already computed above: start_date, end_date
        # entries_by_pair already built above while looping full_qs:
        #    entries_by_pair[(e.client.id, e.account.id)].append((from_date, last_date))

        # prefetch maps
        cli_ids = {cli for (cli, _) in entries_by_pair.keys()}
        acc_ids = {acc for (_, acc) in entries_by_pair.keys()}
        client_map = {c.id: c for c in Client.objects.filter(id__in=cli_ids)}
        account_map = {a.id: a for a in AccountBank.objects.filter(id__in=acc_ids)}

        summary_rows = []
        for (client_pk, account_pk), covered in entries_by_pair.items():
            client = client_map.get(client_pk)
            account = account_map.get(account_pk)
            if not client or not account:
                continue

            # Merge covered ranges and compute gaps within the selected window
            merged = merge_ranges(covered)
            received = [(max(start_date, a), min(end_date, b))
                        for a, b in merged if b >= start_date and a <= end_date]
            received = [(a, b) for a, b in received if a <= b]

            gaps = invert_ranges(start_date, end_date, covered)

            # Format strings
            received_str = rng_list_to_str(received) if received else ""
            gaps_str     = rng_list_to_str(gaps)     if gaps else ""
            # mail = MailLog.objects.filter(msg_id=entry.msg_id).first()

            summary_rows.append({
                "client": client.client_name,
                "account": account.account,
                "format": (getattr(account, "statement_format", None) or getattr(account, "statement_type", None) or "").lower() or (account.statement_type if hasattr(account,"statement_type") else ""),
                "received": received_str,
                "missings": gaps_str,
                "client_pk": client_pk,
                "account_pk": account_pk,
                # "link": mail.statement_link if mail else "",
                
            })

        # Nice stable sort
        summary_rows.sort(key=lambda r: (r["client"], r["account"]))

        return JsonResponse({
            "summary": summary_rows,
            "month_status": account_month_status,  # keep if you still color month pills
        })

    print("JSON PREVIEW:", data[:3])

    return JsonResponse({
        "rows": data,
        "month_status": account_month_status,  # ⬅ Add this
    })


# views.py
import base64
import re

from django.http import HttpResponse, HttpResponseBadRequest, HttpResponseNotFound
from django.shortcuts import redirect

from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build

SCOPES = ["https://www.googleapis.com/auth/gmail.readonly"]

def _decode_part_body(body_dict):
    """Helper: decode Gmail base64 body to text."""
    data = body_dict.get("data")
    if not data:
        return ""
    # Gmail uses URL-safe base64
    missing_padding = 4 - (len(data) % 4)
    if missing_padding and missing_padding < 4:
        data += "=" * missing_padding
    try:
        return base64.urlsafe_b64decode(data).decode("utf-8", "ignore")
    except Exception:
        return ""


def _find_html_part(payload):
    """
    Recursively find the first text/html part in Gmail payload.
    """
    if not payload:
        return ""

    mime_type = payload.get("mimeType", "")
    body = payload.get("body", {})

    if mime_type == "text/html":
        return _decode_part_body(body)

    # multipart: check inner parts
    for part in (payload.get("parts") or []):
        html = _find_html_part(part)
        if html:
            return html

    return ""


def open_mail_and_extract_link(request):
    msg_id = request.GET.get("msg_id", "").strip()

    if not msg_id:
        return HttpResponseBadRequest("Missing msg_id")

    # 🔹 Load Gmail credentials (must run OAuth script once to create this file)
    try:
        creds = Credentials.from_authorized_user_file("gmail_token.json", SCOPES)
    except Exception:
        return HttpResponse(
            "gmail_token.json not found. Run your Gmail OAuth script once on this server.",
            status=500,
        )

    service = build("gmail", "v1", credentials=creds)

    try:
        msg = (
            service.users()
            .messages()
            .get(userId="me", id=msg_id, format="full")
            .execute()
        )
    except Exception as e:
        return HttpResponseNotFound(f"Failed to fetch mail for msg_id={msg_id}: {e}")

    # 🔍 Get HTML body
    html_body = _find_html_part(msg.get("payload"))
    if not html_body:
        return HttpResponseNotFound("No HTML body found in this mail.")

    # 🔍 Extract first https link (you can tighten this regex to your bank domain if needed)
    # Example: href="https://something.com/SmartStatement/xyz..."
    link_match = re.search(r'href="(https?://[^"]+)"', html_body, flags=re.IGNORECASE)

    if not link_match:
        return HttpResponseNotFound("No statement link found in this mail body.")

    statement_url = link_match.group(1)

    # ✅ Redirect browser directly to the statement page (like Shell Chrome ... -url <url>)
    return redirect(statement_url)








# pending details summary
from django.http import JsonResponse
from django.db.models import Q
from .models import Client, AccountBank

ALLOWED_ACCOUNT_GROUPS = ["Bank Accounts", "Credit Card","Bank O/D Account"]

def group_clients(request):
    group_id = request.GET.get("group_id")
    if not group_id:
        return JsonResponse({"ok": False, "error": "group_id required"}, status=400)

    qs = Client.objects.filter(group_id=group_id).order_by("client_name")
    data = [{"client_id": c.client_id, "client_name": c.client_name} for c in qs]
    return JsonResponse({"ok": True, "clients": data})


from accounts.utills import get_db_for_fy

def _build_pending_qs(request):
    fy = request.GET.get("fy")
    db_alias = get_db_for_fy(fy) if fy else get_db_for_fy("2024-25")

    account_ids_str = (request.GET.get("account_ids") or "").strip()
    account_ids = [int(a) for a in account_ids_str.split(",") if a.strip()]

    # ⚠️ IMPORTANT: do NOT return early
    qs = DataEntry.objects.using(db_alias).filter(account_id__in=account_ids)

    column_select = (request.GET.get("column_select") or "all").lower()
    if column_select == "received":
        qs = qs.filter(rec_date__isnull=False)
    elif column_select == "missing":
        qs = qs.filter(rec_date__isnull=True)

    return qs, account_ids, db_alias, fy


def pending_detail_excel(request):
    # qs = _build_pending_qs(request)
    rows, fy_display = _build_pending_ranges(request)

    

    grouped = {}

    for r in rows:
        client  = r["client"]
        account = r["account"]
        period  = f'{r["from_date"]} To {r["to_date"]}'
        kind    = r["kind"]  # "received" or "missing"

        if client not in grouped:
            grouped[client] = {}

        if account not in grouped[client]:
            grouped[client][account] = {
                "received": [],
                "missing": []
            }

        grouped[client][account][kind].append(period)


    selected_group_name = request.GET.get("group_name", "Statement Report")
    column_select = (request.GET.get("column_select") or "all").capitalize()

    headers = ["Client Name", "Account Name", "Statement Period", "Status"]


    html = []
    html.append('<table style="width:100%; border-collapse:collapse;">')

    html.append(f"""
    <tr>
    <td colspan="{len(headers)}"
        style="font-weight:bold; font-size:16px; padding:8px; border:none;">
        {request.GET.get("group_name", "Statement Report")} – {column_select.capitalize()} Summary
    </td>
    </tr>
    """)

    html.append('<tr style="background:#002060; color:#fff; font-weight:bold;">')
    for h in headers:
        html.append(f'<th style="border:1px solid black; padding:5px;">{h}</th>')
    html.append('</tr>')


    for client, accounts in grouped.items():
        # total_rows = sum(len(stmts) for stmts in accounts.values())
        total_rows = sum(
            len(v["received"]) + len(v["missing"])
            for v in accounts.values()
        )

        client_written = False

        # for acc_name, stmts in accounts.items():
        #     rowspan_acc = len(stmts)
        for acc_name, stmts in accounts.items():
            all_rows = (
                [(p, "Received") for p in stmts["received"]] +
                [(p, "Missing") for p in stmts["missing"]]
            )

            if not all_rows:
                continue

            rowspan_acc = len(all_rows)  
            
            # for idx, stmt in enumerate(all_rows):
            #     row_html = "<tr>"

            #     if not client_written:
            #         row_html += (
            #             f'<td style="border:1px solid black; padding:5px; font-weight:bold;" '
            #             f'rowspan="{total_rows}">{client}</td>'
            #         )
            #         client_written = True

            #     if idx == 0:
            #         row_html += (
            #             f'<td style="border:1px solid black; padding:5px;" '
            #             f'rowspan="{rowspan_acc}">{acc_name}</td>'
            #         )

            #     # row_html += f'<td style="border:1px solid black; padding:5px;">{stmt}</td>'
            #     color = "#198754" if status == "Received" else "#dc3545"

            #     row_html += f'''
            #     <td style="border:1px solid black; padding:5px;">{stmt}</td>
            #     <td style="border:1px solid black; padding:5px; font-weight:bold; color:{color};">
            #         {status}
            #     </td>
            #     row_html += "</tr>"'''
            #     html.append(row_html)
            
            for idx, (stmt, status) in enumerate(all_rows):
                row_html = "<tr>"

                if not client_written:
                    row_html += (
                        f'<td style="border:1px solid black; padding:5px; font-weight:bold;" '
                        f'rowspan="{total_rows}">{client}</td>'
                    )
                    client_written = True

                if idx == 0:
                    row_html += (
                        f'<td style="border:1px solid black; padding:5px;" '
                        f'rowspan="{rowspan_acc}">{acc_name}</td>'
                    )

                # 🔶 ORANGE for Missing, GREEN for Received
                color = "#fd7e14" if status == "Missing" else "#198754"

                row_html += (
                    f'<td style="border:1px solid black; padding:5px;">{stmt}</td>'
                    f'<td style="border:1px solid black; padding:5px; font-weight:bold; color:{color};">'
                    f'{status}'
                    f'</td>'
                )

                row_html += "</tr>"
                html.append(row_html)

    html.append("</table>")

    full_html = f"""
    <html xmlns:o="urn:schemas-microsoft-com:office:office"
          xmlns:x="urn:schemas-microsoft-com:office:excel"
          xmlns="http://www.w3.org/TR/REC-html40">
      <head><meta charset="UTF-8"></head>
      <body>{''.join(html)}</body>
    </html>
    """

    response = HttpResponse(full_html, content_type="application/vnd.ms-excel")
    response["Content-Disposition"] = 'attachment; filename="Pending_Detail_Summary.xls"'
    return response

from collections import defaultdict
from datetime import date, timedelta
from collections import defaultdict
from datetime import timedelta
from accounts.utills import get_db_for_fy, fiscal_year_range, get_current_fy
from .models import DataEntry, Client, AccountBank
from datetime import datetime

def _build_pending_ranges(request):
    """
    FINAL FIXED VERSION

    Logic:
    - Matching is done ONLY by account_id
    - Client is derived from AccountBank
    """

    # --------------------------------------------------
    # 1) FY + DB
    # --------------------------------------------------
    fy = request.GET.get("fy") or get_current_fy()
    fy_for_range = fy.replace("_", "-")
    db_alias = get_db_for_fy(fy)

    # --------------------------------------------------
    # SINGLE DATE (acts as LAST DATE)
    # --------------------------------------------------
    single_date_str = request.GET.get("single_date")
    single_date = None

    if single_date_str:
        try:
            single_date = datetime.strptime(single_date_str, "%Y-%m-%d").date()
        except ValueError:
            single_date = None


    fy_start, fy_end = fiscal_year_range(fy_for_range)
    # ✅ If single date selected → treat it as LAST DATE
    if single_date:
        if single_date < fy_start:
            return [], fy_for_range   # selected date before FY → no data

        fy_end = min(single_date, fy_end)

    # --------------------------------------------------
    # 2) Selected accounts
    # --------------------------------------------------
    account_ids_str = (request.GET.get("account_ids") or "").strip()
    try:
        account_ids = [int(a) for a in account_ids_str.split(",") if a.strip()]
    except ValueError:
        account_ids = []

    if not account_ids:
        return [], fy_for_range

    # --------------------------------------------------
    # 3) Account + Client info (DEFAULT DB)
    # --------------------------------------------------
    accounts = (
        AccountBank.objects
        .filter(id__in=account_ids)
        .select_related("client")
    )

    account_info = {
        a.id: {
            "client_name": a.client.client_name if a.client else "Unknown Client",
            "account_name": f"{a.account} ({a.account_id})",
        }
        for a in accounts
    }

    # --------------------------------------------------
    # 4) Load DataEntry rows (FY DB)
    # --------------------------------------------------
    qs = DataEntry.objects.using(db_alias).filter(
        account_id__in=account_ids,
        from_date__isnull=False,
        last_date__isnull=False,
        from_date__lte=fy_end,
        last_date__gte=fy_start,
    )

    # --------------------------------------------------
    # 5) Group covered ranges by account_id ONLY
    # --------------------------------------------------
    covered_map = defaultdict(list)
    for e in qs:
        covered_map[e.account_id].append(
            (max(e.from_date, fy_start), min(e.last_date, fy_end))
        )

    # --------------------------------------------------
    # 6) Build rows
    # --------------------------------------------------
    rows = []
    column_select = (request.GET.get("column_select") or "all").lower()

    for acc_id, info in account_info.items():
        client_name  = info["client_name"]
        account_name = info["account_name"]

        covered = covered_map.get(acc_id, [])

        # ===== NO ENTRY AT ALL =====
        if not covered:
            if column_select in ("missing", "all"):
                rows.append({
                    "client": client_name,
                    "account": account_name,
                    "from_date": fy_start.strftime("%d-%m-%Y"),
                    "to_date": fy_end.strftime("%d-%m-%Y"),
                    "kind": "missing",
                })
            continue

        # ===== HAS ENTRIES =====
        covered.sort()

        merged = []
        for s, e in covered:
            if not merged or s > merged[-1][1] + timedelta(days=1):
                merged.append([s, e])
            else:
                merged[-1][1] = max(merged[-1][1], e)

        # RECEIVED
        if column_select in ("received", "all"):
            for s, e in merged:
                rows.append({
                    "client": client_name,
                    "account": account_name,
                    "from_date": s.strftime("%d-%m-%Y"),
                    "to_date": e.strftime("%d-%m-%Y"),
                    "kind": "received",
                })

        # MISSING (ONLY GAPS)
        if column_select in ("missing", "all"):
            cur = fy_start
            for s, e in merged:
                if cur < s:
                    rows.append({
                        "client": client_name,
                        "account": account_name,
                        "from_date": cur.strftime("%d-%m-%Y"),
                        "to_date": (s - timedelta(days=1)).strftime("%d-%m-%Y"),
                        "kind": "missing",
                    })
                cur = e + timedelta(days=1)

            if cur <= fy_end:
                rows.append({
                    "client": client_name,
                    "account": account_name,
                    "from_date": cur.strftime("%d-%m-%Y"),
                    "to_date": fy_end.strftime("%d-%m-%Y"),
                    "kind": "missing",
                })

    return rows, fy_for_range




def pending_detail_summary(request):
    rows, fy_display = _build_pending_ranges(request)
    return JsonResponse({"ok": True, "rows": rows})






def client_accounts_multi(request):
    client_ids_str = request.GET.get("client_ids", "").strip()
    if not client_ids_str:
        return JsonResponse({"ok": False, "error": "client_ids required"}, status=400)

    client_ids = [cid.strip() for cid in client_ids_str.split(",") if cid.strip()]

    qs = AccountBank.objects.filter(
        client__client_id__in=client_ids,
        account_group__in=ALLOWED_ACCOUNT_GROUPS
    ).order_by("client__client_name", "account")

    data = [{
        "id": a.id,
        "account": a.account,
        "account_id": a.account_id,
        "client_id": a.client.client_id,
        "client_name": a.client.client_name,
    } for a in qs]

    return JsonResponse({"ok": True, "accounts": data})


# views.py
from io import BytesIO
from django.http import HttpResponse
from django.utils.timezone import now
import xlsxwriter

from .models import DataEntry   # adjust if your model is in another app


from django.http import HttpResponse
from django.http import HttpResponse

def pending_detail_excel(request):
    rows, fy_display = _build_pending_ranges(request)

    if not rows:
        # Optional: show an empty but valid Excel, or short message
        html = "<html><body><p>No data for selected filters.</p></body></html>"
        response = HttpResponse(html, content_type="application/vnd.ms-excel")
        response["Content-Disposition"] = 'attachment; filename="Pending_Detail_Summary.xls"'
        return response

    group_name    = request.GET.get("group_name", "Statement Report")
    column_select = (request.GET.get("column_select") or "all").capitalize()

    # Group rows by client -> account -> list of statements
    grouped = {}
    for r in rows:
        client  = r["client"]
        account = r["account"]
        stmt    = f'{r["from_date"]} To {r["to_date"]}'

        grouped.setdefault(client, {}).setdefault(account, [])
        if stmt not in grouped[client][account]:
            grouped[client][account].append(stmt)

    html_parts = []
    html_parts.append(
        f"""
        <table style="width:100%; border-collapse:collapse;">
          <tr>
            <td colspan="3" style="font-weight:bold; font-size:16px; padding:8px; border:none;">
              {group_name} - {column_select} Summary {fy_display}
            </td>
          </tr>
          <tr style="background:#002060; color:#fff; font-weight:bold; text-align:left;">
            <th style="border:1px solid black; padding:5px;">Client Name</th>
            <th style="border:1px solid black; padding:5px;">Account Name</th>
            <th style="border:1px solid black; padding:5px;">Statement Period</th>
          </tr>
        """
    )

    for client, accounts in grouped.items():
        total_rows = sum(len(v) for v in accounts.values())
        client_written = False

        for acc_name, stmts in accounts.items():
            rowspan_acc = len(stmts)
            for idx, stmt in enumerate(stmts):
                row_html = "<tr>"

                if not client_written:
                    row_html += (
                        f'<td style="border:1px solid black; padding:5px; font-weight:bold;" '
                        f'rowspan="{total_rows}">{client}</td>'
                    )
                    client_written = True

                if idx == 0:
                    row_html += (
                        f'<td style="border:1px solid black; padding:5px;" '
                        f'rowspan="{rowspan_acc}">{acc_name}</td>'
                    )

                row_html += f'<td style="border:1px solid black; padding:5px;">{stmt}</td>'
                row_html += "</tr>"
                html_parts.append(row_html)

    html_parts.append("</table>")

    full_html = f"""
    <html xmlns:o="urn:schemas-microsoft-com:office:office"
          xmlns:x="urn:schemas-microsoft-com:office:excel"
          xmlns="http://www.w3.org/TR/REC-html40">
      <head><meta charset="UTF-8"></head>
      <body>{''.join(html_parts)}</body>
    </html>
    """

    response = HttpResponse(full_html, content_type="application/vnd.ms-excel")
    response["Content-Disposition"] = 'attachment; filename="Pending_Detail_Summary.xls"'
    return response




from django.conf import settings
def download_launcher(request):
    file_path = os.path.join(settings.BASE_DIR, 'accounting_app', 'static', 'hta', 'dashboard.hta')
    return FileResponse(open(file_path, 'rb'), as_attachment=True, filename='dashboard.hta')



import os
from urllib.parse import unquote
from django.http import FileResponse, Http404, HttpResponse
from django.conf import settings
import mimetypes

def serve_bank_pdf(request):
    # Get relative path from DB or frontend (e.g. fy_2025_26\ADRPC5164B\1901\34.pdf)
    relative_path = request.GET.get("path", "").strip()
    if not relative_path:
        return HttpResponse("Missing 'path' parameter", status=400)

    # Clean and normalize
    relative_path = os.path.normpath(relative_path)

    # Set base directory where all statements are stored
    base_dir = r"Z:\Kameet Soft Project\Bank Statement"

    # Join base_dir with relative path from DB
    full_path = os.path.abspath(os.path.join(base_dir, relative_path))
    print("📂 Full resolved file path:", full_path)

    # Security: check if it’s inside base_dir
    allowed_base = os.path.abspath(base_dir)
    if not full_path.lower().startswith(allowed_base.lower()):
        return HttpResponse("Access denied to this file path", status=403)

    # Check file existence
    if not os.path.isfile(full_path):
        return HttpResponse("File not found", status=404)

    # Serve the file
    # response = FileResponse(open(full_path, 'rb'), content_type='application/pdf')
    # response['Content-Disposition'] = f'inline; filename="{os.path.basename(full_path)}"'
    # return response
    filename = os.path.basename(full_path)
    ext = os.path.splitext(filename)[1].lower()

    # 🔹 Detect MIME dynamically
    mime_type, _ = mimetypes.guess_type(full_path)
    mime_type = mime_type or "application/octet-stream"

    # 🔹 File types allowed inline
    INLINE_EXTS = {".pdf", ".htm", ".html", ".txt"}

    response = FileResponse(open(full_path, "rb"), content_type=mime_type)

    if ext in INLINE_EXTS:
        response["Content-Disposition"] = f'inline; filename="{filename}"'
    else:
        # Excel, ZIP, etc → force download
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response["Content-Length"] = os.path.getsize(full_path)

    return response


# busy entry status #######
# views.py
import pyodbc
from django.http import JsonResponse
import os
from datetime import datetime

def get_busy_status(request):
    busy_code = request.GET.get("busy_code")
    busy_acc_code = request.GET.get("busy_acc_code")
    fy = request.GET.get("fy")

    print("📥 GET Busy Status Request →", {
        "busy_code": busy_code,
        "busy_acc_code": busy_acc_code,
        "fy": fy
    })

    if not all([busy_code, busy_acc_code, fy]):
        return JsonResponse({"error": "Missing parameters"}, status=400)

    try:
        busy_acc_code = int(busy_acc_code)
    except ValueError:
        return JsonResponse({"error": "Invalid Busy Account Code"}, status=400)

    # base_path = r"\\Dadaji\e\Busy\Data\KameetData"
    base_path = r"Y:\Busy\DATA"

    results = {
        "prev_year": "",
        "curr_year": "",
        "last_date": ""
    }

    start_fy = int(fy[:4]) - 1
    for i in range(2):  # i = 0 for prev year, 1 for curr year
        year = start_fy + i
        db_path = os.path.join(base_path, busy_code, f"db1{year}.bds")
        print(f"🔍 Checking DB path: {db_path}")

        if not os.path.exists(db_path):
            print(f"❌ DB file not found for FY {year}")
            continue

        conn_str = (
            f"DRIVER={{Microsoft Access Driver (*.mdb, *.accdb)}};"
            f"DBQ={db_path};"
            f"PWD=ILoveMyINDIA;"
        )

        try:
            with pyodbc.connect(conn_str, timeout=2) as conn:
                cursor = conn.cursor()
                query = f"""
                    SELECT 
                        (SELECT MAX(Tran2.Date) FROM Tran2 WHERE Tran2.MasterCode1 = M1.Code AND Tran2.Value1 > 0) AS LDate_Pmt,
                        (SELECT COUNT(*) FROM Tran2 WHERE Tran2.MasterCode1 = M1.Code) AS TransactionCount,
                        (SELECT Folio1.D1 FROM Folio1 WHERE Folio1.MasterCode = M1.Code) AS OpBal,
                        (SELECT SUM(IIF(IsNull(DailySum.Dr1), 0, DailySum.Dr1)) FROM DailySum WHERE DailySum.MasterCode1 = M1.Code) AS Debit,
                        (SELECT SUM(IIF(IsNull(DailySum.Cr1), 0, DailySum.Cr1)) FROM DailySum WHERE DailySum.MasterCode1 = M1.Code) AS Credit,
                        IIF(
                            (SELECT SUM(IIF(IsNull(DailySum.Dr1), 0, DailySum.Dr1)) FROM DailySum WHERE DailySum.MasterCode1 = M1.Code) IS NULL,
                            (SELECT Folio1.D1 FROM Folio1 WHERE Folio1.MasterCode = M1.Code),
                            (SELECT Folio1.D1 FROM Folio1 WHERE Folio1.MasterCode = M1.Code)
                            - (SELECT SUM(IIF(IsNull(DailySum.Dr1), 0, DailySum.Dr1)) FROM DailySum WHERE DailySum.MasterCode1 = M1.Code)
                            + (SELECT SUM(IIF(IsNull(DailySum.Cr1), 0, DailySum.Cr1)) FROM DailySum WHERE DailySum.MasterCode1 = M1.Code)
                        ) AS ClBal
                    FROM Master1 AS M1
                    WHERE M1.Code = {busy_acc_code}
                """
                print(f"📄 Executing Query for year {year}:\n{query}")
                row = cursor.execute(query).fetchone()

                if row:
                    print(f"✅ Query Result for FY {year}: Date={row.LDate_Pmt}, Count={row.TransactionCount}, ClBal={row.ClBal}")
                    if i == 0:
                        results["prev_year"] = row.TransactionCount
                    elif i == 1:
                        results["curr_year"] = row.TransactionCount
                        if row.LDate_Pmt:
                            results["last_date"] = row.LDate_Pmt.strftime("%d-%m-%Y")
                        
                        if row.ClBal is not None:
                            cl_bal = float(row.ClBal)

                            if cl_bal < 0:
                                # Dr balance → show positive value
                                results["closing_balance"] = f"{abs(cl_bal):,.2f} Dr"
                            else:
                                # Cr balance
                                results["closing_balance"] = f"{cl_bal:,.2f} Cr"

                        # if row.ClBal is not None:
                        #     cl_bal = row.ClBal
                        #     if cl_bal < 0:
                        #         results["closing_balance"] = f"{(cl_bal):,.2f} Dr"
                        #     else:
                        #         results["closing_balance"] = f"{cl_bal:,.2f} Cr"

                        # if row.ClBal is not None:
                        #     results["closing_balance"] = str(round(row.ClBal, 2))
                else:
                    print(f"⚠️ No result for MasterCode {busy_acc_code} in FY {year}")

        except Exception as e:
            print(f"🚨 Error opening Busy DB for FY {year}: {e}")
            continue

    print("📤 Final Busy Status Response:", results)
    return JsonResponse(results)



# views.py
from django.http import JsonResponse
from .models import Client
# whole client without select the group
def get_clients(request):
    group_id = request.GET.get('group_id')
    if group_id:
        clients = Client.objects.filter(group_id=group_id)
    else:
        clients = Client.objects.all()  # ✅ Return all if no group selected

    data = [
        {'client_id': client.id, 'client_name': client.name}
        for client in clients
    ]
    return JsonResponse(data, safe=False)



def suspend_form(request, suspension_id=None):
    is_modify_mode = suspension_id is not None
    suspension = None

    if suspension_id:
        suspension = get_object_or_404(TaxSuspension, pk=suspension_id)

    if request.method == 'POST':
        print("🔍 RAW POST:", request.POST)

        post_data = request.POST.copy()

        group_id = post_data.get('group_id')
        client_id = post_data.get('client_id')
        account_id = post_data.get('account_id')

        # Resolve group
        if group_id:
            try:
                post_data['group'] = Group.objects.get(pk=group_id).id
            except Group.DoesNotExist:
                post_data['group'] = None

        # Resolve client
        if client_id:
            try:
                if client_id.isdigit():
                    client_obj = Client.objects.get(pk=client_id)
                else:
                    client_obj = Client.objects.get(client_id=client_id)
                post_data['client'] = client_obj.id
            except Client.DoesNotExist:
                post_data['client'] = None

        # Resolve account
        if account_id:
            acc_obj = (
                AccountBank.objects.filter(pk=account_id).first() or
                AccountBank.objects.filter(account_id=account_id).first()
            )
            post_data['account'] = acc_obj.id if acc_obj else None

        # Bind form (with instance in modify mode)
        form = TaxSuspensionForm(post_data, instance=suspension)

        if form.is_valid():
            suspension = form.save(commit=False)

            # Set status
            action = post_data.get('action')
            suspension.is_active = True if action == 'suspend' else False

            suspension.save()
            messages.success(request, "✅ Tax suspension saved successfully!")

            return redirect('suspend_form')  # or redirect to list

        else:
            print("❌ Form errors:", form.errors)
            messages.error(request, "Please fix the errors before submitting.")

    else:
        if suspension:
            # Format dates for initial
            date_format = get_format('DATE_INPUT_FORMATS')[0]
            initial_data = {
                'tax_type': suspension.tax_type,
                'from_date': DateFormat(suspension.from_date).format(date_format),
                'to_date': DateFormat(suspension.to_date).format(date_format) if suspension.to_date else '',
                'remarks': suspension.remarks,
                'action': 'suspend' if suspension.is_active else 'activate',
            }
            if suspension.account:
                initial_data['account_id'] = suspension.account.account_id
                initial_data['account_name'] = suspension.account.account

            if suspension.client:
                initial_data['client_id'] = suspension.client.id
                initial_data['client_name'] = suspension.client.client_name

            if suspension.group:
                initial_data['group_id'] = suspension.group.id
                initial_data['group_name'] = suspension.group.group_name
            print("📦 Initial Data:", initial_data)
            form = TaxSuspensionForm(initial=initial_data, instance=suspension)
        else:
            form = TaxSuspensionForm()

    # Send required data
    groups = Group.objects.all().order_by("group_name")
    clients = Client.objects.all().order_by("client_name")
    accounts = AccountBank.objects.all().order_by("account")

    return render(request, 'suspend_form.html', {
        'form': form,
        'groups': groups,
        'clients': clients,
        'accounts': accounts,
        'suspension': suspension,
        'is_modify_mode': is_modify_mode,
    })


def suspend_list(request):
    suspensions = TaxSuspension.objects.all()

    return render(request,"suspend_list.html",{"suspensions":suspensions})

def delete_tax_suspension(request, pk):
    """
    View to delete a tax suspension record
    """
    suspension = get_object_or_404(TaxSuspension, pk=pk)
    
    try:
        suspension.delete()
        messages.success(request, f"Successfully deleted tax suspension record #{pk}")
    except Exception as e:
        messages.error(request, f"Error deleting suspension: {str(e)}")
    
    return redirect('suspend_list')  # Replace with your list view name


from django.db.models import Count, Q
from django.shortcuts import render


def group_wise_report(request):
    groups = Group.objects.annotate(
            total_it_clients=Count('clients', filter=Q(clients__it_return=True))
        ).order_by('group_name')

    context={
        'groups': groups,
    }
    return render(request, 'group_wise_report.html',context)

##############################################################################33
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction
from django.conf import settings
from .models import IncomeTaxReturn, Client, UserData
from .forms import IncomeTaxReturnForm
from .utills import get_db_for_fy
from django.http import Http404

def resolve_foreign_keys(instance):
    if instance:
        if instance.client_id:
            try:
                instance.client = Client.objects.using('default').get(id=instance.client_id)
            except Client.DoesNotExist:
                instance.client = None
        if instance.alloted_to_id:
            try:
                instance.alloted_to = UserData.objects.using('default').get(id=instance.alloted_to_id)
            except UserData.DoesNotExist:
                instance.alloted_to = None
    return instance


# --- Helpers you already use elsewhere ---
def get_db_for_fy(fy_str_with_underscore):
    # e.g. '2024_25' -> 'fy_2024_25'
    return f"fy_{fy_str_with_underscore}"

def get_fiscal_year_from_date(d):
    """date(2024,4,1) -> '2024_25'"""
    y = d.year
    return f"{y}_{str(y+1)[-2:]}"

def resolve_foreign_keys(instance):
    # If you need to detach FK objects or similar; keep passthrough for now
    return instance

# --- AJAX endpoint to return client details ---
def get_client_details(request, client_id):
    try:
        c = Client.objects.using('default').select_related('group').get(id=client_id)
        group_name = getattr(c.group, 'group_name', '') if getattr(c, 'group_id', None) else ''
        return JsonResponse({
            "group_name": group_name or "",
            "pan_no": getattr(c, 'pan', '') or "",
            "legal_name": getattr(c, 'legal_name', '') or "",
            "dob": c.dob.strftime("%Y-%m-%d") if getattr(c, 'dob', None) else "",
        })
    except Client.DoesNotExist:
        return JsonResponse({}, status=404)


from urllib.parse import urlencode
from django.urls import reverse
from django.utils.http import url_has_allowed_host_and_scheme

from urllib.parse import urlencode
from django.urls import reverse
from django.utils.http import url_has_allowed_host_and_scheme

def _back_to_report_url(request, fy_display: str, it_id: int) -> str:
    """Build a safe URL back to it_return_report with preserved filters and highlight."""
    # Prefer explicit ?next=... if present and safe
    next_url = request.POST.get("next") or request.GET.get("next")
    if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
        return next_url

    params = {}
    for key in ("rtype", "group", "stage", "staff"):
        val = request.POST.get(key) or request.GET.get(key)
        if val:
            params[key] = val

    if fy_display:
        params["fy"] = fy_display

    if it_id:
        params["highlight_id"] = it_id

    return f"{reverse('it_return_report')}?{urlencode(params)}" if params else reverse('it_return_report')


from django.conf import settings
from django.contrib import messages
from django.db import transaction
from django.http import Http404
from django.shortcuts import redirect, render
from django.urls import reverse

# --- helper: normalize FY and build db alias ---
def _fy_to_db(fy_param: str):
    """
    Accepts '2025-26', '2025_26', '2025/26', or 'fy_2025_26'
    Returns (db_alias, fy_display, fy_underscore)
    """
    if not fy_param:
        return None, None, None
    s = str(fy_param).strip().lower()
    if s.startswith("fy_"):
        s = s[3:]
    s_u = s.replace("-", "_").replace("/", "_")
    db_alias = f"fy_{s_u}"
    fy_display = s_u.replace("_", "-")
    return db_alias, fy_display, s_u
def _normalize_fy(fy_param: str):
    """
    Accept '2025-26' / '2025_26' / '2025/26' / 'fy_2025_26'
    Return: (db_alias, fy_display, fy_underscore)
    """
    if not fy_param:
        return None, None, None
    s = str(fy_param).strip()
    if s.lower().startswith("fy_"):
        s = s[3:]
    s_u = s.replace("-", "_").replace("/", "_")
    db_alias = f"fy_{s_u}"
    fy_display = s_u.replace("_", "-")
    return db_alias, fy_display, s_u

from django import forms
def _configure_return_type(form, is_modify_mode):
    # always keep the full model choices available
    form.fields['return_type'].choices = IncomeTaxReturn.RETURN_TYPE_CHOICES

    if is_modify_mode:
        form.fields['return_type'].widget = forms.Select()
        form.fields['return_type'].disabled = False
    else:
        # hide in add mode but still submit “Original”
        form.fields['return_type'].initial = 'Original'
        form.fields['return_type'].widget = forms.HiddenInput()



from django.conf import settings
from django.contrib import messages
from django.db import transaction
from django.http import Http404
from django.shortcuts import redirect, render
from django.urls import reverse


def it_return_form(request):
    is_modify_mode = 'mode=modify' in request.get_full_path() or '/it-return/modify/' in request.path
    selected_it_id = request.GET.get('it_id') or request.POST.get('it_id')
    fy_param = request.GET.get('fy') or request.POST.get('fy')
    instance = None
    db_to_use = None

    # === Modify Mode bootstrap ===
    if is_modify_mode:
        if not fy_param:
            raise Http404("Fiscal year is required in modify mode.")
        db_to_use, fy_display, fy_underscore = _normalize_fy(fy_param)
        if not db_to_use or db_to_use not in settings.DATABASES:
            raise Http404(f"Database '{db_to_use or fy_param}' not found.")
        if selected_it_id:
            try:
                instance = IncomeTaxReturn.objects.using(db_to_use).get(id=selected_it_id)
                if 'resolve_foreign_keys' in globals():
                    instance = resolve_foreign_keys(instance)
            except IncomeTaxReturn.DoesNotExist:
                raise Http404("IT Return not found in selected year.")

        # Guard: prevent editing Filed records
        if instance and str(getattr(instance, "stage", "")).strip().lower() == "filed":
            messages.error(request, "This return is already Filed and cannot be edited.")
            return redirect(f"{reverse('it_return_report')}?rtype=audit&fy={fy_param}")

    # === POST: Save ===
    if request.method == 'POST':
        safe_instance = resolve_foreign_keys(instance) if instance and 'resolve_foreign_keys' in globals() else instance

        # 1) Build form
        form = IncomeTaxReturnForm(request.POST, instance=safe_instance, is_modify_mode=is_modify_mode)
        form.fields['client'].queryset = Client.objects.using('default').all()
        form.fields['alloted_to'].queryset = UserData.objects.using('default').all()
        _configure_return_type(form, is_modify_mode)

        # 2) If user chose "Other…", capture typed custom value
        stage_in_post = request.POST.get('stage')
        final_stage = (request.POST.get('stage_other') or '').strip() if stage_in_post == '__other__' else None

        # 3) TEMP: extend choices on both form + model field so model.full_clean() accepts custom
        orig_model_choices = IncomeTaxReturn._meta.get_field('stage').choices
        try:
            if final_stage:
                # extend form field choices (keeps widget consistent if form re-renders)
                sf = form.fields['stage']
                new_fc = list(sf.choices)
                new_fc.append((final_stage, final_stage))
                sf.choices = new_fc
                sf.widget.choices = new_fc

                # extend *model* field choices (validation uses model._meta)
                IncomeTaxReturn._meta.get_field('stage').choices = list(orig_model_choices) + [(final_stage, final_stage)]

            # 4) Validate and save
            if form.is_valid():
                fy_input = request.POST.get('fy') or fy_param
                db_alias, fy_display, fy_underscore = _normalize_fy(fy_input) if fy_input else (None, None, None)
                if not db_alias or db_alias not in settings.DATABASES:
                    raise Http404(f"Database '{db_alias}' not found.")

                client_id = form.cleaned_data.get('client').id if form.cleaned_data.get('client') else None

                # Prevent duplicate client entry in Add mode
                if not is_modify_mode and client_id:
                    exists = IncomeTaxReturn.objects.using(db_alias).filter(client_id=client_id).exists()
                    if exists:
                        messages.error(request, "This client already has an IT Return in the selected financial year.")
                        return redirect(request.path)

                with transaction.atomic(using=db_alias):
                    it_obj = form.save(commit=False)

                    # clean() already maps __other__ → custom, but keep this safe guard
                    stage_value = form.cleaned_data.get('stage') or ''
                    if stage_value == '__other__':
                        stage_value = final_stage
                    if stage_value:
                        it_obj.stage = stage_value

                    it_obj.save(using=db_alias)

                messages.success(
                    request,
                    f"IT Return {'updated' if is_modify_mode else 'added'} successfully in database: {db_alias}"
                )

                if is_modify_mode:
                    return redirect(_back_to_report_url(request, fy_display, it_obj.id))
                return redirect(request.path)
            else:
                messages.error(request, f"Form errors: {form.errors.as_text()}")
        finally:
            # 5) Always restore original model choices
            IncomeTaxReturn._meta.get_field('stage').choices = orig_model_choices

    # === GET: Build fresh form ===
    else:
        safe_instance = resolve_foreign_keys(instance) if instance and 'resolve_foreign_keys' in globals() else instance
        form = IncomeTaxReturnForm(instance=safe_instance, is_modify_mode=is_modify_mode)
        form.fields['client'].queryset = Client.objects.using('default').all()
        form.fields['alloted_to'].queryset = UserData.objects.using('default').all()
        _configure_return_type(form, is_modify_mode)

        # In add mode default Return Type to Original (and hide if you want)
        if not is_modify_mode:
            form.fields['return_type'].initial = 'Original'
            form.fields['return_type'].widget = forms.HiddenInput()

        # If editing and saved stage is custom (not in choices), preselect Other…
        if instance:
            stage_field = form.fields['stage']
            stage_values = {v for v, _ in stage_field.choices}
            if getattr(instance, 'stage', None) and instance.stage not in stage_values:
                form.initial['stage'] = '__other__'
                # NOTE: to pre-fill the textbox, render {{ stage_other_initial }} in template
                # and pass it in context below.

    # === Context ===
    clients = Client.objects.using('default').all().order_by('client_name')
    users = UserData.objects.using('default').all().order_by('username')

    selected_client_obj = None
    selected_user_obj = None
    if instance:
        if getattr(instance, "client_id", None):
            selected_client_obj = Client.objects.using('default').filter(id=instance.client_id).first()
        if getattr(instance, "alloted_to_id", None):
            selected_user_obj = UserData.objects.using('default').filter(id=instance.alloted_to_id).first()

    cancel_url = _safe_back_url(request, default_name="home")

    # Optional: prefill custom stage text in edit mode when stage is not in choices
    stage_other_initial = ""
    if instance:
        stage_field = form.fields['stage']
        stage_values = {v for v, _ in stage_field.choices}
        if getattr(instance, 'stage', None) and instance.stage not in stage_values:
            stage_other_initial = instance.stage

    return render(request, 'it_return_form.html', {
        'form': form,
        'is_modify_mode': is_modify_mode,
        'clients': clients,
        'users': users,
        'title': 'Modify IT Return' if is_modify_mode else 'Add IT Return',
        'selected_it_id': int(selected_it_id) if selected_it_id else None,
        'selected_client_obj': selected_client_obj,
        'selected_user_obj': selected_user_obj,
        'cancel_url': cancel_url,
        'stage_other_initial': stage_other_initial,  # use in template if you want to prefill the input
    })


# def it_return_list(request):
#     fy = request.GET.get('fy')  # Expecting something like '2024_25'

#     from datetime import date
#     fiscal_years = ['2023_24', '2024_25', '2025_26']

#     if not fy:
#         current_year = date.today().year
#         next_year = current_year + 1
#         fy = f"{current_year}_{str(next_year)[-2:]}"


#     db_to_use = f"fy_{fy}"
#     itrs = IncomeTaxReturn.objects.using(db_to_use).all()
    
#     clients = Client.objects.all()
#     users = UserData.objects.all()

#     context = {
#         'itrs': itrs,
#         'fy': fy,
#          'clients': clients,
#          'users': users,
#         'fiscal_years': fiscal_years,

#     }

#     return render(request,'it_return_list.html',context)

from django.conf import settings
def it_return_list(request):
    # Allowed fiscal years (ONLY existing DBs)
    fiscal_years = ['2023_24', '2024_25', '2025_26']

    # Get fy from URL
    fy = request.GET.get('fy')

    # ✅ FORCE DEFAULT FY
    if not fy:
        fy = '2024_25'

    # ❌ Block invalid / future FY
    if fy not in fiscal_years:
        fy = '2024_25'

    db_to_use = f"fy_{fy}"

    # Extra safety
    if db_to_use not in settings.DATABASES:
        fy = '2024_25'
        db_to_use = 'fy_2024_25'

    itrs = IncomeTaxReturn.objects.using(db_to_use).all()

    context = {
        'itrs': itrs,
        'fy': fy,
        'fy_selected': fy,   # 🔥 IMPORTANT
        'fiscal_years': fiscal_years,
    }

    return render(request, 'it_return_list.html', context)



# views.py

from django.shortcuts import redirect
from django.urls import reverse
from django.contrib import messages
from .models import IncomeTaxReturn

def delete_itr(request, pk):
    fy = request.GET.get('fy')
    if not fy:
        messages.error(request, "Fiscal year not specified.")
        return redirect('/')  # Safe fallback

    db_to_use = f'fy_{fy}'

    try:
        itr = IncomeTaxReturn.objects.using(db_to_use).get(pk=pk)
        itr.delete(using=db_to_use)
        messages.success(request, "ITR deleted successfully.")
    except IncomeTaxReturn.DoesNotExist:
        messages.error(request, "ITR not found.")
    except Exception as e:
        messages.error(request, f"Error deleting ITR: {e}")

    return redirect(f"{reverse('it_return_list')}?fy={fy}")

def get_current_fy():
    today = date.today()
    if today.month >= 4:
        return f"{today.year}-{str(today.year + 1)[2:]}"  # e.g., 2024-25
    else:
        return f"{today.year - 1}-{str(today.year)[2:]}"   # e.g., 2023-24
    
# Define fixed stage options from the model
STAGE_CHOICES = [
    'Sleeping',
    'Call for data',
    'Working',
    'Query to client',
    'Pending for checking',
    'Pending for approve',
    'Pending for audit sent',
    'audit sent',
    'Pending for filing',
    'Filed',
]

import os, re, fitz, json
from decimal import Decimal
from django.conf import settings
from django.http import JsonResponse
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from .models import AISUpload, AISRecord
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.conf import settings
import os, fitz, re, json
from decimal import Decimal
from .models import AISUpload, AISRecord


# views.py
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.conf import settings
import os, fitz, re, json
from decimal import Decimal, InvalidOperation
from .models import AISUpload, AISRecord

def extract_rows_from_blocks(blocks):
    rows = []
    for block in blocks:
        for line in block.get("lines", []):
            row = []
            for span in line.get("spans", []):
                txt = span.get("text", "").strip()
                if txt:
                    row.append(txt)
            if row:
                rows.append(row)
    return rows

@csrf_exempt
def upload_ais_pdf(request):
    if request.method == 'POST':
        pdf_file = request.FILES['pdf_file']
        password = pdf_file.name.replace('.pdf', '')
        file_path = os.path.join(settings.MEDIA_ROOT, 'temp', pdf_file.name)
        os.makedirs(os.path.dirname(file_path), exist_ok=True)

        with open(file_path, 'wb+') as f:
            for chunk in pdf_file.chunks():
                f.write(chunk)

        try:
            doc = fitz.open(file_path)
            if doc.needs_pass and not doc.authenticate(password):
                return JsonResponse({'error': 'Incorrect PDF password'}, status=401)
            existing_upload = AISUpload.objects.filter(file_name=pdf_file.name).first()
            if existing_upload:
                return redirect('review_upload', upload_id=existing_upload.id)

            # Create upload metadata
            current_upload = AISUpload.objects.create(
                file_name=pdf_file.name,
                information_code="",
                description="",
                source="",
                count=0,
                amount=Decimal('0.00')
            )

            current_section = None
            current_headers = []
            current_rows = []
            record_count = 0

            for page in doc:
                blocks = page.get_text("dict")["blocks"]

                for block in blocks:
                    # Combine block text for header check
                    block_text = " ".join([
                        span.get("text", "").strip()
                        for line in block.get("lines", [])
                        for span in line.get("spans", [])
                    ])

                    if not block_text:
                        continue

                    # Detect section start
                    if re.match(r"^SR\.?\s*NO\.", block_text, re.IGNORECASE):
                        # Save previous section's rows
                        if current_section and current_rows:
                            for row in current_rows:
                                data = {f"col_{i}": val for i, val in enumerate(row)}
                                AISRecord.objects.create(
                                    upload=current_upload,
                                    section_title=current_section,
                                    data_json=data
                                )
                                record_count += 1
                            current_rows = []

                        # Start new section
                        current_section = block_text
                        current_headers = extract_rows_from_blocks([block])  # not used currently
                    elif current_section:
                        # It's a data row
                        rows = extract_rows_from_blocks([block])
                        current_rows.extend(rows)

            # Save the last section
            if current_section and current_rows:
                for row in current_rows:
                    data = {f"col_{i}": val for i, val in enumerate(row)}
                    AISRecord.objects.create(
                        upload=current_upload,
                        section_title=current_section,
                        data_json=data
                    )
                    record_count += 1

            # Save count in main upload
            current_upload.count = record_count
            current_upload.total_records = record_count
            current_upload.save()

            return redirect('review_upload', upload_id=current_upload.id)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return render(request, 'upload.html')

def review_upload(request, upload_id):
    upload = AISUpload.objects.get(id=upload_id)
    grouped = {}

    for rec in upload.records.all():
        grouped.setdefault(rec.section_title, []).append(rec)

    return render(request, 'review.html', {
        'upload': upload,
        'grouped': grouped
    })

def summary_view(request):
    uploads = AISUpload.objects.filter(is_approved=True).prefetch_related('records')
    grouped = {}
    for up in uploads:
        for rec in up.records.all():
            grouped.setdefault(rec.section_title, []).append(rec)
    return render(request, 'summary.html', {'grouped': grouped})



# pdf2excel
import os
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .pdf2excel.find_module_name import get_module_name
from .pdf2excel.bank_excel import bank_excel
from .pdf2excel.busy_excel import busy_excel
from .pdf2excel import (
    m_axis_1, m_bob_1, m_bom_1, m_hdfc_1, m_icici_1,
    m_sarvodaya_1, m_spcb_1, m_spcb_2, m_sutex_1,m_sbi_1,
    m_canara_1,m_union_1,m_pnb_1,m_indian_1,m_sbi_2,m_kalupur_1,
    m_kotak_1,m_idbi_choicepoint,m_spcb_3
)
# from urllib.parse import quote
from django.http import JsonResponse
import os
from .pdf2excel import bank_modules
# views.py
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse, FileResponse, Http404
from urllib.parse import quote, unquote   # <-- import BOTH
import os

# make sure these exist in your module or import from where you defined them
# from .pdf2excel.detect import get_module_name, bank_modules

# @csrf_exempt
# def convert_pdfs_to_excel(request):
#     if request.method != "POST":
#         return JsonResponse({"error": "Only POST allowed"}, status=405)
    
#      # ✅ Require account always
#     account_id = request.POST.get("account_id", "").strip()
#     if not account_id:
#         return JsonResponse({"error": "Please select an Account before exporting."}, status=400)


#     # Read arrays sent as multiple keys ...AND... gracefully handle a single comma-joined string fallback
#     file_paths = request.POST.getlist("file_paths[]")
#     if len(file_paths) == 1 and "," in file_paths[0]:
#         file_paths = [p.strip() for p in file_paths[0].split(",") if p.strip()]

#     raw_passwords = request.POST.getlist("passwords[]")
#     if len(raw_passwords) == 1 and "," in raw_passwords[0]:
#         raw_passwords = [p.strip() for p in raw_passwords[0].split(",") if p.strip()]

#     # unique, non-empty passwords
#     passwords = []
#     for entry in raw_passwords:
#         for pw in entry.split(","):
#             clean_pw = pw.strip()
#             if clean_pw and clean_pw not in passwords:
#                 passwords.append(clean_pw)

#     mode = request.POST.get("mode")  # 'bank' or 'busy'
#      # 🔒 keep only real, existing PDF files
#     import os
#     file_paths = [p for p in file_paths
#                   if p and os.path.isfile(p) and p.lower().endswith(".pdf")]

#     if not file_paths:
#         return JsonResponse({"error": "No PDF files selected"}, status=400)

#     try:
#         # Detect module per file
#         detected_modules = {}
#         for path in file_paths:
#             module = get_module_name(path, passwords)
#             print(f"[detect] File: {path} -> Module: {module}")
#             detected_modules[path] = module

#         unique_modules = list(set(detected_modules.values()))
#         print("[detect] Unique modules:", unique_modules)

#         if "Module Not Found" in unique_modules:
#             return JsonResponse({"error": "❌ Some PDFs couldn't be recognized. Check the format."}, status=400)
#         if "Passwords failed" in unique_modules:
#             return JsonResponse({"error": "❌ Password incorrect or missing."}, status=400)
#         if "Check Path" in unique_modules:
#             return JsonResponse({"error": "❌ Invalid file path."}, status=400)

#         # If you want to allow multiple modules at once, you could group & export separately.
#         # For now, enforce a single module to keep behavior predictable.
#         if len(unique_modules) != 1:
#             return JsonResponse({"error": f"❌ Multiple modules found: {unique_modules}. Select the same bank/module."}, status=400)

#         module_name = unique_modules[0]
#         module_func = bank_modules.get(module_name)
#         if not module_func:
#             return JsonResponse({"error": f"❌ Module not implemented: {module_name}"}, status=400)

#         # Parse PDFs -> result (your parser should support multiple files)
#         result = module_func(file_paths, passwords)

#         # Where to save the Excel: same directory as the first PDF
#         output_dir = os.path.dirname(file_paths[0])
#         if mode == "bank":
#             from .pdf2excel.bank_excel import bank_excel
#             output_path = bank_excel(result, [], output_dir)  # <-- MUST return full path
#         elif mode == "busy":
#             from .pdf2excel.busy_excel import busy_excel
#             output_path = busy_excel(result, [], output_dir)  # <-- MUST return full path
#         else:
#             return JsonResponse({"error": "❌ Unknown export mode"}, status=400)

#         if not output_path or not os.path.exists(output_path):
#             return JsonResponse({"error": "❌ Excel generation failed (no file returned)."}, status=500)

#         return JsonResponse({
#             "message": "✅ Excel file generated successfully!",
#             "download_url": f"/download-excel/?path={quote(output_path)}"  # <-- quote imported
#         })

#     except Exception as e:
#         # Log full exception in console; return concise message to client
#         import traceback
#         traceback.print_exc()
#         return JsonResponse({"error": f"❌ Exception: {str(e)}"}, status=500)


# def download_excel(request):
#     raw_path = request.GET.get("path")
#     if not raw_path:
#         raise Http404("File not found")

#     file_path = unquote(raw_path)
#     if not os.path.exists(file_path):
#         raise Http404("File not found")

#     return FileResponse(
#         open(file_path, "rb"),
#         as_attachment=True,
#         filename=os.path.basename(file_path),
#         content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#     )

# views.py
import os
from urllib.parse import quote, unquote

from django.http import JsonResponse, FileResponse, Http404
from django.views.decorators.csrf import csrf_exempt

from .pdf2excel.find_module_name import get_module_name
from .pdf2excel import bank_modules
from accounts.models import AccountBank   # <-- import AccountBank
# at the top of the file with convert_pdfs_to_excel
from accounts.models import AccountBank, Client

# 👇 adjust this import according to where you defined them.
# If _busy_file_path and _fetch_busy_bank_accounts are in the SAME file,
# you don't need these imports.
# from .views_busy_mapping import _busy_file_path, _fetch_busy_bank_accounts  # <-- change path if needed


# at top of file with convert_pdfs_to_excel
from accounts.models import AccountBank
# from .views_bank_mapping import _busy_file_path, _fetch_busy_bank_accounts   # <-- CHANGE module name if different


# def get_busy_bank_display_name(acc_obj: AccountBank, fy_code: str) -> str | None:
#     """
#     Use Busy .bds to get the real Busy bank name for this mapped account.
#     Returns something like 'BOM-CA-60224531554 — 60224531554'.
#     """

#     # Need: mapped Busy code + client Busy company code
#     if not acc_obj.busyacccode:
#         return None
#     if not acc_obj.client or not acc_obj.client.busy_code:
#         return None

#     company_code = acc_obj.client.busy_code.strip()
#     if not company_code:
#         return None

#     # 1) Open that Busy company file
#     bds_path = _busy_file_path(company_code, fy_code)
#     if not os.path.exists(bds_path):
#         return None

#     # 2) Read Busy accounts from this company/year
#     rows = _fetch_busy_bank_accounts(bds_path)
#     if not rows:
#         return None

#     mapped = str(acc_obj.busyacccode).strip()

#     # 3) Find matching Busy row (by Code or Name)
#     for r in rows:
#         code   = str(r.get("Code") or "").strip()
#         name   = str(r.get("Name") or "").strip()
#         acc_no = str(r.get("AccNo") or "").strip()

#         # mapping may store Code (3495) or Name (BOM-CA-60224531554)
#         if mapped in (code, name):
#             # 🔹 This is how you show it in mapping: NAME — ACCNO
#             if name and acc_no:
#                 return f"{name} — {acc_no}"
#             return name or acc_no or code

#     return None
import os

def get_busy_bank_display_name(acc_obj: AccountBank, fy_code: str) -> str | None:
    """
    Use Busy .bds to get the real Busy bank name for this mapped account.
    Returns something like 'BOM-CA-60224531554 — 60224531554'.
    """

    # Need: mapped Busy code + client Busy company code
    if not acc_obj.busyacccode:
        return None
    if not acc_obj.client or not acc_obj.client.busy_code:
        return None

    company_code = (acc_obj.client.busy_code or "").strip()
    if not company_code:
        return None

    # 1) Open that Busy company file
    bds_path = _busy_file_path(company_code, fy_code)
    if not os.path.exists(bds_path):
        return None

    # 2) Read Busy accounts from this company/year
    rows = _fetch_busy_bank_accounts(bds_path)
    if not rows:
        return None

    # --- NORMALIZE mapped code (remove .0 etc.) ---
    mapped_raw = str(acc_obj.busyacccode).strip()
    mapped = mapped_raw.split(".")[0]  # "3404.0" -> "3404"

    # 3) Find matching Busy row
    for r in rows:
        code_raw = str(r.get("Code") or "").strip()
        code = code_raw.split(".")[0]   # normalize Busy code too
        name = str(r.get("Name") or "").strip()
        acc_no = str(r.get("AccNo") or "").strip()

        # mapping may store Code (3495 / 3404.0) or Name (BOM-CA-60224531554)
        if mapped == code or mapped == name or mapped in name:
            # same display as in mapping: NAME — ACCNO
            if name and acc_no:
                return f"{name} — {acc_no}"
            return name or acc_no or code

    return None


@csrf_exempt
def convert_pdfs_to_excel(request):
    if request.method != "POST":
        return JsonResponse({"error": "Only POST allowed"}, status=405)

    # ✅ Require account always
    account_id = (request.POST.get("account_id") or "").strip()
    if not account_id:
        return JsonResponse({"error": "Please select an Account before exporting."}, status=400)

    try:
        # acc_obj = AccountBank.objects.get(id=account_id)
        acc_obj = AccountBank.objects.get(account_id=account_id)

        # FY code – same as your report / mapping dropdown
        fy_code = (request.POST.get("fy") or request.GET.get("fy") or "2024").strip()

        # 1️⃣ first try: real Busy name from Busy data
        busy_ledger_name = get_busy_bank_display_name(acc_obj, fy_code)

        # 2️⃣ fallback: if Busy lookup failed, use DB fields
        if not busy_ledger_name:
            # prefer mapping code, else our own account/bank_name
            busy_ledger_name = acc_obj.busyacccode or acc_obj.account or acc_obj.bank_name

    except AccountBank.DoesNotExist:
        busy_ledger_name = None
    # Read arrays sent as multiple keys ...AND... gracefully handle a single comma-joined string
    file_paths = request.POST.getlist("file_paths[]")
    if len(file_paths) == 1 and "," in file_paths[0]:
        file_paths = [p.strip() for p in file_paths[0].split(",") if p.strip()]

    raw_passwords = request.POST.getlist("passwords[]")
    if len(raw_passwords) == 1 and "," in raw_passwords[0]:
        raw_passwords = [p.strip() for p in raw_passwords[0].split(",") if p.strip()]

    # unique, non-empty passwords
    passwords = []
    for entry in raw_passwords:
        for pw in entry.split(","):
            clean_pw = pw.strip()
            if clean_pw and clean_pw not in passwords:
                passwords.append(clean_pw)

    mode = request.POST.get("mode")  # 'bank' or 'busy'

    # 🔒 keep only real, existing PDF files
    file_paths = [
        p for p in file_paths
        if p and os.path.isfile(p) and p.lower().endswith(".pdf")
    ]

    if not file_paths:
        return JsonResponse({"error": "No PDF files selected"}, status=400)

    try:
        # Detect module per file
        detected_modules = {}
        for path in file_paths:
            module = get_module_name(path, passwords)
            print(f"[detect] File: {path} -> Module: {module}")
            detected_modules[path] = module

        unique_modules = list(set(detected_modules.values()))
        print("[detect] Unique modules:", unique_modules)

        # if "Module Not Found" in unique_modules:
        #     return JsonResponse({"error": "❌ Some PDFs couldn't be recognized. Check the format."}, status=400)
        if "Module Not Found" in unique_modules:

            # ✅ Store failure in DB
            for path in file_paths:
                PdfConvertFailure.objects.create(
                    client=acc_obj.client if acc_obj else None,
                    group=acc_obj.client.group if acc_obj and acc_obj.client else None,
                    account=acc_obj,
                    attachment_name=os.path.basename(path),
                    attachment_path=path,
                    module_name="Unknown",
                    error_message="PDF format not recognized"
                )

            return JsonResponse({"error": "❌ Some PDFs couldn't be recognized. Check the format."}, status=400)


        if "Passwords failed" in unique_modules:

            for path in file_paths:
                PdfConvertFailure.objects.create(
                    client=acc_obj.client if acc_obj else None,
                    group=acc_obj.client.group if acc_obj and acc_obj.client else None,
                    account=acc_obj,
                    attachment_name=os.path.basename(path),
                    attachment_path=path,
                    module_name="Password Error",
                    error_message="Incorrect PDF password"
                )

            return JsonResponse({"error": "❌ Password incorrect or missing."}, status=400)

        # if "Passwords failed" in unique_modules:
        #     return JsonResponse({"error": "❌ Password incorrect or missing."}, status=400)
        if "Check Path" in unique_modules:
            return JsonResponse({"error": "❌ Invalid file path."}, status=400)

        # For simplicity, allow only one module at a time
        # if len(unique_modules) != 1:
        #     return JsonResponse(
        #         {"error": f"❌ Multiple modules found: {unique_modules}. Select the same bank/module."},
        #         status=400
        #     )

        if len(unique_modules) != 1:

            for path in file_paths:
                PdfConvertFailure.objects.create(
                    client=acc_obj.client if acc_obj else None,
                    group=acc_obj.client.group if acc_obj and acc_obj.client else None,
                    account=acc_obj,
                    attachment_name=os.path.basename(path),
                    attachment_path=path,
                    module_name=str(unique_modules),
                    error_message="Multiple bank formats detected"
                )

            return JsonResponse(
                {"error": f"❌ Multiple modules found: {unique_modules}. Select same bank/module."},
                status=400
            )


        module_name = unique_modules[0]
        module_func = bank_modules.get(module_name)
        if not module_func:
            return JsonResponse({"error": f"❌ Module not implemented: {module_name}"}, status=400)

        # Parse PDFs -> result (your parser should support multiple files)
        result = module_func(file_paths, passwords)

        # Where to save the Excel: same directory as the first PDF
        output_dir = os.path.dirname(file_paths[0])

        if mode == "bank":
            from .pdf2excel.bank_excel import bank_excel
            output_path = bank_excel(result, [], output_dir)  # unchanged
        elif mode == "busy":
            from .pdf2excel.busy_excel import busy_excel
            print("DEBUG BUSY EXPORT:",
                "account_id=", acc_obj.id,
                "busyacccode=", acc_obj.busyacccode,
                "busy_ledger_name=", busy_ledger_name)
            # pass Busy ledger name so it appears in the Bank column
            output_path = busy_excel(
                final_data=result,
                detail=[],
                output_dir=output_dir,
                bank_ledger_name=busy_ledger_name,
            )
        else:
            return JsonResponse({"error": "❌ Unknown export mode"}, status=400)

        if not output_path or not os.path.exists(output_path):
            return JsonResponse({"error": "❌ Excel generation failed (no file returned)."}, status=500)

        return JsonResponse({
            "message": "✅ Excel file generated successfully!",
            "download_url": f"/download-excel/?path={quote(output_path)}"
        })

    except Exception as e:
        import traceback
        traceback.print_exc()

         # ✅ save failure for each PDF
        for path in file_paths:
            PdfConvertFailure.objects.create(
                client=acc_obj.client if acc_obj else None,
                group=acc_obj.client.group if acc_obj and acc_obj.client else None,
                account=acc_obj,
                attachment_name=os.path.basename(path),
                attachment_path=path,
                module_name="Runtime Error",
                error_message=str(e)
            )

        return JsonResponse({"error": f"❌ Exception: {str(e)}"}, status=500)


def download_excel(request):
    raw_path = request.GET.get("path")
    if not raw_path:
        raise Http404("File not found")

    file_path = unquote(raw_path)
    if not os.path.exists(file_path):
        raise Http404("File not found")

    return FileResponse(
        open(file_path, "rb"),
        as_attachment=True,
        filename=os.path.basename(file_path),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# accounts/views.py
from datetime import date
from django.conf import settings
from django.db.models import Count, Q
from django.http import JsonResponse, Http404
from django.shortcuts import render
from .models import Client, Group, IncomeTaxReturn

# ---- Inline choices + helpers ----
STAGE_CHOICES = [
    ("Received", "Received"),
    ("Processing", "Processing"),
    ("Pending", "Pending"),
    ("Filed", "Filed"),
    ("Completed", "Completed"),
]

def get_current_fy():
    """Return Indian FY as 'YYYY-YY' (Apr–Mar)."""
    t = date.today()
    if t.month >= 4:
        start, end = t.year, t.year + 1
    else:
        start, end = t.year - 1, t.year
    return f"{start}-{str(end)[-2:]}"

def _available_fys():
    """fy list like ['2025-26','2024-25', ...] based on configured DBs."""
    fys = []
    for key in settings.DATABASES.keys():
        if key.startswith("fy_"):
            fys.append(key.replace("fy_", "").replace("_", "-"))
    return sorted(fys, reverse=True)

def _fy_to_db(fy_param: str) -> str:
    if not fy_param:
        fy_param = get_current_fy()
    db = f"fy_{fy_param.replace('-', '_')}"
    if db not in settings.DATABASES:
        raise Http404(f"Database '{db}' not configured.")
    return db, fy_param

def get_stage_choices(db_to_use: str):
    """Dynamic stages from DB; fallback to static."""
    try:
        vals = (
            IncomeTaxReturn.objects.using(db_to_use)
            .exclude(stage__isnull=True).exclude(stage__exact="")
            .values_list("stage", flat=True).distinct()
        )
        items = sorted({(v, v) for v in vals})
        return items or STAGE_CHOICES
    except Exception:
        return STAGE_CHOICES

def get_staff_choices(db_to_use=None, active_only=True):
    """
    Build staff list from UserData.username.
    If active_only=True, include users whose out_date is null or in the future.
    """
    qs = UserData.objects.using('default').all()
    if active_only:
        today = date.today()
        qs = qs.filter(Q(out_date__isnull=True) | Q(out_date__gte=today))
    usernames = qs.order_by('username').values_list('username', flat=True)
    return [(u, u) for u in usernames]
# ----------------------------------


from django.shortcuts import render
from django.db.models import Count, Q
from django.http import JsonResponse, Http404
from django.urls import reverse
from collections import defaultdict



def it_return_report(request):
    # ---------- basics ----------
    rtype = (request.GET.get("rtype") or "audit").lower()
    if rtype not in ("audit", "group"):
        rtype = "audit"

    db_to_use, fy_param = _fy_to_db(request.GET.get("fy"))
    q = (request.GET.get("q") or "").strip().lower()
    detail_group_id = request.GET.get("group_id")  # for nested group details

    # map user_id -> username for Staff column
    user_map = dict(UserData.objects.order_by('username').values_list('id', 'username'))

    # ---------- helpers ----------
    def staff_name_from(value, m):
        if value is None:
            return ""
        if hasattr(value, "id"):
            value = value.id
        if isinstance(value, str) and value.isdigit():
            value = int(value)
        return m.get(value, "") if isinstance(value, int) else (str(value) if isinstance(value, str) else "")

    def status_badge_html(status_value):
        s = (status_value or "").strip()
        if not s:
            return ""
        cls = "badge-active" if s.lower() == "active" else "badge-hold"
        return f'<span class="badge {cls} badge-status">{s}</span>'

    def audit_applicable_text(obj):
        try:
            txt = obj.get_audit_applicable_display()
            return txt or ""
        except Exception:
            return getattr(obj, "audit_applicable", "") or ""

    def audit_status_text(obj):
        try:
            # if the model field has choices, this returns the human label
            return obj.get_audit_status_display() or ""
        except Exception:
            return getattr(obj, "audit_status", "") or ""


    def edit_btn_html(it_id, stage):
        is_filed = (str(stage or "").strip().lower() == "filed")
        if is_filed:
            # disabled/locked icon, no navigation
            return (
                '<button class="btn btn-secondary btn-sm btn-action" '
                'title="Locked (Filed)" disabled>'
                '<i class="bi bi-lock-fill"></i></button>'
            )
        # normal edit
        url = f"{reverse('it_return_form')}?mode=modify&it_id={it_id}&fy={fy_param}"
        return (
            f'<button class="btn btn-primary btn-sm btn-action" title="Edit" '
            f'onclick="location.href=\'{url}\'"><i class="bi bi-pencil"></i></button>'
        )


    def row_matches_q(row_dict):
        """Global search across visible text columns."""
        if not q:
            return True
        import re

        def as_text(val):
            if val is None:
                return ""
            s = str(val)
            if "<" in s and ">" in s:
                s = re.sub(r"<[^>]*>", "", s)
            return s.lower().strip()

        for key in [
            "group", "legal_name", "ca", "staff", "stage", "audit_applicable",
            "return_type", "status_html", "flag", "form", "foot_print", "actions_html"
        ]:
            if q in as_text(row_dict.get(key, "")):
                return True
        return False

    # ---- Revised number helpers ----
    def _build_rev_map(qs):
        """
        Returns {entry_id: display_revision_no} for all Revised entries.
        If revised_no is missing, assigns sequential numbers per (client_id, year).
        """
        groups = defaultdict(list)
        for e in qs:
            if (getattr(e, "return_type", "") or "").lower() == "revised":
                groups[(e.client_id, e.year)].append(e)

        out = {}
        for _, items in groups.items():
            # numbereds first by their number, then unnumbered by id
            def _to_int(v):
                try:
                    return int(v)
                except Exception:
                    return None

            items.sort(key=lambda x: ((_to_int(x.revise_no) or 10**9), x.id))
            used = set()
            next_no = 1
            for e in items:
                n = _to_int(getattr(e, "revised_no", None))
                if not n:
                    # assign next available
                    while next_no in used:
                        next_no += 1
                    n = next_no
                used.add(n)
                next_no = max(next_no, n + 1)
                out[e.id] = n
        return out

    def _fmt_return_type(e, rev_map):
        # base label from choices
        try:
            label = e.get_return_type_display() or ""
        except Exception:
            label = getattr(e, "return_type", "") or ""

        if (getattr(e, "return_type", "") or "").lower() == "revised":
            n = getattr(e, "revised_no", None)
            try:
                n = int(n) if n not in (None, "", 0) else None
            except Exception:
                n = None
            if not n:
                n = rev_map.get(e.id)
            if n:
                return f"{label}({n})"   # <--- parentheses style
        return label

    fiscal_years = _available_fys()
    groups_qs = (Group.objects
                 .annotate(total_it_clients=Count('clients', filter=Q(clients__it_return=True)))
                 .order_by('group_name'))

    # ---------- AJAX ----------
    if request.headers.get("x-requested-with") == "XMLHttpRequest" or request.GET.get("format") == "json":

        if rtype == "audit":
            columns = [
                "Group", "Legal Name", "CA", "Staff", "Stage",
                "Audit Applicable", "Audit Status", "Return Type",
                "Flag if Any", "Form", "Foot Print", "Edit"
            ]
            fields = [
                "group", "legal_name", "ca", "staff", "stage",
                "audit_applicable", "status_html", "return_type",
                "flag", "form", "foot_print", "actions_html"
            ]

            # preload client info (group + legal name)
            client_map = {
                c.id: {
                    "group": (c.group.group_name if c.group else ""),
                    "legal_name": c.legal_name or "",
                }
                for c in (Client.objects.using('default')
                          .filter(it_return=True)
                          .select_related('group'))
            }

            qs = IncomeTaxReturn.objects.using(db_to_use).all()
            rev_map = _build_rev_map(qs)

            def s_key(e):
                cm = client_map.get(e.client_id) or {}
                return (cm.get("group", ""), cm.get("legal_name", ""))

            rows = []
            for e in sorted(qs, key=s_key):
                info = client_map.get(e.client_id)
                if not info:
                    continue

                row = {
                    "group": info["group"],
                    "legal_name": info["legal_name"],
                    "ca": e.ca or "",
                    "staff": staff_name_from(getattr(e, "alloted_to_id", None), user_map),
                    "stage": e.stage or "",
                    "audit_applicable": audit_applicable_text(e),
                    # "status_html": status_badge_html(e.audit_status or ""),
                    "status_html": status_badge_html(audit_status_text(e)),

                    "return_type": _fmt_return_type(e, rev_map),   # <<< Revised(n)
                    "flag": e.flag or "",
                    "form": "",
                    "foot_print": "",
                    # "actions_html": edit_btn_html(e.id),
                    "actions_html": edit_btn_html(e.id, e.stage),   # ← pass stage here
                }
                if row_matches_q(row):
                    rows.append(row)

            return JsonResponse({"columns": columns, "fields": fields, "rows": rows})

        # ===== GROUP TABLE (summary + expandable details) =====
        # ---- DETAIL for a single group (when clicked) ----
        if detail_group_id:
            try:
                g = Group.objects.get(id=int(detail_group_id))
            except (Group.DoesNotExist, ValueError, TypeError):
                raise Http404("Group not found")

            # clients in this group who have ITR enabled
            client_qs = Client.objects.using('default').filter(group_id=g.id, it_return=True)
            client_ids = list(client_qs.values_list("id", flat=True))

            # pre-map client_id -> legal_name for quick lookup
            client_name = {cid: name for cid, name in client_qs.values_list("id", "legal_name")}

            # all entries for those clients in the selected FY DB
            entries = (IncomeTaxReturn.objects
                       .using(db_to_use)
                       .filter(client_id__in=client_ids))
            rev_map = _build_rev_map(entries)

            columns = ["Legal Name", "CA", "Staff", "Stage",
                       "Audit Applicable", "Status", "Return Type", "Edit"]
            fields = ["legal_name", "ca", "staff", "stage",
                      "audit_applicable", "status_html", "return_type", "actions_html"]

            detail_rows = []
            for e in entries.order_by("client_id", "id"):
                detail_rows.append({
                    "legal_name": client_name.get(e.client_id, ""),
                    "ca": e.ca or "",
                    "staff": staff_name_from(getattr(e, "alloted_to_id", None), user_map),
                    "stage": e.stage or "",
                    "audit_applicable": audit_applicable_text(e),
                    "status_html": status_badge_html(e.audit_status or ""),
                    "return_type": _fmt_return_type(e, rev_map),  # <<< Revised(n)
                    # "actions_html": edit_btn_html(e.id),
                    "actions_html": edit_btn_html(e.id, e.stage),   # ← pass stage here too
                })

            return JsonResponse({
                "detail_for": g.id,
                "group_name": g.group_name,
                "columns": columns,
                "fields": fields,
                "rows": detail_rows,
            })

        # ---- SUMMARY (default group-wise list) ----
        client_ids_with_itr = set(
            IncomeTaxReturn.objects.using(db_to_use).values_list("client_id", flat=True).distinct()
        )
        PRE_WORK_STAGES = ["Sleeping", "Call for data"]

        # map client_id -> group_id from DEFAULT DB (no join)
        client_to_group = dict(
            Client.objects.using('default').values_list('id', 'group_id')
        )


      # helper to count distinct clients per group
        def count_by_group(client_ids):
            per_group = defaultdict(int)
            for cid in client_ids:
                gid = client_to_group.get(cid)
                if gid:
                    per_group[gid] += 1
            return per_group

        # base queryset from FY DB; annotate a trimmed stage to survive trailing spaces
        base = IncomeTaxReturn.objects.using(db_to_use).annotate(stage_clean=Trim("stage"))

        pre_work_clients = set(
            base.filter(stage_clean__in=PRE_WORK_STAGES)
                .values_list("client_id", flat=True).distinct()
        )
        working_clients = set(
            base.filter(stage_clean__iexact="Query to client")
                .values_list("client_id", flat=True).distinct()
        )
        pend_check_clients = set(
            base.filter(stage_clean__iexact="Pending for checking")
                .values_list("client_id", flat=True).distinct()
        )
        pend_approve_clients = set(
            base.filter(stage_clean__iexact="Pending for approve")
                .values_list("client_id", flat=True).distinct()
        )
        pend_filing_clients = set(
            base.filter(stage_clean__iexact="Pending for filing")
                .values_list("client_id", flat=True).distinct()
        )
        filed_clients = set(
            base.filter(stage_clean__iexact="Filed")
                .values_list("client_id", flat=True).distinct()
        )

        # turn each client set into a per-group counter
        pre_work_cnt    = count_by_group(pre_work_clients)
        working_cnt     = count_by_group(working_clients)
        pend_check_cnt  = count_by_group(pend_check_clients)
        pend_approve_cnt= count_by_group(pend_approve_clients)
        pend_filing_cnt = count_by_group(pend_filing_clients)
        filed_cnt       = count_by_group(filed_clients)


        data = []
        for g in groups_qs:
            total_it = g.total_it_clients or 0
            group_clients = list(g.clients.filter(it_return=True).values_list("id", flat=True))
            with_entry = sum(1 for cid in group_clients if cid in client_ids_with_itr)
            # s = stage_by_group.get(g.id, {})

            row = {
                "group": g.group_name,
                "total_it_clients": total_it,
                "with_entry_in_fy": with_entry,
                "pre_work":pre_work_cnt.get(g.id, 0),
                "working":working_cnt.get(g.id, 0),
                "pend_check":pend_check_cnt.get(g.id, 0),
                "pend_approve":pend_approve_cnt.get(g.id, 0),
                "pend_filing": pend_filing_cnt.get(g.id, 0),
                "filed":filed_cnt.get(g.id, 0),
                "group_id":g.id,  # for JS to open detail
            }
            if not q or q in g.group_name.lower():
                data.append(row)

        return JsonResponse({
            "columns": ["Group", "Total IT Clients", "Have Entry in FY","Pre-work", "Working", "Pending Check", "Pending Approve", "Pending Filing", "Filed"],
            "fields": ["group", "total_it_clients", "with_entry_in_fy",   "pre_work", "working", "pend_check", "pend_approve", "pend_filing", "filed"],
            "rows": data,
        })

    # ---------- initial page ----------
    return render(request, "it_return_report.html", {
        "fiscal_years": fiscal_years,
        "selected_fy": fy_param,
        "default_rtype": rtype,
    })



# app/views.py
import os, re, time
from urllib.parse import quote, unquote
import pandas as pd

from django.conf import settings
from django.contrib import messages
from django.shortcuts import render
from django.urls import reverse

# === your existing imports (keep these) ===
from .pdf2excel.find_module_name import get_module_name
from .pdf2excel.bank_excel import bank_excel
from .pdf2excel.busy_excel import busy_excel
from .pdf2excel import bank_modules   # dict: {module_name: callable}

# ---------------- helpers ----------------

def _parse_passwords(raw: str) -> list[str]:
    """Accept comma/newline separated passwords; de-dup; preserve order."""
    if not raw:
        return []
    parts = re.split(r"[\n,]+", raw)
    out, seen = [], set()
    for p in (x.strip() for x in parts):
        if p and p not in seen:
            out.append(p); seen.add(p)
    return out

def _save_uploads(files) -> tuple[str, list[str]]:
    """
    Save uploaded PDFs under MEDIA_ROOT/bank_uploads/{timestamp}/
    Returns (folder, [fullpaths]).
    """
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    folder = os.path.join(settings.MEDIA_ROOT, "bank_uploads", timestamp)
    os.makedirs(folder, exist_ok=True)

    saved = []
    for f in files:
        name = os.path.basename(f.name)
        root, ext = os.path.splitext(name)
        if ext.lower() != ".pdf":
            continue
        safe = re.sub(r"[^A-Za-z0-9_.-]+", "_", root) + ".pdf"
        full = os.path.join(folder, safe)
        with open(full, "wb") as dst:
            for chunk in f.chunks():
                dst.write(chunk)
        saved.append(full)
    return folder, saved

def _excel_preview_tables(xlsx_path: str, limit_rows: int | None = None):
    """
    Read the generated Excel and shape for template rendering.
    Returns: [{"sheet": sheet_name, "columns":[...], "rows":[[...], ...]}, ...]
    """
    tables = []
    xl = pd.ExcelFile(xlsx_path)
    for sheet in xl.sheet_names:
        df = xl.parse(sheet)
        if limit_rows:
            df = df.head(limit_rows)
        df = df.fillna("")
        cols = [str(c) for c in df.columns]
        rows = df.astype(str).values.tolist()
        tables.append({"sheet": sheet, "columns": cols, "rows": rows})
    return tables

# ---------------- main upload view ----------------
#main fully work code 
def bank_pdf_upload(request):
    """
    Upload PDF(s) -> detect bank module -> parse -> write Excel -> preview -> download.
    NO account id required.
    """
    context = {
        "detected_module": None,
        "download_url": None,
        "tables": [],
        "saved_folder": None,
    }

    if request.method == "POST":
        mode = (request.POST.get("mode") or "bank").strip().lower()  # 'bank' or 'busy'
        passwords = _parse_passwords(request.POST.get("passwords", ""))
        files = request.FILES.getlist("pdfs")

        if not files:
            messages.error(request, "Please choose at least one PDF.")
            return render(request, "bank_pdf_upload.html", context)

        # Save uploads
        folder, file_paths = _save_uploads(files)
        context["saved_folder"] = folder

        # Validate PDFs
        file_paths = [p for p in file_paths if os.path.isfile(p) and p.lower().endswith(".pdf")]
        if not file_paths:
            messages.error(request, "No valid PDF files were uploaded.")
            return render(request, "bank_pdf_upload.html", context)

        # Detect parser module for each file
        detected = {}
        for path in file_paths:
            mod = get_module_name(path, passwords)  # tries passwords internally
            detected[path] = mod

        unique = list(set(detected.values()))
        if "Module Not Found" in unique:
            messages.error(request, "❌ Some PDFs couldn't be recognized. Check the format.")
            return render(request, "bank_pdf_upload.html", context)
        if "Passwords failed" in unique:
            messages.error(request, "❌ Password incorrect or missing.")
            return render(request, "bank_pdf_upload.html", context)
        if "Check Path" in unique:
            messages.error(request, "❌ Invalid file path.")
            return render(request, "bank_pdf_upload.html", context)
        if len(unique) != 1:
            messages.error(request, f"❌ Multiple modules found: {unique}. Upload same bank format together.")
            return render(request, "bank_pdf_upload.html", context)

        module_name = unique[0]
        context["detected_module"] = module_name

        module_func = bank_modules.get(module_name)
        if not module_func:
            messages.error(request, f"❌ Module not implemented: {module_name}")
            return render(request, "bank_pdf_upload.html", context)

        try:
            # Parse using your bank module (supports list of files + passwords)
            parsed = module_func(file_paths, passwords)
            # Debug: ensure we got tables and rows
            try:
                print("[DEBUG] tables:", len(parsed))
                if parsed and isinstance(parsed[0], list):
                    print("[DEBUG] first table rows:", len(parsed[0]) - 2)  # minus account row + headers
            except Exception as _:
                pass

            # Write Excel using your existing writers
            if mode == "bank":
                xlsx_path = bank_excel(parsed, [], folder)
            elif mode == "busy":
                xlsx_path = busy_excel(parsed, [], folder)
            else:
                messages.error(request, "Unknown mode selected.")
                return render(request, "bank_pdf_upload.html", context)

            if not xlsx_path or not os.path.exists(xlsx_path):
                messages.error(request, "Excel generation failed.")
                return render(request, "bank_pdf_upload.html", context)

            # Preview the generated Excel
            context["tables"] = _excel_preview_tables(xlsx_path)  # use limit_rows=100 if huge
            context["download_url"] = f"{reverse('download_excel')}?path={quote(xlsx_path)}"
            messages.success(request, "✅ Parsed and generated Excel successfully.")

        except Exception as e:
            import traceback
            traceback.print_exc()
            messages.error(request, f"❌ Error: {e}")

    return render(request, "bank_pdf_upload.html", context)





def footprint_local(request, key="default"):
    """
    key = any string to separate pages (e.g. ITR id, client id, or 'default')
    Each key writes to a different localStorage bucket on the client.
    """
    return render(request, "footprint_local.html", {"storage_key": f"footprint:{key}"})


# views.py
import io, csv, re
from pathlib import Path
from datetime import date
from django.conf import settings
from django.contrib import messages
from django.shortcuts import redirect
from django.utils import timezone
from django.views.decorators.http import require_POST

# Excel support (install: pip install pandas openpyxl)
try:
    import pandas as pd
except Exception:
    pd = None

from .models import IncomeTaxReturn, Client  # adjust import if Client is in another app
from .utills import get_db_for_fy, fiscal_year_range  # <-- change to .utils if that's your filename

# ---- CONFIG: where Client lives and PAN field name there ----
CLIENT_DB_ALIAS = "default"
CLIENT_PAN_FIELD = "pan"

# ---- Helpers ----
PAN_RE = re.compile(r'^[A-Z]{5}[0-9]{4}[A-Z]$')
ACK_DIGITS = re.compile(r"\D+")
MOBILE_10_RE = re.compile(r"^[6-9]\d{9}$")           # common India mobile
ACK_14_16_RE = re.compile(r"\b\d{14,16}\b")          # e-filing ack usually ~15 digits
ACK_12_20_RE = re.compile(r"\b\d{12,20}\b")          # fallback width

def _norm_pan(s: str) -> str:
    return (s or "").strip().upper().replace(" ", "")

def _valid_pan(s: str) -> bool:
    return bool(PAN_RE.match(_norm_pan(s)))

def _clean_ack(val: str) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    if s.lower() in ("nan", "none"):
        return ""
    s = ACK_DIGITS.sub("", s)
    return s[:20]

def _parse_date_flexible(s: str):
    if not s:
        return None
    s = str(s).strip()
    if pd is not None:
        try:
            dt = pd.to_datetime(s, dayfirst=True, errors="raise")
            return date(dt.year, dt.month, dt.day)
        except Exception:
            pass
    m = re.match(r"^\s*(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})\s*$", s)
    if m:
        d, mth, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        y = 2000 + y if y < 100 else y
        try:
            return date(y, mth, d)
        except Exception:
            return None
    return None

def _norm_headers(cols):
    return [str(c or "").strip().lower() for c in cols]

def _pick(headers, candidates):
    for h in headers:
        if h in candidates or any(tok in h for tok in candidates):
            return h
    return None

def _score_ack_series(series):
    """Prefer 14–16 digit numbers; downweight 10-digit mobiles."""
    s14_16 = 0
    s12_20 = 0
    s_mobile = 0
    for v in series.astype(str):
        vv = v.strip()
        if not vv:
            continue
        if ACK_14_16_RE.search(vv):
            s14_16 += 1
        elif ACK_12_20_RE.search(vv):
            s12_20 += 1
        elif MOBILE_10_RE.fullmatch(vv):
            s_mobile += 1
    # Strong weight to 14–16; weak to 12–20; penalize mobiles
    return (3 * s14_16) + (1 * s12_20) - (2 * s_mobile)

@require_POST
def it_upload_status(request):
    fy = (request.POST.get("fy") or "").strip()
    f = request.FILES.get("file")
    if not fy or not f:
        messages.error(request, "Please select a Fiscal Year and choose a file.")
        return redirect(request.META.get("HTTP_REFERER", "/it-returns/"))

    fy_db = get_db_for_fy(fy)
    if fy_db not in settings.DATABASES:
        messages.error(request, f"Database alias '{fy_db}' is not defined.")
        return redirect(request.META.get("HTTP_REFERER", "/it-returns/"))

    fy_start, fy_end = fiscal_year_range(fy)

    rows = []
    ext = Path(f.name).suffix.lower()

    COLS = {
        "pan": {"pan", "p.a.n", "pan no", "pan no.", "panno", "pan_number", "pan number"},
        "ack": {"ack", "ack no", "ack no.", "ack_no", "ack. no.", "acknowledgement",
                "acknowledgement no", "acknowledgement_number", "acknowledgment", "acknowledgment no"},
        "file_date": {"date of filing", "filing date", "ack date", "ack. date",
                      "acknowledgement date", "process date"},
    }

    try:
        if ext in (".xlsx", ".xls"):
            if pd is None:
                messages.error(request, "Excel upload requires 'pandas' and 'openpyxl'.")
                return redirect(request.META.get("HTTP_REFERER", "/it-returns/"))

            # ---------- Robust header row detection ----------
            raw = pd.read_excel(f, header=None, dtype=str, keep_default_na=False)
            header_idx = None
            scan_rows = min(15, len(raw))
            for i in range(scan_rows):
                r = [str(x).strip().lower() for x in raw.iloc[i].tolist()]
                if any("pan" in c for c in r) and any(("ack" in c) or ("ack." in c) or ("ackno" in c) or ("acknow" in c) for c in r):
                    header_idx = i
                    break
            if header_idx is None:
                # fallback: first row as header
                header_idx = 0

            # Data starts after header row
            df = raw.iloc[header_idx + 1:].copy()
            df.columns = _norm_headers(raw.iloc[header_idx].tolist())

            headers = list(df.columns)
            pan_col = _pick(headers, COLS["pan"])
            ack_col = _pick(headers, COLS["ack"])
            date_col = _pick(headers, COLS["file_date"])

            # Fallback by content if needed
            if pan_col is None:
                best, sc = None, -1
                for c in df.columns:
                    s = sum(1 for v in df[c].astype(str) if _valid_pan(v))
                    if s > sc:
                        best, sc = c, s
                pan_col = best if sc > 0 else None

            if ack_col is None:
                best, sc = None, -1
                for c in df.columns:
                    cname = str(c).lower()
                    s = (3 if "ack" in cname else 0) + _score_ack_series(df[c])
                    if s > sc:
                        best, sc = c, s
                ack_col = best if sc > 0 else None

            if pan_col is None or ack_col is None:
                messages.error(request, "Could not detect PAN/Ack columns in the sheet.")
                return redirect(request.META.get("HTTP_REFERER", "/it-returns/"))

            for _, r in df.iterrows():
                pan = _norm_pan(r.get(pan_col, ""))
                ack = _clean_ack(r.get(ack_col, ""))
                ackd = _parse_date_flexible(r.get(date_col, "")) if date_col else None
                if pan:
                    rows.append({"pan": pan, "ack": ack, "ack_date": ackd})

        else:
            # ---------- CSV/TXT path (headers already on first row) ----------
            data = io.TextIOWrapper(f.file, encoding="utf-8", errors="ignore")
            rdr = csv.DictReader(data)
            headers = _norm_headers(rdr.fieldnames or [])
            pan_h = _pick(headers, COLS["pan"])
            ack_h = _pick(headers, COLS["ack"])
            date_h = _pick(headers, COLS["file_date"])
            if pan_h is None or ack_h is None:
                messages.error(request, "CSV must contain PAN and Ack columns.")
                return redirect(request.META.get("HTTP_REFERER", "/it-returns/"))
            for r in rdr:
                pan = _norm_pan(r.get(pan_h, "") or "")
                ack = _clean_ack(r.get(ack_h, "") or "")
                ackd = _parse_date_flexible((r.get(date_h, "") or "")) if date_h else None
                if pan:
                    rows.append({"pan": pan, "ack": ack, "ack_date": ackd})
    except Exception as e:
        messages.error(request, f"Failed to read file: {e}")
        return redirect(request.META.get("HTTP_REFERER", "/it-returns/"))

    if not rows:
        messages.warning(request, "No usable PAN/Ack rows found.")
        return redirect(request.META.get("HTTP_REFERER", "/it-returns/"))

    # --- Build PAN -> [client_id,..] map FROM THE CLIENT DB (no cross-DB join) ---
    pans = sorted({r["pan"] for r in rows if _valid_pan(r["pan"])})
    pan_to_client_ids = {p: [] for p in pans}
    if pans:
        client_qs = Client.objects.using(CLIENT_DB_ALIAS).filter(**{f"{CLIENT_PAN_FIELD}__in": pans})
        for cid, cpan in client_qs.values_list("id", CLIENT_PAN_FIELD):
            pan_to_client_ids[_norm_pan(cpan)].append(cid)

    # --- Update IncomeTaxReturn in FY DB by client_id + FY range ---
    fy_start, fy_end = fiscal_year_range(fy)
    today = timezone.now().date()
    updated = missing = invalid = 0

    itr_mgr = IncomeTaxReturn.objects.using(fy_db)

    for item in rows:
        pan = item["pan"]
        ack_raw = item["ack"]
        ack = _clean_ack(ack_raw)
        ack_date_val = item.get("ack_date") or today

        if not _valid_pan(pan):
            invalid += 1
            continue

        client_ids = pan_to_client_ids.get(pan, [])
        if not client_ids:
            missing += 1
            continue

        qs = itr_mgr.filter(
            client_id__in=client_ids,
            year__gte=fy_start,
            year__lte=fy_end,
        )

        if not qs.exists():
            missing += 1
            continue

        for itr in qs.select_related("client"):
            print(
                f"[ITR Update] FY={fy}, DB={fy_db}, "
                f"ClientID={itr.client_id}, ClientName={itr.client}, "
                f"PAN={pan}, AckRaw='{ack_raw}', AckStored='{ack}', AckDate={ack_date_val}"
            )

        updates = {"stage": "Filed", "ack_date": ack_date_val}
        if ack:
            updates["ack_no"] = ack

        count = qs.update(**updates)
        updated += count

    messages.success(request, f"Updated: {updated}, Missing: {missing}, Invalid PANs: {invalid}.")
    return redirect(request.META.get("HTTP_REFERER", "/it-returns/"))




from django.http import JsonResponse, HttpResponseBadRequest
from .utills import fy_to_daterange   # <-- import the new helper
from django.db.models.functions import Trim

def it_user_list_json(request):
    # Users table is in DEFAULT DB
    users = list(UserData.objects.using("default").order_by("username").values("id", "username"))
    return JsonResponse({"users": users})

# helpers to select the right DB alias for the FY

import re
from django.conf import settings

def normalize_fy(fy: str) -> str:
    # "2025-26" / "2025_26" -> "2025_26"
    fy = (fy or "").strip()
    if not fy:
        return ""
    m = re.match(r"^(\d{4})[ _-]?(\d{2})$", fy)
    if not m:
        return fy
    return f"{m.group(1)}_{m.group(2)}"

def get_db_alias_for_fy(request, fy_str: str) -> str | None:
    """
    Decide which DB alias to use for a given FY.
    Adjust this to match how you name per-year databases.

    Examples:
      FY "2025-26" -> alias "fy_2025_26"
      or "it_2025" (if you keep only the start year),
      or take from session if you store it there.
    """
    # 1) If you store active alias in session (recommended)
    alias = request.session.get("active_db_alias")
    if alias and alias in settings.DATABASES:
        return alias

    # 2) Compute from FY; adjust to your scheme
    fy_norm = normalize_fy(fy_str)  # "2025_26"
    candidates = [
        f"fy_{fy_norm}",           # e.g., "fy_2025_26"
        f"db_{fy_norm}",           # e.g., "db_2025_26"
    ]

    # Or, if your alias uses only start-year (e.g. "fy_2025"), include:
    try:
        start_year = int(fy_norm[:4])
        candidates.append(f"fy_{start_year}")    # e.g., "fy_2025"
        candidates.append(f"db_{start_year}")    # e.g., "db_2025"
    except Exception:
        pass

    for a in candidates:
        if a in settings.DATABASES:
            return a

    # 3) As a last resort, default (but usually wrong for you)
    return None

def _safe_json(obj):
    return json.dumps(obj, default=str, indent=2)

@require_POST
def it_bulk_reassign(request):
    from_id = request.POST.get("from_user")
    to_id   = request.POST.get("to_user")
    fy_str  = (request.POST.get("fy") or "").strip()
    only_active = request.POST.get("only_active") == "1"

    if not from_id or not to_id:
        return HttpResponseBadRequest("from_user and to_user are required.")
    if from_id == to_id:
        return HttpResponseBadRequest("Source and destination users must be different.")
    try:
        from_id = int(from_id); to_id = int(to_id)
    except ValueError:
        return HttpResponseBadRequest("Invalid user id(s).")

    # Users always from DEFAULT DB
    from_user = UserData.objects.using("default").filter(id=from_id).values("id", "username").first()
    to_user   = UserData.objects.using("default").filter(id=to_id).values("id", "username").first()
    if not from_user or not to_user:
        return HttpResponseBadRequest("Invalid user id(s).")

    # FY is required -> date window
    rng = fy_to_daterange(fy_str)
    if not rng:
        return HttpResponseBadRequest("Invalid fiscal year format. Example: 2025-26")
    start, end = rng

    # Resolve which DB alias holds IncomeTaxReturn for this FY
    alias = get_db_alias_for_fy(request, fy_str)
    if not alias:
        return HttpResponseBadRequest("No database configured for this fiscal year.")

    # Build queries against the YEAR DB
    all_rows = IncomeTaxReturn.objects.using(alias).all()

    total_rows = all_rows.count()
    total_for_from_any_fy = all_rows.filter(alloted_to_id=from_id).count()
    total_in_fy_any_user = all_rows.filter(year__gte=start, year__lte=end).count()

    # NO cross-DB joins: only use alloted_to_id here
    # If you want usernames for debug, resolve ids via default DB below
    distinct_users_in_fy_ids = list(
        all_rows.filter(year__gte=start, year__lte=end)
                .values("alloted_to_id")
                .annotate(n=Count("id"))
                .order_by("-n")[:20]
    )
    # add usernames from default DB for diagnostics
    id_list = [row["alloted_to_id"] for row in distinct_users_in_fy_ids if row["alloted_to_id"]]
    id_to_name = dict(UserData.objects.using("default")
                      .filter(id__in=id_list).values_list("id", "username"))
    distinct_users_in_fy = [
        {"alloted_to_id": r["alloted_to_id"], "username": id_to_name.get(r["alloted_to_id"]), "n": r["n"]}
        for r in distinct_users_in_fy_ids
    ]

    # Step filters
    base = all_rows.filter(alloted_to_id=from_id)
    c_from = base.count()
    peek_after_from = list(base.values("id", "year", "alloted_to_id", "return_status")[:10])

    in_fy = base.filter(year__gte=start, year__lte=end)
    c_fy = in_fy.count()
    peek_after_fy = list(in_fy.values("id", "year", "alloted_to_id", "return_status")[:10])

    qs = in_fy
    if only_active:
        qs = qs.annotate(_ret_status=Trim("return_status")).filter(_ret_status__iexact="Active")
    c_status = qs.count()
    peek_after_status = list(qs.values("id", "year", "alloted_to_id", "return_status")[:10])

    # Update takes place in the YEAR DB
    updated = 0
    if c_status > 0:
        updated = qs.update(alloted_to_id=to_id)

    diag = {
        "db_alias_used": alias,
        "post": {"from_id": from_id, "to_id": to_id, "fy": fy_str, "range": [start, end], "only_active": only_active},
        "from_user": from_user, "to_user": to_user,
        "totals": {
            "total_rows": total_rows,
            "total_for_from_any_fy": total_for_from_any_fy,
            "total_in_fy_any_user": total_in_fy_any_user,
        },
        "distinct_users_in_fy": distinct_users_in_fy,
        "matched_after_from": c_from,
        "matched_after_fy": c_fy,
        "matched_after_status": c_status,
        "peek_after_from": peek_after_from,
        "peek_after_fy": peek_after_fy,
        "peek_after_status": peek_after_status,
    }

    print("\n========== BULK REASSIGN (USING DB:", alias, ") ==========")
    print(_safe_json(diag))
    print("==========================================================\n")

    return JsonResponse({"updated": updated, "diag": diag})



# views.py
from django.views.decorators.http import require_POST
from django.http import JsonResponse, HttpResponseBadRequest
from django.db.models.functions import Trim

def it_group_list_json(request):
    fy = (request.GET.get("fy") or "").strip()
    in_fy_only = request.GET.get("in_fy_only") == "1"

    # pick FY database (no joins here)
    db_alias, _fy_param = _fy_to_db(fy)

    # 1) Which clients actually have rows in THIS FY DB?
    #    (group only by client_id; NO FK traversal)
    client_rows = list(
        IncomeTaxReturn.objects.using(db_alias)
        .values("client_id")
        .annotate(n=Count("id"))
    )
    client_ids_in_fy = [r["client_id"] for r in client_rows]

    # Map client_id -> count (how many ITR rows this FY)
    count_by_client = {r["client_id"]: r["n"] for r in client_rows}

    # 2) Pull client -> group from DEFAULT DB
    #    (only the ones we need if filtering by FY)
    client_qs = Client.objects.using("default").filter(it_return=True)
    if in_fy_only:
        client_qs = client_qs.filter(id__in=client_ids_in_fy)

    client_to_group = dict(
        client_qs.values_list("id", "group_id")
    )

    # 3) Aggregate a count per group_id (only if in_fy_only)
    count_by_group = defaultdict(int)
    if in_fy_only:
        for cid, n in count_by_client.items():
            gid = client_to_group.get(cid)
            if gid:
                count_by_group[gid] += n

    # 4) Build the final group list from DEFAULT DB
    group_ids = list(set(client_to_group.values()))
    groups_qs = Group.objects.using("default").filter(id__in=group_ids).order_by("group_name")

    payload = []
    for g in groups_qs:
        if in_fy_only:
            in_fy = count_by_group.get(g.id, 0)
            if in_fy:
                payload.append({"id": g.id, "name": g.group_name, "in_fy": in_fy})
        else:
            payload.append({"id": g.id, "name": g.group_name})

    return JsonResponse({"groups": payload})

@require_POST
def it_group_reassign(request):
    group_id = request.POST.get("group_id")
    to_id    = request.POST.get("to_user")
    fy_str   = (request.POST.get("fy") or "").strip()
    only_active = request.POST.get("only_active") == "1"

    # validate
    if not group_id or not to_id or not fy_str:
        return HttpResponseBadRequest("group_id, to_user, and fy are required.")
    try:
        group_id = int(group_id); to_id = int(to_id)
    except ValueError:
        return HttpResponseBadRequest("Invalid id(s).")

    # ensure user exists
    if not UserData.objects.filter(id=to_id).exists():
        return HttpResponseBadRequest("Invalid user id.")

    # pick FY database
    db_alias, _fy_param = _fy_to_db(fy_str)

    # clients belonging to this group (default DB)
    client_ids = list(
        Client.objects.using("default")
              .filter(group_id=group_id, it_return=True)
              .values_list("id", flat=True)
    )

    if not client_ids:
        return JsonResponse({"updated": 0, "diag": {"group_id": group_id, "clients": 0}})

    # rows to update (FY DB)
    qs = IncomeTaxReturn.objects.using(db_alias).filter(client_id__in=client_ids)
    if only_active:
        qs = qs.annotate(_ret=Trim("return_status")).filter(_ret__iexact="Active")

    updated = qs.update(alloted_to_id=to_id)

    return JsonResponse({
        "updated": updated,
        "diag": {"group_id": group_id, "clients": len(client_ids)}
    })


# your_app/views_pdf.py
from io import BytesIO
from tempfile import NamedTemporaryFile

from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone

import openpyxl
from openpyxl.utils import get_column_letter

from .m_generic_auto import generic_auto



# def bank_pdf_to_excel(request):
#     """
#     Upload 1..N bank statement PDFs (any bank) + optional comma-separated passwords.
#     Returns a single Excel with one sheet per input PDF.
#     """
#     if request.method == "GET":
#         return render(request, "bank_convert.html")

#     # POST
#     files = request.FILES.getlist("pdfs")
#     if not files:
#         return render(request, "bank_convert.html", {"error": "Please choose at least one PDF."})

#     # Comma/space separated passwords are ok
#     raw_pw = (request.POST.get("passwords") or "").strip()
#     passwords = [p.strip() for p in raw_pw.replace(";", ",").split(",") if p.strip()]

#     # Convert each uploaded PDF → rows (list-of-lists)
#     results = []
#     sheet_names = []
#     for up in files:
#         # save to a temp path because pdfplumber needs a file path/handle
#         with NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
#             for chunk in up.chunks():
#                 tmp.write(chunk)
#             tmp_path = tmp.name

#         rows = generic_auto([tmp_path], passwords)  # ← your generic extractor
#         results.append(rows)

#         base = (up.name.rsplit(".", 1)[0] or "Sheet")[:31]
#         sheet_names.append(base)

#     # Build a single workbook with one sheet per input file
#     wb = openpyxl.Workbook()
#     wb.remove(wb.active)

#     for rows, name in zip(results, sheet_names):
#         ws = wb.create_sheet(title=name[:31])
#         for r_idx, row in enumerate(rows, start=1):
#             for c_idx, cell in enumerate(row, start=1):
#                 ws.cell(row=r_idx, column=c_idx, value=cell)

#         # simple auto-width
#         for col in ws.columns:
#             letter = get_column_letter(col[0].column)
#             maxlen = 10
#             for cell in col:
#                 try:
#                     maxlen = max(maxlen, len(str(cell.value or "")))
#                 except Exception:
#                     pass
#             ws.column_dimensions[letter].width = min(60, maxlen + 2)

#     buf = BytesIO()
#     wb.save(buf)
#     buf.seek(0)

#     fname = f"converted_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
#     resp = HttpResponse(
#         buf.getvalue(),
#         content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#     )
#     resp["Content-Disposition"] = f'attachment; filename="{fname}"'
#     return resp


def bank_pdf_to_excel(request):
    """
    Upload 1..N bank statement PDFs (any bank) + optional comma-separated passwords.
    Returns a single Excel with one sheet per input PDF.
    """
    if request.method == "GET":
        return render(request, "bank_convert.html")

    # POST
    files = request.FILES.getlist("pdfs")
    if not files:
        return render(request, "bank_convert.html", {"error": "Please choose at least one PDF."})

    # Comma/space separated passwords are ok
    raw_pw = (request.POST.get("passwords") or "").strip()
    passwords = [p.strip() for p in raw_pw.replace(";", ",").split(",") if p.strip()]

    # Convert each uploaded PDF → rows (list-of-lists)
    results = []
    sheet_names = []
    for up in files:
        # save to a temp path because pdfplumber needs a file path/handle
        with NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            for chunk in up.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name

        rows = generic_auto([tmp_path], passwords)  # ← your generic extractor
        results.append(rows)

        base = (up.name.rsplit(".", 1)[0] or "Sheet")[:31]
        sheet_names.append(base)

    # Build a single workbook with one sheet per input file
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for rows, name in zip(results, sheet_names):
        ws = wb.create_sheet(title=name[:31])
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, cell in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=cell)

        # simple auto-width
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            maxlen = 10
            for cell in col:
                try:
                    maxlen = max(maxlen, len(str(cell.value or "")))
                except Exception:
                    pass
            ws.column_dimensions[letter].width = min(60, maxlen + 2)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"converted_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp



#fetch mail data
import base64
import re
import imaplib, email
from email.header import decode_header, make_header
from email.utils import parsedate_to_datetime
# from datetime import datetime, timedelta, timezone
import datetime as dt


from django.db import transaction
from django.conf import settings
import os, re
from email.utils import getaddresses  # add this
from django.utils import timezone     # add this

from .models import MailLog  # your table
_SAN_RX = re.compile(r"[^a-z0-9._-]+", re.I)

def _sanitize_email_for_path(email: str) -> str:
    s = (email or "").strip().lower().replace("@", "_at_")
    s = _SAN_RX.sub("_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s or "unknown"

def _parse_recipients(raw_header: str) -> list[str]:
    addrs = []
    for _name, addr in getaddresses([raw_header or ""]):
        addr = (addr or "").strip()
        if addr:
            addrs.append(addr)
    seen = set(); uniq = []
    for a in addrs:
        if a.lower() not in seen:
            uniq.append(a)
            seen.add(a.lower())
    return uniq

def _attachment_rel_path(mailbox_email: str, recipient_email: str, rec_dat_utc: dt.datetime, filename: str) -> str:
    """
    fy_<YYYY_YY>/<mailbox>/<recipient>/<filename>  (FY from *local* date)
    """
    local_d = timezone.localtime(rec_dat_utc).date()
    start_year = local_d.year if local_d.month >= 4 else local_d.year - 1
    fy = f"{start_year}_{(start_year + 1) % 100:02d}"         # e.g. 2025_26
    box = _sanitize_email_for_path(mailbox_email)              # acstkameet_gmail.com
    rec = _sanitize_email_for_path(recipient_email)            # abc_gmail.com
    return os.path.join(f"fy_{fy}", box, rec, filename)

def _ensure_parent_dir(path: str):
    os.makedirs(os.path.dirname(path), exist_ok=True)

IMAP_ACCOUNTS = [
    ("acstkameet@gmail.com", "yafbbshkzxbbpwmr"),
    ("itpamitmehta@yahoo.co.in", "earvolzcpbrhaiqw", "imap.mail.yahoo.com"),
    # ("yourmail2@gmail.com", "app_password_2"),
]

# --- helpers ---------------------------------------------------------------

def _imap_host_for(addr: str, explicit_host: str | None = None) -> str:
    """
    Choose the IMAP host. If explicit_host is given in IMAP_ACCOUNTS, use it.
    Otherwise infer from email domain.
    """
    if explicit_host:
        return explicit_host

    dom = (addr or "").split("@")[-1].lower()
    if "gmail.com" in dom or "googlemail.com" in dom:
        return "imap.gmail.com"
    if dom.endswith("yahoo.com") or dom.endswith("yahoo.co.in") or "yahoo." in dom:
        return "imap.mail.yahoo.com"
    # sensible fallback
    return "imap."+dom

def _normalize_mid(v: str | None, fallback: str) -> str:
    """
    Normalize Message-ID to reduce duplicates across providers.
    - Strip <>, whitespace
    - Lowercase
    """
    if not v:
        return fallback
    v = v.strip().strip("<>").strip().lower()
    return v or fallback


def _dh(v: str) -> str:
    """Decode RFC2047 headers safely."""
    if not v:
        return ""
    try:
        return str(make_header(decode_header(v)))
    except Exception:
        return v

def _fy_alias_from_date(d: dt.datetime) -> str:
    d = d.astimezone(dt.timezone.utc)
    start_year = d.year if d.month >= 4 else d.year - 1
    return f"fy_{start_year}_{(start_year + 1) % 100:02d}"


# ---------- IGNORE FILTERS ----------

# Words/phrases to ignore in SUBJECT (case-insensitive, substring match)
SUBJECT_IGNORE_WORDS = {
    "equity","mutual","growth","uti","margin","funds","ipo","retention","ledger",
    "securities","sebi","demat","holding","policy","portfolio","investor","nse",
    "bill","invoice","prudent","cams","broking","folio","enq_fti","camsonline",
    "adityabirlacapital.mf","insurance","bseindia.in","kfintech.com","privacy",
    "life","liquiloans","credit card"
}

# Sender name / email / domain fragments to ignore (case-insensitive)
SENDER_IGNORE_FRAGMENTS = {
    "paytm","partner","aumassociate","njindiaonline","privacy","motilaloswal","sharekhan",
    "ecas@cdslstatement.com","angeltrade","adityabirlacapital","iciciprulife","utimf@kfintech",
    "risk.management","kotaksecuriti","backoffice@rsec","way2wealth","mutual","camsonline",
    "google","mf@ppfas.com","kfintech","zerodha","pgimindia","prudentbroking","groww",
    "angelbroking","nse.co.in"
}

# Exact/near-exact masked numbers or IDs to ignore if they appear in SUBJECT
SUBJECT_IGNORE_NUMBERS = {
    "***6709","85201003005","xxx017xxxxxxx191x","xxx017xxxxxxx523x","xxxxxxxxxx49",
    "xxxxxxxxxx05"  # appears twice in screenshot—one entry is fine
}

_SAN_ALNUM_RX = re.compile(r"[^a-z0-9]+", re.I)

def _norm_text(s: str) -> str:
    """Lowercase; collapse to plain text for robust contains checks."""
    return (s or "").strip().lower()

def _norm_alnum(s: str) -> str:
    """Lowercase and remove non-alphanumerics for 'masked number' comparison."""
    return _SAN_ALNUM_RX.sub("", (s or "").lower())

def _contains_any(text: str, needles: set[str]) -> bool:
    t = _norm_text(text)
    return any(n in t for n in needles)

def _contains_any_alnum(text: str, needles: set[str]) -> bool:
    """Check after removing punctuation/spaces to catch XXX… patterns."""
    t = _norm_alnum(text)
    return any(_norm_alnum(n) in t for n in needles)

def _sender_matches_ignore(h_from: str) -> bool:
    """
    Return True if From header matches any fragment:
    - display name fragments
    - full email
    - domain
    """
    from email.utils import parseaddr
    disp, addr = email.utils.parseaddr(h_from or "")
    disp_n = _norm_text(disp)
    addr_n = _norm_text(addr)

    # domain part
    dom_n = addr_n.split("@")[-1] if "@" in addr_n else addr_n

    # check fragments against display name, full email, and domain
    for frag in SENDER_IGNORE_FRAGMENTS:
        f = _norm_text(frag)
        if f and (f in disp_n or f in addr_n or f in dom_n):
            return True
    return False

def _subject_matches_ignore(h_sub: str) -> bool:
    """Return True if subject matches any of the ignore rules."""
    # plain word/phrase blocks
    if _contains_any(h_sub, SUBJECT_IGNORE_WORDS):
        return True
    # masked numbers / IDs
    if _contains_any_alnum(h_sub, SUBJECT_IGNORE_NUMBERS):
        return True
    return False
# ---------- END IGNORE FILTERS ----------




def _imap_fmt_date(d: dt.date) -> str:
    """Always English month abbrev, regardless of system locale."""
    MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    return f"{d.day:02d}-{MONTHS[d.month-1]}-{d.year}"

def fetch_imap_mailbox(mailbox_email: str,
                       app_password: str,
                       *,
                       host: str | None = None,
                       since_days: int = 30,
                       from_date: dt.date | None = None,
                       to_date: dt.date | None = None,
                       limit: int = 300,
                       auto_fy: bool = False,
                       db_alias: str = "default",
                       message_id_to_find: str | None = None,
                       to_email: str | None = None,               # NEW
                       subject_contains: str | None = None       # NEW
                       ) -> dict:
    inserted = 0
    updated = 0
    existing = 0

    imap_host = _imap_host_for(mailbox_email, host)
    M = imaplib.IMAP4_SSL(imap_host)
    try:
        M.login(mailbox_email, app_password)

        typ, _ = M.select("INBOX")
        if typ != "OK":
            raise RuntimeError("Could not select INBOX")

       
                # ---- Build criterion (supports id / date range / since_days / to_email / subject)
        if message_id_to_find:
            mid = message_id_to_find.strip()
            mid_wo = mid.strip("<>").strip()
            mid_wi = f"<{mid_wo}>"
            # (OR HEADER Message-ID "<id>" HEADER Message-ID "id")
            criterion = f'(OR HEADER Message-ID "{mid_wi}" HEADER Message-ID "{mid_wo}")'
           
        else:
            parts = []

            # Date range
            if from_date or to_date:
                if from_date:
                    parts.append(f"SINCE {_imap_fmt_date(from_date)}")
                if to_date:
                    # BEFORE is exclusive, so add +1 day
                    parts.append(f"BEFORE {_imap_fmt_date(to_date + dt.timedelta(days=1))}")
            else:
                since_date = (dt.datetime.now(dt.timezone.utc) - dt.timedelta(days=since_days)).date()
                parts.append(f"SINCE {_imap_fmt_date(since_date)}")

            # Receiver filter (TO)
            if to_email:
                # IMAP is picky about quotes; strip internal quotes
                to_q = to_email.replace('"', ' ')
                # parts.append(f'TO "{to_q}"')
                parts.append(f'(OR TO "{to_q}" CC "{to_q}")')

            # Subject contains filter
            if subject_contains:
                subj_q = subject_contains.replace('"', ' ')
                # Most servers (Gmail/Yahoo) support SUBJECT keyword
                parts.append(f'SUBJECT "{subj_q}"')

            criterion = f'({" ".join(parts)})' if parts else "ALL"



        # typ, data = M.search(None, criterion)
        # with:
        try:
            typ, data = M.search("UTF-8", criterion)   # ask server to parse as UTF-8
        except imaplib.IMAP4.error:
            # fallback: try ASCII if server doesn't like charset
            typ, data = M.search(None, criterion)
        if typ != "OK":
            raise RuntimeError(f"Search failed: {criterion}")

        ids = data[0].split()
        if not ids:
            return {"inserted": 0, "updated": 0, "existing": 0, "searched": 0}

        ids = list(reversed(ids))[:limit]

        for num in ids:
            typ, msg_data = M.fetch(num, "(RFC822)")
            if typ != "OK" or not msg_data or not msg_data[0]:
                continue
            raw = msg_data[0][1]
            msg = email.message_from_bytes(raw)

            #########################
            mail_body = ""

            if msg.is_multipart():
                for part in msg.walk():
                    ctype = part.get_content_type()
                    disp = (part.get("Content-Disposition") or "").lower()

                    # Skip attachments
                    if "attachment" in disp:
                        continue

                    if ctype in ("text/html", "text/plain"):
                        payload = part.get_payload(decode=True)
                        if payload:
                            try:
                                text = payload.decode("utf-8", errors="ignore")
                            except:
                                text = payload.decode(errors="ignore")
                            mail_body += text
            else:
                payload = msg.get_payload(decode=True)
                if payload:
                    try:
                        mail_body = payload.decode("utf-8", errors="ignore")
                    except:
                        mail_body = payload.decode(errors="ignore")
            
            # ✅ extract first https link from mail
            # link_match = re.search(r"https?://[^\s\"'>]+", mail_body)
            # statement_link = link_match.group(0) if link_match else None
            # ✅ Match only the SmartStatement link (not images, not banners)
            # link_match = re.search(
            #     r"https://smartstatements\.hdfc\.bank\.in/[^\s\"'>]+",
            #     mail_body,
            #     re.IGNORECASE
            # )

            # -----------------------------------------
            # 🔎 Extract Statement Link (HDFC + BOB)
            # -----------------------------------------

            statement_link = None

            patterns = [
                r"https://smartstatements\.hdfc\.bank\.in/[^\s\"'>]+",
                r"https://estmt\.bankofbaroda\.bank\.in/login/[A-Za-z0-9\-_]+[^\s\"'>]*",
            ]

            for pat in patterns:
                match = re.search(pat, mail_body, re.IGNORECASE)
                if match:
                    statement_link = match.group(0)
                    break


            # statement_link = link_match.group(0) if link_match else None


            h_from = _dh(msg.get("From"))
            h_to_all = msg.get_all("To", []) or []
            h_to = ", ".join(_dh(x) for x in h_to_all)
            h_sub = _dh(msg.get("Subject"))
            # --- IGNORE FILTER: subject / sender
            if _subject_matches_ignore(h_sub) or _sender_matches_ignore(h_from):
                # Skip this message entirely
                continue

            raw_mid = msg.get("Message-ID")
            fallback_mid = f"{mailbox_email}-{num.decode(errors='ignore')}"
            h_mid = _normalize_mid(raw_mid, fallback_mid)

            try:
                rec_dt = parsedate_to_datetime(msg.get("Date"))
                if rec_dt.tzinfo is None:
                    rec_dt = rec_dt.replace(tzinfo=dt.timezone.utc)
                rec_dt = rec_dt.astimezone(dt.timezone.utc)
            except Exception:
                rec_dt = dt.datetime.now(dt.timezone.utc)

            # --- collect attachment parts
            attachment_parts = []
            if msg.is_multipart():
                for part in msg.walk():
                    cd = (part.get("Content-Disposition") or "").lower()
                    fn = part.get_filename()
                    if ("attachment" in cd) or fn:
                        attachment_parts.append(part)

            target_db = _fy_alias_from_date(rec_dt) if auto_fy else db_alias

            with transaction.atomic(using=target_db):
                # create/get MailLog (JSON attachments live here now)
                obj, created = MailLog.objects.using(target_db).get_or_create(
                    mailbox=mailbox_email,
                    msg_id=h_mid,
                    defaults=dict(
                        rec_dat=rec_dt,
                        sender_mail=h_from,
                        receiver_mail=h_to,
                        subject=h_sub,
                        statement_link=statement_link,
                        # attachments JSONField has default=list in model; no need to set here
                    ),
                )

                if not created:
                    changed = False
                    if statement_link and obj.statement_link != statement_link:
                        obj.statement_link = statement_link
                        changed = True
                    if obj.rec_dat != rec_dt:
                        obj.rec_dat = rec_dt; changed = True
                    if obj.sender_mail != h_from:
                        obj.sender_mail = h_from; changed = True
                    if obj.receiver_mail != h_to:
                        obj.receiver_mail = h_to; changed = True
                    if obj.subject != h_sub:
                        obj.subject = h_sub; changed = True
                    if changed:
                        obj.save(using=target_db)
                        updated += 1
                    else:
                        existing += 1
                else:
                    inserted += 1

                # --- recipients bucket (use To:; you can include CC if you want)
                to_addrs = _parse_recipients(obj.receiver_mail)
                if not to_addrs:
                    # fallback so we don't lose files
                    to_addrs = [_dh(msg.get("From")).split()[-1] or h_from]

                # current attachment entries to avoid duplicates across re-runs
                current_json = list(obj.attachments or [])
                existing_keys = {
                    ( (a.get("recipient") or "").lower(), a.get("path") or "" )
                    for a in current_json
                }
                first_dir_set = bool(getattr(obj, "attachment_dir", ""))

                # --- write each part once per recipient, record JSON
                any_added = False
                for i, part in enumerate(attachment_parts):
                    data = part.get_payload(decode=True) or b""
                    if not data:
                        continue
                    filename = part.get_filename() or f"part_{i+1}.bin"
                    ctype = part.get_content_type() or ""

                    for recip in to_addrs:
                        rel_path = _attachment_rel_path(obj.mailbox, recip, obj.rec_dat, filename)
                        key = (recip.lower(), rel_path.replace("\\", "/"))
                        if key in existing_keys:
                            continue  # idempotent

                        abs_path = os.path.join(settings.MEDIA_ROOT, rel_path)
                        _ensure_parent_dir(abs_path)
                        with open(abs_path, "wb") as f:
                            f.write(data)

                        size = os.path.getsize(abs_path) if os.path.exists(abs_path) else len(data)

                        meta = {
                            "recipient": recip,
                            "filename": filename,
                            "path": rel_path.replace("\\", "/"),
                            "content_type": ctype,
                            "size_bytes": int(size),
                            "part_index": i,
                        }
                        current_json.append(meta)
                        existing_keys.add(key)
                        any_added = True

                        # set a convenient base dir once (if you added attachment_dir field)
                        if hasattr(obj, "attachment_dir") and not first_dir_set:
                            obj.attachment_dir = "/".join(meta["path"].split("/")[:-1])
                            first_dir_set = True

                if any_added:
                    # save JSON (and dir if present)
                    if hasattr(obj, "attachment_dir") and obj.attachment_dir:
                        obj.attachments = current_json
                        obj.save(using=target_db, update_fields=["attachments", "attachment_dir"])
                    else:
                        obj.attachments = current_json
                        obj.save(using=target_db, update_fields=["attachments"])


        return {"inserted": inserted, "updated": updated, "existing": existing, "searched": len(ids)}

    finally:
        try:
            M.close()
        except Exception:
            pass
        M.logout()


def fetch_imap_all_accounts(*,
                            since_days: int = 30,
                            from_date: dt.date | None = None,
                            to_date: dt.date | None = None,
                            limit: int = 300,
                            auto_fy: bool = False,
                            db_alias: str = "default",
                            message_id_to_find: str | None = None,
                            to_email: str | None = None,               # NEW
                            subject_contains: str | None = None,       # NEW
                            ) -> dict:
    tot = {"inserted": 0, "updated": 0, "existing": 0, "searched": 0}
    for entry in IMAP_ACCOUNTS:
        if len(entry) == 2:
            email_addr, app_pw = entry
            host = None
        else:
            email_addr, app_pw, host = entry[0], entry[1], entry[2]

        res = fetch_imap_mailbox(
            email_addr,
            app_pw,
            host=host,
            since_days=since_days,
            from_date=from_date,
            to_date=to_date,
            limit=limit,
            auto_fy=auto_fy,
            db_alias=db_alias,
            message_id_to_find=message_id_to_find,
            to_email=to_email,                         # pass through
            subject_contains=subject_contains,         # pass through
        )
        for k in tot:
            tot[k] += res.get(k, 0)
    return tot


# views.py
from django.conf import settings
from django.shortcuts import render
from django.contrib import messages

from .models import MailLog
from .utills import get_current_fy, get_db_for_fy, fiscal_year_range
from django.db.models import Q
from django.core.validators import validate_email
from django.core.exceptions import ValidationError


def _fy_alias_map():
    """
    Returns a dict of DB alias -> human FY string
    Example: {'fy_2024_25': '2024-25', 'fy_2025_26': '2025-26'}
    """
    out = {}
    for alias in settings.DATABASES.keys():
        if alias.startswith("fy_"):
            out[alias] = alias.replace("fy_", "").replace("_", "-")
    return out


def _available_fys_for_model(model_cls):
    """
    Return FY strings (e.g. '2025-26') where model has >=1 row.
    Sorted by start year descending.
    """
    alias_map = _fy_alias_map()  # alias -> 'YYYY-YY'
    years = []
    for alias, fy in alias_map.items():
        try:
            if model_cls.objects.using(alias).exists():
                years.append(fy)
        except Exception:
            # model/table not present in that DB; ignore
            continue
    years.sort(key=lambda s: int(s.split("-")[0]), reverse=True)
    return years


def _fy_to_alias(fy_str: str):
    """'2025-26' or '2025_26' -> 'fy_2025_26'"""
    if not fy_str:
        return None
    return f"fy_{fy_str.strip().replace('-', '_')}"


# views.py
import logging
log = logging.getLogger(__name__)

from datetime import datetime, timedelta
from django.utils import timezone
from django.db.models import DateField, DateTimeField
import re
from django.db.models import Q
from accounts.models import AccountBank, Client, Group  # adjust app label if different

# ---- account number candidates ----
FULL_ACC_RX   = re.compile(r"\b(\d{9,18})\b")                     # 9–18 full digits
LAST4_RX      = re.compile(r"\b(\d{4})\b")                        # any plain last-4 candidates
MASK_LAST4_RX = re.compile(r"(?:x{2,}|\*{2,}|#{2,})\s*(\d{4})\b", re.I)  # **1234, xx1234, ##1234
MASK_LAST2_RX = re.compile(r"(?:x{2,}|\*{2,}|#{2,})\s*(\d{2})\b", re.I)  # **12, xx12, ##12
SUBJ_MASK_TAIL_RX = re.compile(r"[Xx#\*]{3,}\s*(\d{3,4})\b")

# --- SOFT OPTIONAL IMPORTS (won't crash if missing) ---
try:
    from .models import Client
except Exception:
    Client = None

try:
    from .models import Group
except Exception:
    Group = None

try:
    from .models import AccountBank
except Exception:
    AccountBank = None

STATUS_FILTERS = [
    ("unknown", "Unknown Email"),
    ("known", "Known Email"),
    ("bank_done", "Bank Done"),
    ("bank_missing", "Bank Not Found"),
    ("has_attachment", "Has Attachment"),
    ("no_attachment", "No Attachment"),
]

def mail_log_list(request):
    import re
    from datetime import datetime, timedelta
    from django.utils import timezone
    from django.db.models import DateField, DateTimeField
    from django.contrib import messages

    years = _available_fys_for_model(MailLog)
    fy_param = (request.GET.get("fy") or request.GET.get("year") or "").strip()
    fy = fy_param.replace("_", "-") if fy_param else ""
    # filter_key = request.GET.get("f", "").strip().lower()
    filter_keys = [f.lower() for f in request.GET.getlist("f")]
    selected_statuses = request.GET.getlist("f")

    mailbox_filter = (request.GET.get("mailbox") or "").strip().lower()

    date_from = request.GET.get("date_from")
    date_to   = request.GET.get("date_to")

     # 🚨 HARD STOP — NO DATE = NO MAILS
    # if not date_from or not date_to:
    #     return render(
    #         request,
    #         "mailLog_list.html",
    #         {
    #             "mails": [],              # ⛔ EMPTY
    #             "years": years,
    #             "year_selected": fy,
    #             "IMAP_ACCOUNTS": IMAP_ACCOUNTS,
    #             "groups": Group.objects.order_by("group_name"),
    #             "clients": Client.objects.order_by("client_name"),
    #             "status_filters": STATUS_FILTERS,
    #             "selected_statuses": [],
    #             "fy": fy,
    #         }
    #     )

    if not fy:
        fy = years[0] if years else get_current_fy()

    try:
        db_alias = _fy_to_alias(fy)
        start_date, end_date = fiscal_year_range(fy)   # date objects
    except Exception as e:
        # SAFETY: define mails before render
        messages.error(request, f"Invalid FY: {fy}. {e}")
        mails = MailLog.objects.none()
        return render(
            request,
            "mailLog_list.html",
            # {"mails": mails, "years": years, "year_selected": fy, "IMAP_ACCOUNTS": IMAP_ACCOUNTS},
            {"mails": mails, "years": years, "year_selected": fy, "IMAP_ACCOUNTS": IMAP_ACCOUNTS,
            "groups": Group.objects.order_by("group_name"),
            "clients": Client.objects.order_by("client_name"),  # <-- add this
            "fy": fy},
        )

    # Build [start_dt, end_dt) range
    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt_exclusive = datetime.combine(end_date + timedelta(days=1), datetime.min.time())

    # Make tz-aware if needed
    if timezone.is_naive(start_dt):
        tz = timezone.get_current_timezone()
        start_dt = timezone.make_aware(start_dt, tz)
        end_dt_exclusive = timezone.make_aware(end_dt_exclusive, tz)

    # Query MailLog in that FY DB
    try:
        rec_field = MailLog._meta.get_field("rec_dat")
        if isinstance(rec_field, DateField) and not isinstance(rec_field, DateTimeField):
            mails = (
                MailLog.objects.using(db_alias)
                .filter(rec_dat__gte=start_date, rec_dat__lte=end_date)
                .order_by("-rec_dat", "-id")
            )
        else:
            mails = (
                MailLog.objects.using(db_alias)
                .filter(rec_dat__gte=start_dt, rec_dat__lt=end_dt_exclusive)
                .order_by("-rec_dat", "-id")
            )

        # --------- DATE VALIDATION (optional but recommended) ----------
        if date_from and date_to:
            if date_from > date_to:
                messages.warning(
                    request,
                    "From-date cannot be greater than To-date."
                )
        # --------------------------------------------------------------

        # ---------------- DATE FILTER (Received Date) ----------------
        if date_from:
            df = datetime.strptime(date_from, "%Y-%m-%d").date()
            if isinstance(rec_field, DateField) and not isinstance(rec_field, DateTimeField):
                mails = mails.filter(rec_dat__gte=df)
            else:
                start_dt_user = datetime.combine(df, datetime.min.time())
                if timezone.is_naive(start_dt_user):
                    start_dt_user = timezone.make_aware(
                        start_dt_user, timezone.get_current_timezone()
                    )
                mails = mails.filter(rec_dat__gte=start_dt_user)

        if date_to:
            dt = datetime.strptime(date_to, "%Y-%m-%d").date()
            if isinstance(rec_field, DateField) and not isinstance(rec_field, DateTimeField):
                mails = mails.filter(rec_dat__lte=dt)
            else:
                end_dt_user = datetime.combine(dt, datetime.max.time())
                if timezone.is_naive(end_dt_user):
                    end_dt_user = timezone.make_aware(
                        end_dt_user, timezone.get_current_timezone()
                    )
                mails = mails.filter(rec_dat__lte=end_dt_user)
        # ------------------------------------------------------------

    except Exception as e:
        messages.error(request, f"{db_alias} not ready: {e}")
        mails = MailLog.objects.none()

    if not mails and years and fy not in years:
        messages.info(request, f"No data in {fy}. Available FYs with data: {', '.join(years)}")


        # ===== DEBUG: inspect MailLog fields once =====
    try:
        if mails:
            m0 = mails[0]
            field_names = [f.name for f in m0._meta.get_fields() if hasattr(f, "name")]
            print("[MAILLOG FIELDS]", field_names)
    except Exception as _e:
        pass
    # ===== end DEBUG =====

    # ---------------- helpers (safe getters & naming) ----------------
    def _safe_get(obj, attr, default=None):
        try:
            v = getattr(obj, attr)
            return v if v is not None else default
        except Exception:
            return default

    # def _client_display_name(c):
    #     for field in ("name", "client_name", "full_name", "display_name"):
    #         v = _safe_get(c, field)
    #         if v:
    #             return v
    #     return f"Client #{_safe_get(c, 'pk', '')}" if c else None

    def _client_display_name(c):
        for field in ("client_name", "legal_name", "trade_name"):
            v = getattr(c, field, None)
            if v:
                return v
        return f"Client #{getattr(c, 'pk', '')}" if c else None


    def _group_display_name(g):
        for field in ("name", "group_name", "title"):
            v = _safe_get(g, field)
            if v:
                return v
        return f"Group #{_safe_get(g, 'pk', '')}" if g else None


    def _subject_tail_candidates(m):
        """
        From subject like '... for XXXXX645', extract {'645'} or {'1234'}.
        Only used AFTER we know the client (to avoid false positives).
        """
        subj = (_safe_get(m, "subject", "") or "")
        return set(SUBJ_MASK_TAIL_RX.findall(subj))

    # === PLACE A: directly after your `_subject_tail_candidates` helper ===

    def _mail_body_text(m):
        """
        Try very hard to get plain text for the mail body/snippet from MailLog row `m`.
        - Checks many common attribute names
        - If still empty, scans ALL attributes whose name contains 'body' or 'snippet'
        - Strips HTML if present
        Returns lowercased, whitespace-collapsed string.
        """
        import re

        # Known candidates first (ordered by likelihood in many pipelines)
        cand_attrs = [
            "text_body", "plain_body", "body_text", "body",  # text variants
            "snippet", "preview", "preview_text", "body_snippet",
            "html_body", "body_html", "html",                # html variants
            "raw_body", "message", "content", "content_text",
            "gmail_snippet", "mail_text", "mail_body",
        ]

        val = ""
        for attr in cand_attrs:
            v = getattr(m, attr, None)
            if v:
                val = str(v)
                break

        # JSON container some pipelines use
        if not val:
            try:
                bjson = getattr(m, "body_json", None) or {}
                val = str(bjson.get("text") or bjson.get("body") or "")
            except Exception:
                pass

        # As a LAST resort, scan all model attributes for anything named like *body*/*snippet*
        if not val:
            try:
                for k, v in m.__dict__.items():
                    if not v or not isinstance(v, (str, bytes)):
                        continue
                    key = k.lower()
                    if "body" in key or "snippet" in key or "preview" in key or "content" in key:
                        val = v.decode("utf-8", "ignore") if isinstance(v, (bytes, bytearray)) else str(v)
                        if val and len(val.strip()) >= 3:
                            break
            except Exception:
                pass

        if not val:
            return ""

        # Strip HTML tags if it looks like HTML
        if "<" in val and ">" in val:
            val = re.sub(r"<[^>]+>", " ", val)

        # quick entity cleanups
        val = val.replace("&nbsp;", " ").replace("&amp;", "&")

        # normalize and lower
        return re.sub(r"\s+", " ", val).strip().lower()




    # def _client_aliases():
    #     """
    #     Build {client_id: {'aliases': set[str], 'client': Client}}.
    #     All aliases are normalized to lowercase and punctuation-less variants.
    #     """
    #     import re
    #     norm = lambda s: re.sub(r"[^a-z0-9 ]+", " ", (s or "").lower()).strip()

    #     data = {}
    #     qs = Client.objects.using("default").select_related("group").only(
    #         "id", "client_name", "name", "display_name", "trade_name", "firm_name"
    #     )
    #     for c in qs:
    #         aliases = set()
    #         for f in ("client_name", "name", "display_name", "trade_name", "firm_name"):
    #             v = getattr(c, f, None)
    #             if v:
    #                 a = norm(v)
    #                 if a:
    #                     aliases.add(a)
    #                     aliases.add(a.replace(" ", ""))  # tight variant
    #         if aliases:
    #             data[c.id] = {"aliases": aliases, "client": c}
    #     return data


    def _client_aliases():
        """
        Build {client_id: {'aliases': set[str], 'client': Client}} using
        only real Client fields (client_name, legal_name, trade_name).
        """
        import re
        norm = lambda s: re.sub(r"[^a-z0-9 ]+", " ", (s or "").lower()).strip()

        use_fields = ["client_name", "legal_name", "trade_name"]

        # ✅ removed .select_related("group") to avoid FieldError
        qs = Client.objects.using("default").only("id", *use_fields)

        data = {}
        for c in qs:
            aliases = set()
            for f in use_fields:
                v = getattr(c, f, None)
                if v:
                    a = norm(v)
                    if a:
                        aliases.add(a)
                        aliases.add(a.replace(" ", ""))  # tight variant
            if aliases:
                data[c.id] = {"aliases": aliases, "client": c}
        return data

    _CLIENT_ALIASES = _client_aliases()


    def _client_from_body(m):
        """
        Scan body text for client names. Return a Client or None.
        """
        import re
        body = _mail_body_text(m)
        if not body:
            return None
        body_tight = re.sub(r"[^a-z0-9]+", "", body)

        hit = None
        best = 0
        for cid, item in _CLIENT_ALIASES.items():
            for alias in item["aliases"]:
                if not alias or len(alias) < 4:
                    continue
                score = 0
                if re.search(rf"\b{re.escape(alias)}\b", body):
                    score = len(alias)
                elif alias.replace(" ", "") in body_tight:
                    score = len(alias) - 1
                if score > best:
                    best = score
                    hit = item["client"]
        return hit if best >= 6 else None


    def _parse_emails_list(raw):
        try:
            from email.utils import getaddresses
            addrs = []
            for _, addr in getaddresses([raw or ""]):
                a = (addr or "").strip()
                if a:
                    addrs.append(a)
            seen = set()
            out = []
            for a in addrs:
                low = a.lower()
                if low not in seen:
                    out.append(a)
                    seen.add(low)
            return out
        except Exception:
            return []


    def _identity_emails(m):
        """
        Decide which email(s) to use for client recognition.
        Rule:
        - If mailbox == receiver → use sender
        - Else → use receiver(s)
        """
        mailbox = (_safe_get(m, "mailbox", "") or "").lower().strip()
        sender  = (_safe_get(m, "sender_mail", "") or "").lower().strip()

        recips = _parse_emails_list(_safe_get(m, "receiver_mail", ""))
        recips = [r.lower().strip() for r in recips if r]

        # 🔑 IMPORTANT RULE
        if mailbox and mailbox in recips:
            if sender:
                return [sender]

        return recips



    # cache
    _email_to_client = {}

    # ---------------- client lookup by email ----------------
    def _find_client_by_email(addr_lower: str):
        """
        Find Client for a recipient email.
        Order:
        1) AccountBank.acc_mail_id → client
        2) Client.* email fields (case-insensitive)
        3) Client.extra_emails (JSON/Array membership)
        All queries run on DEFAULT DB.
        """
        if not addr_lower:
            return None
        if addr_lower in _email_to_client:
            return _email_to_client[addr_lower]

        c = None

        # 1) via AccountBank.acc_mail_id
        if AccountBank is not None:
            try:
                ab = (
                    AccountBank.objects.using("default")
                    .select_related("client")
                    .filter(acc_mail_id__iexact=addr_lower)
                    .first()
                )
                if ab and getattr(ab, "client", None):
                    c = ab.client
            except Exception:
                pass

        # 2) via Client.* "primary" email fields
        if c is None and Client is not None:
            try:
                email_fields = [
                    f for f in ("email", "alt_email", "acc_mail_id",
                                "contact_email", "primary_email")
                    if hasattr(Client, f)
                ]

                q = Q()
                for f in email_fields:
                    q |= Q(**{f"{f}__iexact": addr_lower})

                # 3) ALSO check extra_emails list (JSON/Array)
                # store emails lower-cased when saving; we match with lower too.
                if hasattr(Client, "extra_emails"):
                    # works for Postgres JSONField/ArrayField
                    q |= Q(extra_emails__contains=[addr_lower])

                c = (
                    Client.objects.using("default")
                    .select_related("group")
                    .filter(q)
                    .first()
                )

                # Fallback for DBs that don’t support __contains on JSON/Array (e.g. SQLite without JSON1):
                if c is None and hasattr(Client, "extra_emails"):
                    try:
                        for cand in (
                            Client.objects.using("default")
                            .only("id", "extra_emails")[:5000]  # safety cap
                        ):
                            extras = getattr(cand, "extra_emails", None) or []
                            # ensure items are comparable in lower-case
                            if any((e or "").lower() == addr_lower for e in extras):
                                c = cand
                                break
                    except Exception:
                        pass

            except Exception:
                pass

        _email_to_client[addr_lower] = c
        return c
    def _clients_for_email(addr_lower: str):
        """
        Return ALL Client objects that reference this email.
        We strictly coerce IDs to ints to avoid 'id expected a number' errors.
        """
        if not addr_lower:
            return []

        raw_ids = set()

        # From AccountBank.acc_mail_id → client_id (ints)
        if AccountBank is not None:
            try:
                raw_ids.update(
                    AccountBank.objects.using("default")
                    .filter(acc_mail_id__iexact=addr_lower, client__isnull=False)
                    .values_list("client_id", flat=True)
                )
            except Exception:
                pass

        # From Client primary email fields + extra_emails
        if Client is not None:
            try:
                email_fields = [
                    f for f in ("email", "alt_email", "acc_mail_id", "contact_email", "primary_email")
                    if hasattr(Client, f)
                ]
                q = Q()
                for f in email_fields:
                    q |= Q(**{f"{f}__iexact": addr_lower})
                if hasattr(Client, "extra_emails"):
                    q |= Q(extra_emails__contains=[addr_lower])

                if q:
                    raw_ids.update(
                        Client.objects.using("default")
                        .filter(q)
                        .values_list("id", flat=True)
                    )

                # Fallback for DBs without JSON contains
                if hasattr(Client, "extra_emails") and not raw_ids:
                    for cand in Client.objects.using("default").only("id", "extra_emails")[:5000]:
                        extras = cand.extra_emails or []
                        if any((e or "").lower() == addr_lower for e in extras):
                            raw_ids.add(cand.id)
            except Exception:
                pass

        # ---- sanitize: keep only integers ----
        safe_ids = []
        bad = []
        for x in raw_ids:
            try:
                # if already int this is a no-op
                safe_ids.append(int(x))
            except (TypeError, ValueError):
                bad.append(x)
        # if bad:
            # log.warning("Ignoring non-integer Client IDs for %s: %r", addr_lower, bad)

        if not safe_ids:
            return []

        return list(
            Client.objects.using("default")
            .select_related("group")
            .filter(id__in=safe_ids)
            .order_by("id")
        )
    
    # ---------------- NEW: attachment → account-number extraction ----------------
    def _attachment_text(m):
        v = _safe_get(m, "attachments", None)
        if not v:
            return ""
        try:
            parts = []
            for item in (v or []):
                if isinstance(item, dict):
                    parts.append(str(item.get("filename") or item.get("name") or ""))
                    parts.append(str(item.get("path") or item.get("filepath") or ""))
                else:
                    parts.append(str(item))
            return " ".join([p for p in parts if p])
        except Exception:
            return str(v)

    # 9–18 digit numbers = likely full account numbers
    ACC_NUM_RX = re.compile(r"\b(\d{9,18})\b")
    # common masked-last4 patterns: ****1234, XX1234, ####1234, etc.
    LAST4_RX   = re.compile(r"(?:x{2,}|\*{2,}|X{2,}|#){0,}\s*(\d{4})\b")

    def _extract_candidate_account_numbers(text: str):
        """
        Returns (full_numbers:set[str], last4:set[str]) from any attachment text.
        """
        fulls = set(ACC_NUM_RX.findall(text or ""))
        last4 = set(LAST4_RX.findall(text or ""))
        # also split on separators and rescan
        parts = re.split(r"[\s_\-\.()]+", text or "")
        for p in parts:
            fulls.update(ACC_NUM_RX.findall(p))
            last4.update(LAST4_RX.findall(p))
        return fulls, last4

    # ---------------- NEW: pick ONE account by matching account_no ----------------
    # def _account_for_client_by_attachment(client_obj, email_addr: str, attach_text: str):


    def _account_for_client_by_attachment(client_obj, email_addr, attach_text, subject_text_tails=None):
        if not (AccountBank and client_obj):
            return None

        # 🔑 DO NOT FILTER BY EMAIL HERE
        base_qs = AccountBank.objects.using("default").filter(client=client_obj)

        fulls, last4s = _extract_candidate_account_numbers(attach_text or "")

        if fulls:
            hit = base_qs.filter(account_no__in=list(fulls)) \
                        .values("bank_name", "account_no", "stms_pws") \
                        .first()
            if hit:
                return {
                    "bank_name": hit["bank_name"] or "",
                    "account_no": hit["account_no"] or "",
                    "stms_pws": hit.get("stms_pws") or "",
                }

        for tail in last4s:
            hit = base_qs.filter(account_no__iendswith=tail) \
                        .values("bank_name", "account_no", "stms_pws") \
                        .first()
            if hit:
                return {
                    "bank_name": hit["bank_name"] or "",
                    "account_no": hit["account_no"] or "",
                    "stms_pws": hit.get("stms_pws") or "",
                }

        # subject tail logic stays SAME
        tails = set(subject_text_tails or [])
        for tail in tails:
            if not tail or len(tail) < 3:
                continue
            rows = list(
                base_qs.filter(account_no__iendswith=tail)
                    .values("bank_name", "account_no", "stms_pws")[:2]
            )
            if len(rows) == 1:
                hit = rows[0]
                return {
                    "bank_name": hit["bank_name"] or "",
                    "account_no": hit["account_no"] or "",
                    "stms_pws": hit.get("stms_pws") or "",
                }

        return None

    
        # ====== BLOCK A: build fast indexes & 2 tiny match helpers (BANK first) ======

    # Build Account indexes (once per request)
    acc_rows = list(
        AccountBank.objects.using("default")
        .select_related("client", "client__group")
        .values("id", "account_no", "bank_name",
                "client_id", "client__client_name",
                "client__group_id", "client__group__group_name")
    )

    acc_index_full = {str(r["account_no"]): r for r in acc_rows if r["account_no"]}
    acc_index_last4 = {}
    for r in acc_rows:
        acc = str(r["account_no"] or "")
        if len(acc) >= 4:
            acc_index_last4.setdefault(acc[-4:], []).append(r)

    # Email → Client quick index
    client_email_index = {}
    for c in Client.objects.select_related("group").only(
        "id", "client_name", "email", "extra_emails", "group__group_name"
    ):
        emails = []
        if c.email: emails.append(c.email.strip().lower())
        for e in (c.extra_emails or []):
            if e: emails.append(e.strip().lower())
        for e in set(emails):
            client_email_index[e] = {
                "id": c.id,
                "client_name": c.client_name,
                "group_name": getattr(c.group, "group_name", "") or "",
            }

    # Email → Group quick index
    group_email_index = {}
    for g in Group.objects.only("id", "group_name", "group_email", "extra_emails"):
        emails = []
        if g.group_email: emails.append(g.group_email.strip().lower())
        for e in (g.extra_emails or []):
            if e: emails.append(e.strip().lower())
        for e in set(emails):
            group_email_index[e] = {"id": g.id, "group_name": g.group_name}


    def _collect_mail_text(m):
        """
        ONLY attachment filenames/paths (NO subject) to avoid random 4-digit hits.
        """
        v = _safe_get(m, "attachments", None) or []
        parts = []
        for it in v:
            if isinstance(it, dict):
                parts.append(str(it.get("filename") or it.get("name") or ""))
                parts.append(str(it.get("path") or it.get("filepath") or it.get("file") or ""))
            else:
                parts.append(str(it))
        # collapse spaces; keep only non-empty
        return " ".join(p for p in parts if p).strip()

    def _first_recipient(m):
        recs = (_safe_get(m, "receiver_mail", "") or "")
        if not recs:
            return ""
        tokens = re.split(r"[,\s;]+", recs)
        return (tokens[0] or "").strip().lower()

    def _match_bank_first(m):
        """
        Priority-1: Bank account by ATTACHMENT content only.
        Rules:
          1) Full 9–18 digits → direct match.
          2) Masked last-4 (****1234 / xx1234 / ##1234) → match ONLY if unique.
          3) NO plain last-4, NO last-2 (to prevent false positives).
        Returns: acc-row dict or None.
        """
        text = _collect_mail_text(m)
        if not text:
            return None  # no attachments -> skip

        # 1) full 9–18 digits
        for num in FULL_ACC_RX.findall(text):
            if num in acc_index_full:
                return acc_index_full[num]

        # 2) masked last-4 only
        last4s = set(MASK_LAST4_RX.findall(text))
        for l4 in last4s:
            rows = acc_index_last4.get(l4, [])
            if len(rows) == 1:
                return rows[0]   # unambiguous masked last-4

        # 3) disabled: plain last-4 & masked last-2 (too risky)
        return None

    def _match_email_then_group(m):
        """
        Priority-2: Client by email; Priority-3: Group by email.
        Returns (client_name, group_name, matched_email)
        """
        first_to = _first_recipient(m)
        if not first_to:
            return ("", "", "")
        c = client_email_index.get(first_to)
        if c:
            return (c["client_name"], c["group_name"], first_to)
        g = group_email_index.get(first_to)
        if g:
            return ("", g["group_name"], first_to)
        return ("", "", "")
    
    
    filtered_mails = []

    # ---------------- map render fields per row ----------------
    for m in mails:
        m.matched_email = ""
        m.first_recipient = ""          # NEW: for display if unknown
        m.email_known = False           # NEW: flag
        m.email_status = "Unknown email"  # NEW: human text
        m.client_name = None
        m.group_name = None
        m.accounts = []
        m.bank_account_label = None
        m.bank_status = "Bank not found"
        m.open_url = _open_link_for_mail(m)
        m.stmt_pw = ""



        # recips = _parse_emails_list(_safe_get(m, "receiver_mail", ""))
        recips = _identity_emails(m)

        # fallback to mailbox as a "recipient" if none parsed
        if not recips:
            rb = (_safe_get(m, "mailbox", "") or "").strip()
            if rb:
                recips = [rb]

        # keep for display even if unknown
        m.first_recipient = recips[0] if recips else ""
        
        attach_text = _attachment_text(m)


        found = False
        email_candidates = []

        for r in recips:
            low = r.lower().strip()
            if not low:
                continue

            #     continue
            email_candidates = _clients_for_email(low)

            # ---------- NEW: GROUP EMAIL FALLBACK ----------
            if not email_candidates:
                grp = group_email_index.get(low)
                if grp:
                    m.email_known = True
                    m.email_status = "Known (group)"
                    m.group_name = grp["group_name"]
                    m.matched_email = r
                continue

            # We have at least one candidate ⇒ email is "Known"
            m.email_known = True
            # m.email_status = "Known"
            if not found and len(email_candidates) > 1:
                m.email_status = "Known (ambiguous)"

            m.matched_email = r

            tails = _subject_tail_candidates(m)

            chosen_client = None
            chosen_acc = None

            # 1) try each candidate by attachment/subject
            for cand in email_candidates:
                acc_hit = _account_for_client_by_attachment(
                    cand, low, attach_text, subject_text_tails=tails
                )
               
                if acc_hit:
                    # ✅ CORRECT CLIENT FOUND BY ACCOUNT
                    m.client = cand
                    m.client_name = _client_display_name(cand)
                    g = getattr(cand, "group", None)
                    m.group_name = _group_display_name(g) if g else None

                    m.accounts = [acc_hit]
                    m.bank_status = "Done"

                    pw = acc_hit.get("stms_pws") or ""
                    m.stmt_pw = pw

                    found = True
                    break   # stop checking more clients

            if found:
                break   # stop checking more recipients
                       
          
        if not found and email_candidates:
            #  Do NOT auto-select client
            m.client = None
            m.client_name = None
            m.accounts = []
            m.bank_status = "Bank not found"

            # If all clients belong to SAME group, show that group
            groups = {getattr(c.group, "id", None): c.group for c in email_candidates if c.group}

            if len(groups) == 1:
                grp = list(groups.values())[0]
                m.group_name = _group_display_name(grp)
            else:
                m.group_name = "Multiple Groups"

            # ✅ JUST STORE LIST FOR DISPLAY (READ-ONLY)
            m.candidate_clients = email_candidates


    

           
        # after trying all recipients, ensure it's clean if no match
        if m.bank_status != "Done":
            m.accounts = []
            m.bank_account_label = None            # acc = _account_for_client_by_attachment(c, r, attach_text)

    groups = Group.objects.order_by("group_name")  # <-- provide groups for the modal
   # -------- APPLY FILTERS AFTER PROCESSING --------
    filtered = []

    for m in mails:

        # 🔹 Status filter
        # if filter_key:
        #     if filter_key == "unknown" and m.email_known:
        #         continue

        #     if filter_key == "known" and not m.email_known:
        #         continue

        #     if filter_key == "bank_done" and m.bank_status != "Done":
        #         continue

        #     if filter_key == "bank_missing" and m.bank_status == "Done":
        #         continue

        if filter_keys:
            keep = False

            has_attachment = bool(m.attachments)

            for fk in filter_keys:
                # ---- Email known / unknown ----
                if fk == "unknown" and not m.email_known:
                    keep = True
                elif fk == "known" and m.email_known:
                    keep = True

                # ---- Bank status ----
                elif fk == "bank_done" and m.bank_status == "Done":
                    keep = True
                elif fk == "bank_missing" and m.bank_status != "Done":
                    keep = True

                 # ---- Attachment status ----
                elif fk == "has_attachment" and has_attachment:
                    keep = True
                elif fk == "no_attachment" and not has_attachment:
                    keep = True
                    
            if not keep:
                continue



        # 🔹 Mailbox filter (NEW)
        if mailbox_filter:
            mbox = (_safe_get(m, "mailbox", "") or "").lower()
            if mbox != mailbox_filter:
                continue

        filtered.append(m)

    mails = filtered

    
    return render(
           request,
           "mailLog_list.html",
           {
               "mails": mails,
               "years": years,
               "year_selected": fy,
               "IMAP_ACCOUNTS": IMAP_ACCOUNTS,
               "db_alias": db_alias,
               "groups": groups,   # <-- used to fill Group <select>
                "clients": Client.objects.order_by("client_name"),
               "fy": fy,           # <-- send FY so AJAX can update MailLog in the right DB
               "selected_statuses": selected_statuses,
               "status_filters": STATUS_FILTERS,
           },
    )

###############################################
# views.py (add below your existing imports)
from django.views.decorators.http import require_POST
from django.http import JsonResponse, HttpResponseBadRequest
from django.middleware.csrf import get_token

# helper: build a quick dict for account lookup
def _imap_account_map():
    """
    Returns {email_addr: (app_pw, host_or_None)}
    """
    m = {}
    for entry in IMAP_ACCOUNTS:
        if len(entry) == 2:
            m[entry[0]] = (entry[1], None)
        else:
            m[entry[0]] = (entry[1], entry[2])
    return m


from django.views.decorators.http import require_POST
from django.http import JsonResponse, HttpResponseBadRequest

@require_POST
def mail_log_fetch(request):
    try:
        account = (request.POST.get("account") or "ALL").strip()
        limit = int(request.POST.get("limit") or 300)
        auto_fy = (request.POST.get("auto_fy") or "1") in ("1","true","True","on")
        message_id = (request.POST.get("message_id") or "").strip() or None

        # NEW: optional fetch filters
        to_email = (request.POST.get("to_email") or "").strip() or None
        subject_contains = (request.POST.get("subject_contains") or "").strip() or None
        from_date_s = (request.POST.get("from_date") or "").strip()
        to_date_s   = (request.POST.get("to_date")   or "").strip()
        from_date = dt.date.fromisoformat(from_date_s) if from_date_s else None
        to_date   = dt.date.fromisoformat(to_date_s)   if to_date_s else None

        # keep old fallback if no dates provided
        since_days = int(request.POST.get("since_days") or 30)

        if from_date and to_date and to_date < from_date:
            return HttpResponseBadRequest("to_date cannot be earlier than from_date.")

    except Exception as e:
        return HttpResponseBadRequest(f"Invalid parameters: {e}")

    acct_map = { (e[0]): (e[1], (e[2] if len(e)>2 else None)) for e in IMAP_ACCOUNTS }

    if account != "ALL":
        if account not in acct_map:
            return HttpResponseBadRequest("Unknown account")
        app_pw, host = acct_map[account]
        result = fetch_imap_mailbox(
            mailbox_email=account,
            app_password=app_pw,
            host=host,
            since_days=since_days,
            from_date=from_date,
            to_date=to_date,
            limit=limit,
            auto_fy=auto_fy,
            db_alias="default",
            message_id_to_find=message_id,
            to_email=to_email,                      
            subject_contains=subject_contains,      
        )
    else:
        result = fetch_imap_all_accounts(
            since_days=since_days,
            from_date=from_date,
            to_date=to_date,
            limit=limit,
            auto_fy=auto_fy,
            db_alias="default",
            message_id_to_find=message_id,
            to_email=to_email,                     
            subject_contains=subject_contains, 
        )

    return JsonResponse({
        "ok": True,
        "account": account,
        "from_date": from_date_s,
        "to_date": to_date_s,
        "limit": limit,
        "message_id": message_id or "",
        "to_email": to_email or "",
        "subject_contains": subject_contains or "",
        **result,  # inserted / updated / existing / searched
    })


# accounts/views.py
import os, io, mimetypes
from django.http import FileResponse, Http404, HttpResponse
from django.conf import settings
from django.utils.encoding import iri_to_uri
from django.db import DatabaseError, ProgrammingError
from .models import MailLog

def _inline_name_header(name: str) -> str:
    # RFC 5987 (works across browsers)
    safe = (name or "file.pdf").replace('"', "")
    return f"inline; filename*=UTF-8''{iri_to_uri(safe)}"

def _fetch_mail_from_alias(mail_id: int, alias: str):
    try:
        return MailLog.objects.using(alias).get(pk=mail_id)
    except (MailLog.DoesNotExist, ProgrammingError, DatabaseError):
        return None

def _get_mail_any_db(mail_id: int):
    obj = _fetch_mail_from_alias(mail_id, "default")
    if obj:
        return obj
    for alias in sorted(a for a in settings.DATABASES if a.startswith("fy_")):
        obj = _fetch_mail_from_alias(mail_id, alias)
        if obj:
            return obj
    raise Http404("Mail not found")

def mail_attachment_pdf(request, mail_id: int, idx: int, filename: str):
    """
    Guaranteed inline PDF streamer.
    Template should call {% url 'mail_attachment_pdf' mail.id forloop.counter0 fname %}?db=...
    """
    db_alias = request.GET.get("db")
    mail = _fetch_mail_from_alias(mail_id, db_alias) if db_alias else None
    if not mail:
        mail = _get_mail_any_db(mail_id)

    atts = list(mail.attachments or [])
    try:
        att = atts[int(idx)]
    except Exception:
        raise Http404("Attachment not found")

    rel_path = (att.get("path") or att.get("filepath") or att.get("file") or "").replace("\\", "/")
    if not rel_path:
        raise Http404("Attachment path missing")

    file_path = rel_path if os.path.isabs(rel_path) else os.path.join(settings.MEDIA_ROOT, rel_path)
    if not os.path.exists(file_path):
        raise Http404("File not found on disk")

    # Always return inline PDF
    f = open(file_path, "rb")
    resp = FileResponse(f, content_type="application/pdf")
    resp["Content-Disposition"] = _inline_name_header(os.path.basename(filename) or "document.pdf")
    resp["X-Content-Type-Options"] = "nosniff"
    return resp
    
# inside mail_log_list(), add this helper with your other small helpers
def _open_link_for_mail(m):
    """
    Build a provider-specific web URL to open the mail in webmail.
    Uses only getattr (no _safe_get dependency).
    """
    def _val(obj, name):
        v = getattr(obj, name, None)
        return v if v is not None else ""

    msgid   = (_val(m, "msg_id") or "").strip()
    subject = _val(m, "subject") or ""
    sender  = _val(m, "sender_mail") or ""
    mailbox = (_val(m, "mailbox") or _val(m, "receiver_mail") or "").lower()

    # strip <...> around Message-ID if present
    if msgid.startswith("<") and msgid.endswith(">"):
        msgid = msgid[1:-1]

    # Detect Gmail / Google Workspace
    if "gmail" in mailbox or "googlemail" in mailbox or "google" in mailbox:
        if msgid:
            return f"https://mail.google.com/mail/u/0/#search/rfc822msgid%3A{quote(msgid)}"
        q = " ".join(
            s for s in [
                f'subject:"{subject}"' if subject else "",
                f'from:{sender}' if sender else ""
            ] if s
        )
        return f"https://mail.google.com/mail/u/0/#search/{quote(q)}"

    # Detect Yahoo
    if "yahoo" in mailbox:
        q = msgid or " ".join(s for s in [subject, sender] if s)
        return f"https://mail.yahoo.com/d/search/?q={quote(q)}"

    # Generic fallback (Gmail search often works for Workspace/custom domains)
    if msgid:
        return f"https://mail.google.com/mail/u/0/#search/rfc822msgid%3A{quote(msgid)}"
    if subject or sender:
        q = " ".join(s for s in [subject, sender] if s)
        return f"https://mail.google.com/mail/u/0/#search/{quote(q)}"
    return None



# emal add from mail report
# views.py
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from django.views.decorators.http import require_GET, require_POST
from django.db import transaction
from django.db.models import Q
from django.core.validators import validate_email
from django.core.exceptions import ValidationError

from .models import Group, Client


def _norm_email(e: str) -> str:
    e = (e or "").strip().lower()
    validate_email(e)  # raises ValidationError if invalid
    return e


@require_GET
def ajax_group_clients(request):
    """Return clients for a selected group, including primary+extra emails."""
    group_id = request.GET.get("group_id")
    clients = Client.objects.filter(group_id=group_id).order_by("client_name")
    return JsonResponse({
        "clients": [
            {
                "id": c.id,
                "name": c.client_name,
                "primary": (c.email or ""),
                "extras": (c.extra_emails or []),
            }
            for c in clients
        ]
    })

@require_POST
@transaction.atomic
def ajax_associate_email(request):
    from django.core.exceptions import ValidationError
    from django.shortcuts import get_object_or_404
    from django.db.models import Q

    def bad(msg, code=400):
        return JsonResponse({"ok": False, "error": msg}, status=code)

    # ---- inputs ----
    try:
        email = _norm_email(request.POST.get("email"))
    except ValidationError:
        return bad("Invalid email address.")

    client_id = (request.POST.get("client_id") or "").strip()
    group_id  = (request.POST.get("group_id")  or "").strip()
    apply_to_group = (request.POST.get("apply_to_group") == "true")
    mail_id  = (request.POST.get("mail_id") or "").strip()
    fy       = (request.POST.get("fy") or "").strip()

    if not email or (not client_id and not group_id):
        return bad("Missing email or target entity.")

    # ---- GROUP-ONLY path ----
    if group_id and not client_id:
        try:
            g = get_object_or_404(Group.objects.select_for_update(), id=group_id)

            # avoid duplicates across other groups (fallback if JSON contains not supported)
            try:
                exists_elsewhere = Group.objects.filter(
                    Q(group_email__iexact=email) | Q(extra_emails__contains=[email])
                ).exclude(id=g.id).exists()
            except Exception:
                exists_elsewhere = any(
                    (email == (gg.group_email or "").lower()) or
                    (email in [(e or "").lower() for e in (gg.extra_emails or [])])
                    for gg in Group.objects.exclude(id=g.id).only("id","group_email","extra_emails")
                )
            if exists_elsewhere:
                return bad("This email already belongs to another group.", 409)

            extras = list(g.extra_emails or [])
            primary = (g.group_email or "").strip().lower()
            if not primary:
                g.group_email = email
                if email in extras: extras.remove(email)
            elif email != primary and email not in extras:
                extras.append(email)
            g.extra_emails = sorted(set(extras))
            g.save(update_fields=["group_email", "extra_emails"])

            # optional: mark MailLog row
            if mail_id:
                try:
                    from .models import MailLog
                    qs = MailLog.objects
                    if fy:
                        from .utills import get_db_for_fy
                        qs = qs.using(get_db_for_fy(fy))
                    qs.filter(id=mail_id).update(
                        email_known=True, matched_email=email, client=None, group=g
                    )
                except Exception:
                    pass

            return JsonResponse({
                "ok": True,
                "group_name": g.group_name,
                "primary": g.group_email or "",
                "extras": g.extra_emails or [],
                "email": email,
            })
        except Exception as e:
            return bad(f"Server error (group): {e}", 500)

    # ---- CLIENT path ----
    try:
        client = get_object_or_404(Client.objects.select_for_update(), id=client_id)

        # duplicate across other clients?
        try:
            dup = Client.objects.filter(
                Q(email__iexact=email) | Q(extra_emails__contains=[email])
            ).exclude(id=client.id).exists()
        except Exception:
            dup = any(
                (email == (c.email or "").lower()) or
                (email in [(e or "").lower() for e in (c.extra_emails or [])])
                for c in Client.objects.exclude(id=client.id).only("id","email","extra_emails")
            )
        if dup:
            return bad("This email already belongs to another client.", 409)

        extras  = list(client.extra_emails or [])
        primary = (client.email or "").strip().lower()
        if not primary:
            client.email = email
            if email in extras: extras.remove(email)
        elif email != primary and email not in extras:
            extras.append(email)
        client.extra_emails = sorted(set(extras))
        client.save(update_fields=["email", "extra_emails"])

        group_name = ""
        if apply_to_group and client.group_id:
            g = Group.objects.select_for_update().get(id=client.group_id)
            group_name = g.group_name
            g_extras  = list(g.extra_emails or [])
            g_primary = (g.group_email or "").strip().lower()
            if not g_primary:
                g.group_email = email
                if email in g_extras: g_extras.remove(email)
            elif email != g_primary and email not in g_extras:
                g_extras.append(email)
            g.extra_emails = sorted(set(g_extras))
            g.save(update_fields=["group_email", "extra_emails"])

        if mail_id:
            try:
                from .models import MailLog
                qs = MailLog.objects
                if fy:
                    from .utills import get_db_for_fy
                    qs = qs.using(get_db_for_fy(fy))
                qs.filter(id=mail_id).update(
                    email_known=True, matched_email=email, client=client, group=client.group
                )
            except Exception:
                pass

        return JsonResponse({
            "ok": True,
            "client_name": client.client_name,
            "group_name": group_name,
            "primary": client.email or "",
            "extras": client.extra_emails or [],
            "email": email,
        })
    except Exception as e:
        return bad(f"Server error (client): {e}", 500)



# views.py
import re
from django.conf import settings
from django.db.models import Q
from django.shortcuts import render
from django.core.exceptions import FieldError
from django.utils import timezone

from .models import Group, Client, AccountBank, MailLog
from .utills import get_db_for_fy, fiscal_year_range, get_current_fy  # keep your utils import

# ---------------- Regex helpers ----------------
EMAIL_RX   = re.compile(r'[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}', re.I)
FULL_ACC   = re.compile(r'\b(\d{9,18})\b')                         # 9–18 digits
LAST4_ACC  = re.compile(r'(?:x{2,}|\*{2,}|#\s*)?(\d{4})\b', re.I)  # **1234 / xx1234 / #1234
LAST2_ACC  = re.compile(r'(?:x{2,}|\*{2,}|#\s*)?(\d{2})\b',  re.I) # **12   / xx12   / #12

def _norm_email(e: str) -> str:
    return (e or "").strip().lower()

def _split_emails(val) -> list[str]:
    if isinstance(val, (list, tuple)):
        raw = " ".join(map(str, val))
    else:
        raw = str(val or "")
    return sorted({_norm_email(m.group(0)) for m in EMAIL_RX.finditer(raw)})

def _extract_acc_tokens(names: list[str]) -> dict:
    fulls, last4s, last2s = set(), set(), set()
    for nm in names or []:
        s = str(nm or "")
        fulls.update(m.group(1) for m in FULL_ACC.finditer(s))
        last4s.update(m.group(1) for m in LAST4_ACC.finditer(s))
        last2s.update(m.group(1) for m in LAST2_ACC.finditer(s))
    return {"full": fulls, "last4": last4s, "last2": last2s}

def _to_local(dt):
    if not dt:
        return None
    try:
        return timezone.localtime(dt) if timezone.is_aware(dt) else dt
    except Exception:
        return dt

# ---------------- DB alias discovery ----------------
def _db_aliases_for_fy():
    """
    Use default + all fy_* DBs if present. Adjust here if you have other prefixes.
    """
    keys = list(settings.DATABASES.keys())
    out = []
    if "default" in keys:
        out.append("default")
    out.extend([k for k in keys if k.startswith("fy_")])
    # unique & stable
    seen, ordered = set(), []
    for a in out or keys:
        if a not in seen:
            ordered.append(a)
            seen.add(a)
    return ordered

# ---------------- Email → Client/Group maps ----------------
def _client_email_map():
    out = {}
    qs = Client.objects.all().only("id","client_id","client_name","email","extra_emails","group_id")
    for c in qs:
        if c.email:
            out[_norm_email(c.email)] = c
        for e in (c.extra_emails or []):
            ee = _norm_email(e)
            if ee:
                out[ee] = c
    return out

def _group_email_map():
    out = {}
    qs = Group.objects.all().only("id","group_name","group_email","extra_emails")
    for g in qs:
        if g.group_email:
            out[_norm_email(g.group_email)] = g
        for e in (g.extra_emails or []):
            ee = _norm_email(e)
            if ee:
                out[ee] = g
    return out

def _accs_for_client_id(client_id: str):
    accs = (AccountBank.objects
            .filter(client_id=client_id)
            .values("account_no","account","bank_name","stms_pws"))
    display, pw = [], set()
    for a in accs:
        label = f"{a['bank_name']} — {a['account']} ({a['account_no']})"
        display.append(label)
        if a.get("stms_pws"):
            pw.add(a["stms_pws"])
    return display, sorted(pw)

def _accs_for_client_id_by_tokens(client_id: str, tokens: dict):
    """
    From this client's accounts, return ONLY ONE best match by:
    full > last4 > last2. Returns ([one display string], [pw]).
    """
    fulls = tokens.get("full", set()) or set()
    l4    = tokens.get("last4", set()) or set()
    l2    = tokens.get("last2", set()) or set()

    qs = (AccountBank.objects
          .filter(client_id=client_id)
          .values("account_no", "account", "bank_name", "stms_pws"))

    def strength(acc_no: str) -> int:
        if not acc_no:
            return 0
        if acc_no in fulls:
            return 3
        if len(acc_no) >= 4 and acc_no[-4:] in l4:
            return 2
        if len(acc_no) >= 2 and acc_no[-2:] in l2:
            return 1
        return 0

    winner, win_s = None, 0
    for a in qs:
        acc = (a.get("account_no") or "").strip()
        s = strength(acc)
        if s > win_s:
            win_s = s
            winner = a
            if win_s == 3:
                break

    if not winner:
        return [], []

    label = f"{winner['bank_name']} — {winner['account']} ({winner['account_no']})"
    pws = [winner["stms_pws"]] if winner.get("stms_pws") else []
    return [label], pws


def _accs_via_email_and_tokens(rcv_emails: set[str], tokens: dict, att_recipients: set[str] | None = None):
    """
    Priority #1: acc_mail_id ∈ receiver emails AND account_no matches attachment tokens.
    Return (client, [ONE display string], [pw]) — single best match by:
      strength: full(3) > last4(2) > last2(1),
      then tie-breaker: acc_mail_id in attachment recipients > in receivers > lexicographic.
    """
    att_recipients = {e.strip().lower() for e in (att_recipients or set())}
    rcv_emails = {e.strip().lower() for e in (rcv_emails or set())}
    if not rcv_emails:
        return None, [], []

    qs = (AccountBank.objects
          .filter(acc_mail_id__in=list(rcv_emails))
          .values("client_id", "client__client_name",
                  "account_no", "account", "bank_name", "stms_pws", "acc_mail_id"))

    accs = list(qs)
    if not accs:
        return None, [], []

    fulls = tokens.get("full", set()) or set()
    l4    = tokens.get("last4", set()) or set()
    l2    = tokens.get("last2", set()) or set()

    def strength(acc_no: str) -> int:
        acc_no = (acc_no or "").strip()
        if not acc_no:
            return 0
        if acc_no in fulls:
            return 3
        if len(acc_no) >= 4 and acc_no[-4:] in l4:
            return 2
        if len(acc_no) >= 2 and acc_no[-2:] in l2:
            return 1
        return 0

    # Build scored candidates
    scored = []
    for a in accs:
        acc_no = (a.get("account_no") or "").strip()
        s = strength(acc_no)
        if s == 0:
            continue
        acc_mail = (a.get("acc_mail_id") or "").strip().lower()
        in_att   = 1 if acc_mail and acc_mail in att_recipients else 0
        in_rcv   = 1 if acc_mail and acc_mail in rcv_emails     else 0
        # Sort key: higher is better
        sort_key = (s, in_att, in_rcv, a["bank_name"] or "", a["account"] or "", acc_no or "")
        scored.append((sort_key, a))

    if not scored:
        return None, [], []

    # Choose the single best account overall (this also picks the client)
    scored.sort(reverse=True)  # highest (s, in_att, in_rcv, …)
    winner = scored[0][1]
    chosen_client_id = winner["client_id"]

    label = f"{winner['bank_name']} — {winner['account']} ({winner['account_no']})"
    pws = [winner["stms_pws"]] if winner.get("stms_pws") else []
    client = Client.objects.filter(client_id=chosen_client_id).select_related("group").first()
    return client, [label], pws




# ---------------- Search & order helpers for MailLog ----------------
def _apply_search_filters(qs, q):
    if not q:
        return qs
    cond = Q()
    for field in ("subject",
                  "sender_mail", "receiver_mail",   # your exact names
                  "sender","from_email","receivers","to","to_emails"):
        try:
            cond |= Q(**{f"{field}__icontains": q})
        except FieldError:
            pass
    return qs.filter(cond) if cond else qs

def _safe_order(qs):
    for f in ("rec_dat",   # your exact datetime field first
              "received_at","received_date","rec_date","date",
              "created_at","timestamp","fetched_at"):
        try:
            return qs.order_by(f"-{f}")
        except FieldError:
            continue
    try:
        return qs.order_by("-id")
    except FieldError:
        return qs

# ---------------- MAIN VIEW ----------------
def mail_log_report(request):
    # UI params
    active_fy = (request.GET.get("fy") or "").strip()   # "" or a DB alias like fy_2024_25
    q         = (request.GET.get("q") or "").strip()
    statusF   = (request.GET.get("status") or "").lower()
    flagF     = (request.GET.get("flag") or "").lower()
    detail    = (request.GET.get("detail") or "").lower()
    list_sel  = (request.GET.get("list") or "").lower()

    # build maps once
    client_by_email = _client_email_map()
    group_by_email  = _group_email_map()

    db_aliases  = _db_aliases_for_fy()
    read_aliases = [active_fy] if active_fy else db_aliases

    # fetch MailLog rows
    mail_objs = []
    for alias in read_aliases:
        try:
            qs = MailLog.objects.using(alias).all()
            qs = _apply_search_filters(qs, q)
            qs = _safe_order(qs)
            mail_objs.extend(list(qs[:500]))  # cap per DB
        except Exception:
            continue  # skip aliases without MailLog

    rows = []
    for m in mail_objs:
        # ---- normalize base fields (your names first) ----
        def gf(*names):
            for n in names:
                if hasattr(m, n):
                    return getattr(m, n)
            return None

        rec_dat    = gf("rec_dat","received_at","received_date","rec_date","date","created_at")
        fetched_at = gf("fetched_at","processed_at","created_at")
        sender     = gf("sender_mail","sender","from_email") or ""
        receivers  = gf("receiver_mail","receivers","to","to_emails","recipient","recipients") or ""
        subject    = gf("subject") or ""
        mailbox    = gf("mailbox","mailbox_name") or ""
        msg_id     = gf("msg_id","message_id","internet_message_id") or ""

        # attachments JSON -> [{name,url}] for template
        attachments = gf("attachments","files","files_meta") or []

        att_names, att_render = [], []
        att_recipients = set()
        if isinstance(attachments, list):
            for a in attachments:
                if isinstance(a, dict):
                    nm  = a.get("filename") or a.get("name") or a.get("path") or ""
                    url = a.get("url")
                    rec = (a.get("recipient") or "").strip().lower()
                    if rec:
                        att_recipients.add(rec)
                else:
                    nm, url = str(a), None
                att_names.append(nm)
                att_render.append({"name": nm or "attachment", "url": url})
        else:
            att_names = [str(attachments)]
            att_render = [{"name": str(attachments), "url": None}]


        rcv_emails = set(_split_emails(receivers))
        tokens = _extract_acc_tokens(att_names)

        # ---- Priority 1: acc_mail_id + attachment account ----
        # client, bank_display, stmt_pws = _accs_via_email_and_tokens(rcv_emails, tokens)
        client, bank_display, stmt_pws = _accs_via_email_and_tokens(rcv_emails, tokens, att_recipients)

        group = getattr(client, "group", None) if client else None
        email_match = "Known" if client else "Unknown"

        # ---- Priority 2: receiver → Client ----
        if not client:
            for e in rcv_emails:
                if e in client_by_email:
                    client = client_by_email[e]
                    group = getattr(client, "group", None)
                    # bank_display, stmt_pws = _accs_for_client_id(client.client_id)
                    bank_display, stmt_pws = _accs_for_client_id_by_tokens(client.client_id, tokens)

                    email_match = "Known"
                    break

        # ---- Priority 3: receiver → Group ----
        if (not client) and (not group):
            for e in rcv_emails:
                if e in group_by_email:
                    group = group_by_email[e]
                    email_match = "Known"
                    break

        # ---- flag + status ----
        flag_val = ""
        if email_match == "Unknown":
            flag_val = "mail_not_registered"
        else:
            if client and not bank_display:
                flag_val = "bank_not_added"
        if not att_render or (len(att_render) == 1 and (att_render[0]["name"] in [None, "", "attachment"])):
            flag_val = flag_val or "document_not_added"

        status_val = "done" if (email_match == "Known" and (bank_display or client or group)) else "pending"

        # apply UI filters
        if statusF and status_val != statusF:
            continue
        if flagF and flag_val != flagF:
            continue

        rows.append({
            "id": getattr(m, "id", None),
            "rec_dat": _to_local(rec_dat),   # ← will render in template
            "sender_mail": sender,
            "receiver_mail": ", ".join(sorted(rcv_emails)) if rcv_emails else (receivers or ""),
            "subject": subject,
            "attachments": att_render,
            "mailbox": mailbox,
            "status": status_val,
            "flag": flag_val or "",
            "email_match": email_match,
            "group_name": getattr(group, "group_name", None) if group else None,
            "client_name": getattr(client, "client_name", None) if client else None,
            "bank_accounts": bank_display,
            "stmt_pws": stmt_pws,
            "msg_id": msg_id,
            "fetched_at": _to_local(fetched_at),
        })

    context = {
        "db_aliases": _db_aliases_for_fy(),  # dropdown in your template
        "active_fy": active_fy,
        "q": q,
        "status": statusF,
        "flag": flagF,
        "detail": detail,
        "list_sel": list_sel,
        "rows": rows,
    }
    return render(request, "mail_log.html", context)


# bank mapping
import os, pyodbc, logging, time, gc, multiprocessing
from django.db.models import Q, Prefetch
from django.shortcuts import render
from django.http import JsonResponse
from .models import Client, AccountBank
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed

# ---- CONFIG ----
BUSY_ROOT = r"Y:\Busy\DATA"
BUSY_PASSWORD = "ILoveMyINDIA"
MAX_COMPANIES_PER_REQUEST = 1500
MAX_PARALLEL = max(2, multiprocessing.cpu_count() - 1)
# KAMEET_DB_BASE = r"Z:\Busy\Data\KameetData"   # <-- change if your file is elsewhere
BUSY_LOCK = threading.Lock()

logging.basicConfig(level=logging.INFO)
def normalize_code(v):
    """
    Convert busyacccode to clean string:
    1282.0 → '1282'
    '1282 ' → '1282'
    None → ''
    """
    if v is None:
        return ""
    try:
        # handles float like 1282.0
        iv = int(float(v))
        return str(iv)
    except Exception:
        return str(v).strip()

# --------------------- UTIL ---------------------
def _busy_file_path(company_code: str, fy_code: str) -> str:
    folder = os.path.join(BUSY_ROOT, company_code)
    for name in [f"db1{fy_code}.bds", f"db{fy_code}.bds", "db.bds"]:
        path = os.path.join(folder, name)
        if os.path.exists(path):
            return path
    try:
        all_bds = [os.path.join(folder, f) for f in os.listdir(folder) if f.lower().endswith(".bds")]
        if all_bds:
            all_bds.sort(key=os.path.getmtime, reverse=True)
            return all_bds[0]
    except Exception:
        pass
    return os.path.join(folder, f"db1{fy_code}.bds")


def _fetch_busy_bank_accounts(bds_path: str):

    with BUSY_LOCK:
        if not os.path.exists(bds_path) or os.path.getsize(bds_path) < 100 * 1024:
            return []

        conn_str = f"DRIVER={{Microsoft Access Driver (*.mdb, *.accdb)}};DBQ={bds_path};PWD={BUSY_PASSWORD};"
        conn = cur = None
        result = []

        try:
            conn = pyodbc.connect(conn_str, autocommit=True, timeout=5)
            cur = conn.cursor()

            # Verify Master1 table exists
            existing = {r.table_name.lower() for r in cur.tables()}
            if not any(t in existing for t in ["master1", "master"]):
                logging.warning(f"[Busy] No Master table in {bds_path}")
                return []


            # ✅ ADD THIS TEST HERE
            test_q = "SELECT TOP 1 * FROM masteraddressinfo"
            cur.execute(test_q)
            cols = [c[0] for c in cur.description]
            row = cur.fetchone()
            logging.info(f"[BUSY TEST] Columns: {cols}")
            logging.info(f"[BUSY TEST] Sample row: {row}")
            # ⬆️ this will show actual column names and one row of data


            # ✅ Senior’s subselect query (works in all Busy builds)
            query = """
                SELECT
                    M1.Code,
                    M1.Name,
                    (SELECT Name FROM Master1 WHERE Master1.Code = 
                        (SELECT ParentGrp FROM Master1 WHERE Master1.Code = M1.Code)) AS AcGroup,
                    (SELECT C4 FROM masteraddressinfo WHERE masteraddressinfo.MasterCode = M1.Code) AS BankName,
                    (SELECT AccNo FROM masteraddressinfo WHERE masteraddressinfo.MasterCode = M1.Code) AS AccNo,
                    (SELECT C5 FROM masteraddressinfo WHERE masteraddressinfo.MasterCode = M1.Code) AS IFSC,
                    (SELECT C8 FROM masteraddressinfo WHERE masteraddressinfo.MasterCode = M1.Code) AS SwiftCode
                FROM Master1 AS M1
                WHERE M1.ParentGrp IN (112, 128)
                ORDER BY M1.Code, M1.Name;
            """

            cur.execute(query)
            rows = cur.fetchall()
            logging.info(f"[DEBUG BUSY] ✅ Fetched {len(rows)} rows from {os.path.basename(bds_path)}")
            # ✅ ADD THIS REGEX EXTRACTOR just before looping rows
            import re
            ACCOUNT_RX = re.compile(r'\d{6,}')  # 6+ consecutive digits = likely account number

            # ✅ LOOP START
            for r in rows:
                name = getattr(r, "Name", "") or ""
                acc_no = getattr(r, "AccNo", "") or ""

                # If AccNo field empty, extract digits from Name (e.g., "BANK OF BARODA - 27380100001812")
                if not acc_no:
                    m = ACCOUNT_RX.search(name)
                    if m:
                        acc_no = m.group(0)

                # optional debug
                logging.info(f"[DEBUG BUSY] Row sample: Name={name}, AccNo={acc_no}")

                # store result
                result.append({
                    "Code": getattr(r, "Code", None),
                    "Name": name,
                    "AcGroup": getattr(r, "AcGroup", None),
                    "BankName": getattr(r, "BankName", None),
                    "AccNo": acc_no,
                    "IFSC": getattr(r, "IFSC", None),
                    "SwiftCode": getattr(r, "SwiftCode", None),
                })

        except Exception as e:
            logging.error(f"[Busy] Error reading {bds_path}: {e}")
        finally:
            if cur:
                try:
                    cur.close()
                except:
                    pass
            if conn:
                try:
                    conn.close()
                except:
                    pass
            gc.collect()

        return result

# --------------------- WORKER (no Django imports here) ---------------------
def _read_company_parallel(args):
    company_code, fy_code = args
    try:
        bds_path = _busy_file_path(company_code, fy_code)
        return (company_code, _fetch_busy_bank_accounts(bds_path))
    except Exception as e:
        logging.debug(f"[Busy] Error in {company_code}: {e}")
        return (company_code, [])


# --------------------- AJAX BUSY ---------------------
def ajax_busy_accounts(request):
    busy_code = (request.GET.get("busy_code") or "").strip()
    fy_code = (request.GET.get("fy") or "2024").strip()

    diag = {"busy_code": busy_code, "fy": fy_code}
    if not busy_code:
        diag["error"] = "Missing busy_code"
        return JsonResponse({"ok": False, "diag": diag, "rows": []})

    bds_path = _busy_file_path(busy_code, fy_code)
    diag["bds_path"] = bds_path
    diag["exists"] = os.path.exists(bds_path)
    if not diag["exists"]:
        diag["error"] = "bds file not found"
        return JsonResponse({"ok": False, "diag": diag, "rows": []})
    

    rows = _fetch_busy_bank_accounts(bds_path)

    # 🔒 REMOVE "Do Not Display" BUSY accounts
    hidden_acc_nos = set(
        clean_acc_no(a.account_no)
        for a in AccountBank.objects.using("default")
            .filter(
                client__busy_code=busy_code,
               account_group__in=["Do Not Display", "Secured Loans"]
            )
            .exclude(account_no__isnull=True)
            .exclude(account_no__exact="")
    )

    rows = [
        r for r in rows
        if clean_acc_no(r.get("AccNo") or r.get("Name")) not in hidden_acc_nos
    ]

    return JsonResponse({"ok": True, "diag": diag, "rows": rows})
    # rows = _fetch_busy_bank_accounts(bds_path)
    # return JsonResponse({"ok": True, "diag": diag, "rows": rows})

import threading
from concurrent.futures import ThreadPoolExecutor, as_completed

def safe_threadpool_run(funcs, max_workers=8):
    """Safely run threads inside Django devserver."""
    if not threading.main_thread().is_alive():
        return {}
    results = {}
    try:
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {executor.submit(f, *args): args[0] for f, args in funcs}
            # futures = {executor.submit(f, args): args[0] for f, args in funcs}
            for future in as_completed(futures):
                key = futures[future]
                try:
                    results[key] = future.result()
                except Exception as e:
                    logging.warning(f"[Busy] Thread error in {key}: {e}")
        return results
    except RuntimeError as e:
        logging.warning(f"[Busy] Thread pool shut down early: {e}")
        return results


# --------------------- MAIN VIEW ---------------------
from concurrent.futures import ThreadPoolExecutor, as_completed

def bank_account_mapping(request):
    q = (request.GET.get("q") or "").strip()
    status = (request.GET.get("status") or "All").strip()
    fy_code = (request.GET.get("fy") or "2024").strip()

    # ---- Get Clients ----
    qs = Client.objects.using("default").all()
    if status and status.lower() != "all":
        qs = qs.filter(Q(status__iexact=status) | Q(status__isnull=True) | Q(status=""))
    if q:
        qs = qs.filter(
            Q(client_name__icontains=q)
            | Q(trade_name__icontains=q)
            | Q(client_id__icontains=q)
        )

    qs = qs.prefetch_related(
        Prefetch(
            "accounts",
            queryset=AccountBank.objects.using("default")
            # .filter(bank_name__isnull=False, 
            #         account_no__isnull=False)
            .filter(account_no__isnull=False)
            # .exclude(bank_name__exact="")
            .exclude(account_no__exact="")
            .exclude(account_group__in=["Do Not Display", "Secured Loans"]) 
            .order_by("bank_name", "account_no"),
        )
    ).order_by("client_name").only("id", "client_id", "client_name", "busy_code")

    clients = list(qs)
    company_codes = [(c.busy_code.strip(), fy_code) for c in clients if c.busy_code]

    logging.info(f"[Busy]  Starting threaded fetch for {len(company_codes)} companies")

    # ---- Run in threads safely ----
    funcs = [(_read_company_parallel, args) for args in company_codes]
    company_cache_raw = safe_threadpool_run(funcs, max_workers=8)
    company_cache = {code: rows for code, rows in company_cache_raw.items()}

    logging.info(f"[Busy] ✅ Completed threaded fetch for {len(company_cache)} companies")

    # ---- Prepare mapping ----
    mapping_data = []
    for c in clients:
        code = (c.busy_code or "").strip()

        #  Collect Do Not Display account numbers
        hidden_acc_nos = set(
            clean_acc_no(a.account_no)
            for a in AccountBank.objects.using("default")
                .filter(client=c, account_group__in=["Do Not Display", "Secured Loans"])
                .exclude(account_no__isnull=True)
                .exclude(account_no__exact="")
        )
        mapping_data.append({
            "client": c,
            "busy_company": code,
            # "busy_accounts": company_cache.get(code, []),
           "busy_accounts": [
                b for b in company_cache.get(code, [])
                if clean_acc_no(
                    b.get("AccNo") or b.get("Name")
                ) not in hidden_acc_nos
            ],

            "accountbank_list": getattr(c, "accounts", []).all() if hasattr(c, "accounts") else [],
        })

    return render(
        request,
        "bank_mapping.html",
        {
            "clients": clients,
            "mapping_data": mapping_data,
            "q": q,
            "status": status,
            "fy": fy_code,
            "busy_root": BUSY_ROOT,
        },
    )

# mapping peding in busy 
def clean_acc_no(v):
    if not v:
        return ""
    return "".join(ch for ch in str(v) if ch.isdigit())


def _busy_latest_file(company_code: str, fy: str) -> str:
    """
    Return BUSY .bds file for a specific FY (file-based)
    Example: FY 2023-24 → db12023.bds
    """

    # Extract starting year
    start_year = int(fy.split("-")[0])   # "2023-24" → 2023

    # BUSY file naming convention
    bds_filename = f"db1{start_year}.bds"

    folder = os.path.join(BUSY_ROOT, company_code)
    bds_path = os.path.join(folder, bds_filename)

    if os.path.exists(bds_path):
        return bds_path

    return ""

def _busy_files_from_2023(company_code: str, start_year=2023):
    """
    Return list of BUSY .bds files from start_year till latest
    """
    folder = os.path.join(BUSY_ROOT, company_code)
    if not os.path.isdir(folder):
        return []

    files = []
    for fname in os.listdir(folder):
        if fname.startswith("db1") and fname.endswith(".bds"):
            try:
                year = int(fname.replace("db1", "").replace(".bds", ""))
                if year >= start_year:
                    files.append(os.path.join(folder, fname))
            except ValueError:
                continue

    # Sort oldest → newest
    return sorted(files)

from django.db.models import Prefetch
from django.shortcuts import render
from .models import Client, AccountBank,PendingBusyBankMapping
from django.contrib import messages
from django.utils.dateparse import parse_datetime

def get_busy_fy_folder(fy: str) -> str:
    """
    Convert FY like '2023-24' to folder name.
    Adjust if your folder naming is different.
    """
    return fy



def pending_bank_mapping(request):

    rebuild = request.GET.get("rebuild") == "1"

    # ----------------------------------
    # IF NOT rebuild → load from DB
    # ----------------------------------
    if not rebuild and PendingBusyBankMapping.objects.exists():
      
        # rows = (
        #     PendingBusyBankMapping.objects
        #     .select_related("client", "group")
        #     .order_by("client__client_name", "fy", "bank_name")
        # )


        # -------------------------------------------------
        # LOAD pending rows (EXCLUDE Do Not Display / Loans)
        # -------------------------------------------------

        # 1️⃣ Collect hidden account numbers per client
        hidden_map = {
            (a.client.client_id, clean_acc_no(a.account_no))
            for a in AccountBank.objects.using("default")
                .filter(account_group__in=["Do Not Display", "Secured Loans"])
                .exclude(account_no__isnull=True)
                .exclude(account_no__exact="")
        }

        # 2️⃣ Load pending rows
        rows = (
            PendingBusyBankMapping.objects
            .select_related("client", "group")
            .order_by("client__client_name", "fy", "bank_name")
        )

        # 3️⃣ REMOVE hidden rows
        # rows = [
        #     r for r in rows
        #     if (r.client.client_id, clean_acc_no(r.account_no)) not in hidden_map
        # ]


        rows = [
            r for r in rows
            if (r.client.client_id, clean_acc_no(r.account_no)) not in hidden_map
        ]


        # 🔹 Map busyacccode → account_group from AccountBank
        # account_group_map = {
        #     str(a.busyacccode).strip(): a.account_group
        #     for a in AccountBank.objects.using("default")
        #         .exclude(busyacccode__isnull=True)
        #         .exclude(busyacccode__exact="")
        # }

     
        # account_group_map = {
        #     (a.client.client_id, str(a.busyacccode).strip()): a.account_group
        #     for a in AccountBank.objects
        #         .exclude(busyacccode__isnull=True)
        #         .exclude(busyacccode__exact="")
        # }

        # ----------------------------------
        # BUILD ACCOUNT GROUP LOOKUPS
        # ----------------------------------

        # 1️⃣ Primary match: BUSY Account Code
        # busy_code_group_map = {
        #     (a.client.client_id, str(a.busyacccode).strip()): a.account_group
        #     for a in AccountBank.objects
        #         .exclude(busyacccode__isnull=True)
        #         .exclude(busyacccode__exact="")
        # }

        # # 2️⃣ Fallback match: Account Number
        # acc_no_group_map = {
        #     (a.client.client_id, clean_acc_no(a.account_no)): a.account_group
        #     for a in AccountBank.objects
        #         .exclude(account_no__isnull=True)
        #         .exclude(account_no__exact="")
        # }
        # ✅ SINGLE SOURCE OF TRUTH: account_no → account_group
        account_group_map = {
            (a.client.client_id, clean_acc_no(a.account_no)): a.account_group
            for a in AccountBank.objects
                .exclude(account_no__isnull=True)
                .exclude(account_no__exact="")
        }



        # for r in rows:
        #     key = (
        #             r.client.client_id,
        #             str(r.busy_account_code).strip() if r.busy_account_code else ""
        #         )

        #     r.account_group_display = account_group_map.get(key, "—")

        #     print(
        #         "[PENDING MAP]",
        #         "client_id =", r.client_id,
        #         "acc_no =", clean_acc_no(r.account_no),
        #         "account_group =", account_group_map.get(key)
        #     )

        # ----------------------------------
        # ATTACH ACCOUNT GROUP TO EACH ROW
        # ----------------------------------
        for r in rows:
            key = (r.client.client_id, clean_acc_no(r.account_no))
            r.account_group_display = account_group_map.get(key, "—")

            print(
                "[PENDING MAP]",
                r.client.client_id,
                clean_acc_no(r.account_no),
                "→",
                r.account_group_display
            )

        # for r in rows:
        #     account_group = None

        #     # ✅ First: match using BUSY account code
        #     if r.busy_account_code:
        #         key1 = (
        #             r.client.client_id,
        #             str(r.busy_account_code).strip()
        #         )
        #         account_group = busy_code_group_map.get(key1)

        #     # 🔁 Fallback: match using account number
        #     if not account_group:
        #         key2 = (
        #             r.client.client_id,
        #             clean_acc_no(r.account_no)
        #         )
        #         account_group = acc_no_group_map.get(key2)

        #     r.account_group_display = account_group or "—"

        #     print(
        #         "[PENDING MAP]",
        #         "client_id =", r.client.client_id,
        #         "busy_code =", r.busy_account_code,
        #         "acc_no =", clean_acc_no(r.account_no),
        #         "account_group =", r.account_group_display
        #     )




        # 🔹 FETCH AccountBank busyacccode lookup
        # account_code_map = {
        #     (a.client.id, clean_acc_no(a.account_no)): a.busyacccode
        #     for a in AccountBank.objects.using("default")
        #     .select_related("client")
        #     .exclude(busyacccode="")
        # }

        # # 🔹 Attach busyacccode to each pending row
        # print("\n====== PENDING BUSY BANK MAPPING ROWS ======")

        # for r in rows:
        #     key = (r.client_id, clean_acc_no(r.account_no))
        #     r.busyacccode_display = account_code_map.get(key)

        # for r in rows:
        #     print({
        #         "id": r.id,
        #         "client_id": r.client_id,
        #         "client_name": r.client.client_name if r.client else None,
        #         "group": r.group.group_name if r.group else None,
        #         "busy_company_code": r.busy_company_code,
        #         "bank_name": r.bank_name,
        #         "account_no": r.account_no,
        #         "ifsc": r.ifsc,
        #         "swift": r.swift,
        #         "fy": r.fy,
        #         "busy_updated_at": r.busy_updated_at,
        #         "busyacccode_display": r.busyacccode_display,  # ✅ THIS IS KEY
        #     })
        #     print("====== END ======\n")


        

        return render(
            request,
            "pending_bank_mapping.html",
            {
                "rows": rows,
                "total_pending": len(rows),
                # "total_pending": rows.count(),
                "from_db": True
            }
        )

    # ----------------------------------
    # REBUILD DATA FROM BUSY
    # ----------------------------------
    # PendingBusyBankMapping.objects.all().delete()

    clients = (
        Client.objects.using("default")
        .select_related("group")
        .prefetch_related(
            Prefetch(
                "accounts",
                queryset=AccountBank.objects.using("default")
            )
        )
        .only("id", "client_name", "busy_code", "group")
    )

    # objects_to_create = []

    for c in clients:
        busy_code = (c.busy_code or "").strip()
        if not busy_code:
            continue
        

        # ❌ Accounts that must NEVER appear in pending list
        excluded_acc_nos = set(
            clean_acc_no(a.account_no)
            for a in AccountBank.objects.using("default")
                .filter(
                    client=c,
                    account_group__in=[
                        "Do Not Display",
                        "Secured Loans"
                    ]
                )
                .exclude(account_no__isnull=True)
                .exclude(account_no__exact="")
        )

        # ✅ Already mapped bank accounts (Bank + OD)
        mapped_bank_acc_nos = set(
            clean_acc_no(a.account_no)
            for a in AccountBank.objects.using("default")
                .filter(
                    client=c,
                    account_group__in=[
                        "Bank Accounts",
                        "Bank O/D Account"
                    ]
                )
                .exclude(account_no__isnull=True)
                .exclude(account_no__exact="")
        )

        # ✅ Busy account codes already existing in KAMEET (ONLY truth)
        existing_busy_codes = {
            str(a.busyacccode).strip()
            for a in AccountBank.objects.using("default")
                .filter(client=c)
                .exclude(busyacccode__isnull=True)
                .exclude(busyacccode__exact="")
        }


        # mapped_acc_nos = {
        #     clean_acc_no(a.account_no)
        #     for a in c.accounts.all()
        #     if a.account_no
        # }


        
        # 🔥 HARD CLEAN: remove pending rows already mapped in AccountBank
        PendingBusyBankMapping.objects.filter(
            client=c
        ).filter(
            Q(account_no__in=mapped_bank_acc_nos) |
            Q(busy_account_code__in=existing_busy_codes)
        ).delete()
    
        # 🔥 HARD CLEAN: remove Do Not Display / Secured Loans (ACCOUNT NO + BUSY CODE)
        excluded_busy_codes = {
            str(a.busyacccode).strip()
            for a in AccountBank.objects.using("default")
                .filter(
                    client=c,
                    account_group__in=["Do Not Display", "Secured Loans"]
                )
                .exclude(busyacccode__isnull=True)
                .exclude(busyacccode__exact="")
        }

        PendingBusyBankMapping.objects.filter(
            client=c
        ).filter(
            Q(account_no__in=excluded_acc_nos) |
            Q(busy_account_code__in=excluded_busy_codes)
        ).delete()

        # 🔥 HARD CLEAN: remove Do Not Display / Secured Loans
        PendingBusyBankMapping.objects.filter(
            client=c,
            account_no__in=excluded_acc_nos
        ).delete()

        bds_files = _busy_files_from_2023(busy_code, start_year=2023)
        if not bds_files:
            continue

        # for bds in bds_files:
        #     year = int(os.path.basename(bds).replace("db1", "").replace(".bds", ""))
        #     fy_label = f"{year}-{str(year + 1)[-2:]}"

        #     modified_dt = datetime.fromtimestamp(os.path.getmtime(bds))

        #     busy_accounts = _fetch_busy_bank_accounts(bds)

        #     for b in busy_accounts:
        #         busy_acc = clean_acc_no(
        #             b.get("AccNo") or b.get("AccountNo") or b.get("AcNo")
        #         )
        #         if not busy_acc or busy_acc in mapped_acc_nos:
        #             continue
        #         obj, created = PendingBusyBankMapping.objects.update_or_create(
        #             client=c,
        #             busy_company_code=busy_code,
        #             account_no=busy_acc,
        #             fy=fy_label,
        #             defaults={
        #                 "group": c.group,
        #                 "bank_name": b.get("BankName") or b.get("Name") or "",
        #                 "ifsc": b.get("IFSC"),
        #                 "swift": b.get("SwiftCode"),
        #                 "busy_updated_at": modified_dt,
        #             }
        #         )

        seen_busy_accounts = set()  # ✅ ADD THIS

        for bds in bds_files:
            modified_dt = datetime.fromtimestamp(os.path.getmtime(bds))
            busy_accounts = _fetch_busy_bank_accounts(bds)

            for b in busy_accounts:
                busy_acc = clean_acc_no(
                    b.get("AccNo") or b.get("AccountNo") or b.get("AcNo")
                )

                if not busy_acc:
                    continue

                # ✅ Extract BUSY account code (SAFE)
                busy_code_val = (
                    b.get("AccCode") or
                    b.get("AccountCode") or
                    b.get("Code")
                )
                busy_code_val = str(busy_code_val).strip() if busy_code_val else ""

                # ❌ Skip if BUSY account code already exists in AccountBank
                if busy_code_val and busy_code_val in existing_busy_codes:
                    continue


                # ❌ Skip if already mapped in KAMEET
                # if busy_acc in mapped_acc_nos:
                #     continue

                # ❌ Skip if Do Not Display or Secured Loan
                if busy_acc in excluded_acc_nos:
                    continue

                # ❌ Skip if already mapped as Bank / OD account
                if busy_acc in mapped_bank_acc_nos:
                    continue


                # ❌ Skip if already added once (FY duplication fix)
                unique_key = (c.id, busy_acc)
                if unique_key in seen_busy_accounts:
                    continue

                seen_busy_accounts.add(unique_key)

                PendingBusyBankMapping.objects.update_or_create(
                    client=c,
                    busy_company_code=busy_code,
                    account_no=busy_acc,
                    defaults={          # ✅ FY REMOVED
                        "group": c.group,
                        "bank_name": b.get("BankName") or b.get("Name") or "",
                        "ifsc": b.get("IFSC"),
                        "swift": b.get("SwiftCode"),
                        "busy_account_code": (
                            b.get("AccCode") or
                            b.get("AccountCode") or
                            b.get("Code")
                        ),
                        "busy_updated_at": modified_dt,
                    }
                )


                # objects_to_create.append(
                #     PendingBusyBankMapping(
                #         client=c,
                #         group=c.group,
                #         busy_company_code=busy_code,
                #         bank_name=b.get("BankName") or b.get("Name") or "",
                #         account_no=busy_acc,
                #         ifsc=b.get("IFSC"),
                #         swift=b.get("SwiftCode"),
                #         fy=fy_label,
                #         busy_updated_at=modified_dt,
                #     )
                # )

    # ----------------------------------
    # BULK INSERT
    # ----------------------------------
    # PendingBusyBankMapping.objects.bulk_create(
    #     objects_to_create,
    #     ignore_conflicts=True
    # )

    messages.success(
        request,
        "BUSY bank mapping refreshed successfully (data merged, not deleted)."
    )

    rows = (
        PendingBusyBankMapping.objects
        .select_related("client", "group")
        .order_by("client__client_name", "fy", "bank_name")
    )

    # 🔹 FETCH AccountBank busyacccode lookup
    # account_code_map = {
    #     (a.client.id, clean_acc_no(a.account_no)): a.busyacccode
    #     for a in AccountBank.objects.using("default")
    #     .select_related("client")
    #     .exclude(busyacccode="")
    # }

    # # 🔹 Attach busyacccode to each pending row
    # for r in rows:
    #     key = (r.client_id, clean_acc_no(r.account_no))
    #     r.busyacccode_display = account_code_map.get(key)
        
    return render(
        request,
        "pending_bank_mapping.html",
        {
            "rows": rows,
            "total_pending": rows.count(),
            "from_db": False
        }
    )




from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
import json

@csrf_exempt
def ajax_add_account(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "Invalid method"})

    try:
        data = json.loads(request.body.decode("utf-8"))
        client_id = data.get("client_id")
        bank_name = data.get("bank_name")
        account_no = data.get("account_no")

        if not (client_id and bank_name and account_no):
            return JsonResponse({"ok": False, "error": "Missing data"})

        client = Client.objects.filter(client_id=client_id).first()
        if not client:
            return JsonResponse({"ok": False, "error": "Client not found"})

        # Check for duplicates
        if AccountBank.objects.filter(client=client, account_no=account_no).exists():
            return JsonResponse({"ok": False, "error": "Account already exists"})

        acc = AccountBank.objects.create(
            client=client,
            bank_name=bank_name,
            account_no=account_no,
        )

        return JsonResponse({"ok": True, "id": acc.id})
    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)})

from django.views.decorators.http import require_GET
from django.http import JsonResponse
import logging

@require_GET
def ajax_filtered_mapping(request):
    client_id = request.GET.get("client_id")
    busy_code = (request.GET.get("busy_code") or "").strip()
    fy_code   = (request.GET.get("fy") or "2024").strip()
    status    = (request.GET.get("status") or "all").lower()

    logging.info(f"[FILTER] 🔹 client_id={client_id}, busy_code={busy_code}, fy={fy_code}, status={status}")

    if not client_id:
        return JsonResponse({"ok": False, "error": "Missing client ID", "bank_accounts": [], "busy_accounts": []})

    try:
        client = Client.objects.using("default").prefetch_related("accounts").get(id=client_id)
    except Client.DoesNotExist:
        return JsonResponse({"ok": False, "error": "Client not found", "bank_accounts": [], "busy_accounts": []})

    # --- Busy accounts (right) ---
    busy_accounts = []
    if busy_code:
        bds_path = _busy_file_path(busy_code, fy_code)
        busy_accounts = _fetch_busy_bank_accounts(bds_path)
        # 🔒 Remove BUSY accounts that are Do Not Display in AccountBank
        hidden_accnos = {
            clean_acc_no(a.account_no)
            for a in client.accounts
                .filter(account_group__in=["Do Not Display", "Secured Loans"])
                .exclude(account_no__isnull=True)
                .exclude(account_no__exact="")
        }

        busy_accounts = [
            b for b in busy_accounts
            if clean_acc_no(b.get("AccNo") or b.get("Name")) not in hidden_accnos
        ]
    # --- Local accounts (left) ---
    bank_accounts = list(
        # client.accounts.filter(bank_name__isnull=False)
        client.accounts
        .exclude(account_group__in=["Do Not Display", "Secured Loans"]) 
        # .exclude(bank_name__exact="")
        .exclude(account_no__exact="")
        .values("bank_name","account", "account_no", "busyacccode")
    )

    # Sets reused by all branches
    # mapped_codes  = {str(a.get("busyacccode", "")).strip() for a in bank_accounts if (a.get("busyacccode") or "").strip()}
    # mapped_accnos = {str(a.get("account_no", "")).strip()  for a in bank_accounts if (a.get("busyacccode") or "").strip()}
    mapped_codes = {
        normalize_code(a.get("busyacccode"))
        for a in bank_accounts
        if normalize_code(a.get("busyacccode"))
    }

    mapped_accnos = {
        str(a.get("account_no", "")).strip()
        for a in bank_accounts
        if normalize_code(a.get("busyacccode"))
    }


    # --- Filtering logic ---
    if status == "done":
        # show only mapped pairs (exists in Busy AND local has busyacccode)
        # busy_accnos = {str(b.get("AccNo", "")).strip() for b in busy_accounts if b.get("AccNo")}
        # bank_accounts = [
        #     a for a in bank_accounts
        #     # if (a.get("busyacccode") or "").strip() and str(a["account_no"]).strip() in busy_accnos
        #     if normalize_code(a.get("busyacccode")) and str(a["account_no"]).strip() in busy_accnos

        # ]
        # busy_accounts = [
        #     b for b in busy_accounts
        #     if str(b.get("AccNo", "")).strip() in {str(a["account_no"]).strip() for a in bank_accounts}
        # ]
        # DONE = mapped by Busy CODE (primary key)
        mapped_codes = {
            normalize_code(a.get("busyacccode"))
            for a in bank_accounts
            if normalize_code(a.get("busyacccode"))
        }

        bank_accounts = [
            a for a in bank_accounts
            if normalize_code(a.get("busyacccode"))
        ]

        busy_accounts = [
            b for b in busy_accounts
            if normalize_code(b.get("Code")) in mapped_codes
        ]

    elif status == "pending":
        # LEFT: all UNMAPPED local accounts (ignore whether their number exists in Busy)
        pending_bank_accounts = [a for a in bank_accounts if not (a.get("busyacccode") or "").strip()]

        # RIGHT: Busy rows that are NOT mapped to any local (by code OR by accno)
        # def is_mapped_busy(row):
        #     code = str(row.get("Code",  "")).strip()
        #     acc  = str(row.get("AccNo", "")).strip()
        #     return (code and code in mapped_codes) or (acc and acc in mapped_accnos)
        def is_mapped_busy(row):
            code = normalize_code(row.get("Code"))
            acc  = str(row.get("AccNo", "")).strip()
            return (code in mapped_codes) or (acc in mapped_accnos)


        pending_busy_accounts = [b for b in busy_accounts if not is_mapped_busy(b)]

        bank_accounts = pending_bank_accounts
        busy_accounts = pending_busy_accounts

    else:
        # "all" → no extra filtering
        pass

    # Build mapping pairs for UI (based on current filtered lists)
    mappings = []
    try:
        # for a in bank_accounts:
        #     acc = str(a["account_no"]).strip()
        #     linked = next((r for r in busy_accounts if str(r.get("AccNo", "")).strip() == acc), None)
        #     if not linked:
        #         linked = next((r for r in busy_accounts if str(r.get("AccNo", ""))[-2:] == acc[-2:]), None)
        #     if linked:
        #         mappings.append({
        #             "bank_name": a["bank_name"],
        #             "bank_acc": acc,
        #             "busy_name": linked.get("Name") or linked.get("BankName"),
        #             "busy_acc": linked.get("AccNo"),
        #             "busy_code": linked.get("Code", ""),
        #         })
        for a in bank_accounts:
            busy_code = normalize_code(a.get("busyacccode"))

            linked = next(
                (r for r in busy_accounts if normalize_code(r.get("Code")) == busy_code),
                None
            )

            if linked:
                mappings.append({
                    "bank_name": a["bank_name"],
                    "bank_acc": a["account_no"],
                    "busy_name": linked.get("Name") or linked.get("BankName"),
                    "busy_acc": linked.get("AccNo") or "",
                    "busy_code": linked.get("Code"),
                })

    except Exception as e:
        logging.exception("[FILTER] mapping error")

    return JsonResponse({
        "ok": True,
        "bank_accounts": bank_accounts,
        "busy_accounts": busy_accounts,
        "mappings": mappings,
    })

# --------------------- AJAX SUBMIT MAPPING ---------------------
from django.views.decorators.csrf import csrf_exempt
import json
import pyodbc
import os
import logging
from datetime import datetime
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from datetime import datetime
import json, pyodbc, logging

def _escape_access(s: str) -> str:
    return s.replace("'", "''") if s is not None else ""

def _access_date_literal(dt: datetime) -> str:
    # Access date/time literal format
    return f"#{dt.strftime('%m/%d/%Y %H:%M:%S')}#"

def _load_acc_tbl_types(cursor):
    """
    Return dict: {COLNAME_UPPER: type_name_upper} for AccTbl columns.
    """
    types = {}
    for c in cursor.columns(table="AccTbl"):
        if getattr(c, "column_name", None):
            types[c.column_name.upper()] = (c.type_name or "").upper()
    return types

def _as_literal(value, type_name_upper):
    """
    Build a correct Access literal for the given value and target column type.
    Handles Text/LongText, Numeric types, Date/Time, and NULLs.
    """
    t = (type_name_upper or "").upper()

    # NULLs
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return "NULL"

    # Text types
    if "CHAR" in t or "TEXT" in t or "MEMO" in t or "LONGCHAR" in t:
        return f"'{_escape_access(str(value))}'"

    # Date/Time types
    if "DATE" in t or "TIME" in t:
        # value may be str or datetime
        if isinstance(value, datetime):
            return _access_date_literal(value)
        # attempt parse? simple fallback: quote as #...# if parseable else as string
        try:
            # try common formats
            dt = datetime.fromisoformat(str(value).replace("Z", "+00:00"))  # tolerate ISO-ish
            return _access_date_literal(dt)
        except Exception:
            # If column really is Date/Time but value not parseable -> let Access throw;
            # otherwise you can change the column to Text.
            return _access_date_literal(datetime.now())

    # Boolean
    if "BIT" in t or "YESNO" in t or "BOOLEAN" in t:
        v = str(value).strip().lower()
        if v in ("1", "true", "yes", "y", "on"):
            return "True"   # Access boolean literal
        return "False"

    # Numeric (INTEGER/SMALLINT/BYTE/LONG/COUNTER/DECIMAL/NUMERIC/DOUBLE/SINGLE)
    # Try int then float; if fails, return NULL to avoid 22018
    s = str(value).strip()
    try:
        if "." in s:
            return str(float(s))
        return str(int(s))
    except Exception:
        # as last resort, NULL (or change your column to text)
        return "NULL"


# views.py
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods, require_GET
from django.db import transaction
from .models import AccountBank, Client
import json
import logging
from django.db.models import Q
from django.http import Http404


import re

_DIGIT_RX = re.compile(r"\D+")

def _digits(s: str) -> str:
    return _DIGIT_RX.sub("", str(s or ""))

def _find_accountbank_or_404(
    accountbank_id=None,
    client_id=None,        # numeric Client.id OR left empty
    client_code=None,      # business code like ALTPS2328D OR left empty
    account_no=None,
):
    qs = AccountBank.objects.using("default").all()

    # 1) Direct by PK
    if accountbank_id:
        try:
            return qs.get(id=accountbank_id)
        except AccountBank.DoesNotExist:
            raise Http404("AccountBank not found for id")

    # 2) By client + account number (tolerant)
    if not account_no or not (client_id or client_code):
        raise ValueError("Provide accountbank_id OR (client_id/client_code and account_no)")

    # Build a base queryset limited to this client (by numeric FK OR business code)
    client_filter = Q()
    if client_id and str(client_id).isdigit():
        client_filter |= Q(client_id=int(client_id))
    if client_code:
        client_filter |= Q(client__client_id=str(client_code).strip())

    base = qs.filter(client_filter)
    if not base.exists():
        raise Http404("AccountBank not found for client")

    acct = str(account_no).strip()
    acct_digits = _digits(acct)

    # Try strict equals first
    hit = base.filter(account_no=acct).first()
    if hit:
        return hit

    # i) case-insensitive equals
    hit = base.filter(account_no__iexact=acct).first()
    if hit:
        return hit

    # ii) contains (handles extra spaces/dashes in DB)
    hit = base.filter(account_no__icontains=acct).first()
    if hit:
        return hit

    # iii) endswith (often account numbers are stored with prefixes)
    hit = base.filter(account_no__iendswith=acct).first()
    if hit:
        return hit

    # iv) digits-only comparison among this client's accounts (Python side)
    for a in base.only("id", "account_no", "bank_name"):
        if _digits(a.account_no) == acct_digits:
            return a

    # Not found → include a small sample in logs
    sample = [{"id": a.id, "account_no": a.account_no, "bank": a.bank_name} for a in base[:5]]
    logging.error(
        "[finder] ❌ not found: client_id=%s client_code=%s account_no=%s (digits=%s). "
        "First candidates: %s",
        client_id, client_code, acct, acct_digits, sample
    )
    raise Http404("AccountBank not found for client/account_no")


@csrf_exempt
@require_http_methods(["POST"])
@transaction.atomic(using="default")
def ajax_submit_mapping(request):
    logging.info("[ajax_submit_mapping] 🚚 POST received")
    try:
        body = request.body.decode("utf-8")
        logging.info("[ajax_submit_mapping] Raw body: %s", body)
        payload = json.loads(body)
    except Exception:
        logging.exception("Invalid JSON body")
        return JsonResponse({"ok": False, "error": "Invalid JSON body"}, status=400)

    try:
        acc = _find_accountbank_or_404(
            accountbank_id=payload.get("accountbank_id"),
            client_id=payload.get("client_id"),
            client_code=payload.get("client_code"),
            account_no=payload.get("account_no"),
        )

        busy_code = (payload.get("BusyAccCode") or "").strip()
        if not busy_code:
            return JsonResponse({"ok": False, "error": "BusyAccCode is required"}, status=400)

        fields_to_update = []
        if acc.busyacccode != busy_code:
            acc.busyacccode = busy_code
            fields_to_update.append("busyacccode")

        acc_group = (payload.get("AccGroup") or "").strip()
        if acc_group and acc.account_group != acc_group:
            acc.account_group = acc_group
            fields_to_update.append("account_group")

        busy_name = (payload.get("BusyAccName") or "").strip()
        if busy_name and acc.account != busy_name:
            acc.account = busy_name
            fields_to_update.append("account")

        if not fields_to_update:
            return JsonResponse({"ok": True, "message": "No changes"})

        acc.save(update_fields=fields_to_update)
        logging.info("[ajax_submit_mapping] ✅ Saved %s for id=%s", fields_to_update, acc.id)
        return JsonResponse({"ok": True,
                             "message": f"Updated mapping for AccountBank id={acc.id}",
                             "updated": fields_to_update})

    except Http404 as e:
        # ✅ Return JSON, not an HTML 404 page
        return JsonResponse({"ok": False, "error": str(e)}, status=404)
    except Exception as e:
        logging.exception("ajax_submit_mapping error")
        return JsonResponse({"ok": False, "error": str(e)}, status=500)

# views.py
from django.http import JsonResponse
from .models import Client, AccountBank

def ajax_clear_mapping(request):
    client_id = (request.GET.get("client_id") or "").strip()
    account_no = (request.GET.get("account_no") or "").strip()

    if not client_id or not account_no:
        return JsonResponse({"ok": False, "error": "Missing client_id or account_no"})

    # If a numeric PK was sent, translate it to the business client_id
    if client_id.isdigit():
        try:
            client_id = Client.objects.using("default").get(id=client_id).client_id
        except Client.DoesNotExist:
            return JsonResponse({"ok": False, "error": "Client not found"})

    try:
        ab = AccountBank.objects.using("default").get(
            client__client_id=client_id, account_no=account_no
        )
    except AccountBank.DoesNotExist:
        return JsonResponse({
            "ok": False,
            "error": f"AccountBank not found for client_id={client_id} & account_no={account_no}"
        })

    # clear mapping fields
    ab.busyacccode = ""
    ab.save(update_fields=["busyacccode"])

    return JsonResponse({"ok": True, "message": "Mapping cleared"})



# import mail to data entry
# views.py
import json
import re
import datetime as dt
from django.http import JsonResponse
from django.db.models import Q
from django.utils import timezone

from .models import MailLog, DataEntry, Client, AccountBank
from .utills import (
    get_fiscal_year_from_date,  # returns '2024_25'
    get_db_for_fy,              # '2024_25' -> 'fy_2024_25'
)


import os, re, shutil
from pathlib import Path
from django.conf import settings

def _client_pan(client):
    for f in ("pan", "pan_no", "pan_number", "client_pan"):
        v = getattr(client, f, None)
        if v:
            return re.sub(r"[^A-Z0-9]", "", str(v).upper())
    return "UNKNOWNPAN"

def _client_code(client):
    # Prefer your Busy/ERP code if present; fallback to numeric id
    for f in ("busy_code", "client_code", "code"):
        v = getattr(client, f, None)
        if v not in (None, ""):
            return re.sub(r"[^A-Za-z0-9]", "", str(v))
    return str(getattr(client, "id", "0"))


def _gather_attachment_sources(mail_obj):
    """
    Returns a list of (source_path, filename_guess) from MailLog.attachments JSON.
    Supports dict entries or plain strings.
    """
    out = []
    items = getattr(mail_obj, "attachments", None) or []
    for item in items:
        if isinstance(item, dict):
            src = item.get("path") or item.get("filepath") or ""
            name = item.get("filename") or item.get("name") or ""
            out.append((src, name))
        else:
            out.append((str(item), ""))  # might already be a path
    return out



# --- NEW helper: path builder for your scheme ----------------
def _fy_folder_from_alias_or_date(de_alias: str, rec_date):
    """
    If de_alias looks like 'fy_YYYY_YY', use it (matches your DB alias).
    Otherwise compute from the rec_date via get_fiscal_year_from_date.
    Returns a folder name like 'fy_2025_26'.
    """
    if de_alias and de_alias.startswith("fy_"):
        return de_alias
    lbl = get_fiscal_year_from_date(rec_date)  # 'YYYY_YY'
    return f"fy_{lbl}"


MAIL_ATTACH_ROOT = getattr(settings, "MAIL_ATTACH_ROOT", None)  # e.g. r'Z:\Kameet Soft Project\Mail Attachments'

def _resolve_attachment_source(mail_obj, src: str, fname: str) -> str | None:
    """
    Resolve an attachment file path using MEDIA_ROOT only.
    Tries (in order):
      1) src if absolute & exists
      2) MEDIA_ROOT/src                   (if src is a relative path like 'acstkameet_at_gmail.com/file.pdf')
      3) MEDIA_ROOT/mail_attachments/<mail_id>/<fname>
      4) MEDIA_ROOT/<mailbox-folder>/<fname>
      5) MEDIA_ROOT/<mailbox-folder>/<mail_id>/<fname>
    """
    import os
    # 1) absolute as-is
    if src and os.path.isabs(src) and os.path.exists(src):
        return src

    # 2) MEDIA_ROOT/src (src was stored as a relative path)
    if src:
        cand = os.path.join(settings.MEDIA_ROOT, src)
        if os.path.exists(cand):
            return cand

    # mailbox folder normalized (e.g. 'acstkameet@gmail.com' → 'acstkameet_at_gmail.com')
    mbox = (getattr(mail_obj, "mailbox", "") or "").strip().lower().replace("@", "_at_")
    mail_id = str(getattr(mail_obj, "id", "") or "")
    filename = fname or (os.path.basename(src) if src else "")

    # 3) MEDIA_ROOT/mail_attachments/<mail_id>/<fname>
    if filename:
        cand = os.path.join(settings.MEDIA_ROOT, "mail_attachments", mail_id, filename)
        if os.path.exists(cand):
            return cand

    # 4) MEDIA_ROOT/<mailbox-folder>/<fname>
    if filename and mbox:
        cand = os.path.join(settings.MEDIA_ROOT, mbox, filename)
        if os.path.exists(cand):
            return cand

    # 5) MEDIA_ROOT/<mailbox-folder>/<mail_id>/<fname>
    if filename and mbox and mail_id:
        cand = os.path.join(settings.MEDIA_ROOT, mbox, mail_id, filename)
        if os.path.exists(cand):
            return cand

    return None


def select_best_attachment(attachments):
    """
    Prefer PDF > Excel > Others
    """
    if not attachments:
        return None

    # 1️⃣ Prefer PDF
    for att in attachments:
        if att.get("content_type") == "application/pdf":
            return att

    # 2️⃣ Prefer Excel
    for att in attachments:
        if att.get("content_type") in (
            "application/vnd.ms-excel",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ):
            return att

    # 3️⃣ Fallback → first attachment
    return attachments[0]


def _save_mail_attachments_to_dataentry(de, mail_obj, de_alias, rec_date):
    """
    Save to: MEDIA_ROOT/<fy_db>/<client_id>/<account_id>/<dataentryid>.<ext>
    Also updates de.attach_file to the FIRST saved file (if any).
    """
    from pathlib import Path
    import os, shutil, logging
    log = logging.getLogger(__name__)

    fy_folder   = _fy_folder_from_alias_or_date(de_alias, rec_date)  # e.g. "fy_2025_26"
    client_part = str(getattr(de.client, "client_id", "") or "").strip() or "UNKNOWN-CLIENT"

   
    account_part = "unknown-account"
    try:
        acc_pk = getattr(de, "account_id", None)
        if acc_pk:
            acc = AccountBank.objects.using("default").only("account_id").get(pk=acc_pk)
            account_part = (acc.account_id or "").strip() or account_part
    except Exception:
        pass
    base_rel_dir = os.path.join(fy_folder, client_part, account_part)
    base_abs_dir = os.path.join(settings.MEDIA_ROOT, base_rel_dir)
    os.makedirs(base_abs_dir, exist_ok=True)

    # pairs = _gather_attachment_sources(mail_obj)  # [(src, filename_guess), ...]
    raw_items = list(getattr(mail_obj, "attachments", []) or [])

    best = select_best_attachment(raw_items)

    # Reorder attachments → BEST FIRST
    ordered_items = []
    if best:
        ordered_items.append(best)
        ordered_items.extend([a for a in raw_items if a is not best])
    else:
        ordered_items = raw_items

    pairs = []
    for item in ordered_items:
        if isinstance(item, dict):
            pairs.append((
                item.get("path") or item.get("filepath") or "",
                item.get("filename") or item.get("name") or ""
            ))
        else:
            pairs.append((str(item), ""))

   
    log.info("[IMPORT DEBUG] Mail #%s attachments: %r",
         getattr(mail_obj, "id", None), pairs)


    # saved_relpaths = []
    # attempted = []
    saved_relpaths, attempted = [], []
    seq, first_rel = 0, None

    # seq = 0
    # first_rel = None


    for raw_src, fname in pairs:
        # ✅ resolve an absolute source first
        src = _resolve_attachment_source(mail_obj, raw_src, fname)
        ext = Path(fname).suffix or (Path(src).suffix if src else "") or ".pdf"
        seq += 1
        file_name = f"{de.id}{ext}" if seq == 1 else f"{de.id}-{seq}{ext}"

        relpath = os.path.join(base_rel_dir, file_name)
        abspath = os.path.join(settings.MEDIA_ROOT, relpath)

        exists = bool(src and os.path.exists(src))
        attempted.append({"src": src or raw_src, "exists": exists, "dest": relpath})

        if not exists:
            log.warning("[IMPORT DEBUG] Attachment source missing for MailLog id=%s: %r",
                        getattr(mail_obj, "id", "?"), raw_src)
            continue

        try:
            shutil.copy2(src, abspath)
            rel_fwd = relpath.replace("\\", "/")
            saved_relpaths.append(rel_fwd)
            # if seq == 1:
            #     first_rel = rel_fwd
            if not first_rel:
                first_rel = rel_fwd   # first saved = BEST attachment

            log.info("[IMPORT DEBUG] Copied attachment → %s", rel_fwd)
        except Exception as e:
            log.exception("Failed to copy attachment from %r to %r: %s", src, abspath, e)

    # Update attach_file for FIRST saved file
    if first_rel and hasattr(de, "attach_file"):
        de.attach_file.name = first_rel
        de.save(using=de._state.db, update_fields=["attach_file"])
        log.info("[IMPORT DEBUG] attach_file set → %s (db=%s, de_id=%s)",
                 first_rel, de._state.db, de.id)

    # Optional JSON list of all files
    if saved_relpaths and hasattr(de, "attachments_json"):
        try:
            current = list(getattr(de, "attachments_json") or [])
            current.extend(saved_relpaths)
            de.attachments_json = current
            de.save(using=de._state.db, update_fields=["attachments_json"])
        except Exception:
            log.exception("Failed updating attachments_json for DE#%s", de.id)

    return {
        "base_dir": base_rel_dir.replace("\\", "/"),
        "saved": saved_relpaths,
        "attempted": attempted,
    }

# ----------------- helpers -----------------
def _to_date(s: str | None):
    try:
        return dt.datetime.strptime((s or "").strip(), "%Y-%m-%d").date()
    except Exception:
        return None

def _local_date(dt_utc):
    try:
        return timezone.localtime(dt_utc).date()
    except Exception:
        return dt_utc.date() if hasattr(dt_utc, "date") else None

def _fy_labels_overlapping(start_date: dt.date, end_date: dt.date):
    """Return FY labels like '2025_26' for every month crossing the window."""
    if not start_date or not end_date:
        return []
    labels, seen = [], set()
    cur = start_date.replace(day=1)
    while cur <= end_date:
        lbl = get_fiscal_year_from_date(cur)  # 'YYYY_YY'
        if lbl not in seen:
            labels.append(lbl)
            seen.add(lbl)
        # next month
        y = cur.year + (1 if cur.month == 12 else 0)
        m = 1 if cur.month == 12 else cur.month + 1
        cur = cur.replace(year=y, month=m, day=1)
    return labels

def _maillog_aliases_for_window(start_date: dt.date, end_date: dt.date):
    """Map window to MailLog DB aliases (fy_YYYY_YY)."""
    return [get_db_for_fy(lbl.replace("_", "-")) for lbl in _fy_labels_overlapping(start_date, end_date)]

def _parse_recipients(raw):
    try:
        from email.utils import getaddresses
        addrs = []
        for _, addr in getaddresses([raw or ""]):
            if addr:
                addrs.append(addr.strip().lower())
        # uniq preserve order
        seen, out = set(), []
        for a in addrs:
            if a not in seen:
                out.append(a); seen.add(a)
        return out
    except Exception:
        return []

def _safe(obj, name, default=None):
    try:
        v = getattr(obj, name)
        return v if v is not None else default
    except Exception:
        return default


def _mail_body_text(mail_obj) -> str:
    # Try common field names without crashing
    for field in ("body_text", "body", "text", "content", "description", "mail_text"):
        v = getattr(mail_obj, field, None)
        if v:
            return str(v)
    return ""



# -------- improved account patterns --------
# Full 9–18 continuous digits (unchanged)
FULL_ACC_RX = re.compile(r"\b(\d{9,18})\b")

# Masked last-4 when Xs precede the digits: e.g. XXXXXXXX8217 / XXXX 8217 / ****8217 / ###8217
# Accept X, x, *, #, sometimes separated by spaces or hyphens
MASK_LAST4_AFTER_XS_RX = re.compile(r"(?:[Xx\*#][\s\-]*){3,}(\d{4})\b")

# Masked last-4 when digits precede Xs (rare but requested): e.g. 8217XXXXXX / 8217 ****
MASK_LAST4_BEFORE_XS_RX = re.compile(r"\b(\d{3,4})(?:[\s\-]*[Xx\*#]){3,}")

# Subject-tail “XXX 8217” pattern (keep but tighten)
SUBJ_MASK_TAIL_RX = re.compile(r"(?:[Xx\*#]{3,}\s*)(\d{3,4})\b")

# Kotak-style: "A/c X0831", "A/C X 0831"
SUBJ_AC_X_TAIL_RX = re.compile(r"\bA/?c\.?\s*[Xx]\s*([0-9]{3,4})\b")

# Cautious single-X prefix, subject-only token: " X0831", "(X 0831)"
# Word boundary + optional space after X + 3-4 digits
SUBJ_SINGLE_X_TAIL_RX = re.compile(r"\b[Xx]\s*([0-9]{3,4})\b")

LEADING4_FROM_FILENAME_RX = re.compile(r"^(\d{4})")  # first 4 chars if digits


def _attachments_text_plus_names(attachments_json):
    parts = []
    for item in (attachments_json or []):
        if isinstance(item, dict):
            for k in ("filename", "name", "path", "filepath"):
                v = item.get(k)
                if v: parts.append(str(v))
        else:
            parts.append(str(item))
    return " ".join([p for p in parts if p]).strip()

def _attachment_text(attachments_json):
    parts = []
    for item in (attachments_json or []):
        if isinstance(item, dict):
            parts.append(str(item.get("filename") or item.get("name") or ""))
            parts.append(str(item.get("path") or item.get("filepath") or ""))
        else:
            parts.append(str(item))
    return " ".join([p for p in parts if p]).strip()

def _clients_for_email(addr_lower: str):
    """Return ALL Clients referencing this email (via AccountBank.acc_mail_id, Client email fields, extra_emails)."""
    if not addr_lower:
        return []
    raw_ids = set()

    # 1) AccountBank.acc_mail_id -> client_id
    try:
        raw_ids.update(
            AccountBank.objects.using("default")
            .filter(acc_mail_id__iexact=addr_lower, client__isnull=False)
            .values_list("client_id", flat=True)
        )
    except Exception:
        pass

    # 2) Client email fields + extra_emails
    email_fields = [f for f in ("email","alt_email","acc_mail_id","contact_email","primary_email") if hasattr(Client, f)]
    q = Q()
    for f in email_fields:
        q |= Q(**{f"{f}__iexact": addr_lower})
    if hasattr(Client, "extra_emails"):
        q |= Q(extra_emails__contains=[addr_lower])

    try:
        if q:
            raw_ids.update(
                Client.objects.using("default")
                .filter(q)
                .values_list("id", flat=True)
            )
        # Fallback if DB can't JSON-contains
        if hasattr(Client, "extra_emails") and not raw_ids:
            for cand in Client.objects.using("default").only("id","extra_emails")[:5000]:
                extras = cand.extra_emails or []
                if any((e or "").lower() == addr_lower for e in extras):
                    raw_ids.add(cand.id)
    except Exception:
        pass

    ids = []
    for x in raw_ids:
        try: ids.append(int(x))
        except: pass

    if not ids:
        return []

    return list(
        Client.objects.using("default")
        .filter(id__in=ids)
        .order_by("id")
    )

def _subject_tail_candidates_text(subj: str) -> set[str]:
    return set(SUBJ_MASK_TAIL_RX.findall(subj or ""))


def _leading4_from_any_filename(attach_text: str) -> set[str]:
    """
    From attachment name/path blob, return 4-digit prefixes that appear
    at the *very beginning* of a filename (before any non-digit).
    Filters out obvious years like 2019..2099 to avoid confusion.
    """
    out = set()
    for token in re.split(r"[,\s]+", attach_text or ""):
        name = os.path.basename(token)
        m = LEADING4_FROM_FILENAME_RX.match(name)
        if not m:
            continue
        first4 = m.group(1)
        # Avoid common '20YY' year prefixes
        if re.fullmatch(r"20\d\d", first4):
            continue
        out.add(first4)
    return out


def _digits(s: str) -> str:
    return re.sub(r"\D+", "", str(s or ""))


def _account_for_client_by_attachment(client_obj, email_addr: str, attach_text: str, subject_text=""):
    if not client_obj:
        return None

    base_qs = AccountBank.objects.using("default").filter(client=client_obj)

    if email_addr:
        qs_email = base_qs.filter(acc_mail_id__iexact=email_addr)
        if qs_email.exists():
            base_qs = qs_email

    blob = f"{attach_text or ''}\n{subject_text or ''}"
    blob_clean = re.sub(r"[^A-Za-z0-9\s*#Xx\-]", " ", blob)

    # 1) full 9–18 digits
    for num in set(FULL_ACC_RX.findall(blob_clean)):
        hit = base_qs.filter(account_no__icontains=num).values("id","bank_name","account_no").first()
        if hit:
            return hit

    # 2) masked tails
    tails = set(MASK_LAST4_AFTER_XS_RX.findall(blob_clean))
    tails.update(MASK_LAST4_BEFORE_XS_RX.findall(blob_clean))
    subj = subject_text or ""
    tails.update(SUBJ_MASK_TAIL_RX.findall(subj))
    tails.update(SUBJ_AC_X_TAIL_RX.findall(subj))
    tails.update(SUBJ_SINGLE_X_TAIL_RX.findall(subj))
    tails = {t[-4:] for t in tails if t and t[-4:].isdigit()}

    if tails:
        # prefetch once and normalize for Python-side compare
        candidates = list(base_qs.only("id","bank_name","account_no"))
        norm = [(acc, _digits(acc.account_no)) for acc in candidates]

        # 3) unique last-4 on digits-only
        for l4 in tails:
            matches = [acc for acc, d in norm if d.endswith(l4)]
            if len(matches) == 1:
                acc = matches[0]
                return {"id": acc.id, "bank_name": acc.bank_name, "account_no": acc.account_no}

        # 4) tie-break with bank-name hint if multiple
        hint = (subject_text + " " + attach_text).lower()
        for l4 in tails:
            matches = [acc for acc, d in norm if d.endswith(l4)]
            if len(matches) > 1:
                for acc in matches:
                    bank_words = (acc.bank_name or "").lower().split()
                    if any(w and w in hint for w in bank_words):
                        return {"id": acc.id, "bank_name": acc.bank_name, "account_no": acc.account_no}

    # --- 3) NEW: filename-leading-4 → match FIRST-4 of account number ---
    # Example: "0667202510030500535323.pdf" → use "0667"
    # We compare against digits-only account_no and require a UNIQUE match.
    leading4s = _leading4_from_any_filename(attach_text)
    if leading4s:
        candidates = list(base_qs.only("id","bank_name","account_no"))
        norm = [(acc, _digits(acc.account_no)) for acc in candidates]

        for f4 in leading4s:
            # unique match on *prefix*
            matches = [acc for acc, d in norm if d.startswith(f4)]
            if len(matches) == 1:
                acc = matches[0]
                return {"id": acc.id, "bank_name": acc.bank_name, "account_no": acc.account_no}

        # (Optional) tie-break by bank-name hint if multiple prefix matches
        hint = (subject_text + " " + attach_text).lower()
        for f4 in leading4s:
            matches = [acc for acc, d in norm if d.startswith(f4)]
            if len(matches) > 1:
                for acc in matches:
                    bank_words = (acc.bank_name or "").lower().split()
                    if any(w and w in hint for w in bank_words):
                        return {"id": acc.id, "bank_name": acc.bank_name, "account_no": acc.account_no}

    
    
    return None


# ----------------- MAIN VIEW (minimal import) -----------------
# views.py
import json, re, datetime as dt, logging
from django.http import JsonResponse
from django.db.models import Q
from django.utils import timezone

from .models import MailLog, DataEntry, Client, AccountBank
from .utills import get_fiscal_year_from_date, get_db_for_fy

log = logging.getLogger(__name__)  # e.g. 'accounts.views'
# helpers (put near the other helpers)
def _received_by_from_mailbox(mailbox: str) -> str:
    """
    Map the MailLog.mailbox (the account you fetched from) to your
    DataEntry.RECEIVED_BY_CHOICES value.
    """
    m = (mailbox or "").strip().lower()

    # Explicit mappings first
    if "acstkameet" in m:     # e.g. acstkameet@gmail.com
        return "Acstkameet"
    # Default fallback
    return "Mail"

def _pick_account_for_client_fallback(client_obj, recips, attach_text, subject_text, sender_text=""):
    """
    When attachment/subject matching didn't return a unique account, try:
      1) accounts with acc_mail_id == any recipient (unique);
      2) if the client has EXACTLY one account, use it;
      3) bank-name heuristic from sender/subject;
      4) last-4 unique match across client accounts from attachment/subject.
    Returns an AccountBank instance or None.
    """
    base_qs = AccountBank.objects.using("default").filter(client=client_obj)

    # 1) email-linked accounts (unique)
    for r in (recips or []):
        qs = base_qs.filter(acc_mail_id__iexact=r)
        if qs.count() == 1:
            return qs.first()

    # 2) single-account client
    if base_qs.count() == 1:
        return base_qs.first()

    # 3) bank name heuristic
    bank_hint = f"{sender_text or ''} {subject_text or ''}".lower()
    BANK_KEYS = [
        ("hdfc", "HDFC"),
        ("icici", "ICICI"),
        ("sbi", "SBI"),
        ("state bank", "SBI"),
        ("kotak", "KOTAK"),
        ("axis", "AXIS"),
        ("bank of maharashtra", "MAHARASHTRA"),
        ("canara", "CANARA"),
        ("pnb", "PUNJAB NATIONAL"),
        ("punjab national", "PUNJAB NATIONAL"),
        ("surat national", "SURAT NATIONAL"),
        ("spcb", "SPCB"),
        ("sutex", "SUTEX"),
    ]
    for key, token in BANK_KEYS:
        if key in bank_hint:
            qs = base_qs.filter(bank_name__icontains=token)
            if qs.count() == 1:
                return qs.first()

    # 4) last-4 unique across all accounts (attachments/subject)
    tails = set()
    tails.update(MASK_LAST4_AFTER_XS_RX.findall(subject_text or ""))
    tails.update(MASK_LAST4_BEFORE_XS_RX.findall(subject_text or ""))
    tails.update(SUBJ_MASK_TAIL_RX.findall(subject_text or ""))
    tails.update(MASK_LAST4_AFTER_XS_RX.findall(attach_text or ""))
    tails.update(MASK_LAST4_BEFORE_XS_RX.findall(attach_text or ""))

    tails = {t[-4:] for t in tails if str(t)[-4:].isdigit()}
    for l4 in tails:
        # if len(l4) < 3:
        #     continue
        qs = base_qs.filter(account_no__iendswith=l4)
        if qs.count() == 1:
            return qs.first()

    return None


# from date  recived date helper
import datetime as dt
import re

# Accept many date formats like: 22-Sep-2025, 01.09.2025, 30/09/25, 1 Sep 2025, etc.
_DATE_FORMATS = [
    "%d-%b-%Y", "%d-%B-%Y", "%d-%b-%y", "%d-%B-%y",
    "%d-%m-%Y", "%d/%m/%Y", "%d.%m.%Y",
    "%d-%m-%y", "%d/%m/%y", "%d.%m.%y",
    "%d %b %Y", "%d %B %Y", "%d %b %y", "%d %B %y",
]

def _parse_any_date(s: str) -> dt.date | None:
    s = (s or "").strip().replace("–", "-").replace("—", "-")
    # Normalize multiple spaces
    s = re.sub(r"\s+", " ", s)
    # Make month names Title Case for %b/%B
    s = re.sub(r"\b([A-Za-z]{3,})\b", lambda m: m.group(1).title(), s)
    for fmt in _DATE_FORMATS:
        try:
            return dt.datetime.strptime(s, fmt).date()
        except Exception:
            continue
    return None

_PERIOD_RXES = [
    re.compile(
        r"(?:for\s+the\s+)?period\s+"
        r"([0-9]{1,2}[./\-\s][A-Za-z0-9]{3,9}[./\-\s][0-9]{2,4})\s*"
        r"(?:to|till|till date|upto|up to|[-–—])\s*"
        r"([0-9]{1,2}[./\-\s][A-Za-z0-9]{3,9}[./\-\s][0-9]{2,4})",
        re.I),
    re.compile(
        r"\bfrom\s+([0-9]{1,2}[./\-\s][A-Za-z0-9]{3,9}[./\-\s][0-9]{2,4})\s*"
        r"(?:to|till|till date|upto|up to|[-–—])\s*"
        r"([0-9]{1,2}[./\-\s][A-Za-z0-9]{3,9}[./\-\s][0-9]{2,4})",
        re.I),
]

def _first_day_prev_month(d: dt.date) -> dt.date:
    first_this = d.replace(day=1)
    prev_last = first_this - dt.timedelta(days=1)
    return prev_last.replace(day=1)

def _extract_period_window(subject_text: str, body_text: str, rec_date: dt.date) -> tuple[dt.date, dt.date]:
    """
    Try to read 'period' from subject/body.
    Fallback: (first day of previous month, rec_date - 1 day)
    """
    blob = f"{subject_text or ''}\n{body_text or ''}"
    for rx in _PERIOD_RXES:
        m = rx.search(blob)
        if m:
            d1 = _parse_any_date(m.group(1))
            d2 = _parse_any_date(m.group(2))
            if d1 and d2:
                # order them if reversed
                if d1 > d2:
                    d1, d2 = d2, d1
                return d1, d2

    # Fallback rule
    # return _first_day_prev_month(rec_date), (rec_date - dt.timedelta(days=1))
    # Fallback → full previous month (1st..last)
    prev_first = _first_day_prev_month(rec_date)
    prev_last  = _last_day_prev_month(rec_date)
    return prev_first, prev_last

def _last_day_prev_month(d: dt.date) -> dt.date:
    first_this = d.replace(day=1)
    return first_this - dt.timedelta(days=1)


# ---- Message-ID helpers ----
RFC_MSGID_RX = re.compile(r"<[^<>]+>")  # minimal RFC-ish check

def _get_message_id(mail_obj) -> str | None:
    """
    Return the verbatim RFC822 Message-ID with angle brackets if available.
    Tries attributes: msg_id/message_id, then headers/raw_headers. Never strips brackets.
    """
    # 1) direct attributes on the model
    for field in ("msg_id", "message_id", "messageid", "MessageID"):
        v = getattr(mail_obj, field, None)
        if v:
            v = str(v).strip()
            if v.startswith("<") and v.endswith(">"):
                return v
            m = RFC_MSGID_RX.search(v)
            return m.group(0) if m else f"<{v}>"

    # 2) headers dict (if your MailLog stores headers)
    try:
        headers = getattr(mail_obj, "headers", None) or {}
        for hdr in ("Message-ID", "Message-Id", "MessageID"):
            v = headers.get(hdr)
            if v:
                v = str(v).strip()
                if v.startswith("<") and v.endswith(">"):
                    return v
                m = RFC_MSGID_RX.search(v)
                return m.group(0) if m else f"<{v}>"
    except Exception:
        pass

    # 3) raw headers text (optional)
    raw = getattr(mail_obj, "raw_headers", None)
    if raw:
        m = re.search(r"^Message-[Ii]d:\s*(<[^<>]+>)", str(raw), re.M)
        if m:
            return m.group(1).strip()

    return None

def _is_probably_valid_msgid(s: str | None) -> bool:
    if not s:
        return False
    # very relaxed: looks like <something@something>
    return bool(re.fullmatch(r"<[^<>\s]+@[^<>\s]+>", s))


# def ajax_import_mail_to_dataentry(request):
#     if request.method != "POST":
#         return JsonResponse({"ok": False, "error": "POST required"})

#     try:
#         body = json.loads(request.body.decode("utf-8"))
#     except Exception as e:
#         print("[IMPORT DEBUG] Invalid JSON:", e)
#         log.exception("Invalid JSON for import")
#         return JsonResponse({"ok": False, "error": "Invalid JSON"})

#     mailbox = (body.get("mailbox") or "").strip()
#     mail_rec_from = _to_date(body.get("mail_rec_from"))
#     mail_rec_to   = _to_date(body.get("mail_rec_to"))
#     auto_fy = str(body.get("auto_fy")).lower() in ("1","true","yes","on","")

#     print(f"[IMPORT DEBUG] HIT ajax_import_mail_to_dataentry mailbox={mailbox} mail_rec_from={mail_rec_from} mail_rec_to={mail_rec_to} auto_fy={auto_fy}")
#     log.info("IMPORT start mailbox=%s from=%s to=%s auto_fy=%s", mailbox, mail_rec_from, mail_rec_to, auto_fy)

#     if not mailbox:
#         print("[IMPORT DEBUG] Missing mailbox")
#         return JsonResponse({"ok": False, "error": "mailbox is required"})

#     # default window if missing
#     if not mail_rec_from or not mail_rec_to:
#         today = timezone.localdate()
#         mail_rec_to = mail_rec_to or today
#         mail_rec_from = mail_rec_from or (today - dt.timedelta(days=30))
#         print(f"[IMPORT DEBUG] Defaulted window → {mail_rec_from} to {mail_rec_to}")

#     if mail_rec_from > mail_rec_to:
#         mail_rec_from, mail_rec_to = mail_rec_to, mail_rec_from
#         print(f"[IMPORT DEBUG] Swapped window → {mail_rec_from} to {mail_rec_to}")

#     ml_aliases = _maillog_aliases_for_window(mail_rec_from, mail_rec_to)
#     print("[IMPORT DEBUG] MailLog DBs to scan:", ml_aliases)
#     log.info("MailLog aliases: %s", ml_aliases)

#     tz = timezone.get_current_timezone()
#     start_dt = timezone.make_aware(dt.datetime.combine(mail_rec_from, dt.time.min), tz)
#     end_dt_ex = timezone.make_aware(dt.datetime.combine(mail_rec_to + dt.timedelta(days=1), dt.time.min), tz)

#     inserted = 0
#     skipped_no_client = 0
#     scanned = 0
#     errors = 0


#     per_mail = []  # <-- put this before the alias loop


#     for alias in ml_aliases:

#         # per_mail = []  # <-- put this before the alias loop

#         try:
#             mails = (
#                 MailLog.objects.using(alias)
#                 .filter(mailbox__iexact=mailbox, rec_dat__gte=start_dt, rec_dat__lt=end_dt_ex)
#                 .order_by("rec_dat", "id")
#             )
#             count = mails.count()
#             print(f"[IMPORT DEBUG] Query {alias}: found {count} mails")
#             log.info("Alias %s → %s mails", alias, count)
#         except Exception as e:
#             print(f"[IMPORT DEBUG] Skip alias={alias} (error: {e})")
#             log.exception("Skipping alias %s due to error", alias)
#             continue

#         for m in mails:
#             scanned += 1
#             try:
#                 rec_local = _local_date(_safe(m, "rec_dat")) or mail_rec_from
#                 if auto_fy:
#                     fy_lbl = get_fiscal_year_from_date(rec_local)           # 'YYYY_YY'
#                     de_alias = get_db_for_fy(fy_lbl.replace("_", "-"))      # 'fy_YYYY_YY'
#                 else:
#                     de_alias = "default"

#                 # resolve client
#                 recips = _parse_recipients(_safe(m, "receiver_mail", "")) or []
#                 if not recips:
#                     rb = (_safe(m, "mailbox", "") or "").strip().lower()
#                     if rb: recips = [rb]
#                                 # ---------- RESOLVE CLIENT ----------
#                 chosen_client = None

#                 # Try recipients → clients
#                 for raddr in recips:
#                     candidates = _clients_for_email(raddr)
#                     if not candidates:
#                         continue

#                     # Prefer a candidate if we can already uniquely match its account from SUBJECT
#                     subj_txt_try = _safe(m, "subject", "") or ""
#                     attach_txt_try = _attachments_text_plus_names(_safe(m, "attachments", []))

#                     picked = None
#                     for cand in candidates:
#                         hit = _account_for_client_by_attachment(cand, raddr, "", subj_txt_try)
#                         if hit:
#                             picked = cand
#                             break
#                     # else pick the first candidate; we’ll still try to match account below
#                     chosen_client = picked or candidates[0]
#                     break

#                 if not chosen_client:
#                     # As a final fallback, bind this mail to the mailbox owner if you keep such mapping,
#                     # otherwise skip (same behavior you had before).
#                     skipped_no_client += 1
#                     print(f"[IMPORT DEBUG] Skipped mail id={m.id} (no client)")
#                     continue

#                 # ---------- BUILD TEXTS ----------
#                 subject_txt = _safe(m, "subject", "") or ""
#                 attach_text = _attachments_text_plus_names(_safe(m, "attachments", []))
#                 body_txt    = _mail_body_text(m)

#                 # Statement window (prev month fallback)
#                 from_d, last_d = _extract_period_window(subject_txt, body_txt, rec_local)

#                 # ---------- ACCOUNT RESOLUTION IN STRICT ORDER ----------
#                 account_obj = None
#                 raddr = recips[0] if recips else ""

#                 # (1) SUBJECT first
#                 # hit = _account_for_client_by_attachment(chosen_client, raddr, "", subject_txt)
#                 hit = _account_for_client_by_attachment(chosen_client, raddr, attach_text, subject_txt)

#                 if hit:
#                     print(f"[DEBUG] Subject match hit={hit}")
#                     account_obj = (
#                         AccountBank.objects.using("default")
#                         .filter(pk=hit["id"], client=chosen_client)
#                         .first()
#                     )
#                 else:
#                     print(f"[DEBUG] No subject match for subj={subject_txt[:100]!r}")

#                 # if hit:
#                 #     try:
#                 #         account_obj = AccountBank.objects.using("default").get(
#                 #             client=chosen_client, account_no__iexact=hit["account_no"]
#                 #         )
#                 #     except AccountBank.DoesNotExist:
#                 #         account_obj = None

#                 # (2) ATTACHMENT names if not found via subject
#                 # if account_obj is None and attach_text:
#                 #     hit = _account_for_client_by_attachment(chosen_client, raddr, attach_text, "")
#                 #     if hit:
#                 #         try:
#                 #             account_obj = AccountBank.objects.using("default").get(
#                 #                 client=chosen_client, account_no__iexact=hit["account_no"]
#                 #             )
#                 #         except AccountBank.DoesNotExist:
#                 #             account_obj = None

#                 # (3) FALLBACKS: acc_mail_id unique / single account / bank hint / last-4 unique
#                 if account_obj is None:
#                     account_obj = _pick_account_for_client_fallback(
#                         chosen_client,
#                         recips,
#                         attach_text,
#                         subject_txt,
#                         _safe(m, "sender_name", "") or _safe(m, "sender_mail", "") or ""
#                     )

#                 # ---------- LOG ----------
#                 log.info(
#                     "[IMPORT MATCH] mail=%s client=%s → account=%s",
#                     getattr(m, "id", "?"),
#                     getattr(chosen_client, "id", "?"),
#                     getattr(account_obj, "account_no", None) if account_obj else None
#                 )

#                 # ---- Message-ID: always store angle-bracketed RFC form ----

#                 # chosen_acc = None
#                 # # attach_text = _attachment_text(_safe(m, "attachments", []))
#                 # # subject_txt = _safe(m, "subject", "") or ""

#                 # attach_text = _attachments_text_plus_names(_safe(m, "attachments", []))
#                 # subject_txt = _safe(m, "subject", "") or ""

#                 # # NEW: derive statement window from subject/body with fallback
#                 # body_txt = _mail_body_text(m)
#                 # from_d, last_d = _extract_period_window(subject_txt, body_txt, rec_local)
#                 # for raddr in recips:
#                 #     candidates = _clients_for_email(raddr)
#                 #     if not candidates:
#                 #         continue
#                 #     for cand in candidates:
#                 #         # acc_hit = _account_for_client_by_attachment(cand, raddr, attach_text, subject_txt)
#                 #         hit = _account_for_client_by_attachment(cand, raddr, attach_text, subject_txt)
#                 #         # if acc_hit:
#                 #         #     chosen_client = cand
#                 #         #     chosen_acc = acc_hit
#                 #         #     break
#                 #         if hit:
#                 #             chosen_client = cand
#                 #             chosen_acc = hit
#                 #             break
#                 #     if not chosen_client:
#                 #         chosen_client = candidates[0]
#                 #     break

#                 # if not chosen_client:
#                 #     skipped_no_client += 1
#                 #     print(f"[IMPORT DEBUG] Skipped mail id={m.id} (no client)")
#                 #     continue

#                 # account_obj = None
#                 # if account_obj:
#                 #     log.info("[IMPORT] Using account #%s (%s) for client #%s",
#                 #             account_obj.id, account_obj.account_no, chosen_client.id)
#                 # else:
#                 #     log.warning("[IMPORT] No account resolved for client #%s (mail id=%s, recips=%s, subj=%r)",
#                 #                 chosen_client.id, getattr(m, "id", "?"), recips, subject_txt[:140])

#                 # if chosen_acc:
#                 #     try:
                      
#                 #         account_obj = (
#                 #             AccountBank.objects.using("default")
#                 #             .get(client=chosen_client, account_no__iexact=chosen_acc["account_no"])
#                 #         )
#                 #     except AccountBank.DoesNotExist:
#                 #         account_obj = None

                        
#                 # if account_obj is None:
#                 #     # try fallbacks using email/subject/sender/bank hints
#                 #     account_obj = _pick_account_for_client_fallback(
#                 #         chosen_client,
#                 #         recips,
#                 #         attach_text,
#                 #         subject_txt,
#                 #         _safe(m, "sender_name", "") or _safe(m, "sender_mail", "") or ""
#                 #     )
#                 # # log after resolving:
#                 # if account_obj:
#                 #     log.info("[IMPORT] Using account #%s (%s) for client #%s",
#                 #             account_obj.id, account_obj.account_no, chosen_client.id)
#                 # else:
#                 #     log.warning("[IMPORT] No account resolved for client #%s (mail id=%s, recips=%s, subj=%r)",
#                 #                 chosen_client.id, getattr(m, "id", "?"), recips, subject_txt[:140])
                       
               
#                 # ---- Message-ID: always store angle-bracketed RFC form ----
#                 mail_msg_id = _get_message_id(m) or f"<maillog-{_safe(m, 'id')}>"
#                 if not _is_probably_valid_msgid(mail_msg_id):
#                     log.warning("[IMPORT] Suspicious Message-ID for MailLog id=%s: %r",
#                                 getattr(m, "id", "?"), mail_msg_id)

#                 # Build defaults for a fresh DataEntry
#                 defaults = {
#                     "client": chosen_client,
#                     "account_id": (account_obj.id if account_obj else None),
#                     "rec_date": rec_local,
#                     "received_by": _received_by_from_mailbox(mailbox),
#                 }
#                 # Optional fields if present on your model
#                 if any(f.name == "from_date" for f in DataEntry._meta.fields):
#                     defaults["from_date"] = from_d
#                 if any(f.name == "last_date" for f in DataEntry._meta.fields):
#                     defaults["last_date"] = last_d
#                 if hasattr(DataEntry, "_meta") and any(f.name == "group" for f in DataEntry._meta.fields):
#                     defaults["group"] = getattr(chosen_client, "group", None)

#                 # Idempotent insert by msg_id
#                 de, created = DataEntry.objects.using(de_alias).get_or_create(
#                     msg_id=mail_msg_id,
#                     defaults=defaults,
#                 )

#                 if created:
#                     log.info("[IMPORT] Inserted DE#%s in %s for msg_id=%s", de.id, de_alias, mail_msg_id)

#                     # Save attachments only for new rows
#                     try:
#                         save_info = _save_mail_attachments_to_dataentry(de, m, de_alias, rec_local)
#                         per_mail.append({
#                             "mail_id": getattr(m, "id", None),
#                             "de_id": de.id,
#                             "db": de_alias,
#                             "base_dir": save_info.get("base_dir"),
#                             "saved": save_info.get("saved", []),
#                             "attempted": save_info.get("attempted", []),
#                         })
#                     except Exception:
#                         log.exception("Failed saving attachments for DE#%s", de.id)

#                     inserted += 1
#                     print(f"[IMPORT DEBUG] + Inserted DE#{de.id} to {de_alias} (client={chosen_client.id}, rec_date={rec_local})")
#                 else:
#                     log.info("[IMPORT] Duplicate prevented in %s for msg_id=%s", de_alias, mail_msg_id)
#                     # If you still want to record attempts WITHOUT copying files, keep a minimal trace:
#                     per_mail.append({
#                         "mail_id": getattr(m, "id", None),
#                         "de_id": de.id,
#                         "db": de_alias,
#                         "base_dir": None,
#                         "saved": [],
#                         "attempted": [],
#                     })


#             except Exception as e:
#                 errors += 1
#                 log.exception("Error processing mail id=%s", getattr(m, "id", "?"))
#                 print(f"[IMPORT DEBUG] ERROR on mail id={getattr(m,'id','?')}: {e}")

#     print(f"[IMPORT DEBUG] DONE: scanned={scanned}, inserted={inserted}, skipped_no_client={skipped_no_client}, errors={errors}")
#     log.info("IMPORT DONE scanned=%s inserted=%s skipped_no_client=%s errors=%s", scanned, inserted, skipped_no_client, errors)

#     return JsonResponse({
#         "ok": True,
#         "message": f"{inserted} DataEntry imported successfully.",
#         "inserted": inserted,
#         "skipped_no_client": skipped_no_client,
#         "errors": errors,
#         "scanned": scanned,
#         "window": {"from": str(mail_rec_from), "to": str(mail_rec_to)},
#         # show first 10 mails’ file results
#         "saved_samples": per_mail[:10],
#         # quick total counts to display
#         "total_saved_files": sum(len(row.get("saved", [])) for row in per_mail),
#     })

def _client_from_body(m):
    """
    Try to detect the Client by scanning the mail body text
    for known client names or legal names.
    Returns a Client instance or None.
    """
    import re

    # Ensure _mail_body_text and _CLIENT_ALIASES exist in this file
    body = _mail_body_text(m)
    if not body:
        return None

    # Build tight lowercase version for fallback search
    body_tight = re.sub(r"[^a-z0-9]+", "", body.lower())

    # Lazy global cache (so we don’t query DB every mail)
    global _CLIENT_ALIASES
    if "_CLIENT_ALIASES" not in globals():
        _CLIENT_ALIASES = {}
        norm = lambda s: re.sub(r"[^a-z0-9 ]+", " ", (s or "").lower()).strip()
        qs = Client.objects.using("default").only("id", "client_name", "legal_name", "trade_name")
        for c in qs:
            aliases = set()
            for f in ("client_name", "legal_name", "trade_name"):
                v = getattr(c, f, None)
                if v:
                    a = norm(v)
                    if a:
                        aliases.add(a)
                        aliases.add(a.replace(" ", ""))  # tight variant
            if aliases:
                _CLIENT_ALIASES[c.id] = {"aliases": aliases, "client": c}

    # Scan body for any alias
    best = 0
    hit = None
    for cid, item in _CLIENT_ALIASES.items():
        for alias in item["aliases"]:
            if not alias or len(alias) < 4:
                continue
            score = 0
            if re.search(rf"\b{re.escape(alias)}\b", body):
                score = len(alias)
            elif alias in body_tight:
                score = len(alias) - 1
            if score > best:
                best = score
                hit = item["client"]

    # Require reasonable match length to avoid false positives
    return hit if best >= 6 else None



def ajax_import_mail_to_dataentry(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "POST required"})

    try:
        body = json.loads(request.body.decode("utf-8"))
    except Exception as e:
        print("[IMPORT DEBUG] Invalid JSON:", e)
        log.exception("Invalid JSON for import")
        return JsonResponse({"ok": False, "error": "Invalid JSON"})

    mailbox       = (body.get("mailbox") or "").strip()
    mail_rec_from = _to_date(body.get("mail_rec_from"))
    mail_rec_to   = _to_date(body.get("mail_rec_to"))
    auto_fy       = str(body.get("auto_fy")).lower() in ("1", "true", "yes", "on", "")

    print(f"[IMPORT DEBUG] HIT ajax_import_mail_to_dataentry mailbox={mailbox} mail_rec_from={mail_rec_from} mail_rec_to={mail_rec_to} auto_fy={auto_fy}")
    log.info("IMPORT start mailbox=%s from=%s to=%s auto_fy=%s", mailbox, mail_rec_from, mail_rec_to, auto_fy)

    if not mailbox:
        print("[IMPORT DEBUG] Missing mailbox")
        return JsonResponse({"ok": False, "error": "mailbox is required"})

    # default window if missing
    if not mail_rec_from or not mail_rec_to:
        today = timezone.localdate()
        mail_rec_to   = mail_rec_to or today
        mail_rec_from = mail_rec_from or (today - dt.timedelta(days=30))
        print(f"[IMPORT DEBUG] Defaulted window → {mail_rec_from} to {mail_rec_to}")

    if mail_rec_from > mail_rec_to:
        mail_rec_from, mail_rec_to = mail_rec_to, mail_rec_from
        print(f"[IMPORT DEBUG] Swapped window → {mail_rec_from} to {mail_rec_to}")

    ml_aliases = _maillog_aliases_for_window(mail_rec_from, mail_rec_to)
    print("[IMPORT DEBUG] MailLog DBs to scan:", ml_aliases)
    log.info("MailLog aliases: %s", ml_aliases)

    tz        = timezone.get_current_timezone()
    start_dt  = timezone.make_aware(dt.datetime.combine(mail_rec_from, dt.time.min), tz)
    end_dt_ex = timezone.make_aware(dt.datetime.combine(mail_rec_to + dt.timedelta(days=1), dt.time.min), tz)

    inserted = 0
    skipped_no_client = 0
    scanned  = 0
    errors   = 0

    # keep small audit for UI
    per_mail = []

    for alias in ml_aliases:
        try:
            mails = (
                MailLog.objects.using(alias)
                .filter(mailbox__iexact=mailbox, rec_dat__gte=start_dt, rec_dat__lt=end_dt_ex)
                .order_by("rec_dat", "id")
            )
            count = mails.count()
            print(f"[IMPORT DEBUG] Query {alias}: found {count} mails")
            log.info("Alias %s → %s mails", alias, count)
        except Exception as e:
            print(f"[IMPORT DEBUG] Skip alias={alias} (error: {e})")
            log.exception("Skipping alias %s due to error", alias)
            continue

        for m in mails:
            scanned += 1
            try:
                # choose target DB per row
                rec_local = _local_date(_safe(m, "rec_dat")) or mail_rec_from
                if auto_fy:
                    fy_lbl   = get_fiscal_year_from_date(rec_local)           # 'YYYY_YY'
                    de_alias = get_db_for_fy(fy_lbl.replace("_", "-"))        # 'fy_YYYY_YY'
                else:
                    de_alias = "default"

                # ---------- recipients ----------
                recips = _parse_recipients(_safe(m, "receiver_mail", "")) or []
                if not recips:
                    rb = (_safe(m, "mailbox", "") or "").strip().lower()
                    if rb:
                        recips = [rb]

                # ---------- build texts ----------
                subject_txt = _safe(m, "subject", "") or ""
                attach_text = _attachments_text_plus_names(_safe(m, "attachments", []))
                body_txt    = _mail_body_text(m)

                # Statement window (prev-month full fallback inside)
                from_d, last_d = _extract_period_window(subject_txt, body_txt, rec_local)

                # ---------- CLIENT RESOLUTION (same order as mail_log_list) ----------
                chosen_client   = None
                chosen_acc_hit  = None  # dict with at least id or account_no

                for raddr in recips:
                    candidates = _clients_for_email(raddr)
                    if not candidates:
                        continue

                    # Prefer a candidate that already matches via subject/attachment
                    picked = None
                    for cand in candidates:
                        hit = _account_for_client_by_attachment(
                            cand, raddr, attach_text, subject_text=subject_txt
                        )
                        if hit:
                            picked = cand
                            chosen_acc_hit = hit
                            break

                    # else pick first candidate; account will be resolved below
                    chosen_client = picked or candidates[0]
                    break

                # Optional tiebreaker: detect client name from body if still none
                if not chosen_client:
                    try:
                        body_client = _client_from_body(m)  # your helper from mail_log_list
                    except Exception:
                        body_client = None
                    if body_client:
                        chosen_client = body_client
                        chosen_acc_hit = _account_for_client_by_attachment(
                            chosen_client, (recips[0] if recips else ""), attach_text, subject_text=subject_txt
                        )

                if not chosen_client:
                    skipped_no_client += 1
                    print(f"[IMPORT DEBUG] Skipped mail id={getattr(m,'id','?')} (no client)")
                    continue

                # ---------- ACCOUNT RESOLUTION (same order as mail_log_list) ----------
                account_obj = None
                raddr = recips[0] if recips else ""

                # 1) Subject+Attachment: full/ masked-tail / leading-4 (unique)
                if not chosen_acc_hit:
                    chosen_acc_hit = _account_for_client_by_attachment(
                        chosen_client, raddr, attach_text, subject_text=subject_txt
                    )

                if chosen_acc_hit:
                    # Prefer primary-key when _account_for_client_by_attachment returns id
                    account_obj = (
                        AccountBank.objects.using("default")
                        .filter(pk=chosen_acc_hit.get("id"), client=chosen_client)
                        .first()
                    )
                    if not account_obj and chosen_acc_hit.get("account_no"):
                        account_obj = (
                            AccountBank.objects.using("default")
                            .filter(client=chosen_client, account_no__iexact=chosen_acc_hit["account_no"])
                            .first()
                        )

                # 2) Fallbacks: acc_mail_id unique / single account / bank hint / last-4 unique
                if account_obj is None:
                    account_obj = _pick_account_for_client_fallback(
                        chosen_client,
                        recips,
                        attach_text,
                        subject_txt,
                        _safe(m, "sender_name", "") or _safe(m, "sender_mail", "") or ""
                    )

                log.info(
                    "[IMPORT MATCH] mail=%s client=%s → account=%s",
                    getattr(m, "id", "?"),
                    getattr(chosen_client, "id", "?"),
                    getattr(account_obj, "account_no", None) if account_obj else None
                )

                # ---------- Message-ID idempotency ----------
                mail_msg_id = _get_message_id(m) or f"<maillog-{_safe(m, 'id')}>"
                if not _is_probably_valid_msgid(mail_msg_id):
                    log.warning("[IMPORT] Suspicious Message-ID for MailLog id=%s: %r",
                                getattr(m, "id", "?"), mail_msg_id)

                # ---------- build DataEntry defaults ----------
                defaults = {
                    "client": chosen_client,
                    "account_id": (account_obj.id if account_obj else None),
                    "rec_date": rec_local,
                    "received_by": _received_by_from_mailbox(mailbox),
                }
                # Optional fields
                if any(f.name == "from_date" for f in DataEntry._meta.fields):
                    defaults["from_date"] = from_d
                if any(f.name == "last_date" for f in DataEntry._meta.fields):
                    defaults["last_date"] = last_d
                if hasattr(DataEntry, "_meta") and any(f.name == "group" for f in DataEntry._meta.fields):
                    defaults["group"] = getattr(chosen_client, "group", None)

                # ---------- insert (idempotent on msg_id) ----------
                de, created = DataEntry.objects.using(de_alias).get_or_create(
                    msg_id=mail_msg_id,
                    defaults=defaults,
                )

                if created:
                    log.info("[IMPORT] Inserted DE#%s in %s for msg_id=%s", de.id, de_alias, mail_msg_id)
                    try:
                        save_info = _save_mail_attachments_to_dataentry(de, m, de_alias, rec_local)
                        per_mail.append({
                            "mail_id": getattr(m, "id", None),
                            "de_id": de.id,
                            "db": de_alias,
                            "base_dir": save_info.get("base_dir"),
                            "saved": save_info.get("saved", []),
                            "attempted": save_info.get("attempted", []),
                        })
                    except Exception:
                        log.exception("Failed saving attachments for DE#%s", de.id)
                    inserted += 1
                    print(f"[IMPORT DEBUG] + Inserted DE#{de.id} to {de_alias} (client={getattr(chosen_client,'id',None)}, rec_date={rec_local})")
                else:
                    log.info("[IMPORT] Duplicate prevented in %s for msg_id=%s", de_alias, mail_msg_id)
                    per_mail.append({
                        "mail_id": getattr(m, "id", None),
                        "de_id": de.id,
                        "db": de_alias,
                        "base_dir": None,
                        "saved": [],
                        "attempted": [],
                    })

            except Exception as e:
                errors += 1
                log.exception("Error processing mail id=%s", getattr(m, "id", "?"))
                print(f"[IMPORT DEBUG] ERROR on mail id={getattr(m,'id','?')}: {e}")

    print(f"[IMPORT DEBUG] DONE: scanned={scanned}, inserted={inserted}, skipped_no_client={skipped_no_client}, errors={errors}")
    log.info("IMPORT DONE scanned=%s inserted=%s skipped_no_client=%s errors=%s", scanned, inserted, skipped_no_client, errors)

    return JsonResponse({
        "ok": True,
        "message": f"{inserted} DataEntry imported successfully.",
        "inserted": inserted,
        "skipped_no_client": skipped_no_client,
        "errors": errors,
        "scanned": scanned,
        "window": {"from": str(mail_rec_from), "to": str(mail_rec_to)},
        "saved_samples": per_mail[:10],
        "total_saved_files": sum(len(row.get("saved", [])) for row in per_mail),
    })


# unmaooed bank account with busy
# accounts/views.py

# from django.shortcuts import render
# from django.db.models import Q
# from .models import AccountBank

# def unmapped_busy_accounts_report(request):
#     """
#     Show accounts where busyacccode is NULL or empty
#     """

#     unmapped_accounts = (
#         AccountBank.objects.using("default")
#         .filter(
#             Q(busyacccode__isnull=True) |
#             Q(busyacccode__exact="")
#         )
#         .select_related("client")
#         .order_by("client__client_name", "bank_name")
#     )

#     context = {
#         "accounts": unmapped_accounts,
#         "total": unmapped_accounts.count(),
#     }

#     return render(
#         request,
#         "unmapped_busy_accounts.html",
#         context
#     )

from django.shortcuts import render
from django.db.models import Q
from .models import AccountBank

def unmapped_busy_accounts_report(request):
    """
    Show ONLY BANK accounts where BUSY mapping is missing
    """

    unmapped_accounts = (
        AccountBank.objects.using("default")
        .filter(
            # 🔴 Unmapped BUSY
            Q(busyacccode__isnull=True) | Q(busyacccode__exact=""),

            # 🏦 ONLY BANK ACCOUNTS
            Q(account_group__iexact="Bank Accounts")
        )
        .select_related("client")
        .order_by("client__client_name", "bank_name")
    )

    context = {
        "accounts": unmapped_accounts,
        "total": unmapped_accounts.count(),
    }

    return render(
        request,
        "unmapped_busy_accounts.html",
        context
    )


#hard copy recieved report
# from datetime import date
# from django.shortcuts import render
# from django.db.models import Count

# from accounts.models import Client


# def hardcopy_received_report(request):

#     # ----------------------------
#     # 1️⃣ Auto-select current FY
#     # ----------------------------
#     current_fy = get_current_fy()   # e.g. '2025-26'
#     fy = request.GET.get("fy") or current_fy

#     # Convert FY → DB name
#     # '2025-26' → 'fy_2025_26'
#     db_fy = fy.replace("-", "_")
#     db_name = f"fy_{db_fy}"

#     # ----------------------------
#     # 2️⃣ HARD COPY entries
#     # ----------------------------
#     qs = (
#         DataEntry.objects.using(db_name)
#         .filter(format="Hard Copy")
#         .values("client_id")
#         .annotate(total=Count("id"))
#         .order_by("-total")
#     )

#     # ----------------------------
#     # 3️⃣ Resolve Client / Group
#     # ----------------------------
#     client_ids = [r["client_id"] for r in qs]

#     clients = (
#         Client.objects.using("default")
#         .filter(id__in=client_ids)
#         .select_related("group")
#     )

#     client_map = {
#         c.id: {
#             "client": c.client_name,
#             "group": c.group.group_name if c.group else "-"
#         }
#         for c in clients
#     }

#     rows = []
#     grand_total = 0
#     for r in qs:
#         info = client_map.get(r["client_id"], {})
#         rows.append({
#             "group": info.get("group", "-"),
#             "client": info.get("client", f"Client ID {r['client_id']}"),
#             "count": r["total"],
#         })
#         grand_total += r["total"]

#     # ----------------------------
#     # 4️⃣ FY dropdown list
#     # ----------------------------
#     START_YEAR = 2020   # system start FY
#     now = date.today().year + 1

#     fy_list = [
#         f"{y}-{str(y+1)[-2:]}"
#         for y in range(START_YEAR, now)
#     ]

#     return render(
#         request,
#         "hardcopy_received_report.html",
#         {
#             "rows": rows,
#             "fy": fy,
#             "current_fy": current_fy,
#             "fy_list": fy_list,
#             "grand_total": grand_total,
#         }
#     )


# hard copy report
from datetime import date
from django.shortcuts import render
from django.db.models import Count
from accounts.models import Client, AccountBank




def hardcopy_received_report(request):

    exclude_virtual = request.GET.get("exclude_virtual") == "1"

    # ----------------------------
    # FY handling
    # ----------------------------
    current_fy = get_current_fy()
    fy = request.GET.get("fy") or current_fy
    db_name = f"fy_{fy.replace('-', '_')}"

    # ----------------------------
    # MULTI filters (IMPORTANT)
    # ----------------------------
    group_ids  = request.GET.getlist("group")      # ✅ MULTI
    client_ids = request.GET.getlist("client")     # ✅ MULTI
    account_id = request.GET.get("account")        # single (optional)

    # ----------------------------
    # Resolve clients under selected groups
    # ----------------------------
    group_client_ids = None

    if group_ids:
        group_client_ids = list(
            Client.objects.using("default")
            .filter(group_id__in=group_ids)
            .values_list("id", flat=True)
        )

    # ----------------------------
    # Base queryset (Hard Copy only)
    # ----------------------------
    # qs = DataEntry.objects.using(db_name).filter(format="Hard Copy")
    qs = DataEntry.objects.using(db_name).filter(format="Hard Copy")

    if exclude_virtual:
        qs = qs.exclude(virtual_account_type__in=["1", "2"])


    if group_client_ids:
        qs = qs.filter(client_id__in=group_client_ids)

    if client_ids:
        qs = qs.filter(client_id__in=client_ids)

    if account_id:
        qs = qs.filter(account_id=account_id)

    # ----------------------------
    # Aggregate Hard Copy count
    # ----------------------------
    qs = (
        qs.values("client_id", "account_id", "virtual_account_type")
        .annotate(total=Count("id"))
        .order_by("-total")
    )

    # ----------------------------
    # TOTAL FY entries (same filters)
    # ----------------------------
    # total_qs = DataEntry.objects.using(db_name).all()
    total_qs = DataEntry.objects.using(db_name).all()

    if exclude_virtual:
        total_qs = total_qs.exclude(virtual_account_type__in=["1", "2"])


    if group_client_ids:
        total_qs = total_qs.filter(client_id__in=group_client_ids)

    if client_ids:
        total_qs = total_qs.filter(client_id__in=client_ids)

    if account_id:
        total_qs = total_qs.filter(account_id=account_id)

    total_qs = (
        total_qs
        .values("client_id", "account_id", "virtual_account_type")
        .annotate(total_entries=Count("id"))
    )

    total_map = {
        (r["client_id"], r["account_id"], r["virtual_account_type"]): r["total_entries"]
        for r in total_qs
    }

    # ----------------------------
    # Resolve Clients & Accounts
    # ----------------------------
    used_client_ids  = {r["client_id"] for r in qs}
    used_account_ids = {r["account_id"] for r in qs if r["account_id"]}

    clients = (
        Client.objects.using("default")
        .filter(id__in=used_client_ids)
        .select_related("group")
    )

    accounts = (
        AccountBank.objects.using("default")
        .filter(id__in=used_account_ids)
    )

    client_map = {
        c.id: {
            "client": c.client_name,
            "group": c.group.group_name if c.group else "-"
        }
        for c in clients
    }

    account_map = {
        a.id: f"{a.bank_name} - {a.account_no}"
        for a in accounts
    }

    # ----------------------------
    # Build rows
    # ----------------------------
    rows = []
    grand_total = 0
    grand_total_all = 0

    for r in qs:
        cinfo = client_map.get(r["client_id"], {})

        total_entries = total_map.get(
            (r["client_id"], r["account_id"], r["virtual_account_type"]), 0
        )

        # Account name resolution
        if r["account_id"]:
            account_name = account_map.get(r["account_id"], "-")
        else:
            if r["virtual_account_type"] == "1":
                account_name = "Sales (Virtual)"
            elif r["virtual_account_type"] == "2":
                account_name = "Purchase (Virtual)"
            else:
                account_name = "-"

        rows.append({
            "group": cinfo.get("group", "-"),
            "client": cinfo.get("client", "-"),
            "account": account_name,
            "count": r["total"],
            "total_entries": total_entries,
        })

        grand_total += r["total"]
        grand_total_all += total_entries

    # ----------------------------
    # Dropdown data
    # ----------------------------
    groups = (
        Client.objects.using("default")
        .select_related("group")
        .values("group__id", "group__group_name")
        .distinct()
        .order_by("group__group_name")
    )

    clients_dd = Client.objects.using("default")
    if group_ids:
        clients_dd = clients_dd.filter(group_id__in=group_ids)

    clients_dd = clients_dd.order_by("client_name")

    # ----------------------------
    # FY dropdown
    # ----------------------------
    START_YEAR = 2020
    now = date.today().year + 1
    fy_list = [f"{y}-{str(y+1)[-2:]}" for y in range(START_YEAR, now)]

    return render(request, "hardcopy_received_report.html", {
        "rows": rows,
        "grand_total": grand_total,
        "grand_total_all": grand_total_all,

        "fy": fy,
        "fy_list": fy_list,

        "groups": groups,
        "clients_dd": clients_dd,

        # ✅ REQUIRED for checkbox checked state
        "sel_groups": group_ids,
        "sel_clients": client_ids,
        "sel_account": account_id,
        "exclude_virtual": exclude_virtual,
    })




# def hardcopy_received_report(request):

#     # ----------------------------
#     # FY handling
#     # ----------------------------
#     current_fy = get_current_fy()
#     fy = request.GET.get("fy") or current_fy
#     db_name = f"fy_{fy.replace('-', '_')}"

#     # ----------------------------
#     # Filters
#     # ----------------------------
#     # group_id  = request.GET.get("group")
#     # client_id = request.GET.get("client")
#     # account_id = request.GET.get("account")


#     group_ids  = request.GET.getlist("group")     # ✅ MULTI
#     client_ids = request.GET.getlist("client")   # ✅ MULTI
#     account_id = request.GET.get("account")      
#     # ----------------------------
#     # Resolve clients under selected group
#     # ----------------------------
#     # group_client_ids = None

#     # if group_id:
#     #     group_client_ids = list(
#     #         Client.objects.using("default")
#     #         .filter(group_id=group_id)
#     #         .values_list("id", flat=True)
#     #     )

#     group_client_ids = None

#     if group_ids:
#         group_client_ids = list(
#             Client.objects.using("default")
#             .filter(group_id__in=group_ids)
#             .values_list("id", flat=True)
#         )



#     # ----------------------------
#     # Base queryset (FY DB)
#     # ----------------------------
#     qs = DataEntry.objects.using(db_name).filter(format="Hard Copy")
#     if group_client_ids:
#         qs = qs.filter(client_id__in=group_client_ids)

#     if client_ids:
#         qs = qs.filter(client_id__in=client_ids)


#     # 🔴 IMPORTANT FIX
#     # if group_client_ids is not None:
#     #     qs = qs.filter(client_id__in=group_client_ids)

#     # if client_id:
#     #     qs = qs.filter(client_id=client_id)

#     if account_id:
#         qs = qs.filter(account_id=account_id)

#     # ----------------------------
#     # Aggregate by client + account
#     # ----------------------------
#     qs = (
#         qs.values("client_id", "account_id","virtual_account_type")
#         .annotate(total=Count("id"))
#         .order_by("-total")
#     )
#     # ----------------------------
#     # TOTAL FY entries per client + account
#     # ----------------------------
#     total_qs = DataEntry.objects.using(db_name).all()
#     if group_client_ids:
#         total_qs = total_qs.filter(client_id__in=group_client_ids)

#     if client_ids:
#         total_qs = total_qs.filter(client_id__in=client_ids)

#     # apply SAME filters
#     # if group_client_ids is not None:
#     #     total_qs = total_qs.filter(client_id__in=group_client_ids)

#     # if client_id:
#     #     total_qs = total_qs.filter(client_id=client_id)

#     if account_id:
#         total_qs = total_qs.filter(account_id=account_id)

#     total_qs = (
#         total_qs
#         .values("client_id", "account_id", "virtual_account_type")
#         .annotate(total_entries=Count("id"))
#     )

#     # lookup map → {(client_id, account_id): total_entries}
#     total_map = {
#         (r["client_id"], r["account_id"], r["virtual_account_type"]): r["total_entries"]
#         for r in total_qs
#     }

#     # ----------------------------
#     # Resolve Clients & Groups
#     # ----------------------------
#     client_ids = {r["client_id"] for r in qs}
#     account_ids = {r["account_id"] for r in qs}

#     clients = (
#         Client.objects.using("default")
#         .filter(id__in=client_ids)
#         .select_related("group")
#     )

#     accounts = (
#         AccountBank.objects.using("default")
#         .filter(id__in=account_ids)
#     )

#     client_map = {
#         c.id: {
#             "client": c.client_name,
#             "group": c.group.group_name if c.group else "-"
#         }
#         for c in clients
#     }

#     account_map = {
#         a.id: f"{a.bank_name} - {a.account_no}"
#         for a in accounts
#     }

#     # ----------------------------
#     # Build rows
#     # ----------------------------
#     # rows = []
#     # grand_total = 0

#     # for r in qs:
#     #     cinfo = client_map.get(r["client_id"], {})
#     #     rows.append({
#     #         "group": cinfo.get("group", "-"),
#     #         "client": cinfo.get("client", "-"),
#     #         "account": account_map.get(r["account_id"], "-"),
#     #         "count": r["total"]
#     #     })
#     #     grand_total += r["total"]

#     rows = []
#     grand_total = 0
#     grand_total_all = 0   # NEW

#     for r in qs:
#         cinfo = client_map.get(r["client_id"], {})

#         total_entries = total_map.get(
#             (r["client_id"], r["account_id"], r["virtual_account_type"]), 0        
#         )

        
#     # account name resolution
#         if r["account_id"]:
#             account_name = account_map.get(r["account_id"], "-")
#         else:
#             if r["virtual_account_type"] == "1":
#                 account_name = "Sales (Virtual)"
#             elif r["virtual_account_type"] == "2":
#                 account_name = "Purchase (Virtual)"
#             else:
#                 account_name = "-"

#         rows.append({
#             "group": cinfo.get("group", "-"),
#             "client": cinfo.get("client", "-"),
#             "account": account_name,
#             "count": r["total"],                # Hard Copy
#             "total_entries": total_entries      # TOTAL FY
#         })

#         grand_total += r["total"]
#         grand_total_all += total_entries

#     # ----------------------------
#     # Dropdown data
#     # ----------------------------
#     groups = (
#         Client.objects.using("default")
#         .select_related("group")
#         .values("group__id", "group__group_name")
#         .distinct()
#     )

#     # clients_dd = Client.objects.using("default")
#     # if group_id:
#     #     clients_dd = clients_dd.filter(group_id=group_id)
#     clients_dd = Client.objects.using("default")

#     if group_ids:
#         clients_dd = clients_dd.filter(group_id__in=group_ids)

#     accounts_dd = AccountBank.objects.using("default")
#     if client_id:
#         accounts_dd = accounts_dd.filter(client_id=client_id)

#     # ----------------------------
#     # FY dropdown
#     # ----------------------------
#     START_YEAR = 2020
#     now = date.today().year + 1
#     fy_list = [f"{y}-{str(y+1)[-2:]}" for y in range(START_YEAR, now)]

#     return render(request, "hardcopy_received_report.html", {
#         "rows": rows,
#         "grand_total": grand_total,
#         "fy": fy,
#         "fy_list": fy_list,
#          "grand_total_all": grand_total_all, 
#         "groups": groups,
#         "clients_dd": clients_dd,
#         "accounts_dd": accounts_dd,

#         "sel_group": group_id,
#         "sel_client": client_id,
#         "sel_account": account_id,
#     })




# Document specify utility
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from django.views.decorators.http import require_GET, require_POST
from django.db import transaction
from django.db.models import Q, Count
from django.contrib.auth.decorators import login_required
from accounts.utills import get_all_fy_aliases
from .models import DocCategory, DocSubType, ClientDocItem, ClientDocFile, DocPasswordHistory,ClientWiseCategoryName


def fy_from_datestr(from_year: int) -> str:
    # helper: 2024 -> "2024_25"
    return f"{from_year}_{str(from_year+1)[-2:]}"

def doc_dashboard(request):

    client_id = request.GET.get("client_id", "").strip()
    # fy = request.GET.get("fy", "").strip()
    # fy = request.GET.get("fy", "").strip().replace("-", "_")
    raw_fy = request.GET.get("fy", "").strip()
    fy = raw_fy.replace("-", "_")


    category_id = request.GET.get("category_id", "").strip()
    status = request.GET.get("status", "").strip()
    search = request.GET.get("search", "").strip()
    data_entry_status = request.GET.get("data_entry_status", "").strip()

    db_alias = None   # 🔑 IMPORTANT
   
    # if fy:
    #     db_alias = _fy_to_alias(fy)
    # else:
    #     db_alias = "default"   # fallback only
    # qs = (
    #     ClientDocItem.objects
    #     .using(db_alias)
    #     .select_related("client", "category", "subtype", "category_name")
    #     .prefetch_related("files")
    #     .annotate(file_count=Count("files"))
    # )


    items = []

    if fy:
        # ✅ SINGLE FY
        db_alias = _fy_to_alias(fy)
        qs = (
            ClientDocItem.objects
            .using(db_alias)
            .select_related("client", "category", "subtype", "category_name")
            .prefetch_related("files")
            .annotate(file_count=Count("files"))
        )
        items = list(qs)

    else:
        # ✅ ALL FYs
        for fy_db in get_all_fy_aliases():
            qs = (
                ClientDocItem.objects
                .using(fy_db)
                .select_related("client", "category", "subtype", "category_name")
                .prefetch_related("files")
                .annotate(file_count=Count("files"))
            )
            for obj in qs:
                obj._db_alias = fy_db   # ⭐ REQUIRED
                items.append(obj)



    # 🔹 Client filter
    # if client_id:
    #     qs = qs.filter(client_id=client_id)

    # # 🔹 FY filter (STRICT)
    # # if fy:
    # #     qs = qs.filter(financial_year=fy)

    # # 🔹 Category filter
    # if category_id:
    #     qs = qs.filter(category_id=category_id)

    # # 🔹 Status filter
    # if status:
    #     qs = qs.filter(status=status)

    # if data_entry_status:
    #     qs = qs.filter(data_entry_status=data_entry_status)
    # # 🔹 Search filter
    # if search:
    #     qs = qs.filter(
    #         Q(client__client_name__icontains=search) |
    #         Q(client__pan__icontains=search) |
    #         # Q(category__name__icontains=search) |
    #         Q(category__category_type__icontains=search) |
    #         Q(subtype__name__icontains=search) |
    #         Q(event_name__icontains=search)
    #     )

    if client_id:
        items = [i for i in items if str(i.client_id) == client_id]

    if category_id:
        items = [i for i in items if str(i.category_id) == category_id]

    if status:
        items = [i for i in items if i.status == status]

    if data_entry_status:
        items = [i for i in items if i.data_entry_status == data_entry_status]

    if search:
        s = search.lower()
        items = [
            i for i in items
            if s in i.client.client_name.lower()
            or s in (i.client.pan or "").lower()
            or s in (i.category.category_type or "").lower()
            or s in (i.event_name or "").lower()
        ]


    clients = Client.objects.order_by("client_name")
    # categories = DocCategory.objects.filter(is_active=True).order_by("sort_order", "name")
    # categories = DocCategory.objects.filter(is_active=True).order_by("sort_order", "category_type")
    categories = DocCategory.objects.order_by("sort_order", "category_type")

    return render(request, "documents/dashboard.html", {
        "items": sorted(
            items,
            key=lambda x: (
                x.client.client_name,
                x.category.category_type,
                x.financial_year,
                x.event_name or ""
            )
        ),
        "clients": clients,
        "categories": categories,
        "fy": fy,
    })


    # return render(request, "documents/dashboard.html", {
    #     "items": qs.order_by(
    #         "client__client_name",
    #         # "category__name",
    #         "category__category_type",
    #         "financial_year",
    #         "event_name",
    #     ),
    #     "clients": clients,
    #     "categories": categories,
    #     "fy": fy,
    #     "filters": {
    #         "client_id": client_id,
    #         "category_id": category_id,
    #         "status": status,
    #         "search": search,
    #     },
    # })

# live data search
from django.template.loader import render_to_string
from django.http import JsonResponse
from django.http import JsonResponse

# def api_dashboard_rows(request):
#     client_id = request.GET.get("client_id", "")
#     # fy = request.GET.get("fy", "")
#     fy = request.GET.get("fy", "").replace("-", "_")
#     category_id = request.GET.get("category_id", "")
#     status = request.GET.get("status", "")
#     search = request.GET.get("search", "")

#     # qs = ClientDocItem.objects.select_related(
#     #     "client", "category", "subtype"
#     # ).prefetch_related("files")
#     if fy:
#         db_alias = _fy_to_alias(fy)
#     else:
#         db_alias = "default"   # fallback (optional)

#     # qs = (
#     #     ClientDocItem.objects
#     #     .using(db_alias)
#     #     .prefetch_related("files")
#     # )
#     qs = (
#         ClientDocItem.objects
#         .using(db_alias)
#         .select_related("client", "category", "subtype", "category_name")
#         .prefetch_related("files")
#         .annotate(file_count=Count("files"))
#     )



#     if client_id:
#         qs = qs.filter(client_id=client_id)
#     # if fy:
#     #     qs = qs.filter(financial_year=fy)
#     # if fy:
#     #     # qs = qs.filter(doc_kind="ANNUAL", financial_year=fy)
#     #     qs = qs.filter(financial_year=fy)
#     if category_id:
#         qs = qs.filter(category_id=category_id)
#     if status:
#         qs = qs.filter(status=status)
#     if search:
#         qs = qs.filter(
#             Q(client__client_name__icontains=search) |
#             Q(client__pan__icontains=search) |
#             # Q(category__name__icontains=search) |
#             Q(category__category_type__icontains=search) |
#             Q(subtype__name__icontains=search) |
#             Q(event_name__icontains=search)
#         )

#     items = []
#     for it in qs:
#         items.append({
#             "id": it.id, 
#             "client": it.client.client_name,
#             # "pan": it.client.pan,
#              # ✅ ADD THIS (key fix)
#             "category_name": (
#                 it.category_name.name
#                 if hasattr(it, "category_name") and it.category_name
#                 else "-"
#             ),
#             # "category": it.category.name,
#             "category_type": it.category.category_type,
#             "subtype": it.subtype.name if it.subtype else "",
#             "financial_year": it.financial_year,
#             "data_entry_status": it.data_entry_status,
#             "event_name": it.event_name,
#             "status": it.status,
#             # "file_count": it.files.count(),
#             "file_count": it.files.all().using(db_alias).count(),

#         })

#     return JsonResponse({"ok": True, "items": items})


def _fy_to_alias(fy: str) -> str:
    """
    '2024_25' → 'fy_2024_25'
    """
    fy = fy.replace("-", "_").strip()
    return f"fy_{fy}"

from django.http import JsonResponse
from django.db.models import Q, Count

# def api_dashboard_rows(request):
#     client_id = request.GET.get("client_id", "").strip()
#     fy = request.GET.get("fy", "").replace("-", "_").strip()
#     category_id = request.GET.get("category_id", "").strip()
#     status = request.GET.get("status", "").strip()
#     search = request.GET.get("search", "").strip()

#     rows = []

#     # ❌ NEVER USE default DB
#     if not fy:
#         return JsonResponse({"ok": True, "items": []})

#     db_alias = _fy_to_alias(fy)
#     print("🔥 FY:", fy)
#     print("🔥 DB ALIAS:", db_alias)
#     print(
#         "🔥 COUNT:",
#         ClientDocItem.objects.using(db_alias).count()
#     )

#     qs = (
#         ClientDocItem.objects
#         .using(db_alias)
#         .filter(financial_year=fy)   # 🔥 KEY FIX
#         .select_related("client", "category", "subtype", "category_name")
#         .prefetch_related("files")
#     )

#     if client_id:
#         qs = qs.filter(client_id=client_id)

#     if category_id:
#         qs = qs.filter(category_id=category_id)

#     if status:
#         qs = qs.filter(status=status)

#     if search:
#         qs = qs.filter(
#             Q(client__client_name__icontains=search) |
#             Q(client__pan__icontains=search) |
#             Q(category__category_type__icontains=search) |
#             Q(subtype__name__icontains=search) |
#             Q(event_name__icontains=search)
#         )

#     for it in qs:
#         rows.append({
#             "id": it.id,
#             "client": it.client.client_name,
#             "category_type": it.category.category_type,
#             "category_name": it.category_name.name if it.category_name else "-",
#             "subtype": it.subtype.name if it.subtype else "-",
#             "financial_year": it.financial_year,
#             "status": it.status,
#             "data_entry_status": it.data_entry_status,
#             "event_name": it.event_name or "",
#             "file_count": it.files.all().using(db_alias).count(),
#             "fy": it.financial_year,
#         })

#     return JsonResponse({"ok": True, "items": rows})

def api_dashboard_rows(request):
    client_id = request.GET.get("client_id", "").strip()
    fy = request.GET.get("fy", "").replace("-", "_").strip()
    category_id = request.GET.get("category_id", "").strip()
    status = request.GET.get("status", "").strip()
    search = request.GET.get("search", "").strip()

    if not fy:
        return JsonResponse({"ok": True, "items": []})

    db_alias = _fy_to_alias(fy)

    print("🔥 FY:", fy)
    print("🔥 DB ALIAS:", db_alias)
    print("🔥 COUNT:", ClientDocItem.objects.using(db_alias).count())

    qs = (
        ClientDocItem.objects
        .using(db_alias)
        .filter(financial_year=fy)
        .prefetch_related("files")   # ✅ SAFE (same DB)
    )

    if client_id:
        qs = qs.filter(client_id=client_id)

    if category_id:
        qs = qs.filter(category_id=category_id)

    if status:
        qs = qs.filter(status=status)

    if search:
        qs = qs.filter(
            Q(event_name__icontains=search)
        )

    # 🔁 FETCH RELATED DATA MANUALLY (FROM DEFAULT DB)
    from accounts.models import Client, DocCategory, DocSubType, ClientWiseCategoryName

    rows = []

    for it in qs:
        client = Client.objects.filter(id=it.client_id).first()
        category = DocCategory.objects.filter(id=it.category_id).first()
        subtype = DocSubType.objects.filter(id=it.subtype_id).first() if it.subtype_id else None
        cat_name = (
            ClientWiseCategoryName.objects.filter(id=it.category_name_id).first()
            if it.category_name_id else None
        )

        rows.append({
            "id": it.id,
            "client": client.client_name if client else "-",
            "category_type": category.category_type if category else "-",
            "category_name": cat_name.name if cat_name else "-",
            "subtype": subtype.name if subtype else "-",
            "financial_year": it.financial_year,
            "status": it.status,
            "data_entry_status": it.data_entry_status,
            "event_name": it.event_name or "",
            "file_count": it.files.all().using(db_alias).count(),
            "fy": it.financial_year,
        })

    return JsonResponse({"ok": True, "items": rows})


@require_GET
def api_subtypes(request):
    category_id = request.GET.get("category_id")
    if not category_id:
        return JsonResponse({"ok": False, "error": "category_id required"}, status=400)

    subs = DocSubType.objects.filter(category_id=category_id, is_active=True).order_by("name")
    data = [{"id": s.id, "name": s.name} for s in subs]
    return JsonResponse({"ok": True, "subtypes": data})



@require_POST
@transaction.atomic
def api_upload_doc(request):
    print("🔵 api_upload_doc POST DATA:")
    for k, v in request.POST.items():
        print(f"   {k} = {v}")
    item_id = request.POST.get("item_id")
    # old_fy = request.POST.get("old_fy", "").replace("-", "_")
    def normalize_fy(fy):
        return (fy or "").replace("-", "_").strip()
    
    
    client = get_object_or_404(Client, pk=request.POST.get("client_id"))
    category = get_object_or_404(DocCategory, pk=request.POST.get("category_id"))

    subtype_id = request.POST.get("subtype_id") or None
    subtype = get_object_or_404(DocSubType, pk=subtype_id) if subtype_id else None

    # doc_kind = request.POST.get("doc_kind")
    doc_kind = request.POST.get("doc_kind") or request.POST.get("doc_kind_hidden")

    if not doc_kind:
        return JsonResponse({"ok": False, "error": "Document Type is required"})

    # financial_year = request.POST.get("financial_year", "")
    # financial_year = request.POST.get("financial_year", "").replace("-", "_")
    # financial_year = raw_fy.replace("-", "_")
    # raw_fy = request.POST.get("financial_year", "")
    old_fy = normalize_fy(request.POST.get("old_fy"))
    financial_year = normalize_fy(request.POST.get("financial_year"))
    # 🔥 DETERMINE SOURCE DB FOR EDIT
    if item_id and old_fy:
        source_fy = old_fy
    else:
        source_fy = financial_year

    source_db = _fy_to_alias(source_fy)

    
    db_alias = _fy_to_alias(financial_year)


    event_name = request.POST.get("event_name", "")
    remarks = request.POST.get("remarks", "")
    status = request.POST.get("status", "PENDING")
    data_entry_status = request.POST.get("data_entry_status", "PENDING")
    category_name_id = request.POST.get("category_name_id") or None
    upload = request.FILES.get("file")


    # =====================================================
    # 🔒 BACKEND SAFETY: RECEIVED requires attachment
    # =====================================================
    if status == "RECEIVED":

        # ➕ CREATE MODE
        if not item_id and not upload:
            return JsonResponse({
                "ok": False,
                "error": "Cannot mark RECEIVED without attachment"
            })

        # ✏️ EDIT MODE
        if item_id:
            # item = get_object_or_404(ClientDocItem, pk=item_id)
            # db_alias = _fy_to_alias(request.POST.get("financial_year"))
            # financial_year = request.POST.get("financial_year", "").replace("-", "_")


            # item = get_object_or_404(
            #     ClientDocItem.objects.using(db_alias),
            #     pk=item_id
            # )

            item = get_object_or_404(
                ClientDocItem.objects.using(source_db),
                pk=item_id
            )


            if not upload and not item.files.all().using(db_alias).exists():
                return JsonResponse({
                    "ok": False,
                    "error": "Cannot mark RECEIVED without attachment"
                })

    # =====================================================
    # SAVE DATA
    # =====================================================
    if item_id:

        # ===============================
        # 🔁 FY CHANGE → MOVE RECORD
        # ===============================
        print("🟡 EDIT CHECK:",
        "item_id =", item_id,
        "old_fy =", old_fy,
        "financial_year =", financial_year)

        if old_fy and old_fy != financial_year:

            old_db = _fy_to_alias(old_fy)
            new_db = _fy_to_alias(financial_year)

            # 1️⃣ Get old item from OLD DB
            old_item = get_object_or_404(
                ClientDocItem.objects.using(old_db),
                pk=item_id
            )

            # 2️⃣ Create new item in NEW DB
            new_item = ClientDocItem.objects.using(new_db).create(
                client=client,
                category=category,
                subtype=subtype,
                doc_kind=doc_kind,
                financial_year=financial_year,
                event_name=event_name,
                remarks=remarks,
                status=status,
                data_entry_status=data_entry_status,
                category_name_id=category_name_id,
            )
            from django.core.files.base import ContentFile

            # 3️⃣ Move files
            for f in old_item.files.all():
                ClientDocFile.objects.using(new_db).create(
                    doc_item=new_item,
                    # file=f.file
                    file=ContentFile(f.file.read(), name=f.file.name)
                )

            # 4️⃣ Delete old item (AUTO deletes links)
            # old_item.delete(using=db_alias)
            old_item.delete(using=old_db)

            item = new_item
            db_alias = new_db

        # ===============================
        # ✏️ NORMAL EDIT (FY SAME)
        # ===============================
        else:
            # db_alias = _fy_to_alias(financial_year)
            db_alias = source_db
            item = get_object_or_404(
                ClientDocItem.objects.using(db_alias),
                pk=item_id
            )
            print("🟢 BEFORE SAVE:",
                "DB =", db_alias,
                "item.id =", item.id,
                "OLD data_entry_status =", item.data_entry_status)


            item.client = client
            item.category = category
            item.subtype = subtype
            item.doc_kind = doc_kind
            item.financial_year = financial_year
            item.event_name = event_name
            item.remarks = remarks
            item.status = status
            item.data_entry_status = data_entry_status
            item.category_name_id = category_name_id
            item.save()
            print(
                "🧠 FINAL SAVE:",
                "DB =", db_alias,
                "item_id =", item.id,
                "status =", status,
                "data_entry_status =", data_entry_status,
                "remarks =", remarks
            )

            item.refresh_from_db(using=db_alias)
            print("✅ AFTER SAVE:",
                "NEW data_entry_status =", item.data_entry_status)


    else:
        db_alias = _fy_to_alias(financial_year)

        # ➕ CREATE (ONLY ONCE)
        item = ClientDocItem.objects.using(db_alias).create(
            client=client,
            category=category,
            subtype=subtype,
            doc_kind=doc_kind,
            financial_year=financial_year,
            event_name=event_name,
            remarks=remarks,
            status=status,
            data_entry_status=data_entry_status,
            category_name_id=category_name_id,
        )
     
    # =====================================================
    # FILE UPLOAD
    # =====================================================
    if upload:
        ClientDocFile.objects.using(db_alias).create(
            doc_item=item,
            file=upload
        )

    return JsonResponse({"ok": True, "item_id": item.id, "financial_year": item.financial_year})

@require_POST
@transaction.atomic
def api_delete_doc_item(request):
    item_id = request.POST.get("id")
    fy = request.POST.get("fy")

    if not item_id or not fy or fy == "undefined":
        return JsonResponse({"ok": False, "error": "Invalid delete request"}, status=400)

    fy = fy.replace("-", "_")
    db_alias = _fy_to_alias(fy)

    # 🔒 SAFETY: ensure correct FY DB
    item = ClientDocItem.objects.using(db_alias).filter(
        pk=item_id,
        financial_year=fy
    ).first()

    if not item:
        return JsonResponse({
            "ok": False,
            "error": "Item not found in this FY database"
        }, status=404)

    # 🔥 Delete files FIRST
    # for f in item.files.all():
    #     f.file.delete(save=False)
    #     f.delete()
    for f in item.files.all().using(db_alias):
        f.file.delete(save=False)
        f.delete(using=db_alias)


    # 🔥 Delete document
    # item.delete()
    item.delete(using=db_alias)


    return JsonResponse({
        "ok": True,
        "deleted_id": item_id,
        "fy": fy
    })


@require_GET
def api_doc_item_detail(request):
    item_id = request.GET.get("id")
    fy = request.GET.get("fy")   #  send fy from frontend

    # if not fy:
    #     return JsonResponse({"ok": False, "error": "FY missing"}, status=400)
    # 🔒 SAFETY CHECK (ADD THIS BLOCK)
    if not fy or fy == "undefined":
        return JsonResponse({
            "ok": False,
            "error": "Invalid Financial Year"
        }, status=400)
    db_alias = _fy_to_alias(fy.replace("-", "_"))

    item = get_object_or_404(
        ClientDocItem.objects.using(db_alias),
        pk=item_id
    )

    files = [
        {
            "id": f.id,
            "name": f.file.name.split("/")[-1],
            "url": f.file.url,
        }
        for f in item.files.all().using(db_alias)
    ]

    return JsonResponse({
        "ok": True,
        "item": {
            "id": item.id,
            "client_id": item.client_id,
            "category_id": item.category_id,
            "subtype_id": item.subtype_id,
            "doc_kind": item.doc_kind,
            "financial_year": item.financial_year,
            "event_name": item.event_name,
            "status": item.status,
            "data_entry_status": item.data_entry_status,
            "remarks": item.remarks,
            "category_name_id": item.category_name_id,
            "files": files,
        }
    })



# delete existing file in edit mode@require_POST
# @require_POST
# @transaction.atomic
# def api_delete_doc_file(request):
#     file_id = request.POST.get("file_id")
#     if not file_id:
#         return JsonResponse({"ok": False, "error": "file_id required"}, status=400)

#     file = get_object_or_404(ClientDocFile, pk=file_id)

#     # delete physical file
#     file.file.delete(save=False)

#     # delete db row
#     file.delete()

#     return JsonResponse({"ok": True})

@require_POST
@transaction.atomic
def api_delete_doc_file(request):
    file_id = request.POST.get("file_id")
    fy = request.POST.get("fy")
    if not file_id or not fy:
        return JsonResponse({"ok": False, "error": "Invalid request"}, status=400)

    db_alias = _fy_to_alias(fy.replace("-", "_"))

    file = get_object_or_404(
        ClientDocFile.objects.using(db_alias),
        pk=file_id
    )

    file.file.delete(save=False)
    file.delete(using=db_alias)

    return JsonResponse({"ok": True})


def annual_pending_report(request):
    # fy = request.GET.get("fy", "").strip()
    raw_fy = request.GET.get("fy", "").strip()
    fy = raw_fy.replace("-", "_")
    db_alias = _fy_to_alias(fy)

    if not fy:
        return render(request, "documents/report_annual_pending.html", {"rows": [], "fy": ""})

    rows = (ClientDocItem.objects.using(db_alias)
            .select_related("client", "category", "subtype")
            .filter(doc_kind="ANNUAL", financial_year=fy)
            .values(
                "client__client_name",
                "client__pan",
                "category__category_type",
                "subtype__name",
                "status"
            )
            .order_by("client__client_name", "category__category_type"))

            # .values("client__client_name", "client__pan", "category__name", "subtype__name", "status")
            # .order_by("client__client_name", "category__name"))

    return render(request, "documents/report_annual_pending.html", {"rows": rows, "fy": fy})

def event_pending_report(request):
    # fy = request.GET.get("fy", "").strip()
    raw_fy = request.GET.get("fy", "").strip()
    fy = raw_fy.replace("-", "_")
    db_alias = _fy_to_alias(fy)


    if not fy:
        return render(request, "documents/report_event_pending.html", {
            "rows": [],
            "fy": "",
        })

    rows = (
        ClientDocItem.objects.using(db_alias)
        .select_related("client", "category", "subtype")
        .filter(
            doc_kind="EVENT",
            financial_year=fy
        )
        .values(
            "client__client_name",
            "client__pan",
            "category__category_type",
            "subtype__name",
            "event_name",
            "status"
        )
        .order_by(
            "client__client_name",
            "category__category_type",
            "event_name"
        )

        # .values(
        #     "client__client_name",
        #     "client__pan",
        #     "category__name",
        #     "subtype__name",
        #     "event_name",
        #     "status"
        # )
        # .order_by("client__client_name", "category__name", "event_name")
    )

    return render(request, "documents/report_event_pending.html", {
        "rows": rows,
        "fy": fy,
    })

@require_GET
def api_doc_files(request):
    item_id = request.GET.get("item_id")
    fy = request.GET.get("fy")

    
    if not fy or fy == "undefined":
        return JsonResponse({"ok": False, "error": "Invalid FY"}, status=400)

    db_alias = _fy_to_alias(fy.replace("-", "_"))

    item = get_object_or_404(
        ClientDocItem.objects.using(db_alias),
        pk=item_id
    )
    files = [{
        "id": f.id,
        "name": f.file.name.split("/")[-1],
        "uploaded": f.uploaded_at.strftime("%d-%m-%Y"),
        "url": f.file.url
    } for f in item.files.all().using(db_alias)]

    return JsonResponse({"ok": True, "files": files})
    # files = (
    #     ClientDocFile.objects
    #     .filter(doc_item_id=item_id)
    #     .order_by("-uploaded_at")   # correct
    # )

    # data = [{
    #     "name": f.file.name.split("/")[-1],
    #     "url": f.file.url,
    #     "uploaded": f.uploaded_at.strftime("%d-%b-%Y")  # ✅ FIX
    # } for f in files]

    # return JsonResponse({"ok": True, "files": data})


# CATEGORY LIST
def category_list(request):
    categories = DocCategory.objects.all().order_by("sort_order", "category_type")
    return render(request, "documents/category_list.html", {
        "categories": categories
    })


# Add /Update category
@require_POST
@transaction.atomic
def api_save_category(request):
    cat_id = request.POST.get("id")
    category_type = (request.POST.get("category_type") or "").strip()

    if not category_type:
        return JsonResponse({"ok": False, "error": "Category type required"})

    annual = request.POST.get("annual_allowed") == "on"
    event = request.POST.get("event_allowed") == "on"
    cont = request.POST.get("continue_till_closed") == "on"
    active = request.POST.get("is_active") == "on"

    if cat_id:
        cat = get_object_or_404(DocCategory, pk=cat_id)
        cat.category_type = category_type
        cat.annual_allowed = annual
        cat.event_allowed = event
        cat.continue_till_closed = cont
        cat.is_active = active
        cat.save()
    else:
        if DocCategory.objects.filter(category_type__iexact=category_type).exists():
            return JsonResponse({"ok": False, "error": "Category already exists"})

        cat = DocCategory.objects.create(
            category_type=category_type,
            annual_allowed=annual,
            event_allowed=event,
            continue_till_closed=cont,
            # is_active=active,
        )

    return JsonResponse({"ok": True})


# Delete Category
@require_POST
@transaction.atomic
def api_delete_category(request):
    cat_id = request.POST.get("id")
    cat = get_object_or_404(DocCategory, pk=cat_id)

    if cat.clientdocitem_set.exists():
        return JsonResponse({
            "ok": False,
            "error": "Category is used in documents"
        })

    cat.delete()
    return JsonResponse({"ok": True})


# sub category list
def subdoc_list(request):
    category_id = request.GET.get("category_id", "")
    subdocs = DocSubType.objects.select_related("category")

    if category_id:
        subdocs = subdocs.filter(category_id=category_id)

    subdocs = subdocs.order_by("category__category_type", "name")

    # categories = DocCategory.objects.filter(is_active=True).order_by("category_type")
    categories = DocCategory.objects.all().order_by("category_type")

    return render(request, "documents/subdoc_list.html", {
        "subdocs": subdocs,
        "categories": categories,
        "selected_category": category_id
    })


# addd/update sub category
@require_POST
@transaction.atomic
def api_save_subdoc(request):
    sub_id = request.POST.get("id")
    category_id = request.POST.get("category_id")
    name = (request.POST.get("name") or "").strip()
    active = request.POST.get("is_active") == "on"

    if not category_id or not name:
        return JsonResponse({"ok": False, "error": "Category and name required"})

    category = get_object_or_404(DocCategory, pk=category_id)

    if sub_id:
        sub = get_object_or_404(DocSubType, pk=sub_id)
        sub.category = category
        sub.name = name
        sub.is_active = active
        sub.save()
    else:
        if DocSubType.objects.filter(category=category, name__iexact=name).exists():
            return JsonResponse({"ok": False, "error": "Sub-document already exists"})
        sub = DocSubType.objects.create(
            category=category,
            name=name,
            is_active=active,
            is_system=False
        )

    return JsonResponse({"ok": True})


# delete sub category
@require_POST
@transaction.atomic
def api_delete_subdoc(request):
    sub_id = request.POST.get("id")
    sub = get_object_or_404(DocSubType, pk=sub_id)

    if ClientDocItem.objects.filter(subtype=sub).exists():
        return JsonResponse({
            "ok": False,
            "error": "Sub-document is used in documents"
        })

    sub.delete()
    return JsonResponse({"ok": True})


@require_GET
def api_client_category_names(request):
    client_id = request.GET.get("client_id")
    category_id = request.GET.get("category_id")

    if not category_id:
        return JsonResponse({"ok": False, "items": []})

    qs = ClientWiseCategoryName.objects.filter(
        category_id=category_id,
        is_active=True
    )

    if client_id:
        qs = qs.filter(client_id=client_id)

    qs = qs.order_by("name")

    return JsonResponse({
        "ok": True,
        "items": [{"id": o.id, "name": o.name} for o in qs]
    })



@require_POST
@transaction.atomic
def api_create_category_name(request):
    client_id = request.POST.get("client_id")
    category_id = request.POST.get("category_id")
    name = (request.POST.get("name") or "").strip()
    passwords = (request.POST.get("passwords") or "").strip()

    if not (client_id and category_id and name):
        return JsonResponse({"ok": False, "error": "Missing data"})

    obj, created = ClientWiseCategoryName.objects.get_or_create(
        client_id=client_id,
        category_id=category_id,
        name=name,
        defaults={"passwords": passwords}

    )

    # 🔁 If exists & password provided → update
    if not created and passwords:
        obj.passwords = passwords
        obj.save(update_fields=["passwords"])

    return JsonResponse({
        "ok": True,
        "id": obj.id,
        "name": obj.name,
        "passwords": obj.passwords,
        "created": created
    })



# Client wise category list
@require_GET
def api_all_category_names(request):
    qs = (
        ClientWiseCategoryName.objects
        .select_related("client", "category")
        .order_by(
            "client__client_name",
            "category__category_type",
            "name"
        )
    )

    # 🔹 If request is AJAX → return JSON
    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        data = []
        for obj in qs:
            data.append({
                "id": obj.id,
                "client": obj.client.client_name,
                "category_type": obj.category.category_type,
                "name": obj.name,
                "passwords": obj.passwords,
                "is_active": obj.is_active,
            })

        return JsonResponse({"ok": True, "items": data})

    # 🔹 Normal browser request → render HTML page
    return render(
        request,
        "documents/client_category_names.html"
    )

# delete category name
@require_POST
def delete_category_name(request, pk):
    ClientWiseCategoryName.objects.filter(pk=pk).delete()
    return JsonResponse({"ok": True})


# edit category name
@require_POST
@transaction.atomic
def edit_category_name(request, pk):
    obj = get_object_or_404(ClientWiseCategoryName, pk=pk)

    name = (request.POST.get("name") or "").strip()
    passwords = (request.POST.get("passwords") or "").strip()

    if not name:
        return JsonResponse({"ok": False, "error": "Category name is required"})

    # 🔒 Only allowed fields
    obj.name = name
    obj.passwords = passwords
    obj.save(update_fields=["name", "passwords"])

    return JsonResponse({"ok": True})


# upload document from mailLog

from django.core.files import File

@require_POST
def save_mail_docs(request):
    payload = json.loads(request.body)

    mail_id = payload["mail_id"]
    # fy = payload.get("fy")
    rows = payload["rows"]

    # if not fy:
    #     return JsonResponse({"ok": False, "error": "FY missing"}, status=400)

    # db_alias = _fy_to_alias(fy)
    raw_fy = payload.get("fy", "")
    financial_year = raw_fy.replace("-", "_")

    if not financial_year:
        return JsonResponse({"ok": False, "error": "FY missing"}, status=400)

    db_alias = _fy_to_alias(financial_year)


    # try:
    #     db_alias = _fy_to_alias(fy)
    # except Exception:
    #     return JsonResponse({"ok": False, "error": "Invalid FY"}, status=400)

    try:
        mail = MailLog.objects.using(db_alias).get(id=mail_id)
    except MailLog.DoesNotExist:
        return JsonResponse({
            "ok": False,
            "error": "Mail not found in FY DB"
        }, status=404)

    saved = skipped = 0

    for row in rows:

        if row["action"] != "save":
            skipped += 1
            continue

        client_id = row["client"]
        category_id = row["category"]
        subtype_id = row["sub_type"] or None
        doc_name_id = row["doc_name"] or None
        filename = row["filename"]
        remarks = row.get("notes", "")
        password = row.get("password", "")

        # -------- CREATE ClientDocItem ----------
        doc_item = ClientDocItem.objects.using(db_alias).create(
            client_id=client_id,
            category_id=category_id,
            subtype_id=subtype_id,
            category_name_id=doc_name_id,
            doc_kind="ANNUAL",
            # financial_year=getattr(mail, "financial_year", ""),
            financial_year=financial_year,
            remarks=remarks,
            status="RECEIVED",
            data_entry_status="",
        )

        # -------- FIND ATTACHMENT ----------
        att = next(
            (a for a in (mail.attachments or []) if a.get("filename") == filename),
            None
        )

        if not att or not att.get("path"):
            skipped += 1
            continue

        # -------- CREATE ClientDocFile ----------
        abs_path = os.path.join(settings.MEDIA_ROOT, att["path"])

        if not os.path.exists(abs_path):
            skipped += 1
            continue

        with open(abs_path, "rb") as f:
            ClientDocFile.objects.using(db_alias).create(
                doc_item=doc_item,
                file=File(f, name=filename),
                file_password=password,
                password_validity_year=doc_item.financial_year
            )


        saved += 1

    return JsonResponse({
        "ok": True,
        "saved": saved,
        "skipped": skipped
    })


import json
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from django.views.decorators.http import require_GET
from .models import MailLog, Client, DocCategory



@require_GET
def ajax_mail_attachments(request, mail_id):
    fy = request.GET.get("fy")  # <-- IMPORTANT

    if not fy:
        return JsonResponse({
            "ok": False,
            "error": "FY missing"
        }, status=400)

    try:
        db_alias = _fy_to_alias(fy)
    except Exception:
        return JsonResponse({
            "ok": False,
            "error": "Invalid FY"
        }, status=400)

    try:
        mail = MailLog.objects.using(db_alias).get(id=mail_id)
    except MailLog.DoesNotExist:
        return JsonResponse({
            "ok": False,
            "error": "Mail not found in FY DB"
        }, status=404)

    attachments = []
    for a in (mail.attachments or []):
        if isinstance(a, dict):
            attachments.append({
                "filename": a.get("filename", "file.pdf")
            })

    clients_html = "".join(
        f"<option value='{c.id}'>{c.client_name}</option>"
        for c in Client.objects.order_by("client_name")
    )

    # categories = DocCategory.objects.filter(is_active=True).order_by("category_type")
    categories = DocCategory.objects.all().order_by("category_type")

    categories_html = "".join(
        f"<option value='{c.id}'>{c.category_type}</option>"
        for c in categories
    )

    # doc_types_html = "".join(
    #     f"<option value='{d.id}'>{d.name}</option>"
    #     for d in DocCategory.objects.filter(is_active=True)
    # )
    doc_types_html = "".join(
        f"<option value='{d.id}'>{d.category_type}</option>"
        for d in DocCategory.objects.filter(is_active=True)
    )


    return JsonResponse({
        "ok": True,
        "mail_id": mail.id,
        "from": mail.sender_mail,
        "date": mail.rec_dat.strftime("%Y-%m-%d %H:%M"),
        "group": getattr(mail, "group_name", ""),
        "attachments": attachments,
        "clients_html": clients_html,
        "categories_html": categories_html,
        "doc_types_html": doc_types_html,
    })
